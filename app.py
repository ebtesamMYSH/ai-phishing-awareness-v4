# =============================================================
# AI Phishing Awareness Training Tool
# -------------------------------------------------------------
# Project   : Study 3 - AI Tutor-Based Phishing Awareness
# Purpose   : Bilingual (Arabic/English) web platform for
#             phishing awareness training and assessment
#             designed for Saudi healthcare employees.
# Tech Stack: Python 3.9, Streamlit, Multi-Provider AI (Groq/Claude/OpenAI/Gemini)
# AI Models : Groq LLaMA 3.3-70b | Claude claude-sonnet-4-6 | GPT-4o | Gemini 1.5 Pro
# Admin     : Hidden Admin Panel at /?admin=true (password protected)
#             Compare 4 AI providers with 8-metric scoring system.
# -------------------------------------------------------------
# App Flow:
#   HOME -> LEARNING (6 AI-generated phishing examples)
#        -> COMPLETE -> ASSESSMENT (10 questions)
#        -> RESULTS -> PERFORMANCE REPORT
# =============================================================
#
# EN — QUICK CODE MAP (where to look for what):
#   • ROLE_MAP                  → job-role display names → (description, context, role_type)
#   • PHISHING_SCENARIOS        → legacy/unused list, kept for reference only
#   • FORCED_SCENARIOS          → the actual scenario pool per role_type
#                                 (admin/clinical/it have 10 each, "other" has 6
#                                 tied 1:1 to OTHER_JOB_PROFILES)
#   • OTHER_JOB_PROFILES        → the 6 sub-job profiles used when role="Other"
#                                 (now also carries name_en/name_ar for greeting binding)
#   • RECIPIENT_POOLS           → fixed (name, email) pairs per role_type, used to
#                                 keep the email greeting and the "to" address consistent
#   • get_session_random_order  → shared helper: shuffles indices once per session
#                                 so the same "example slot" doesn't always get the
#                                 same scenario/recipient across different sessions
#   • build_prompt               → builds the prompt for the LEARNING phase (6 examples)
#   • build_assess_prompt        → builds the prompt for the ASSESSMENT phase (10 questions)
#   • call_ai / call_groq        → sends the prompt to whichever provider is active
#                                  (groq/anthropic/openai/gemini) and normalizes the reply
#   • parse_json_response        → robust JSON parsing with repair fallbacks
#   • page_results                → renders "مراجعة الإجابات" (review answers) — has the
#                                  <bdi> bidi-isolation fix for mixed Arabic/English text
#   • load_persistent_provider /
#     save_persistent_provider /
#     set_active_provider        → keeps the admin's chosen AI provider saved to disk
#                                  (provider_config.json) so it survives logout/new sessions
#
# AR — خريطة سريعة للكود (وين تدورين على كل شي):
#   • ROLE_MAP                  → أسماء الأدوار الوظيفية ← (الوصف، السياق، نوع الدور)
#   • PHISHING_SCENARIOS        → قائمة قديمة غير مستخدمة، باقية للمرجعية فقط
#   • FORCED_SCENARIOS          → مسبح السيناريوهات الفعلي لكل دور
#                                 (إداري/سريري/تقني فيها 10 لكل واحد، و"Other" فيها 6
#                                 مرتبطة 1:1 مع OTHER_JOB_PROFILES)
#   • OTHER_JOB_PROFILES        → الـ6 بروفايلات الفرعية المستخدمة لما الدور = "Other"
#                                 (الآن فيها كمان name_en/name_ar لربط اسم التحية)
#   • RECIPIENT_POOLS           → أزواج (اسم، بريد) ثابتة لكل دور، نستخدمها حتى تتطابق
#                                 التحية بالبريد مع عنوان "to" دايمًا
#   • get_session_random_order  → دالة مشتركة: تخلط الفهارس مرة واحدة لكل جلسة، حتى
#                                 "خانة المثال" نفسها ما تطلع لها نفس السيناريو/المستلم
#                                 بكل الجلسات
#   • build_prompt               → يبني تعليمات مرحلة التعلم (الـ6 أمثلة)
#   • build_assess_prompt        → يبني تعليمات مرحلة الاختبار (الـ10 أسئلة)
#   • call_ai / call_groq        → يرسل التعليمات لأي بيئة نشطة حاليًا
#                                  (groq/anthropic/openai/gemini) ويوحّد شكل الرد
#   • parse_json_response        → تحليل JSON قوي مع محاولات تصحيح تلقائية
#   • page_results                → يعرض صفحة "مراجعة الإجابات" — فيها إصلاح اتجاه
#                                  النص (<bdi>) للنصوص المختلطة عربي/إنجليزي
#   • load_persistent_provider /
#     save_persistent_provider /
#     set_active_provider        → يحفظ اختيار الأدمن لمزوّد الذكاء الاصطناعي على القرص
#                                  (provider_config.json) حتى يفضل بعد تسجيل خروج/جلسة جديدة
# =============================================================

import streamlit as st
import json
import requests
import os
import re
import html as html_lib
import random
import urllib.parse
import ast

st.set_page_config(
    page_title="AI Phishing Awareness",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="auto"
)

# =============================================================
# Persistent AI Provider Config (survives logout / new session)
# -------------------------------------------------------------
# ai_provider used to live only in st.session_state, which is
# reset on every logout/login or new session. We now store the
# admin's chosen provider in a small JSON file on disk so it
# stays the single source of truth across all sessions.
# =============================================================
_PROVIDER_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "provider_config.json")
_VALID_PROVIDERS = {"groq", "anthropic", "openai", "gemini"}

def load_persistent_provider(default="openai"):
    try:
        with open(_PROVIDER_CONFIG_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        pk = data.get("ai_provider", default)
        if pk in _VALID_PROVIDERS:
            return pk
    except Exception:
        pass
    return default

def save_persistent_provider(pk):
    if pk not in _VALID_PROVIDERS:
        return
    try:
        with open(_PROVIDER_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"ai_provider": pk}, f)
    except Exception:
        # If the filesystem is read-only (some hosting setups), fail silently
        # and keep relying on session_state for the current session only.
        pass

def set_active_provider(pk):
    """Single entrypoint to change the active AI provider.
    Updates session_state AND persists to disk so it survives
    logout/login and new sessions/devices."""
    st.session_state["ai_provider"] = pk
    save_persistent_provider(pk)

# =============================================================
# PERSISTENT RESEARCH DATA — survives refresh/new sessions
# -------------------------------------------------------------
# Two JSON files on disk next to the app:
#  - runs.json     : one record per FULL CYCLE (6 learning + 10
#                     assessment) the researcher rated, holistically.
#  - metrics.json   : raw auto-tracked API call stats (speed, JSON
#                     success, errors, uniqueness) — accumulated
#                     across every call regardless of session.
# Both are best-effort: if the filesystem is read-only on a given
# host, writes fail silently and behaviour falls back to the old
# session-only behaviour (no crash either way).
# =============================================================
_RUNS_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "runs.json")
_METRICS_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "metrics.json")

def load_runs():
    local = []
    try:
        with open(_RUNS_FILE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            local = data
    except Exception:
        local = []
    try:
        return merge_local_and_gsheet_runs(local)
    except Exception:
        return local

def save_run(record):
    """Append one holistic run-rating record and persist to disk."""
    runs = load_runs()
    runs.append(record)
    try:
        with open(_RUNS_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(runs, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    try:
        push_run_to_gsheet(record)
    except Exception:
        pass
    return runs

def delete_all_runs():
    try:
        with open(_RUNS_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump([], f)
    except Exception:
        pass

def load_metrics_file():
    try:
        with open(_METRICS_FILE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}

def save_metrics_file(metrics_dict):
    try:
        with open(_METRICS_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(metrics_dict, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# =============================================================
# GOOGLE SHEETS SYNC — durable copy independent of this app's
# container/host, so research data survives redeploys/reboots.
# -------------------------------------------------------------
# Fully best-effort and optional: if gspread isn't installed, or
# st.secrets["gcp_service_account"] / st.secrets["GSHEET_ID"]
# aren't configured yet, every function below quietly no-ops.
# Nothing here can ever crash the app or block local saving —
# the local JSON files remain the primary, always-on storage;
# this is purely an extra durable copy on top.
#
# ONE-TIME SETUP (done outside this code, in Google Cloud /
# Google Sheets):
#   1. Google Cloud Console → create a Service Account → create
#      a JSON key for it → copy its "client_email".
#   2. Create a new Google Sheet (any name) under the
#      researcher's own Google account.
#   3. Share that Sheet with the service account's client_email,
#      giving it "Editor" access.
#   4. In Streamlit Cloud → App settings → Secrets, add:
#        GSHEET_ID = "<the sheet's ID from its URL>"
#        [gcp_service_account]
#        type = "service_account"
#        ... (every field from the downloaded JSON key, as TOML)
#   5. Add "gspread" and "google-auth" to requirements.txt.
# That's it — every future "Save Ratings" / metrics update will
# also land in the Sheet automatically, with three tabs:
#   - "Cycle Ratings" : one row per manually-rated cycle
#   - "Auto Metrics"  : one row per periodic metrics snapshot
# =============================================================
_GSHEET_CLIENT_CACHE = {"client": None, "tried": False}

def _get_gsheet_client():
    """Lazily build (and cache) an authorized gspread client from
    st.secrets. Returns None — silently — if the library isn't
    installed or secrets aren't configured, so callers never need
    to special-case "not set up yet" themselves."""
    if _GSHEET_CLIENT_CACHE["tried"]:
        return _GSHEET_CLIENT_CACHE["client"]
    _GSHEET_CLIENT_CACHE["tried"] = True
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_dict = dict(st.secrets["gcp_service_account"])
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        _GSHEET_CLIENT_CACHE["client"] = client
    except Exception:
        _GSHEET_CLIENT_CACHE["client"] = None
    return _GSHEET_CLIENT_CACHE["client"]

def _get_gsheet_id():
    try:
        sid = st.secrets.get("GSHEET_ID")
        return sid.strip() if isinstance(sid, str) and sid.strip() else None
    except Exception:
        return None

def gsheet_setup_status():
    """Human-readable status for the admin panel — never raises."""
    try:
        import gspread  # noqa: F401
    except Exception:
        return False, "مكتبة gspread غير مثبتة (أضيفيها لـ requirements.txt)" if st.session_state.get("language") == "Arabic" else "gspread library not installed (add it to requirements.txt)"
    if not _get_gsheet_id():
        return False, "GSHEET_ID غير موجود بالـ Secrets" if st.session_state.get("language") == "Arabic" else "GSHEET_ID missing from Secrets"
    try:
        has_creds = "gcp_service_account" in st.secrets
    except Exception:
        has_creds = False
    if not has_creds:
        return False, "بيانات اعتماد gcp_service_account غير موجودة بالـ Secrets" if st.session_state.get("language") == "Arabic" else "gcp_service_account credentials missing from Secrets"
    client = _get_gsheet_client()
    if client is None:
        return False, "فشل الاتصال — تأكدي من صحة بيانات الاعتماد ومشاركة الشيت مع الحساب الآلي" if st.session_state.get("language") == "Arabic" else "Connection failed — check credentials and that the Sheet is shared with the service account"
    return True, "متصل وشغال ✅" if st.session_state.get("language") == "Arabic" else "Connected and working ✅"

def _get_or_create_worksheet(sheet, tab_name, headers):
    try:
        ws = sheet.worksheet(tab_name)
    except Exception:
        ws = sheet.add_worksheet(title=tab_name, rows=1000, cols=max(10, len(headers)))
        ws.append_row(headers)
        return ws
    try:
        existing_header = ws.row_values(1)
        if not existing_header:
            ws.append_row(headers)
    except Exception:
        pass
    return ws

def push_run_to_gsheet(record):
    """Append one manually-rated cycle to the 'Cycle Ratings' tab.
    Best-effort: any failure (offline, not configured, quota, etc.)
    is swallowed so the researcher's local save never fails because
    of this extra step."""
    client = _get_gsheet_client()
    sheet_id = _get_gsheet_id()
    if not client or not sheet_id:
        return
    try:
        sheet = client.open_by_key(sheet_id)
        headers = ["timestamp", "provider", "language", "overall",
                   "auto_difficulty", "auto_arabic", "auto_quality", "auto_medical",
                   "n_auto_emails", "avg_speed", "json_rate", "error_rate",
                   "diversity", "note"]
        ws = _get_or_create_worksheet(sheet, "Cycle Ratings", headers)
        ws.append_row([record.get(h, "") for h in headers], value_input_option="USER_ENTERED")
    except Exception:
        pass

def push_metrics_snapshot_to_gsheet(provider, m):
    """Append one timestamped snapshot row per provider to the
    'Auto Metrics' tab, using the SAME raw shape _record_metric keeps
    in st.session_state['metrics'][provider] (speed list, json_ok/
    json_fail counts, errors, calls, hashes) — summarized here into
    the same percentages shown on the Score Card."""
    client = _get_gsheet_client()
    sheet_id = _get_gsheet_id()
    if not client or not sheet_id:
        return
    try:
        sheet = client.open_by_key(sheet_id)
        headers = ["timestamp", "provider", "calls", "avg_speed_s",
                   "json_success_rate_pct", "error_rate_pct", "unique_responses"]
        ws = _get_or_create_worksheet(sheet, "Auto Metrics", headers)
        m = m or {}
        speeds = m.get("speed") or []
        avg_speed = round(sum(speeds) / len(speeds), 2) if speeds else ""
        json_total = (m.get("json_ok", 0) + m.get("json_fail", 0))
        json_rate = round(100 * m.get("json_ok", 0) / json_total, 1) if json_total else ""
        calls = m.get("calls", 0)
        err_rate = round(100 * m.get("errors", 0) / calls, 1) if calls else ""
        unique = len(m.get("hashes") or [])
        ws.append_row([
            __import__("datetime").datetime.now().isoformat(timespec="seconds"),
            provider, calls, avg_speed, json_rate, err_rate, unique,
        ], value_input_option="USER_ENTERED")
    except Exception:
        pass

def pull_latest_auto_metrics_from_gsheet():
    """Read the 'Auto Metrics' tab and return the LATEST snapshot row per
    provider (since each save appends a new timestamped row rather than
    updating in place). Used as a durable fallback for the Score Card's
    live performance boxes (Avg Speed / JSON% / Error% / Unique Responses)
    when the local in-session 'metrics' dict is empty for that provider —
    e.g. right after the app container restarts and metrics.json is wiped,
    even though real generation activity already happened and was synced."""
    cache = st.session_state.get("_gsheet_auto_metrics_cache")
    now = __import__("time").time()
    if cache and (now - cache.get("ts", 0) < 60):
        return cache["latest"]
    client = _get_gsheet_client()
    sheet_id = _get_gsheet_id()
    latest = {}
    if not client or not sheet_id:
        return latest
    try:
        sheet = client.open_by_key(sheet_id)
        ws = sheet.worksheet("Auto Metrics")
        records = ws.get_all_records()
        for rec in records:
            p = str(rec.get("provider", "")).strip()
            if not p:
                continue
            latest[p] = rec  # later rows overwrite earlier ones -> last wins
    except Exception:
        latest = {}
    st.session_state["_gsheet_auto_metrics_cache"] = {"ts": now, "latest": latest}
    return latest

def pull_runs_from_gsheet():
    """Read every row back from the 'Cycle Ratings' tab and reshape it into
    the exact same record dicts load_runs()/save_run() use locally. This is
    what lets the Score Card / Manual Ratings counts / Excel export keep
    working even after the app's own container restarts and wipes
    runs.json — Google Sheets is the durable source of truth, the local
    file is just a fast first-read cache. Cached for 60s per session so
    we don't re-fetch from the Sheets API on every single rerun."""
    cache = st.session_state.get("_gsheet_runs_cache")
    now = __import__("time").time()
    if cache and (now - cache.get("ts", 0) < 60):
        return cache["rows"]
    client = _get_gsheet_client()
    sheet_id = _get_gsheet_id()
    if not client or not sheet_id:
        return []
    rows = []
    try:
        sheet = client.open_by_key(sheet_id)
        ws = sheet.worksheet("Cycle Ratings")
        records = ws.get_all_records()
        numeric_fields = ["overall", "auto_difficulty", "auto_arabic", "auto_quality",
                           "auto_medical", "n_auto_emails", "avg_speed", "json_rate",
                           "error_rate"]
        for rec in records:
            row = dict(rec)
            for f in numeric_fields:
                v = row.get(f)
                if v in ("", None):
                    row[f] = None
                else:
                    try:
                        row[f] = float(v) if f in ("avg_speed",) else int(float(v))
                    except (ValueError, TypeError):
                        row[f] = None
            rows.append(row)
    except Exception:
        rows = []
    st.session_state["_gsheet_runs_cache"] = {"ts": now, "rows": rows}
    return rows

def merge_local_and_gsheet_runs(local_runs):
    """Combine local runs.json entries with whatever's in the durable
    Google Sheet copy, deduplicating by (timestamp, provider, language)
    so the same cycle saved locally AND already synced to the sheet
    doesn't get double-counted."""
    sheet_runs = pull_runs_from_gsheet()
    if not sheet_runs:
        return local_runs
    seen = {(r.get("timestamp"), r.get("provider"), r.get("language")) for r in local_runs}
    merged = list(local_runs)
    for r in sheet_runs:
        key = (r.get("timestamp"), r.get("provider"), r.get("language"))
        if key not in seen:
            seen.add(key)
            merged.append(r)
    return merged

# =============================================================
# SYSTEMATIC ROTATION PLAN — 10 cycles, balanced role x difficulty
# -------------------------------------------------------------
# Purely informational for the researcher: she runs cycles 1-10
# manually following this table (role + difficulty to pick on the
# main app) so that, across 10 cycles, all 4 roles and 3 difficulty
# levels get reasonably balanced coverage instead of relying on
# pure chance with a small sample size.
# =============================================================
ROTATION_PLAN = [
    {"cycle": 1,  "role_en": "Clinical",   "role_ar": "سريري",  "difficulty": "easy",   "language": "English"},
    {"cycle": 2,  "role_en": "Admin",      "role_ar": "إداري",  "difficulty": "medium", "language": "English"},
    {"cycle": 3,  "role_en": "IT",         "role_ar": "تقني",   "difficulty": "hard",   "language": "English"},
    {"cycle": 4,  "role_en": "Other",      "role_ar": "أخرى",  "difficulty": "easy",   "language": "English"},
    {"cycle": 5,  "role_en": "Clinical",   "role_ar": "سريري",  "difficulty": "medium", "language": "English"},
    {"cycle": 6,  "role_en": "Admin",      "role_ar": "إداري",  "difficulty": "hard",   "language": "Arabic"},
    {"cycle": 7,  "role_en": "IT",         "role_ar": "تقني",   "difficulty": "easy",   "language": "Arabic"},
    {"cycle": 8,  "role_en": "Other",      "role_ar": "أخرى",  "difficulty": "medium", "language": "Arabic"},
    {"cycle": 9,  "role_en": "Clinical",   "role_ar": "سريري",  "difficulty": "hard",   "language": "Arabic"},
    {"cycle": 10, "role_en": "Admin",      "role_ar": "إداري",  "difficulty": "easy",   "language": "Arabic"},
]

# =============================================================
# AUTOMATIC EVALUATION — difficulty conformance, Arabic quality,
# general quality, medical relevance. All computed from the
# generated text itself, no human judgement involved.
# =============================================================
_DIRECT_PASSWORD_RE = re.compile(r"\b(password|otp|one[- ]time code|reply with your)\b|كلمة\s*(السر|المرور)|الرمز\s*المؤقت", re.I)
_DIRECT_THREAT_RE    = re.compile(r"\b(suspend(ed)?|terminat(e|ed|ion)|legal action|account.*delet|deactivat)\b|تعليق|إيقاف|إنهاء|حذف\s*الحساب", re.I)
_IMMEDIATE_URGENCY_RE = re.compile(r"\b(immediately|right now|within\s*(1|2|3)\s*hour|asap)\b|فورًا|الآن|خلال\s*ساعت", re.I)
_WINDOW_URGENCY_RE    = re.compile(r"\b(24|48|72)\s*hours?\b|٢٤|٤٨|٧٢\s*ساعة", re.I)
_GENERIC_GREETING_RE  = re.compile(r"^(dear (staff|team|healthcare professional|doctor|valued)|dear sir|عزيزي الموظف|إلى من يهمه)", re.I)
_PERSONAL_GREETING_RE = re.compile(r"dear dr\.?\s+[a-z\u0600-\u06ff]+\s+[a-z\u0600-\u06ff]+|عزيزي\s+د\.?\s*[\u0600-\u06ff]+\s+[\u0600-\u06ff]+", re.I)
_MEDICAL_KEYWORDS = [
    "patient","hospital","clinical","emr","medical","pharmacy","lab","radiology",
    "doctor","dr.","nurse","ministry of health","moh","healthcare","health record",
    "مريض","المستشفى","سريري","نظام طبي","صيدلية","مخبر","أشعة","طبيب","ممرض",
    "وزارة الصحة","صحي","سجل صحي","طبية","عيادة",
]
_PLACEHOLDER_LEFTOVER_RE = re.compile(r"\[QR(?:\s*Code)?\s*:?|suspicious_link\s*:|suspicious_text\s*:", re.I)

def check_difficulty_conformance(result, difficulty, is_phishing=True):
    """Score 0-100 against the documented progressive-difficulty plan.

    Easy: 4-5 obvious indicators, visible URL, generic greeting, direct request.
    Medium: 3-4 moderately subtle indicators, plausible look-alike workflow.
    Hard: 1-2 subtle indicators, polished/personalised content and near-official cues.
    """
    if not isinstance(result, dict):
        return None
    body = str(result.get("body") or "")
    subject = str(result.get("subject") or "")
    frm = str(result.get("from") or "")
    attachment = str(result.get("attachment") or "").strip()
    link = str(result.get("suspicious_link") or "").strip()
    indicators = result.get("indicators", []) if isinstance(result.get("indicators"), list) else []
    text = f"{subject} {body} {frm} {link}"

    domain_match = re.search(r"@([\w.-]+)>?", frm)
    domain = (domain_match.group(1) if domain_match else "").lower()
    domain_obvious = any(w in domain for w in ADVANCED_BANNED_DOMAIN_WORDS)
    has_qr = bool(re.search(r"\[\s*QR\s*:", body, re.I))
    has_raw_url = bool(re.search(r"https?://", body, re.I))
    has_button_marker = bool(re.search(r"\[[^\]]+\]\(https?://", body, re.I))

    checks = []
    if not is_phishing:
        checks.extend([
            not _DIRECT_PASSWORD_RE.search(text),
            not _DIRECT_THREAT_RE.search(text),
            not domain_obvious,
        ])
        return round(sum(checks) / len(checks) * 100)

    if difficulty == "easy":
        checks.extend([
            bool(_GENERIC_GREETING_RE.search(body.strip())) or _has_generic_greeting(body),
            domain_obvious,
            bool(_DIRECT_PASSWORD_RE.search(text)),
            bool(_DIRECT_THREAT_RE.search(text) or _IMMEDIATE_URGENCY_RE.search(text)),
            has_raw_url and not has_button_marker,
            not attachment,
            not has_qr,
            4 <= len(indicators) <= 5,
        ])
    elif difficulty == "medium":
        checks.extend([
            not _has_generic_greeting(body),
            not _DIRECT_PASSWORD_RE.search(text),
            not _DIRECT_THREAT_RE.search(text),
            bool(_WINDOW_URGENCY_RE.search(text)) or not _IMMEDIATE_URGENCY_RE.search(text),
            not has_qr,
            3 <= len(indicators) <= 4,
            bool(link or attachment),
        ])
    else:
        checks.extend([
            bool(_PERSONAL_GREETING_RE.search(body.strip())) or not _has_generic_greeting(body),
            not domain_obvious,
            not _DIRECT_PASSWORD_RE.search(text),
            not _DIRECT_THREAT_RE.search(text),
            not _IMMEDIATE_URGENCY_RE.search(text),
            1 <= len(indicators) <= 2,
            bool(attachment or has_qr or has_button_marker or link),
        ])

    caps_words = re.findall(r"\b[A-Z]{4,}\b", body)
    excl_count = body.count("!")
    checks.append((len(caps_words) + excl_count) >= 1 if difficulty == "easy" else (len(caps_words) + excl_count) == 0)
    return round(sum(checks) / len(checks) * 100) if checks else None

def check_arabic_quality(result, is_ar):
    """Score 0-100: cheap proxy for Arabic text cleanliness — leftover
    Latin words, broken/garbled characters, obviously truncated text."""
    if not is_ar or not isinstance(result, dict):
        return None
    body = str((result.get("body") or ""))
    if not body.strip():
        return 0
    words = body.split()
    if not words:
        return 0
    latin_words = [w for w in words if re.fullmatch(r"[A-Za-z]+", w)]
    latin_ratio = len(latin_words) / len(words)
    score = 100
    score -= min(60, int(latin_ratio * 200))            # too many stray English words
    if re.search(r"\ufffd|\\u[0-9a-fA-F]{4}", body):     # mojibake / unescaped unicode
        score -= 30
    if not re.search(r"[.!؟،]\s*$", body.strip()):       # doesn't end cleanly
        score -= 10
    if re.search(r"(\b\w+\b)(\s+\1){2,}", body):          # same word repeated 3+ times in a row
        score -= 20
    return max(0, min(100, score))

def check_general_quality(result):
    """Score 0-100: structural completeness proxy — required fields
    present, reasonable length, enough indicators, no leftover template
    artifacts that should have been rendered (e.g. raw '[QR:' text)."""
    if not isinstance(result, dict):
        return 0
    score = 100
    for field in ["from", "subject", "body"]:
        if not str(result.get(field, "")).strip():
            score -= 25
    body = str((result.get("body") or ""))
    wc = len(body.split())
    if wc < 15:
        score -= 25
    elif wc > 220:
        score -= 10
    indicators = result.get("indicators", []) if isinstance(result.get("indicators"), list) else []
    if len(indicators) < 2:
        score -= 20
    if _PLACEHOLDER_LEFTOVER_RE.search(body):
        score -= 25
    return max(0, min(100, score))

def check_medical_relevance(result):
    """Score 0/100 (with partial credit) — does this email actually sit
    in a healthcare/hospital context, based on keyword presence."""
    if not isinstance(result, dict):
        return None
    text = " ".join(str(result.get(k, "")) for k in ["from", "subject", "body"]).lower()
    hits = sum(1 for kw in _MEDICAL_KEYWORDS if kw in text)
    if hits == 0:
        return 0
    if hits == 1:
        return 60
    return 100

# =============================================================
# PERSISTENT AUTO-EVAL LOG + PENDING-CYCLE BUCKET
# -------------------------------------------------------------
# auto_eval.json   : permanent raw log, one row per generated email,
#                    with its 4 automatic scores — used for long-term
#                    analysis / CSV export regardless of cycles.
# pending_cycle.json: per (provider, language) bucket of scores
#                    collected SINCE the last saved cycle rating, so
#                    that when the researcher finishes a full cycle
#                    (6 learning + 10 assessment) and rates it, we can
#                    snapshot the average automatic scores for THAT
#                    specific cycle into the run record.
# =============================================================
_AUTO_EVAL_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "auto_eval.json")
_PENDING_CYCLE_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pending_cycle.json")
_PENDING_PERF_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pending_perf.json")

def _load_json_list(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
    except Exception:
        pass
    return []

def _save_json_list(path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# =============================================================
# FULL EXCEL EXPORT — one workbook, several sheets:
#   - Summary: one row per (provider, language) with all 9 metrics
#   - One sheet PER PROVIDER (Groq/Claude/OpenAI/Gemini) with its own
#     10-cycle detailed breakdown — this is the part that keeps the
#     four providers clearly separated instead of mixed in one table.
#   - Raw_Email_Log: every individually auto-scored email (not just
#     cycle averages), useful for deeper statistical analysis.
#   - Rotation_Plan: the systematic 10-cycle role/difficulty/language plan.
# Built fully in-memory and offered via st.download_button, so once the
# researcher downloads it, the file lives on her own computer — totally
# independent of the app's server storage from that point on.
# =============================================================
def build_excel_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    runs = load_runs()
    auto_log = _load_json_list(_AUTO_EVAL_FILE_PATH)

    PROV_ORDER_X = ["groq", "anthropic", "openai", "gemini"]
    PROV_LABELS_X = {"groq": "Groq", "anthropic": "Claude", "openai": "OpenAI", "gemini": "Gemini"}

    HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def autosize(ws, ncols, width=16):
        for c in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(c)].width = width

    wb = Workbook()

    # ---- Sheet 1: Summary across all 4 providers x 2 languages ----
    ws = wb.active
    ws.title = "Summary"
    headers = ["Provider", "Language", "Cycles", "Avg Speed (s)", "JSON %", "Error %",
               "Difficulty %", "Arabic %", "Quality %", "Medical %", "Overall (avg/5)"]
    ws.append(headers)
    for p in PROV_ORDER_X:
        for lang in ["English", "Arabic"]:
            p_runs = [r for r in runs if r.get("provider") == p and r.get("language") == lang]
            def avgf(field):
                vals = [r.get(field) for r in p_runs if r.get(field) is not None]
                return round(sum(vals) / len(vals), 2) if vals else None
            ws.append([
                PROV_LABELS_X[p], lang, len(p_runs),
                avgf("avg_speed"), avgf("json_rate"), avgf("error_rate"),
                avgf("auto_difficulty"), avgf("auto_arabic"), avgf("auto_quality"),
                avgf("auto_medical"), avgf("overall"),
            ])
    style_header(ws, len(headers))
    autosize(ws, len(headers))

    # ---- One sheet per provider: its own 10-cycle breakdown ----
    cycle_headers = ["#", "Timestamp", "Language", "Avg Speed (s)", "JSON %", "Error %",
                      "Diversity", "Difficulty %", "Arabic %", "Quality %", "Medical %",
                      "Overall /5", "Note"]
    for p in PROV_ORDER_X:
        ws_p = wb.create_sheet(PROV_LABELS_X[p])
        ws_p.append(cycle_headers)
        p_runs_en = [r for r in runs if r.get("provider") == p and r.get("language") == "English"]
        p_runs_ar = [r for r in runs if r.get("provider") == p and r.get("language") == "Arabic"]
        for i, r in enumerate(p_runs_en + p_runs_ar, 1):
            ws_p.append([
                i, r.get("timestamp"), r.get("language"), r.get("avg_speed"),
                r.get("json_rate"), r.get("error_rate"), r.get("diversity"),
                r.get("auto_difficulty"), r.get("auto_arabic"), r.get("auto_quality"),
                r.get("auto_medical"), r.get("overall"), r.get("note"),
            ])
        style_header(ws_p, len(cycle_headers))
        autosize(ws_p, len(cycle_headers))

    # ---- Raw per-email auto-evaluation log (all providers together, filterable) ----
    ws_raw = wb.create_sheet("Raw_Email_Log")
    raw_headers = ["Timestamp", "Provider", "Language", "Difficulty Level",
                   "Difficulty Score %", "Arabic Score %", "Quality Score %", "Medical Score %"]
    ws_raw.append(raw_headers)
    for rec in auto_log:
        ws_raw.append([
            rec.get("timestamp"), PROV_LABELS_X.get(rec.get("provider"), rec.get("provider")),
            rec.get("language"), rec.get("difficulty"),
            rec.get("difficulty_score"), rec.get("arabic_score"),
            rec.get("quality_score"), rec.get("medical_score"),
        ])
    style_header(ws_raw, len(raw_headers))
    autosize(ws_raw, len(raw_headers))

    # ---- Rotation plan reference ----
    ws_rot = wb.create_sheet("Rotation_Plan")
    rot_headers = ["Cycle #", "Role", "Difficulty", "Language"]
    ws_rot.append(rot_headers)
    for plan in ROTATION_PLAN:
        ws_rot.append([plan["cycle"], plan["role_en"], plan["difficulty"].capitalize(), plan["language"]])
    style_header(ws_rot, len(rot_headers))
    autosize(ws_rot, len(rot_headers), width=14)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _load_json_dict(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}

def _save_json_dict(path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _load_pending_buckets():
    return _load_json_dict(_PENDING_CYCLE_FILE_PATH)

def _save_pending_buckets(buckets):
    _save_json_dict(_PENDING_CYCLE_FILE_PATH, buckets)

def snapshot_and_clear_pending_perf(provider, language):
    """Average the per-cycle performance bucket (speed/JSON/error/
    diversity) collected since the last saved cycle, for THIS provider
    and language, then clear it so the next cycle starts fresh — mirrors
    snapshot_and_clear_pending_cycle() but for the 4 performance metrics."""
    buckets = _load_json_dict(_PENDING_PERF_FILE_PATH)
    key = f"{provider}__{language}"
    items = buckets.get(key, [])

    speeds = [it["speed"] for it in items if it.get("speed") is not None]
    json_ok = sum(1 for it in items if it.get("json_success") is True)
    json_fail = sum(1 for it in items if it.get("json_success") is False)
    errors = sum(1 for it in items if it.get("is_error"))
    calls = len(items)
    hashes = list({it["hash"] for it in items if it.get("hash")})

    snap = {
        "n_calls": calls,
        "avg_speed": round(sum(speeds)/len(speeds), 2) if speeds else None,
        "json_rate": round(json_ok/(json_ok+json_fail)*100) if (json_ok+json_fail) > 0 else None,
        "error_rate": round(errors/calls*100) if calls > 0 else None,
        "diversity": f"{len(hashes)}/{calls}" if calls > 0 else None,
    }
    if key in buckets:
        del buckets[key]
        _save_json_dict(_PENDING_PERF_FILE_PATH, buckets)
    return snap

def evaluate_and_log_auto_scores(result, difficulty, language, is_phishing=True):
    """Called right after a learning/assessment email is generated.
    Computes the 4 automatic scores and logs them both permanently
    (auto_eval.json) and into the current pending-cycle bucket so the
    next saved manual rating can snapshot this cycle's averages."""
    if not isinstance(result, dict) or "error" in result:
        return
    provider = st.session_state.get("ai_provider", "groq")
    is_ar = (language == "Arabic")
    rec = {
        "timestamp": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "provider": provider,
        "language": language,
        "difficulty": difficulty,
        "difficulty_score": check_difficulty_conformance(result, difficulty, is_phishing),
        "arabic_score": check_arabic_quality(result, is_ar),
        "quality_score": check_general_quality(result),
        "medical_score": check_medical_relevance(result),
    }
    log = _load_json_list(_AUTO_EVAL_FILE_PATH)
    log.append(rec)
    _save_json_list(_AUTO_EVAL_FILE_PATH, log)

    buckets = _load_pending_buckets()
    key = f"{provider}__{language}"
    buckets.setdefault(key, []).append(rec)
    _save_pending_buckets(buckets)

def snapshot_and_clear_pending_cycle(provider, language):
    """Average the pending bucket for this provider+language into one
    snapshot (used when the researcher saves a holistic cycle rating),
    then clear that bucket so the next cycle starts fresh."""
    buckets = _load_pending_buckets()
    key = f"{provider}__{language}"
    items = buckets.get(key, [])

    def _avg(field):
        vals = [it[field] for it in items if it.get(field) is not None]
        return round(sum(vals)/len(vals), 1) if vals else None

    snapshot = {
        "n_emails": len(items),
        "difficulty_score": _avg("difficulty_score"),
        "arabic_score": _avg("arabic_score"),
        "quality_score": _avg("quality_score"),
        "medical_score": _avg("medical_score"),
    }
    if key in buckets:
        del buckets[key]
        _save_pending_buckets(buckets)
    return snapshot

for k, v in [("language","English"),("page","home"),("role",""),
              ("example_index",0),("emails",{}),("difficulty","medium"),
              ("user_name",""),("user_email",""),
              ("ai_provider", load_persistent_provider("openai")),
              ("admin_authenticated",False),
              ("metrics", load_metrics_file()),  # {provider: {speed:[], json_ok:int, json_fail:int, errors:int, calls:int, hashes:[]}} — loaded from disk so it survives refresh
              ("manual_ratings",{}),  # legacy in-session structure, kept for backward compatibility
             ]:
    if k not in st.session_state:
        st.session_state[k] = v

_nav = (st.query_params.get("nav") or "")
if _nav in ("login", "register"):
    st.session_state["login_mode"] = _nav
    st.session_state["page"] = "login"
    _lang = (st.query_params.get("lang") or "")
    if _lang in ("Arabic", "English"):
        st.session_state["language"] = _lang
    st.query_params.clear()

def set_language(lang):
    st.session_state["language"] = lang
    st.session_state["lang_explicitly_chosen"] = True

def t(en, ar):
    return ar if st.session_state["language"] == "Arabic" else en

def _safe_error_text(err, language):
    """Never show raw API/debug dicts to trainees. Always return a short,
    human-readable sentence, and stash whatever technical detail we have in
    the hidden debug log (visible only to the admin panel / developer)."""
    is_ar = (language == "Arabic")
    try:
        _store_debug("ui_display", err)
    except Exception:
        pass
    msg = err.get("message") if isinstance(err, dict) else str(err)
    msg = str(msg or "")
    # If, despite everything, a low-level/raw message slipped through (very
    # long, or looks like a Python dict/JSON dump), replace it with a clean
    # generic sentence instead of exposing internals.
    looks_raw = len(msg) > 450 or msg.strip().startswith("{") or "Parsed keys/values" in msg
    if looks_raw or not msg.strip():
        return ("تعذّر توليد هذا المحتوى حالياً. يرجى الضغط على (حاول مرة أخرى)."
                if is_ar else
                "This content couldn't be generated right now. Please tap Try Again.")
    return msg


def clean_foreign_only(text):
    if not text: return text
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'[\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff\u3400-\u4dbf\uac00-\ud7af\u1100-\u11ff]', '', text)
    text = re.sub(r'[\u0400-\u04ff]', '', text)
    text = re.sub(r'[\u0100-\u017f]', '', text)
    text = re.sub(r'[\u1ea0-\u1ef9]', '', text)
    text = re.sub(r'[\u0900-\u097f]', '', text)
    text = re.sub(r'[\u0e00-\u0e7f]', '', text)
    text = re.sub(r'[\u10a0-\u10ff\u0530-\u058f\u05d0-\u05ff]', '', text)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    text = re.sub(r'  +', ' ', text).strip()
    return text

_ALLOWED_LATIN_RE = re.compile(
    r'^(https?://[^\s]+|[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}|\.(?:pdf|xlsx|docx|txt|csv|zip|exe)|[0-9]+)$',
    re.IGNORECASE
)

def remove_foreign_latin_words(text):
    if not text: return text
    arabic_chars = len(re.findall(r'[\u0600-\u06ff]', text))
    total_chars  = len(re.sub(r'\s', '', text))
    if total_chars == 0 or arabic_chars / total_chars < 0.25:
        return text
    def keep_token(tok):
        if _ALLOWED_LATIN_RE.match(tok): return tok
        if re.search(r'[\u0600-\u06ff]', tok): return tok
        if re.match(r'^[\u060c\u061b\u061f،؛؟!.,;:\-\u2013\u2014()\[\]{}\'"]+$', tok): return tok
        if re.match(r'^[a-zA-Z\u00c0-\u024f]+$', tok): return ''
        return tok
    tokens  = re.split(r'(\s+)', text)
    cleaned = ''.join(keep_token(t) for t in tokens)
    cleaned = re.sub(r'[a-zA-Z]{1,}[-_](?=[\u0600-\u06ff])', '', cleaned)
    cleaned = re.sub(r'(?<=[\u0600-\u06ff])[-_]?[a-zA-Z]{1,}', '', cleaned)
    cleaned = re.sub(r'[a-zA-Z]{1,}(?=[\u0600-\u06ff])', '', cleaned)
    cleaned = re.sub(r'  +', ' ', cleaned).strip()
    cleaned = re.sub(r'\s([،؛،,.;:])', r'\1', cleaned)
    return cleaned

def clean_email_field(addr):
    if not addr: return addr
    addr = re.sub(r'[\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff\u0400-\u04ff\u0100-\u017f]', '', addr)
    addr = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', addr)
    return addr.strip()

def extract_to_email(to_val):
    if not to_val: return 'employee@hospital.org'
    m = re.search(r'[\w.+-]+@[\w.-]+\.[a-zA-Z]{2,}', to_val)
    return m.group(0) if m else 'employee@hospital.org'

def fix_json_newlines(s):
    result, in_string, i = [], False, 0
    while i < len(s):
        c = s[i]
        if c == '"' and (i == 0 or s[i-1] != '\\'):
            in_string = not in_string
        if in_string and c == '\n':   result.append('\\n')
        elif in_string and c == '\r': result.append('\\r')
        elif in_string and c == '\t': result.append('\\t')
        else:                         result.append(c)
        i += 1
    return ''.join(result)

ROLE_MAP = {
    "سريري": (
        "ممرض أو طبيب يعمل في مستشفى",
        "السجلات الطبية وجداول العمل السريرية والأنظمة الطبية وبيانات المرضى",
        "clinical"
    ),
    "إداري / إدارة": (
        "موظف إداري في مستشفى (سكرتارية طبية، استقبال، إدارة ملفات المرضى، التأمين الصحي، الفوترة الطبية)",
        "ملفات المرضى وحجز المواعيد والتأمين الصحي والفوترة الطبية وسكرتارية الأطباء وإدارة المستشفى والمشتريات الطبية",
        "admin"
    ),
    "تقنية المعلومات / المعلوماتية": (
        "متخصص تقنية معلومات في مستشفى",
        "الشبكة والخوادم وتحديثات البرامج والدعم التقني والأمن السيبراني",
        "it"
    ),
    "Clinical": (
        "a nurse or doctor in a hospital",
        "patient records, EMR systems, clinical schedules, medical data",
        "clinical"
    ),
    "Admin / Management": (
        "a healthcare administrative staff (medical secretary, receptionist, patient records manager, insurance coordinator, billing specialist)",
        "patient files, appointment scheduling, medical insurance, hospital billing, doctor's secretary, hospital management, medical procurement",
        "admin"
    ),
    "IT / Informatics": (
        "an IT specialist in a hospital",
        "VPN, network, servers, software updates, IT helpdesk, security systems",
        "it"
    ),
    "Other": (
        "a general hospital employee in Saudi Arabia",
        "any hospital department — clinical, administrative, or technical",
        "other"
    ),
    "أخرى": (
        "موظف عام في مستشفى سعودي",
        "أي قسم في المستشفى — سريري أو إداري أو تقني",
        "other"
    ),
}

EN_NAMES = {
    "clinical": [
        "dr.sarah.almutairi@hospital.org",
        "dr.ahmed.alotaibi@hospital.org",
        "n.noura.alshamri@hospital.org",
        "dr.fahad.aldosari@hospital.org",
        "n.mona.alharbi@hospital.org",
        "dr.khalid.alanazi@hospital.org",
    ],
    "admin": [
        "m.reem.alsabiei@hospital.org",
        "m.abdullah.alqahtani@hospital.org",
        "m.hind.alrashidi@hospital.org",
        "m.sultan.alghamdi@hospital.org",
        "m.dalal.alzahrani@hospital.org",
        "m.omar.albaqami@hospital.org",
    ],
    "it": [
        "t.mohammed.alshahri@hospital.org",
        "t.rania.almalki@hospital.org",
        "t.yusuf.aljuhani@hospital.org",
        "t.lama.alumari@hospital.org",
        "t.bandar.althubaiti@hospital.org",
        "t.nadia.alsalmi@hospital.org",
    ],
    "other": [
        "s.khalid.alharbi@hospital.org",
        "s.sara.alqahtani@hospital.org",
        "s.faisal.alzahrani@hospital.org",
        "s.nora.alotaibi@hospital.org",
        "s.ahmed.alshamri@hospital.org",
        "s.hessa.aldosari@hospital.org",
    ],
}





def get_recipient(role, index, language, phase="learn"):
    # EN: THIS is the real root cause of the recurring "Dear Nurse John"
    # name-mismatch bug. The model was correctly told in the prompt to use
    # a specific (name, email) pair — but AFTER generation, this function
    # used to silently overwrite the "to" field with a DIFFERENT, totally
    # unrelated email picked from EN_NAMES below. The greeting name inside
    # the body (bound to the prompt's pair) and the final "to" address
    # (overwritten here) ended up belonging to two different people.
    #
    # Fix: read from the SAME RECIPIENT_POOLS + the SAME session-randomized
    # order key that build_prompt/build_assess_prompt used to pick the pair
    # for this exact role_type + index + phase, so the override always
    # matches what the model was actually told to write in the greeting.
    #
    # AR: هذا هو السبب الحقيقي لخلل "Dear Nurse John" المتكرر. الموديل كان
    # يُطلب منه صراحة استخدام زوج (اسم، بريد) معيّن بالتعليمات — لكن بعد
    # التوليد، هذي الدالة كانت "تستبدل" حقل "to" بصمت بعنوان مختلف كليًا
    # من قائمة EN_NAMES تحت، فيصير اسم التحية بالنص (المرتبط بالزوج الأصلي)
    # وعنوان "to" النهائي (المُستبدل هنا) يخصان شخصين مختلفين تمامًا.
    #
    # الحل: نقرأ من نفس RECIPIENT_POOLS وبنفس مفتاح الترتيب العشوائي للجلسة
    # اللي استخدمته build_prompt/build_assess_prompt لاختيار الزوج لنفس
    # role_type + index + المرحلة، حتى الاستبدال هنا يطابق دائمًا ما طُلب
    # من الموديل فعليًا يكتبه بالتحية.
    role_info = ROLE_MAP.get(role, ROLE_MAP.get("Clinical"))
    _, _, role_type = role_info
    if role_type == "other":
        # EN: in practice, role_type "other" is handled entirely by
        # generate_other_email/generate_other_assess_email (static
        # templates) and never reaches this function. This is just a
        # safe fallback so we never crash if that ever changes.
        # AR: عمليًا دور "Other" تتم معالجته بالكامل بدوال القوالب الثابتة
        # المنفصلة ومايصل لهذي الدالة أبدًا. هذا احتياط آمن بس لو تغيّر هذا مستقبلًا.
        cached = st.session_state.get(f"_other_recipient_{index}")
        if cached:
            return cached
        legacy = EN_NAMES.get("other", EN_NAMES["clinical"])
        return legacy[index % len(legacy)]
    pool = RECIPIENT_POOLS.get(role_type)
    if not pool:
        legacy = EN_NAMES.get(role_type, EN_NAMES["clinical"])
        return legacy[index % len(legacy)]
    order_key = f"recipient_order_{role_type}" if phase == "learn" else f"assess_recipient_order_{role_type}"
    order = get_session_random_order(len(pool), order_key)
    return pool[order[index % len(order)]]["email"]



# =============================================================
# EN: SESSION-RANDOMIZED ORDER HELPER (variety fix)
# ---------------------------------------------------------------
# Problem we found during testing: "Example #3 of 6" was always
# mapped to the SAME scenario (e.g. "MOH protocol update") no
# matter how many times a user logged in again, because the code
# picked scenarios with `forced[index % len(forced)]` — a fixed,
# deterministic mapping. Across two different sessions, slot #3
# came out nearly identical.
#
# Fix: instead of a fixed mapping, we shuffle the list of indices
# ONCE per session and cache that shuffled order in session_state.
# Slot #3 now points to a DIFFERENT scenario each new session,
# while still guaranteeing no two of the 6 examples in the SAME
# session repeat each other (since it's a permutation, not a
# random pick-with-replacement).
#
# AR: دالة لعشوائية الترتيب لكل جلسة (إصلاح مشكلة التنوع)
# ---------------------------------------------------------------
# المشكلة اللي لاحظناها بالاختبار: "مثال 3 من 6" كان دايمًا يطلع
# له نفس السيناريو (مثل "تحديث بروتوكول وزارة الصحة") بغض النظر
# عن عدد مرات تسجيل الدخول، لأن الكود القديم كان يحدد السيناريو
# بمعادلة ثابتة `forced[index % len(forced)]` — يعني ربط دائم.
#
# الحل: نخلط ترتيب الفهارس مرة واحدة فقط لكل جلسة ونحفظه في
# session_state. الآن "مثال 3" يطلع له سيناريو مختلف كل جلسة جديدة،
# مع ضمان عدم تكرار أي سيناريو مرتين داخل نفس الجلسة (لأنه تبديل
# ترتيب "permutation"، مش اختيار عشوائي قد يتكرر).
# =============================================================
def get_session_random_order(pool_size, session_key):
    if session_key not in st.session_state:
        order = list(range(pool_size))
        random.shuffle(order)
        st.session_state[session_key] = order
    return st.session_state[session_key]

# =============================================================
# =============================================================
# FIX 7: FORCED SCENARIO DIVERSITY PER INDEX
# كل مثال له سيناريو محدد مسبقاً — يمنع التكرار نهائياً
# =============================================================

# EN: RECIPIENT_POOLS — fixed (name, email) pairs for the generic roles
# (clinical/admin/it), shared by both the learning-phase prompt
# (build_prompt) and the assessment prompt (build_assess_prompt). We pick
# ONE pair per example slot using a session-randomized order, then tell
# the model explicitly to use that exact name in the greeting and that
# exact email in "to". This is the fix for the "Dear Nurse John" vs a
# completely different "to" address bug found repeatedly during testing.
# AR: قوائم مستلمين ثابتة (اسم + بريد) للأدوار العامة (سريري/إداري/تقني)،
# تستخدمها دالتا التوليد (التعلم والاختبار) معًا. نختار زوج واحد لكل خانة
# مثال بترتيب عشوائي خاص بالجلسة، ثم نطلب من الموديل صراحة يستخدم هذا الاسم
# بالتحية وهذا البريد بحقل "to" — يحل خلل "Dear Nurse John" مع عنوان مختلف
# كليًا اللي تكرر بالاختبار.
RECIPIENT_POOLS = {
    "clinical": [
        {"en": "Dr. Sarah Almutairi",  "ar": "د. سارة المطيري",      "email": "dr.sarah.almutairi@hospital.org"},
        {"en": "Dr. Ahmed Alotaibi",   "ar": "د. أحمد العتيبي",      "email": "dr.ahmed.alotaibi@hospital.org"},
        {"en": "Nurse Fatima Alharbi", "ar": "الممرضة فاطمة الحربي", "email": "n.fatima.alharbi@hospital.org"},
        {"en": "Nurse Khalid Alqahtani","ar": "الممرض خالد القحطاني","email": "n.khalid.alqahtani@hospital.org"},
        {"en": "Dr. Noura Alshamri",   "ar": "د. نورة الشمري",       "email": "dr.noura.alshamri@hospital.org"},
        {"en": "Dr. Faisal Aldosari",  "ar": "د. فيصل الدوسري",      "email": "dr.faisal.aldosari@hospital.org"},
        {"en": "Nurse Maha Alsubaie",  "ar": "الممرضة مها السبيعي", "email": "n.maha.alsubaie@hospital.org"},
        {"en": "Dr. Omar Alharthy",    "ar": "د. عمر الحارثي",       "email": "dr.omar.alharthy@hospital.org"},
        {"en": "Nurse Reem Alzahrani", "ar": "الممرضة ريم الزهراني","email": "n.reem.alzahrani@hospital.org"},
        {"en": "Dr. Yousef Alghamdi",  "ar": "د. يوسف الغامدي",      "email": "dr.yousef.alghamdi@hospital.org"},
        {"en": "Pharmacist Lama Alqahtani", "ar": "الصيدلانية لمى القحطاني", "email": "ph.lama.alqahtani@hospital.org"},
        {"en": "Pharmacist Ziad Alharbi",   "ar": "الصيدلاني زياد الحربي",   "email": "ph.ziad.alharbi@hospital.org"},
        {"en": "Lab Technician Huda Alsalmi","ar": "فنية المختبر هدى السالمي","email": "lab.huda.alsalmi@hospital.org"},
        {"en": "Radiology Technician Nasser Aldosari","ar": "فني الأشعة ناصر الدوسري","email": "rad.nasser.aldosari@hospital.org"},
    ],
    "admin": [
        {"en": "Sultan Alghamdi",  "ar": "سلطان الغامدي",  "email": "m.sultan.alghamdi@hospital.org"},
        {"en": "Reem Alsabiei",    "ar": "ريم السبيعي",     "email": "m.reem.alsabiei@hospital.org"},
        {"en": "Nadia Alsalmi",    "ar": "نادية السالمي",   "email": "t.nadia.alsalmi@hospital.org"},
        {"en": "Bandar Althubaiti","ar": "بندر الثبيتي",    "email": "t.bandar.althubaiti@hospital.org"},
        {"en": "Hessa Alqahtani",  "ar": "حصة القحطاني",    "email": "m.hessa.alqahtani@hospital.org"},
        {"en": "Majed Alharbi",    "ar": "ماجد الحربي",     "email": "m.majed.alharbi@hospital.org"},
        {"en": "Lama Alshehri",    "ar": "لمى الشهري",      "email": "t.lama.alshehri@hospital.org"},
        {"en": "Turki Aldosari",   "ar": "تركي الدوسري",    "email": "m.turki.aldosari@hospital.org"},
        {"en": "Amal Alzahrani",   "ar": "أمل الزهراني",    "email": "t.amal.alzahrani@hospital.org"},
        {"en": "Faisal Alotaibi",  "ar": "فيصل العتيبي",    "email": "m.faisal.alotaibi@hospital.org"},
    ],
    "it": [
        {"en": "Bandar Althubaiti","ar": "بندر الثبيتي",    "email": "t.bandar.althubaiti@hospital.org"},
        {"en": "Nadia Alsalmi",    "ar": "نادية السالمي",   "email": "t.nadia.alsalmi@hospital.org"},
        {"en": "Mohammed Alshahri","ar": "محمد الشهري",     "email": "t.mohammed.alshahri@hospital.org"},
        {"en": "Rania Almalki",    "ar": "رانية المالكي",   "email": "t.rania.almalki@hospital.org"},
        {"en": "Khalid Alharbi",   "ar": "خالد الحربي",     "email": "t.khalid.alharbi@hospital.org"},
        {"en": "Sara Alqahtani",   "ar": "سارة القحطاني",   "email": "t.sara.alqahtani@hospital.org"},
        {"en": "Yazeed Aldosari",  "ar": "يزيد الدوسري",    "email": "t.yazeed.aldosari@hospital.org"},
        {"en": "Hanan Alzahrani",  "ar": "حنان الزهراني",   "email": "t.hanan.alzahrani@hospital.org"},
        {"en": "Abdullah Alotaibi","ar": "عبدالله العتيبي", "email": "t.abdullah.alotaibi@hospital.org"},
        {"en": "Lina Alsubaie",    "ar": "لينا السبيعي",    "email": "t.lina.alsubaie@hospital.org"},
    ],
}

# =============================================================
# EN: OPEN-ENDED SCENARIO GENERATION (replaces the fixed scenario
# pool for clinical/admin/it — "other" keeps its existing
# profile-paired scenarios since changing those risks breaking the
# 1:1 link with OTHER_JOB_PROFILES).
# ---------------------------------------------------------------
# Instead of picking one of N pre-written scenarios, we now give
# the model a broad CATEGORY (e.g. "VPN / remote network access")
# and explicitly tell it to INVENT a brand-new, specific, realistic
# scenario within that category — different from anything already
# generated this session. This makes the variety effectively
# unlimited instead of capped at a fixed list size, while the
# category + role context still keeps it realistic and on-topic.
#
# AR: توليد سيناريوهات مفتوح (بدل القائمة الثابتة لسريري/إداري/تقني
# — دور "Other" يبقى على سيناريوهاته المرتبطة بالبروفايلات لأن
# تغييرها يهدد الربط 1:1 مع OTHER_JOB_PROFILES).
# ---------------------------------------------------------------
# بدل اختيار سيناريو جاهز من قائمة محدودة، الآن نعطي الموديل فئة
# عامة (مثل "VPN / وصول شبكة عن بُعد") ونطلب منه صراحة يخترع
# سيناريو جديد ومحدد وواقعي ضمن هذي الفئة — مختلف عن أي شي تولّد
# هذي الجلسة. هذا يخلي التنوع فعليًا لا محدود بدل ما يكون مسقوف
# بحجم قائمة ثابتة، مع بقاء الواقعية بفضل الفئة وسياق الدور.
# =============================================================






# FIX 1: build_prompt — upgraded to llama-3.3-70b-versatile
# and enhanced difficulty rules with more detail
# =============================================================

def extract_domains_from_result(result):
    """Extract domains from generated email fields for session-level anti-repeat memory."""
    if not isinstance(result, dict):
        return []
    text = " ".join(str(result.get(k, "")) for k in ["from", "body", "suspicious_link"])
    domains = re.findall(r'(?:https?://)?(?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}', text)
    clean = []
    for d in domains:
        d = re.sub(r'^https?://', '', d).split('/')[0].strip().lower()
        if d and d not in clean and d not in {"hospital.org", "moh.gov.sa"}:
            clean.append(d)
    return clean[:5]



def _domain_root(domain):
    domain = (domain or "").lower()
    domain = re.sub(r'^https?://', '', domain).split('/')[0]
    parts = domain.split('.')
    return ".".join(parts[-2:]) if len(parts) >= 2 else domain

_OBVIOUS_ADVANCED_DOMAIN_WORDS = {
    "secure", "update", "verify", "verification", "login", "reset",
    "password", "urgent", "alert", "account", "emr"
}

def _contains_long_all_caps(s):
    return bool(re.search(r'\b[A-Z][A-Z\s]{18,}\b', s or ""))

def _has_generic_greeting(body):
    b = (body or "").lower()
    generic = [
        "dear staff", "dear team", "dear employee", "dear user",
        "عزيزي الموظف", "عزيزتي الموظفة", "عزيزي المستخدم", "عزيزي الفريق",
        "السادة الموظفين", "الموظف العزيز"
    ]
    return any(g in b for g in generic)

def _domain_has_obvious_advanced_word(domain):
    d = (domain or "").lower()
    return any(w in d for w in _OBVIOUS_ADVANCED_DOMAIN_WORDS)



# =============================================================
# STRICT ROLE + DIFFICULTY GUARDRAILS (patched)
# -------------------------------------------------------------
# These checks reject AI outputs that drift away from the selected
# job role (clinical/admin/IT) or violate the documented difficulty
# framework before the email is shown to the trainee.
# =============================================================









# =============================================================
# HEALTHCARE SCENARIO LIBRARY (300 Scenario Cards)
# -------------------------------------------------------------
# These are NOT email templates. They are compact content directions.
# The AI still writes the full email, analysis and assessment text via API,
# but it no longer invents the core idea from nothing. This prevents repeated
# "password/account" emails and keeps the content healthcare-relevant.
# =============================================================






# Dynamic content-shape engine. These are NOT fixed email templates.
# They are writing constraints that make the API generate different forms of email content each time.







# =============================================================
# UNBOUNDED LEARNING PROMPT
# No fixed templates. No fixed scenario pool. No example domains.
# =============================================================
def get_medium_presentation_mode(phase, index):
    """Return a session-stable Medium delivery style.

    Learning (6): 2 buttons, 3 visible links, 1 attachment/no-link.
    Assessment (10): 4 buttons, 4 visible links, 2 attachment/no-link.
    The order is shuffled once per session to avoid a predictable pattern.
    """
    phase_key = "learn" if phase == "learn" else "assess"
    size = 6 if phase_key == "learn" else 10
    key = f"medium_presentation_order_{phase_key}"
    modes = st.session_state.get(key)
    if not isinstance(modes, list) or len(modes) != size:
        modes = (["button"] * 2 + ["link"] * 3 + ["none"]) if size == 6 else (["button"] * 4 + ["link"] * 4 + ["none"] * 2)
        random.shuffle(modes)
        st.session_state[key] = modes
    try:
        return modes[int(index) % size]
    except Exception:
        return modes[0]


def get_medium_channel_instruction(mode, is_ar=False):
    if is_ar:
        return {
            "button": "صيغة العرض لهذه الرسالة: زر واحد فقط داخل موضع طبيعي في النص. ممنوع كتابة الرابط الخام في body، وممنوع إظهار زر ورابط معاً.",
            "link": "صيغة العرض لهذه الرسالة: رابط نصي واحد ظاهر فقط داخل موضع طبيعي في النص. ممنوع إنشاء زر أو تكرار الرابط.",
            "none": "صيغة العرض لهذه الرسالة: بدون زر وبدون رابط وبدون QR. استخدم مرفق PDF بسيط فقط أو اطلب التحقق من النظام المعتاد بشكل مستقل.",
        }.get(mode, "")
    return {
        "button": "Presentation for this email: use exactly ONE inline button. Do not show the raw URL in body and never combine a button with a visible link.",
        "link": "Presentation for this email: use exactly ONE visible text URL in a natural position. Do not create a button and do not repeat the URL.",
        "none": "Presentation for this email: use NO button, NO URL, and NO QR. Use only a simple PDF attachment or ask the recipient to verify through the usual internal system independently.",
    }.get(mode, "")



# =============================================================
# UNBOUNDED ASSESSMENT PROMPT
# Phishing and legitimate questions are generated dynamically.
# =============================================================


def _init_provider_metrics(provider):
    """Initialize metrics dict for a provider if not present"""
    if "metrics" not in st.session_state:
        st.session_state["metrics"] = {}
    if provider not in st.session_state["metrics"]:
        st.session_state["metrics"][provider] = {
            "speed": [],        # list of response times in seconds
            "json_ok": 0,       # successful JSON parses
            "json_fail": 0,     # failed JSON parses
            "errors": 0,        # API errors
            "calls": 0,         # total API calls
            "hashes": [],       # content hashes for diversity check
        }

def _record_metric(provider, speed_sec, json_success, content_hash=None, is_error=False):
    """Record a single API call metric, and persist it to disk so it
    survives refresh / new sessions (it used to live only in session_state).
    Also feeds the per-cycle pending-performance bucket so each saved
    cycle can carry its OWN speed/JSON/error/diversity snapshot, instead
    of only the all-time cumulative provider stats."""
    _init_provider_metrics(provider)
    m = st.session_state["metrics"][provider]
    m["calls"] += 1

    language = st.session_state.get("language", "English")
    perf_buckets = _load_json_dict(_PENDING_PERF_FILE_PATH)
    key = f"{provider}__{language}"
    perf_buckets.setdefault(key, []).append({
        "speed": None if is_error else round(speed_sec, 2),
        "json_success": None if is_error else bool(json_success),
        "is_error": is_error,
        "hash": None if is_error else content_hash,
    })
    _save_json_dict(_PENDING_PERF_FILE_PATH, perf_buckets)

    if is_error:
        m["errors"] += 1
        save_metrics_file(st.session_state["metrics"])
        try:
            push_metrics_snapshot_to_gsheet(provider, m)
        except Exception:
            pass
        return
    m["speed"].append(round(speed_sec, 2))
    if json_success:
        m["json_ok"] += 1
    else:
        m["json_fail"] += 1
    if content_hash and content_hash not in m["hashes"]:
        m["hashes"].append(content_hash)
    save_metrics_file(st.session_state["metrics"])
    try:
        push_metrics_snapshot_to_gsheet(provider, m)
    except Exception:
        pass

def call_ai(prompt, max_tokens=1600):
    import time
    provider = st.session_state.get("ai_provider", "groq")
    system_prompt = get_system_prompt()

    def get_secret(key):
        try:
            return st.secrets[key]
        except Exception:
            return os.environ.get(key, "")

    _init_provider_metrics(provider)
    start_time = time.time()

    try:
        if provider == "groq":
            resp = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={
                    "Content-Type":  "application/json",
                    "Authorization": f"Bearer {get_secret('GROQ_API_KEY')}"
                },
                json={
                    "model":       "llama-3.3-70b-versatile",
                    "max_tokens":  max_tokens,
                    "temperature": 0.85,
                    "messages":    [
                        {"role": "system", "content": system_prompt},
                        {"role": "user",   "content": prompt}
                    ]
                },
                timeout=45
            )
            data = resp.json()
            elapsed = time.time() - start_time
            _record_metric(provider, elapsed, "choices" in data, str(hash(str(data)))[:8])
            return data

        elif provider == "openai":
            resp = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers={
                    "Content-Type":  "application/json",
                    "Authorization": f"Bearer {get_secret('OPENAI_API_KEY')}"
                },
                json={
                    "model":       "gpt-4o",
                    "max_tokens":  max_tokens,
                    "temperature": 0.85,
                    "response_format": {"type": "json_object"},
                    "messages":    [
                        {"role": "system", "content": system_prompt},
                        {"role": "user",   "content": prompt}
                    ]
                },
                timeout=60
            )
            data = resp.json()
            elapsed = time.time() - start_time
            _record_metric(provider, elapsed, "choices" in data, str(hash(str(data)))[:8])
            return data

        elif provider == "anthropic":
            # Claude needs a bigger token budget than the other providers for
            # the same request: longer/more detailed answers (as observed)
            # plus any internal reasoning blocks both draw from max_tokens.
            # Without enough headroom, Arabic responses in particular were
            # getting cut off before any real text was produced.
            anthropic_max_tokens = max(max_tokens, 3500)
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "Content-Type":      "application/json",
                    "x-api-key":         get_secret("ANTHROPIC_API_KEY"),
                    "anthropic-version": "2023-06-01"
                },
                json={
                    "model":      "claude-sonnet-4-6",
                    "max_tokens": anthropic_max_tokens,
                    "system":     system_prompt,
                    "messages":   [
                        {"role": "user", "content": prompt}
                    ]
                },
                timeout=60
            )
            raw = resp.json()
            elapsed = time.time() - start_time
            if "content" in raw and len(raw["content"]) > 0:
                # Claude can return multiple content blocks (e.g. a "thinking"
                # block followed by the actual "text" block, especially for
                # longer/more complex responses such as Arabic generations).
                # Reading only content[0] silently returned an empty string
                # in those cases. We now join ALL text-type blocks instead.
                text = "".join(
                    (block.get("text") or "")
                    for block in raw["content"]
                    if block.get("type") == "text"
                )
                if not text.strip():
                    # No usable text block found at all — surface as an error
                    # instead of returning an empty string that later fails
                    # JSON parsing with a confusing "char 0" message.
                    _record_metric(provider, elapsed, False, is_error=True)
                    return {"error": {"message": f"Claude returned no text content (stop_reason={raw.get('stop_reason')}): {str(raw)[:300]}"}}
                _record_metric(provider, elapsed, True, str(hash(text))[:8])
                return {"choices": [{"message": {"content": text}}]}
            _record_metric(provider, elapsed, False, is_error=True)
            return {"error": {"message": str(raw)[:300]}}

        elif provider == "gemini":
            api_key = get_secret("GEMINI_API_KEY")
            # Gemini 2.5 Flash is a "thinking" model: by default it spends
            # part of maxOutputTokens on internal reasoning before writing
            # the visible answer. With a small token budget, thinking alone
            # could consume it all, leaving an empty text part (which is
            # exactly the "char 0" JSON parse error seen above). We disable
            # thinking (not needed for this task) and give a safe token
            # floor, mirroring the same fix already applied for Claude.
            gemini_max_tokens = max(max_tokens, 2400)
            resp = requests.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}",
                headers={"Content-Type": "application/json"},
                json={
                    "contents": [{"parts": [{"text": system_prompt + "\n\n" + prompt}]}],
                    "generationConfig": {
                        "maxOutputTokens": gemini_max_tokens,
                        "temperature":     0.85,
                        "responseMimeType": "application/json",
                        "thinkingConfig":  {"thinkingBudget": 0}
                    }
                },
                timeout=60
            )
            raw = resp.json()
            elapsed = time.time() - start_time
            try:
                parts = raw["candidates"][0]["content"]["parts"]
                text = "".join((p.get("text") or "") for p in parts if not p.get("thought"))
                if not text.strip():
                    _record_metric(provider, elapsed, False, is_error=True)
                    return {"error": {"message": f"Gemini returned no text content (finishReason={raw['candidates'][0].get('finishReason')}): {str(raw)[:300]}"}}
                _record_metric(provider, elapsed, True, str(hash(text))[:8])
                return {"choices": [{"message": {"content": text}}]}
            except (KeyError, IndexError):
                _record_metric(provider, elapsed, False, is_error=True)
                return {"error": raw}

        else:
            return {"error": f"Unknown provider: {provider}"}

    except Exception as e:
        elapsed = time.time() - start_time
        _record_metric(provider, elapsed, False, is_error=True)
        return {"error": str(e)}


def _escape_stray_inner_quotes(s):
    """Best-effort repair for a common AI-generation failure: a literal
    double-quote character used INSIDE a JSON string value (e.g. the model
    quoting a word in Arabic/English text) instead of a properly escaped
    \\" or a single quote. We walk the string and re-escape any double-quote
    that doesn't actually look like a string delimiter (i.e. it isn't
    immediately followed by a JSON structural character like , : } ] or
    whitespace+one of those)."""
    out = []
    in_string = False
    i, n = 0, len(s)
    while i < n:
        c = s[i]
        if c == '"' and (i == 0 or s[i-1] != '\\'):
            if not in_string:
                in_string = True
                out.append(c)
            else:
                # We're inside a string — decide if this quote is the
                # closing delimiter or a stray quote inside the text.
                j = i + 1
                while j < n and s[j] in ' \t\r\n':
                    j += 1
                if j >= n or s[j] in ',:}]':
                    in_string = False
                    out.append(c)
                else:
                    out.append('\\"')
        else:
            out.append(c)
        i += 1
    return ''.join(out)

def parse_json_response(raw):
    if "```" in raw:
        parts = raw.split("```")
        raw = parts[1] if len(parts) > 1 else parts[0]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()
    raw = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', raw)
    raw = fix_json_newlines(raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    match = re.search(r'\{.*\}', raw, re.DOTALL)
    if match:
        candidate = match.group(0)
        candidate = re.sub(r"(?<=\w)'(?=\w)", "\u2019", candidate)
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            pass
        # Last resort: repair stray unescaped quotes inside string values
        try:
            return json.loads(_escape_stray_inner_quotes(candidate))
        except json.JSONDecodeError:
            pass
    raise json.JSONDecodeError("Cannot parse JSON", raw, 0)

def clean_result(result, is_arabic):
    for f in ["body","suspicious_text","why_risky","learning_tip","subject","email_type"]:
        if result.get(f):
            result[f] = clean_foreign_only(result[f])
            if is_arabic:
                result[f] = remove_foreign_latin_words(result[f])
    for ind in result.get("indicators",[]):
        for k in ["title","description"]:
            if ind.get(k):
                ind[k] = clean_foreign_only(ind[k])
                if is_arabic:
                    ind[k] = remove_foreign_latin_words(ind[k])
    result["from"] = clean_email_field((result.get("from") or ""))
    result["to"] = extract_to_email((result.get("to") or ""))
    if result.get("suspicious_link"):
        sl = result["suspicious_link"]
        sl = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]','',sl).strip()
        sl = re.sub(r'[\u0600-\u06ff\s]','',sl)
        result["suspicious_link"] = sl
    return result













def render_email_window(email, is_arabic, show_badges=False):
    bd = 'rtl' if is_arabic else 'ltr'
    ta = 'right' if is_arabic else 'left'
    email_font = 'Tahoma,Arial,sans-serif' if is_arabic else "'Courier New',monospace"

    # FINAL TYPE-SAFETY NET: no matter what slipped through every earlier
    # parsing/recovery step, these fields must be plain strings here, or
    # every re.sub()/.strip() below will crash with "expected string or
    # bytes-like object, got 'dict'". Coerce defensively at the last mile.
    def _as_text(v):
        if isinstance(v, str):
            return v
        if v is None:
            return ""
        return str(v)
    email = {**email, **{
        k: _as_text(email.get(k)) for k in
        ["body", "subject", "from", "to", "suspicious_text", "suspicious_link", "attachment"]
        if k in email
    }}

    body_raw        = re.sub(r'<[^>]+>','', (email.get("body") or ""))
    suspicious_text = re.sub(r'<[^>]+>','', (email.get("suspicious_text") or ""))
    suspicious_link = re.sub(r'<[^>]+>','', (email.get("suspicious_link") or "")).strip()


    # Badge numbers come from the same grounded indicators shown in the tutor.
    # This removes the old hard-coded 1/2/3/4 mismatch.
    _inds = email.get("indicators") if isinstance(email.get("indicators"), list) else []
    def _badge_for_target(target):
        for _it in _inds:
            if isinstance(_it, dict) and _it.get("target") == target:
                try:
                    return int(_it.get("number"))
                except Exception:
                    return None
        return None
    def _badge_for_key(key):
        for _it in _inds:
            if isinstance(_it, dict) and _it.get("key") == key:
                try:
                    return int(_it.get("number"))
                except Exception:
                    return None
        return None

    body_raw = re.sub(r'suspicious_link\s*:\s*', '', body_raw, flags=re.IGNORECASE)
    body_raw = re.sub(r'suspicious_text\s*:\s*', '', body_raw, flags=re.IGNORECASE)
    # Remove placeholder phone numbers like +966-XX-XXXXXXX or +1-XXX-XXX-XXXX
    body_raw = re.sub(r'\+?\d{1,4}[-\s]?(?:XX|\d{2,3})[-\s]?(?:XX+|\d{3,4})(?:[-\s]?(?:XX+|\d{3,4}))*', '', body_raw)
    body_raw = re.sub(r'Contact:\s*\n', '', body_raw)
    body_raw = re.sub(r'[ \t]*\n[ \t]*\n[ \t]*\n+', '\n\n', body_raw).strip()

    # --------------------------------------------------------
    # NEW: detect a "[QR Code: label]" / "[QR: label]" placeholder
    # inside the body. Instead of deleting it, swap it for a unique
    # token that stays exactly where the model put it — so the real
    # QR image ends up rendered IN PLACE (not dumped at the very
    # bottom after the signature).
    # --------------------------------------------------------
    qr_label, has_qr = "", False
    _difficulty = st.session_state.get("difficulty", "medium")
    _medium_channel = str(email.get("medium_channel", "")).strip().lower()
    if _difficulty == "medium" and _medium_channel not in {"button", "link", "none"}:
        # Backward-compatible fallback for old cached/generated items.
        _medium_channel = "button" if re.search(r'\[[^\]]+\]\s*\(\s*https?://', body_raw) else ("link" if suspicious_link else "none")

    qr_match = re.search(r'\[\s*QR(?:\s*Code)?\s*:?\s*([^\]]*)\]', body_raw, re.I)
    if qr_match:
        if _difficulty in ("easy", "medium"):
            # QR is FORBIDDEN in easy/medium — remove it completely
            body_raw = body_raw[:qr_match.start()] + "" + body_raw[qr_match.end():]
            body_raw = re.sub(r'[ \t]*\n[ \t]*\n[ \t]*\n+', '\n\n', body_raw).strip()
        else:
            # Hard/Advanced — QR is now OPTIONAL (varies per example, not
            # mandatory in every single one) — keep it only if the model
            # actually wrote a [QR:...] marker naturally in the body.
            has_qr = True
            qr_label = qr_match.group(1).strip()
            body_raw = body_raw[:qr_match.start()] + "@@QR_TOKEN@@" + body_raw[qr_match.end():]
    # NOTE: no forced/auto-injected QR anymore for hard/advanced — if the
    # model didn't naturally include one, this example simply has no QR,
    # which is the intended variation (some Advanced emails have QR,
    # some don't, per the updated requirement).

    # --------------------------------------------------------
    # NEW: detect a markdown-style "[Button label](https://...)"
    # link inside the body. Same idea: swap for a token so the real
    # clickable button renders IN PLACE of where the link was
    # written, not always at the bottom of the email.
    # --------------------------------------------------------
    link_label, link_url, has_link_button = "", "", False
    link_match = re.search(r'\[([^\]]{1,80})\]\s*\(\s*(https?://[^\)\s]+)\s*\)', body_raw)
    if link_match:
        if _difficulty == "medium" and _medium_channel != "button":
            _raw_url = link_match.group(2).strip()
            if _medium_channel == "link":
                body_raw = body_raw[:link_match.start()] + _raw_url + body_raw[link_match.end():]
                suspicious_link = suspicious_link or _raw_url
                email["suspicious_link"] = suspicious_link
            else:  # no-link mode
                body_raw = body_raw[:link_match.start()] + "" + body_raw[link_match.end():]
                suspicious_link = ""
                email["suspicious_link"] = ""
        elif _difficulty == "easy":
            # Easy level: NO button allowed — convert to plain text URL
            _raw_url = link_match.group(2).strip()
            body_raw = body_raw[:link_match.start()] + _raw_url + body_raw[link_match.end():]
            # Also capture for suspicious_link display if not already set
            if not email.get("suspicious_link"):
                email["suspicious_link"] = _raw_url
        else:
            has_link_button = True
            link_label = link_match.group(1).strip()
            link_url   = link_match.group(2).strip()
            body_raw   = body_raw[:link_match.start()] + "@@LINK_TOKEN@@" + body_raw[link_match.end():]

    # Safety net: if the MODEL itself wrote the [QR:...] or [Label](url)
    # marker as the very last line (after the closing signature) instead
    # of inline where it's referenced, reposition it before the
    # signature — same fix as for a bare suspicious_link, generalized to
    # the token placeholders used here.
    for _marker_token in ("@@QR_TOKEN@@", "@@LINK_TOKEN@@"):
        if _marker_token in body_raw:
            body_raw = _reposition_trailing_lone_link(body_raw, _marker_token)

    # Also for easy: if suspicious_link exists but no visible URL in body, append it
    if _difficulty == "easy" and (email.get("suspicious_link") or "").strip():
        _sl = (email.get("suspicious_link") or "").strip()
        # Always run _place_link_in_body — it handles all cases:
        # link missing, link after signature, link already correct.
        if _sl and not has_link_button:
            body_raw = _place_link_in_body(body_raw, _sl)

    # Tidy up extra blank lines left behind after removing the placeholders above.
    body_raw = re.sub(r'[ \t]*\n[ \t]*\n[ \t]*\n+', '\n\n', body_raw).strip()

    # --------------------------------------------------------
    # NEW: providers (Groq/Llama especially) often write a trailing
    # punctuation mark right after the markdown link/QR syntax, e.g.
    # "...here: [Button](url). Your input is crucial..." — once the
    # bracket+url is swapped for a token, that stray "." (or ",")
    # is left dangling right next to the token and renders as an
    # ugly lone punctuation line. Strip stray punctuation that sits
    # immediately before/after either token (allowing whitespace).
    # --------------------------------------------------------
    body_raw = re.sub(r'[ \t]*[.,؛;:]+[ \t]*(@@(?:QR|LINK)_TOKEN@@)', r'\1', body_raw)
    body_raw = re.sub(r'(@@(?:QR|LINK)_TOKEN@@)[ \t]*[.,؛]+(?=[ \t]*(\n|$))', r'\1', body_raw)
    body_raw = re.sub(r'(@@(?:QR|LINK)_TOKEN@@)[ \t]*[.,؛]+[ \t]*', r'\1 ', body_raw)
    body_raw = body_raw.strip()

    # --------------------------------------------------------
    # Remove duplicate URL — strip any standalone line that is exactly the suspicious_link
    # This applies ALWAYS (not just when has_qr or has_link_button) to prevent double display
    # NOTE: For easy level the link was already placed correctly by _place_link_in_body above.
    # This block only removes truly redundant duplicate occurrences.
    if suspicious_link and (has_qr or has_link_button):
        bare_link_pattern = re.escape(suspicious_link)
        bare_no_scheme    = re.escape(re.sub(r'^https?://', '', suspicious_link))
        body_raw = re.sub(rf'^[ \t]*{bare_link_pattern}[ \t]*$\n?', '', body_raw, flags=re.MULTILINE)
        body_raw = re.sub(rf'^[ \t]*{bare_no_scheme}[ \t]*$\n?', '', body_raw, flags=re.MULTILINE)
        body_raw = re.sub(r'[ \t]*\n[ \t]*\n[ \t]*\n+', '\n\n', body_raw).strip()

    # Medium uses exactly one planned delivery style: button OR visible link OR no link.
    if _difficulty == "medium":
        if _medium_channel == "none":
            if suspicious_link:
                body_raw = body_raw.replace(suspicious_link, "")
                body_raw = body_raw.replace(re.sub(r'^https?://', '', suspicious_link), "")
            suspicious_link = ""
            email["suspicious_link"] = ""
            has_link_button = False
            body_raw = body_raw.replace("@@LINK_TOKEN@@", "")
        elif _medium_channel == "link" and suspicious_link and not has_link_button:
            link_bare = re.sub(r'^https?://', '', suspicious_link)
            if suspicious_link not in body_raw and link_bare not in body_raw:
                body_raw = _place_link_in_body(body_raw, suspicious_link)
        elif _medium_channel == "button":
            # A button is rendered from the markdown token; remove any duplicate raw URL.
            if has_link_button and suspicious_link:
                body_raw = re.sub(rf'^\s*{re.escape(suspicious_link)}\s*$', '', body_raw, flags=re.MULTILINE)


    # Keep the suspicious URL exactly once and always before the signature.
    # Provider output may include it inline and again as a trailing standalone line.
    if suspicious_link and not has_link_button and not has_qr and not (_difficulty == "medium" and _medium_channel in {"none", "button"}):
        _url_re = re.compile(re.escape(suspicious_link), re.I)
        _seen = [0]
        def _keep_first_url(m):
            _seen[0] += 1
            return m.group(0) if _seen[0] == 1 else ""
        body_raw = _url_re.sub(_keep_first_url, body_raw)
        body_raw = re.sub(r'[ \t]+\n', '\n', body_raw)
        body_raw = re.sub(r'[ \t]*\n[ \t]*\n[ \t]*\n+', '\n\n', body_raw).strip()
        if _seen[0] == 0:
            body_raw = _place_link_in_body(body_raw, suspicious_link)
        body_raw = _reposition_trailing_lone_link(body_raw, suspicious_link)

    has_attachment  = bool((email.get("attachment") or "").strip())

    # Remove duplicate attachment filename that sometimes appears as plain text at end
    _att_name = (email.get("attachment") or "").strip()
    if _att_name:
        att_escaped = re.escape(_att_name)
        body_raw = re.sub(rf'^\s*Attachment\s*:\s*{att_escaped}\s*$', '', body_raw, flags=re.MULTILINE|re.IGNORECASE)
        body_raw = re.sub(rf'^\s*{att_escaped}\s*$', '', body_raw, flags=re.MULTILINE)
        body_raw = re.sub(r'[ \t]*\n[ \t]*\n[ \t]*\n+', '\n\n', body_raw).strip()

    if _difficulty == "medium":
        body_raw = re.sub(r'\n\s*\n+', '\n', body_raw).strip()

    body_html = html_lib.escape(body_raw)

    def make_badge(n, color="#DC2626"):
        return (f'<span style="display:inline-flex;align-items:center;justify-content:center;'
                f'width:20px;height:20px;border-radius:50%;background:{color};color:white;'
                f'font-size:.7rem;font-weight:800;margin:0 3px;vertical-align:middle;">{n}</span>')

    if show_badges:
        # Mark every body/greeting/link indicator with its own tutor number.
        for _it in _inds:
            if not isinstance(_it, dict):
                continue
            _target = _it.get("target")
            _evidence = str(_it.get("evidence") or "")
            try:
                _num = int(_it.get("number"))
            except Exception:
                continue
            if _target not in ("body", "greeting", "link") or not _evidence:
                continue
            _safe = html_lib.escape(_evidence)
            if _safe in body_html:
                _style = ('color:#60A5FA;text-decoration:underline;' if _target == "link" else 'color:#FCA5A5;')
                body_html = body_html.replace(
                    _safe,
                    f'<span style="border:2px solid rgba(239,68,68,.6);border-radius:7px;'
                    f'padding:.2rem .5rem;background:rgba(239,68,68,.08);{_style}'
                    f'box-decoration-break:clone;-webkit-box-decoration-break:clone;">'
                    f'{make_badge(_num)}{_safe}</span>', 1)

    body_html = body_html.replace("\n","<br>")

    # --------------------------------------------------------
    # NEW: real, scannable QR-code image (rendered after body text).
    # Uses the suspicious_link (or the button URL, as a fallback) as
    # the QR payload so the badge/number still points to a real risk.
    # --------------------------------------------------------
    qr_block_html = ""
    if has_qr:
        qr_data    = suspicious_link or link_url or "https://example-training-only.invalid/qr"
        qr_img_url = f"https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={urllib.parse.quote(qr_data, safe='')}"
        # Use the grounded tutor indicator number for the QR/link.
        _qr_num = (
            _badge_for_target("qr")
            or _badge_for_key("qr")
            or _badge_for_target("link")
            or _badge_for_key("link")
        )
        qr_badge   = make_badge(_qr_num) if show_badges and _qr_num else ""
        qr_caption = html_lib.escape(qr_label) if qr_label else t("Scan with your phone","امسح بهاتفك")
        qr_block_html = f"""
<div style="margin:1rem 0;display:flex;align-items:center;gap:.8rem;direction:{bd};flex-wrap:wrap;">
  {qr_badge}
  <div style="background:#ffffff;padding:8px;border-radius:10px;border:2px solid rgba(239,68,68,.55);display:inline-block;line-height:0;">
    <img src="{qr_img_url}" width="120" height="120" alt="QR code" style="display:block;border-radius:2px;"/>
  </div>
  <div style="color:#94A3B8;font-size:.85rem;">{qr_caption}</div>
</div>"""

    # --------------------------------------------------------
    # NEW: a real clickable link "button" — same visual language as
    # the attachment chip — instead of leaving "[Label] (url)" as
    # inert bracketed text.
    # --------------------------------------------------------
    # Warning page trigger ID
    _warn_key = f"phishing_warn_{id(email)}"

    link_block_html = ""
    if has_link_button:
        # Use the existing grounded link indicator number.
        _link_num = _badge_for_target("link") or _badge_for_key("link")
        link_badge = make_badge(_link_num) if show_badges and _link_num else ""
        # The button's markdown label is swapped for a token BEFORE the generic
        # body-badge loop runs, so a "workflow"/"sharepoint"/"button" indicator
        # (target="body", describing the button itself) can never be matched by
        # text search here — its evidence text no longer exists in body_html by
        # this point. Look it up directly by key instead and badge the button
        # with both numbers, so that indicator isn't silently dropped from the UI.
        _workflow_num = _badge_for_key("workflow") or _badge_for_key("sharepoint")
        _already_shown = _workflow_num and f'>{_workflow_num}</span>' in body_html
        workflow_badge = make_badge(_workflow_num) if show_badges and _workflow_num and _workflow_num != _link_num and not _already_shown else ""
        # Replace generic "Open Link" label with descriptive fallback
        _display_label = link_label
        if _display_label.strip().lower() in ("open link", "click here", "link", "رابط", "اضغط هنا", "فتح الرابط"):
            _display_label = t("View Document", "عرض المستند")
        safe_label = html_lib.escape(_display_label)
        link_block_html = f"""
<div style="margin:.8rem 0;direction:{bd};">
  <button onclick="window.parent.postMessage({{type:'phishing_click',element:'link',label:'{safe_label}'}},\'*\')"
     style="display:inline-flex;align-items:center;gap:.5rem;border:1px solid #0078D4;
            border-radius:6px;padding:.5rem 1.2rem;background:#0078D4;color:white;
            font-size:.92rem;font-weight:700;cursor:pointer;font-family:inherit;
            box-shadow:0 2px 6px rgba(0,120,212,.4);"
     onmouseover="this.style.background='#006CBE'" onmouseout="this.style.background='#0078D4'">
    {workflow_badge}{link_badge}🔗 {safe_label}
  </button>
</div>"""

    # --------------------------------------------------------
    # NEW: inject the QR image / link button INTO the body at the
    # exact spot the model originally placed it (via the tokens
    # above), instead of always dumping it after the signature.
    # If, for any reason, the token didn't survive into body_html
    # (e.g. the model wrote the placeholder oddly), fall back to
    # appending the block at the end so it never gets silently lost.
    # --------------------------------------------------------
    if has_qr:
        if "@@QR_TOKEN@@" in body_html:
            body_html = body_html.replace("@@QR_TOKEN@@", qr_block_html)
        else:
            body_html += qr_block_html
    if has_link_button:
        if "@@LINK_TOKEN@@" in body_html:
            body_html = body_html.replace("@@LINK_TOKEN@@", link_block_html)
        else:
            body_html += link_block_html

    from_val = html_lib.escape((email.get("from") or ""))
    to_val   = html_lib.escape(email.get("to","employee@hospital.org"))
    subj_val = html_lib.escape((email.get("subject") or ""))
    att_val  = html_lib.escape((email.get("attachment") or ""))

    fl = t("From:","من:")
    tl = t("To:","إلى:")
    sl = t("Subject:","الموضوع:")

    _from_nums = [int(_it.get("number")) for _it in _inds if isinstance(_it, dict) and _it.get("target") == "from" and str(_it.get("number","")).strip()]
    _n_subj = _badge_for_target("subject")
    _n_att  = _badge_for_target("attachment")
    b_from = "".join(make_badge(_n) for _n in _from_nums) if show_badges and _from_nums else ""
    b_subj = make_badge(_n_subj) if show_badges and _n_subj else ""
    b_att  = make_badge(_n_att) if show_badges and _n_att else ""

    att_html = ""
    if att_val:
        att_html = (f'<div style="display:inline-flex;align-items:center;gap:.5rem;'
                    f'border:1px solid #0078D4;border-radius:4px;padding:.4rem .8rem;'
                    f'background:rgba(0,120,212,.12);color:#60A5FA;font-size:.88rem;margin:.4rem 0;'
                    f'cursor:pointer;" onclick="window.parent.postMessage({{type:\'phishing_click\',element:\'attachment\'}},\'*\')">'
                    f'{b_att}📎 {att_val}</div>')

    # Outlook-style toolbar buttons (visual only)
    reply_lbl    = "رد" if is_arabic else "Reply"
    forward_lbl  = "إعادة توجيه" if is_arabic else "Forward"
    delete_lbl   = "حذف" if is_arabic else "Delete"
    toolbar_dir  = "rtl" if is_arabic else "ltr"

    # Research v15: use generated display_time when available so dates vary realistically
    # and remain stable across Streamlit reruns. Fallback keeps old behavior.
    _email_time = (email.get("display_time") or "").strip()
    if not _email_time:
        _rand_hour   = random.randint(7, 16)
        _rand_min    = random.choice([0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55])
        _period_en   = "AM" if _rand_hour < 12 else "PM"
        _hour_12     = _rand_hour if _rand_hour <= 12 else _rand_hour - 12
        _hour_12     = 12 if _hour_12 == 0 else _hour_12
        _period_ar   = "ص" if _rand_hour < 12 else "م"
        _time_en     = f"Today, {_hour_12}:{_rand_min:02d} {_period_en}"
        _time_ar     = f"اليوم، {_hour_12}:{_rand_min:02d} {_period_ar}"
        _email_time  = _time_ar if is_arabic else _time_en

    st.markdown(f"""
<div style="background:#D8DCE1;border:1px solid #C5CAD0;border-radius:12px 12px 0 0;overflow:hidden;font-family:'Segoe UI',Arial,sans-serif;box-shadow:0 4px 24px rgba(0,0,0,.5);">
  <!-- Outlook-style title bar -->
  <div style="background:#CDD2D8;padding:.45rem 1rem;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid #B5BAC0;">
    <div style="display:flex;gap:6px;align-items:center;">
      <div style="width:12px;height:12px;border-radius:50%;background:#FF5F57;"></div>
      <div style="width:12px;height:12px;border-radius:50%;background:#FFBD2E;"></div>
      <div style="width:12px;height:12px;border-radius:50%;background:#28C840;"></div>
    </div>
    <div style="color:#555;font-size:.78rem;letter-spacing:.5px;">📧 {"صندوق الوارد — Microsoft Outlook" if is_arabic else "Inbox — Microsoft Outlook"}</div>
    <div style="width:60px;"></div>
  </div>
  <!-- Outlook-style action toolbar -->
  <div style="background:#DDE1E6;padding:.35rem 1rem;display:flex;gap:.5rem;align-items:center;border-bottom:1px solid #C5CAD0;direction:{toolbar_dir};">
    <button style="background:#0078D4;color:white;border:none;border-radius:4px;padding:.28rem .85rem;font-size:.78rem;cursor:pointer;font-family:inherit;font-weight:600;box-shadow:0 1px 3px rgba(0,0,0,.2);transition:background .15s;" onmouseover="this.style.background='#006CBE'" onmouseout="this.style.background='#0078D4'">↩ {reply_lbl}</button>
    <button style="background:#F3F4F6;color:#374151;border:1px solid #CBD5E1;border-radius:4px;padding:.28rem .85rem;font-size:.78rem;cursor:pointer;font-family:inherit;font-weight:500;" onmouseover="this.style.background='#E5E7EB'" onmouseout="this.style.background='#F3F4F6'">→ {forward_lbl}</button>
    <button style="background:#F3F4F6;color:#374151;border:1px solid #CBD5E1;border-radius:4px;padding:.28rem .85rem;font-size:.78rem;cursor:pointer;font-family:inherit;font-weight:500;" onmouseover="this.style.background='#FEE2E2';this.style.color='#DC2626'" onmouseout="this.style.background='#F3F4F6';this.style.color='#374151'">🗑 {delete_lbl}</button>
    <div style="flex:1;"></div>
    <span style="color:#6B7280;font-size:.75rem;font-style:italic;">{_email_time}</span>
  </div>
  <!-- Email header -->
  <div style="padding:.9rem 1.6rem .5rem;font-size:.92rem;color:#CBD5E1;direction:{bd};text-align:{ta};background:#DDE1E6;border-bottom:1px solid #C5CAD0;">
    <table style="width:100%;border-collapse:collapse;direction:{bd};">
      <tr style="vertical-align:top;">
        <td style="color:#6B7280;font-weight:600;padding:0 8px 6px 0;white-space:nowrap;width:80px;font-size:.85rem;">{fl}</td>
        <td style="color:#111827;padding:0 0 6px 0;word-break:break-all;">{b_from}{from_val}</td>
      </tr>
      <tr style="vertical-align:middle;">
        <td style="color:#6B7280;font-weight:600;padding:0 8px 6px 0;white-space:nowrap;font-size:.85rem;">{tl}</td>
        <td style="color:#60A5FA;padding:0 0 6px 0;direction:ltr;text-align:{('right' if bd=='rtl' else 'left')};overflow:hidden;text-overflow:ellipsis;">{to_val}</td>
      </tr>
      <tr style="vertical-align:top;">
        <td style="color:#6B7280;font-weight:600;padding:0 8px 6px 0;white-space:nowrap;font-size:.85rem;">{sl}</td>
        <td style="color:#111827;padding:0 0 6px 0;word-break:break-word;font-weight:700;">{b_subj}{subj_val}</td>
      </tr>
    </table>
    {att_html}
  </div>
</div>
<div style="background:#E6EAF2;border:1px solid #CDD1D6;border-top:none;
            border-radius:0 0 12px 12px;padding:1.2rem 1.6rem 1.6rem;
            font-family:'Segoe UI',Arial,sans-serif;
            font-size:.93rem;color:#1F2937;background:#E6EAF2;
            line-height:1.9;direction:{bd};text-align:{ta};
            box-shadow:0 20px 60px rgba(0,0,0,.3);">
  {body_html}
</div>""", unsafe_allow_html=True)


def page_phishing_caught():
    """Warning page shown when user clicks a phishing link, button, or scans QR."""
    is_arabic = st.session_state.get("language") == "Arabic"
    _dir = "rtl" if is_arabic else "ltr"
    _align = "right" if is_arabic else "left"

    if is_arabic:
        title     = "⚠️ تنبيه! لقد وقعت في فخ التصيد الاحتيالي"
        subtitle  = "هذا بريد تصيد تدريبي — في الواقع الحقيقي كنت ستتعرض للاختراق"
        msg       = "الرابط أو الزر الذي ضغطت عليه كان يؤدي إلى موقع مزيف مصمم لسرقة بياناتك. في بيئة حقيقية، كان يمكن للمهاجم الحصول على معلوماتك الشخصية أو بيانات دخولك."
        tip_title = "💡 ماذا كان يجب أن تفعل؟"
        tips = [
            "تحقق من نطاق البريد الإلكتروني للمرسل قبل الضغط على أي رابط.",
            "لا تضغط على روابط مجهولة المصدر — تواصل مع قسم تقنية المعلومات للتحقق.",
            "تحقق من عنوان URL الفعلي قبل إدخال أي بيانات.",
            "الجهات الرسمية لا تطلب كلمة المرور عبر البريد الإلكتروني.",
        ]
        btn_label = "← العودة ومتابعة التدريب"
    else:
        title     = "⚠️ Alert! You clicked a phishing link"
        subtitle  = "This was a training phishing email — in a real scenario you would have been compromised"
        msg       = "The link or button you clicked would have led to a fake website designed to steal your credentials. In a real attack, the attacker could have obtained your personal information or login details."
        tip_title = "💡 What should you have done?"
        tips = [
            "Always verify the sender's email domain before clicking any link.",
            "Never click unknown links — contact IT to verify first.",
            "Check the actual URL before entering any information.",
            "Official organizations never ask for your password via email.",
        ]
        btn_label = "← Return and Continue Training"

    tips_html = "".join(f'<li style="margin-bottom:.5rem;color:#E2E8F0;">{tip}</li>' for tip in tips)

    st.markdown(f"""
<div dir="{_dir}" style="max-width:700px;margin:3rem auto;text-align:{_align};">
  <div style="background:linear-gradient(135deg,rgba(127,29,29,.95),rgba(69,10,10,.9));
              border:2px solid #EF4444;border-radius:20px;padding:2.5rem;
              box-shadow:0 0 60px rgba(239,68,68,.3);">
    <div style="font-size:3rem;text-align:center;margin-bottom:1rem;">🎣</div>
    <h2 style="color:#FCA5A5;font-size:1.6rem;font-weight:900;text-align:center;margin-bottom:.5rem;">{title}</h2>
    <p style="color:#FECACA;font-size:1rem;text-align:center;margin-bottom:1.5rem;font-style:italic;">{subtitle}</p>
    <div style="background:rgba(0,0,0,.3);border-radius:12px;padding:1.2rem;margin-bottom:1.5rem;">
      <p style="color:#E2E8F0;line-height:1.8;margin:0;">{msg}</p>
    </div>
    <div style="margin-bottom:1.5rem;">
      <p style="color:#FCD34D;font-weight:700;font-size:1rem;margin-bottom:.7rem;">{tip_title}</p>
      <ul style="padding-{('right' if is_arabic else 'left')}:1.2rem;margin:0;line-height:1.9;">
        {tips_html}
      </ul>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button(btn_label, use_container_width=True, type="primary"):
            st.session_state["page"] = "learning"
            st.session_state.pop("phishing_caught", None)
            st.rerun()


def page_home():
    is_arabic      = st.session_state.get("language") == "Arabic"
    dir_attr       = 'rtl' if is_arabic else 'ltr'
    text_align     = 'right' if is_arabic else 'left'
    flex_justify   = 'flex-end' if is_arabic else 'flex-start'
    hero_grid_cols = '1fr 230px' if is_arabic else '230px 1fr'

    st.markdown(f"""<style>
#MainMenu,header,footer{{visibility:hidden;}}
.stApp{{background:radial-gradient(circle at top left,#0B2E68 0%,#020617 35%,#020617 100%);color:white;}}
.block-container{{max-width:1160px;padding-top:2rem;}}
.hero-card{{border:1px solid rgba(37,99,235,.55);border-radius:24px;padding:2.2rem 2.4rem;background:linear-gradient(135deg,rgba(2,6,23,.96),rgba(8,47,73,.88));box-shadow:0 24px 70px rgba(0,0,0,.42);margin-bottom:1.3rem;}}
.hero-grid{{display:grid;grid-template-columns:{hero_grid_cols};gap:2rem;align-items:center;}}
.shield-orb{{width:180px;height:180px;border-radius:50%;margin:auto;display:flex;align-items:center;justify-content:center;background:radial-gradient(circle,rgba(37,99,235,.45),rgba(2,6,23,.15) 65%);border:1px solid rgba(56,189,248,.35);box-shadow:0 0 45px rgba(37,99,235,.36);position:relative;overflow:visible;}}
.shield-orb::before{{content:"";position:absolute;width:215px;height:215px;border-radius:50%;border:1px dashed rgba(56,189,248,.34);}}
.hero-content{{text-align:center;}}
.hero-title{{font-size:3.4rem;font-weight:900;color:#F8FAFC;margin-bottom:.7rem;}}
.hero-tagline{{font-size:1.45rem;font-weight:800;color:#1EA7FF;margin-bottom:.9rem;}}
.hero-desc{{font-size:1rem;color:#DCEBFF;line-height:1.7;max-width:620px;margin:0 auto;}}
.features-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;margin-bottom:1.4rem;direction:{dir_attr};}}
.feature-card{{border:1px solid rgba(37,99,235,.55);background:rgba(2,6,23,.60);border-radius:18px;padding:1.5rem 1rem;min-height:175px;text-align:center;cursor:pointer;transition:.25s ease;}}
.feature-card:hover{{transform:translateY(-6px);border-color:#1EA7FF;box-shadow:0 0 28px rgba(30,167,255,.22);}}
.feature-icon{{height:60px;margin-bottom:.8rem;display:flex;justify-content:center;align-items:center;}}
.feature-title{{font-size:1rem;font-weight:800;color:white;margin-bottom:.5rem;}}
.feature-text{{color:#BFD7F5;font-size:.9rem;line-height:1.55;}}
.form-section{{direction:{dir_attr};text-align:{text_align};margin-bottom:.5rem;}}
.form-title{{font-size:1.35rem;font-weight:900;color:white;margin-bottom:1rem;}}
.section-label{{font-weight:800;color:white;margin-bottom:.5rem;direction:{dir_attr};text-align:{text_align};}}
[data-testid="column"]{{direction:{dir_attr};}}
.stButton>button{{width:100%;min-height:48px;background:rgba(15,23,42,.78);color:#EAF4FF;border:1px solid rgba(37,99,235,.55);border-radius:12px;font-weight:800;direction:{dir_attr};}}
.stButton>button:hover{{background:linear-gradient(90deg,#0B4FA8,#0284C7);color:white;border-color:#1EA7FF !important;}}
.stButton>button:focus,.stButton>button:focus-visible,.stButton>button:focus:not(:hover){{outline:none !important;box-shadow:none !important;}}
.start-btn>button,.start-btn button[kind="primary"]{{min-height:56px !important;background:rgba(15,23,42,.88) !important;color:white !important;border:1px solid rgba(37,99,235,.65) !important;font-size:1.05rem !important;font-weight:900 !important;border-radius:14px !important;box-shadow:none !important;}}
.start-btn>button:hover,.start-btn button[kind="primary"]:hover{{background:linear-gradient(90deg,#0B4FA8,#0284C7) !important;border-color:#1EA7FF !important;}}
div[data-baseweb="select"] *{{color:#EAF4FF!important;-webkit-text-fill-color:#EAF4FF!important;}}
div[data-baseweb="select"] > div{{background:rgba(15,23,42,.82)!important;border:1px solid rgba(37,99,235,.65)!important;border-radius:12px!important;}}
div[data-baseweb="popover"] *{{color:#EAF4FF!important;-webkit-text-fill-color:#EAF4FF!important;}}
.stSelectbox>div>div,.stTextInput>div>div>input{{background-color:rgba(15,23,42,.88) !important;color:white !important;border:1px solid rgba(37,99,235,.55) !important;border-radius:12px !important;min-height:48px;direction:{dir_attr};text-align:{text_align};}}
div[data-baseweb="select"] span{{color:white !important;-webkit-text-fill-color:white!important;}}
div[data-baseweb="single-value"]{{color:white !important;-webkit-text-fill-color:white!important;}}
div[data-baseweb="select"] [data-value]{{color:white !important;-webkit-text-fill-color:white!important;}}
.stSelectbox div[class*="ValueContainer"] span,
.stSelectbox div[class*="singleValue"],
.stSelectbox div[class*="placeholder"]{{color:white !important;-webkit-text-fill-color:white!important;opacity:1!important;}}
div[data-baseweb="select"] input{{color:white !important;caret-color:white;-webkit-text-fill-color:white!important;}}
div[data-baseweb="popover"] ul li{{text-align:{text_align} !important;direction:{dir_attr} !important;}}
div[data-baseweb="popover"] [role="option"]{{text-align:{text_align} !important;direction:{dir_attr} !important;justify-content:{flex_justify} !important;}}
div[data-baseweb="popover"] [role="listbox"]{{direction:{dir_attr} !important;}}
div[data-baseweb="menu"] [role="option"]{{text-align:{text_align} !important;direction:{dir_attr} !important;justify-content:{flex_justify} !important;}}
div[data-baseweb="popover"] li > div{{text-align:{text_align} !important;direction:{dir_attr} !important;width:100%;}}
.footer-bar{{margin-top:2rem;padding:1.5rem 0;border-top:1px solid rgba(37,99,235,.35);display:flex;justify-content:space-between;align-items:center;color:#7DD3FC;font-size:.95rem;direction:{dir_attr};}}
.footer-side{{display:flex;align-items:center;gap:.8rem;}}
button[kind="secondary"]{{width:100% !important;min-height:52px !important;border-radius:14px !important;font-weight:800 !important;font-size:.95rem !important;transition:.2s ease !important;background:rgba(2,6,23,.55) !important;border:2px solid rgba(37,99,235,.35) !important;color:#94A3B8 !important;}}
button[kind="secondary"]:hover{{background:rgba(11,79,168,.25) !important;border-color:#1EA7FF !important;color:#FFFFFF !important;}}
button[kind="primary"]{{width:100% !important;min-height:52px !important;border-radius:14px !important;font-weight:800 !important;font-size:.95rem !important;background:linear-gradient(135deg,#0B4FA8,#0284C7) !important;border:2px solid #1EA7FF !important;color:#FFFFFF !important;box-shadow:0 0 18px rgba(30,167,255,.3) !important;opacity:1 !important;}}
button[kind="primary"]:hover,button[kind="primary"]:focus{{background:linear-gradient(135deg,#0B4FA8,#0284C7) !important;border:2px solid #1EA7FF !important;color:#FFFFFF !important;}}
@media(max-width:950px){{.hero-grid{{grid-template-columns:1fr;}}.features-grid{{grid-template-columns:1fr;}}.footer-bar{{flex-direction:column;gap:1rem;text-align:center;}}}}
</style>""", unsafe_allow_html=True)

    SHIELD_MAIN_SVG = """<svg width="130" height="148" viewBox="0 0 130 148" fill="none" xmlns="http://www.w3.org/2000/svg"><defs><linearGradient id="lock_g" x1="38" y1="72" x2="92" y2="132" gradientUnits="userSpaceOnUse"><stop offset="0%" stop-color="#FFFFFF"/><stop offset="100%" stop-color="#C8E6FF"/></linearGradient><filter id="sh_glow" x="-25%" y="-25%" width="150%" height="150%"><feGaussianBlur stdDeviation="4" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter></defs><path d="M65 4L124 24V72C124 108 98 136 65 144C32 136 6 108 6 72V24L65 4Z" fill="rgba(30,90,180,0.15)" stroke="#4A9EFF" stroke-width="3" filter="url(#sh_glow)"/><path d="M65 13L114 30V72C114 103 91 128 65 136C39 128 16 103 16 72V30L65 13Z" fill="none" stroke="rgba(120,190,255,0.3)" stroke-width="1.5"/><path d="M44 82V67C44 52 86 52 86 67V82" stroke="url(#lock_g)" stroke-width="10" stroke-linecap="round" fill="none"/><rect x="34" y="79" width="62" height="48" rx="9" fill="url(#lock_g)"/><circle cx="65" cy="100" r="8" fill="#1558A8"/><rect x="61.5" y="100" width="7" height="13" rx="2.5" fill="#1558A8"/></svg>"""
    BRAIN_SVG  = """<svg width="52" height="52" viewBox="0 0 52 52" fill="none"><path d="M26 8C20 8 17 12 15 20C12 20 9 22 9 26C9 30 12 32 15 32C14 37 17 40 26 44" stroke="#1EA7FF" stroke-width="2.5" stroke-linecap="round" fill="none"/><path d="M26 8C32 8 35 12 37 20C40 20 43 22 43 26C43 30 40 32 37 32C38 37 35 40 26 44" stroke="#1EA7FF" stroke-width="2.5" stroke-linecap="round" fill="none"/><line x1="26" y1="8" x2="26" y2="44" stroke="#1EA7FF" stroke-width="2"/><path d="M15 20C18 18 22 20 26 18C30 20 34 18 37 20" stroke="#1EA7FF" stroke-width="2" fill="none"/><path d="M15 32C18 30 22 32 26 30C30 32 34 30 37 32" stroke="#1EA7FF" stroke-width="2" fill="none"/></svg>"""
    TARGET_SVG = """<svg width="52" height="52" viewBox="0 0 52 52" fill="none"><circle cx="26" cy="28" r="18" stroke="#1EA7FF" stroke-width="2.5" fill="none"/><circle cx="26" cy="28" r="10" stroke="#1EA7FF" stroke-width="2.5" fill="none"/><circle cx="26" cy="28" r="3" fill="#1EA7FF"/><path d="M34 10L42 10L42 18" stroke="#1EA7FF" stroke-width="2.5" stroke-linecap="round" fill="none"/><line x1="42" y1="10" x2="29" y2="23" stroke="#1EA7FF" stroke-width="2.5"/></svg>"""
    CHART_SVG  = """<svg width="52" height="52" viewBox="0 0 52 52" fill="none"><line x1="10" y1="44" x2="10" y2="8" stroke="#1EA7FF" stroke-width="2.5" stroke-linecap="round"/><line x1="10" y1="44" x2="46" y2="44" stroke="#1EA7FF" stroke-width="2.5" stroke-linecap="round"/><rect x="15" y="32" width="7" height="12" rx="2" fill="#1EA7FF" opacity=".85"/><rect x="26" y="22" width="7" height="22" rx="2" fill="#1EA7FF" opacity=".85"/><rect x="37" y="12" width="7" height="32" rx="2" fill="#1EA7FF" opacity=".85"/></svg>"""
    SHIELD_SVG = """<svg width="52" height="56" viewBox="0 0 52 56" fill="none"><path d="M26 4L46 12V28C46 39 36 50 26 52C16 50 6 39 6 28V12L26 4Z" stroke="#1EA7FF" stroke-width="2.5" fill="none"/><path d="M18 28L23 33L34 22" stroke="#1EA7FF" stroke-width="2.5" stroke-linecap="round" fill="none"/></svg>"""
    SHFOOT     = """<svg width="42" height="48" viewBox="0 0 42 48" fill="none"><path d="M21 3L39 10V24C39 34 31 43 21 46C11 43 3 34 3 24V10L21 3Z" stroke="#1EA7FF" stroke-width="2.5" fill="none"/></svg>"""
    ECG_SVG    = """<svg width="80" height="28" viewBox="0 0 80 28" fill="none"><polyline points="0,14 15,14 20,4 25,24 30,4 35,20 40,14 80,14" stroke="#1EA7FF" stroke-width="2.5" stroke-linecap="round" fill="none"/></svg>"""

    nav_login    = t("Login","تسجيل الدخول")
    nav_register = t("Register","إنشاء حساب")
    nav_brand    = t("AI Phishing Awareness","التوعية بالتصيد الإلكتروني")
    user_name    = (st.session_state.get("user_name") or "")
    shield_small = SHIELD_SVG.replace('width="52"','width="20"').replace('height="56"','height="22"')
    flex_dir     = "row-reverse" if is_arabic else "row"

    if user_name:
        st.markdown(f"""
<div style="background:rgba(11,46,104,0.55);border:1px solid rgba(37,99,235,.4);
     border-radius:14px;padding:8px 20px;margin-bottom:1.2rem;
     display:flex;align-items:center;justify-content:space-between;
     flex-direction:{flex_dir};min-height:52px;">
  <div style="display:flex;align-items:center;gap:8px;flex-direction:{"row-reverse" if is_arabic else "row"};">
    {shield_small}
    <span style="font-size:15px;font-weight:800;color:#F8FAFC;white-space:nowrap;">{nav_brand}</span>
  </div>
  <div style="display:inline-flex;align-items:center;gap:6px;
      background:rgba(37,99,235,.15);border:1px solid rgba(37,99,235,.4);
      border-radius:20px;padding:5px 10px 5px 6px;">
    <div style="width:24px;height:24px;border-radius:50%;
        background:linear-gradient(135deg,#0B4FA8,#0284C7);
        display:flex;align-items:center;justify-content:center;font-size:11px;">👤</div>
    <span style="font-size:12px;color:#7DD3FC;font-weight:700;
        white-space:nowrap;max-width:140px;overflow:hidden;text-overflow:ellipsis;">
      {html_lib.escape(user_name)}</span>
  </div>
</div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"""
<style>
.nb-btn {{height:34px;padding:0 16px;border-radius:9px;font-size:12px;font-weight:700;cursor:pointer;display:inline-flex;align-items:center;white-space:nowrap;text-decoration:none;}}
.nb-btn-ghost {{background:rgba(15,23,42,.88);color:#EAF4FF !important;border:1px solid rgba(37,99,235,.5);}}
.nb-btn-ghost:hover {{background:rgba(37,99,235,.25);border-color:#1EA7FF;color:#fff !important;}}
.nb-btn-solid {{background:linear-gradient(90deg,#0B4FA8,#0284C7);color:white !important;border:none;}}
.nb-btn-solid:hover {{background:linear-gradient(90deg,#1560C0,#0396E0);}}
</style>
<div style="background:rgba(11,46,104,0.55);border:1px solid rgba(37,99,235,.4);
     border-radius:14px;padding:8px 20px;margin-bottom:1.2rem;
     display:flex;align-items:center;justify-content:space-between;
     flex-direction:{flex_dir};min-height:52px;">
  <div style="display:flex;align-items:center;gap:8px;flex-direction:{"row-reverse" if is_arabic else "row"};">
    {shield_small}
    <span style="font-size:15px;font-weight:800;color:#F8FAFC;white-space:nowrap;">{nav_brand}</span>
  </div>
  <div style="display:flex;gap:8px;align-items:center;">
    <a href="?nav=login&lang={st.session_state.get('language','English')}" class="nb-btn nb-btn-ghost">{nav_login}</a>
    <a href="?nav=register&lang={st.session_state.get('language','English')}" class="nb-btn nb-btn-solid">{nav_register}</a>
  </div>
</div>""", unsafe_allow_html=True)

    title   = t("AI Phishing Awareness","التوعية بالتصيد الإلكتروني بالذكاء الاصطناعي")
    tagline = t("Smart, Personalised, Protective.","ذكي، مخصص، وقائي")
    desc    = t("AI-powered training and assessment to help healthcare employees recognise and avoid phishing threats.",
                "تدريب وتقييم مدعوم بالذكاء الاصطناعي لمساعدة الموظفين الصحيين على التعرف على تهديدات التصيد الإلكتروني وتجنبها")
    time_badge = t("⏱️ ~15 minutes","⏱️ ~١٥ دقيقة")
    sh = f'<div class="shield-orb">{SHIELD_MAIN_SVG}</div>'
    co = f'''<div class="hero-content">
      <div class="hero-title">{title}</div>
      <div class="hero-tagline">{tagline}</div>
      <div class="hero-desc">{desc}</div>
      <div style="display:inline-flex;align-items:center;gap:6px;margin-top:.8rem;
                  background:rgba(37,99,235,.15);border:1px solid rgba(37,99,235,.35);
                  border-radius:8px;padding:5px 14px;">
        <span style="font-size:.9rem;color:#7DD3FC;font-weight:600;">{time_badge}</span>
      </div>
    </div>'''
    gi = co+sh if is_arabic else sh+co
    st.markdown(f'<div class="hero-card"><div class="hero-grid">{gi}</div></div>', unsafe_allow_html=True)

    cards = [
        (BRAIN_SVG, t("AI-Powered Learning","تعلم بالذكاء الاصطناعي"), t("Personalised content adapted to your role.","محتوى تعليمي مخصص حسب دورك الوظيفي")),
        (TARGET_SVG,t("Smart Assessment","تقييم ذكي"),                 t("Short, focused assessments to test your awareness.","تقييمات قصيرة ومركزة لاختبار وعيك")),
        (CHART_SVG, t("Personalised Feedback","تغذية راجعة مخصصة"),   t("Detailed results with insights and recommendations.","نتائج مفصلة تتضمن ملاحظات وتوصيات مخصصة")),
        (SHIELD_SVG,t("Stronger Together","معًا أكثر أمانًا"),         t("Building a secure healthcare environment for everyone.","بناء بيئة صحية آمنة للجميع")),
    ]
    st.markdown('<div class="features-grid">'+"".join(f'<div class="feature-card"><div class="feature-icon">{i}</div><div class="feature-title">{tt}</div><div class="feature-text">{tx}</div></div>' for i,tt,tx in cards)+'</div>', unsafe_allow_html=True)

    form_col, panel_col = st.columns([3, 1], gap="large")

    with form_col:
        form_title_txt = t("Let's personalise your experience","لنخصص تجربتك")
        st.markdown(f'<div class="form-section"><div class="form-title">👤 {form_title_txt}</div></div>', unsafe_allow_html=True)

        def step_label(n, txt):
            return f'''<div style="font-size:.85rem;color:#94A3B8;margin-bottom:.5rem;
                        display:flex;align-items:center;gap:6px;direction:{dir_attr};">
              <span style="display:inline-flex;align-items:center;justify-content:center;
                           width:18px;height:18px;border-radius:50%;
                           background:rgba(37,99,235,.5);color:#7DD3FC;
                           font-size:10px;font-weight:800;">{n}</span>
              {txt}
            </div>'''

        st.markdown(step_label("1", t("Select your preferred language","اختر اللغة المفضلة")), unsafe_allow_html=True)
        cur_lang     = (st.session_state.get("language") or "")
        lang_chosen0 = st.session_state.get("lang_explicitly_chosen", False)
        col1,col2 = st.columns(2)
        with col1:
            st.button("English", key="english", on_click=set_language, args=("English",),
                       use_container_width=True, type=("primary" if lang_chosen0 and cur_lang == "English" else "secondary"))
        with col2:
            st.button("العربية", key="arabic",  on_click=set_language, args=("Arabic",),
                       use_container_width=True, type=("primary" if lang_chosen0 and cur_lang == "Arabic" else "secondary"))

        st.markdown(step_label("2", t("Select your role","اختر دورك الوظيفي")), unsafe_allow_html=True)
        opts = [t("Choose your role","اختر دورك الوظيفي"),t("Clinical","سريري"),t("Admin / Management","إداري / إدارة"),t("IT / Informatics","تقنية المعلومات / المعلوماتية"),t("Other","أخرى")]
        sel  = st.selectbox("role",opts,index=0,label_visibility="collapsed")
        other_role = ""
        if sel==opts[-1]: other_role=st.text_input(t("Please specify your role","يرجى كتابة دورك الوظيفي"),placeholder=t("Type your role here","اكتب دورك الوظيفي هنا"))

        if is_arabic:
            # CSS alone doesn't reliably reach this dropdown's rendered option rows
            # (BaseWeb renders them as a body-level portal), so force it directly.
            import streamlit.components.v1 as components
            components.html("""
            <script>
            function fixRoleDropdownRTL() {
                const doc = window.parent.document;
                const opts = doc.querySelectorAll('div[data-baseweb="popover"] li, div[data-baseweb="popover"] [role="option"], div[data-baseweb="menu"] li, ul[role="listbox"] li');
                opts.forEach(function(el) {
                    el.style.setProperty('direction', 'rtl', 'important');
                    el.style.setProperty('text-align', 'right', 'important');
                    el.style.setProperty('justify-content', 'flex-end', 'important');
                    el.style.setProperty('unicode-bidi', 'plaintext', 'important');
                    const inner = el.querySelector('div, span');
                    if (inner) {
                        inner.style.setProperty('direction', 'rtl', 'important');
                        inner.style.setProperty('text-align', 'right', 'important');
                        inner.style.setProperty('width', '100%', 'important');
                    }
                });
                const listboxes = doc.querySelectorAll('ul[role="listbox"], div[role="listbox"]');
                listboxes.forEach(function(el) { el.style.setProperty('direction', 'rtl', 'important'); });
            }
            fixRoleDropdownRTL();
            const roleObserver = new MutationObserver(fixRoleDropdownRTL);
            roleObserver.observe(window.parent.document.body, {childList:true, subtree:true});
            setInterval(fixRoleDropdownRTL, 300);
            </script>
            """, height=0, width=0)

        st.markdown(step_label("3", t("Select difficulty level","اختر مستوى الصعوبة")), unsafe_allow_html=True)

    with panel_col:
        ph_label  = t("Learning phase","مرحلة التعلم")
        as_label  = t("Assessment","الاختبار")
        rep_label = t("Performance report","تقرير الأداء")
        exp_title = t("WHAT TO EXPECT","ماذا تتوقع")
        diff_title= t("DIFFICULTY","الصعوبة")
        beg_lbl   = t("Beginner","مبتدئ")
        mid_lbl   = t("Intermediate","متوسط")
        adv_lbl   = t("Advanced","متقدم")
        small_brain  = BRAIN_SVG.replace('width="52"','width="18"').replace('height="52"','height="18"')
        small_target = TARGET_SVG.replace('width="52"','width="18"').replace('height="52"','height="18"')
        small_chart  = CHART_SVG.replace('width="52"','width="18"').replace('height="52"','height="18"')
        st.markdown(f"""
<div style="background:rgba(8,47,73,.2);border:1px solid rgba(37,99,235,.25);border-radius:14px;padding:1.2rem 1rem;margin-top:1rem;direction:{dir_attr};">
  <div style="font-size:.75rem;font-weight:800;color:#7DD3FC;letter-spacing:.06em;margin-bottom:14px;">{exp_title}</div>
  <div style="display:flex;flex-direction:column;gap:9px;margin-bottom:16px;">
    <div style="display:flex;align-items:center;gap:9px;padding:9px 10px;background:rgba(15,23,42,.7);border:1px solid rgba(37,99,235,.3);border-radius:9px;">{small_brain}<span style="font-size:.82rem;font-weight:700;color:#E2E8F0;">{ph_label}</span></div>
    <div style="display:flex;align-items:center;gap:9px;padding:9px 10px;background:rgba(15,23,42,.7);border:1px solid rgba(37,99,235,.3);border-radius:9px;">{small_target}<span style="font-size:.82rem;font-weight:700;color:#E2E8F0;">{as_label}</span></div>
    <div style="display:flex;align-items:center;gap:9px;padding:9px 10px;background:rgba(15,23,42,.7);border:1px solid rgba(37,99,235,.3);border-radius:9px;">{small_chart}<span style="font-size:.82rem;font-weight:700;color:#E2E8F0;">{rep_label}</span></div>
  </div>
  <div style="border-top:1px solid rgba(37,99,235,.2);padding-top:12px;">
    <div style="font-size:.75rem;font-weight:800;color:#7DD3FC;letter-spacing:.05em;margin-bottom:8px;direction:{dir_attr};text-align:{text_align};">{diff_title}</div>
    <div style="display:flex;flex-direction:column;gap:5px;direction:{dir_attr};text-align:{text_align};">
      <div style="font-size:.8rem;color:#94A3B8;">🟢 {beg_lbl}</div>
      <div style="font-size:.8rem;color:#94A3B8;">🟡 {mid_lbl}</div>
      <div style="font-size:.8rem;color:#94A3B8;">🔴 {adv_lbl}</div>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

    with form_col:
        current_diff  = st.session_state.get("difficulty","medium")
        if st.session_state.get("language","English") == "Arabic":
            ordered = [("easy","🟢  مبتدئ"),("medium","🟡  متوسط"),("hard","🔴  متقدم")]
        else:
            ordered = [("easy","🟢  Beginner"),("medium","🟡  Intermediate"),("hard","🔴  Advanced")]

        diff_cols = st.columns(3)
        if st.session_state.get("language","English") == "Arabic":
            ordered_display = list(reversed(ordered))
        else:
            ordered_display = ordered

        for i,(dk,lbl) in enumerate(ordered_display):
            with diff_cols[i]:
                is_sel  = current_diff == dk and st.session_state.get("diff_explicitly_chosen", False)
                if st.button(lbl, key=f"diff_{dk}", use_container_width=True,
                             type=("primary" if is_sel else "secondary")):
                    st.session_state["difficulty"] = dk
                    st.session_state["diff_explicitly_chosen"] = True
                    st.rerun()

        if st.query_params.get("mode") == "researcher":
            st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)
            st.markdown(f'<div style="font-size:.75rem;font-weight:800;color:#F59E0B;letter-spacing:.06em;margin-bottom:.5rem;direction:{dir_attr};">🔬 RESEARCHER MODE — AI Provider</div>', unsafe_allow_html=True)
            provider_options = {
                "groq":      "🟠 Groq  (LLaMA 3.3-70b) — Baseline v3",
                "openai":    "🟢 ChatGPT  (GPT-4o) — Most used globally",
                "anthropic": "🟣 Claude  (claude-sonnet-4-6) — Best writing quality",
                "gemini":    "🔵 Gemini  (1.5 Pro) — Fastest growing",
            }
            cur_provider = st.session_state.get("ai_provider", "openai")
            prov_cols = st.columns(2)
            prov_items = list(provider_options.items())
            for i, (pk, plbl) in enumerate(prov_items):
                with prov_cols[i % 2]:
                    is_psel = cur_provider == pk
                    if st.button(plbl, key=f"prov_{pk}", use_container_width=True,
                                 type=("primary" if is_psel else "secondary")):
                        set_active_provider(pk)
                        st.rerun()
            st.markdown(f'<div style="font-size:.72rem;color:#64748B;margin-top:.3rem;direction:{dir_attr};">Active: <b style="color:#F59E0B;">{provider_options.get(cur_provider,"")}</b></div>', unsafe_allow_html=True)

        st.markdown('<div class="start-btn" style="margin-top:.8rem;">',unsafe_allow_html=True)
        if st.button(t("Start Personalised Training","ابدأ التدريب المخصص"),key="start_training", use_container_width=True, type="primary"):
            fr = other_role.strip() if sel==opts[-1] else sel
            lang_chosen = st.session_state.get("lang_explicitly_chosen", False)
            diff_chosen = st.session_state.get("diff_explicitly_chosen", False)
            user_logged = (st.session_state.get("user_name") or "").strip() != ""

            if not user_logged:
                st.warning(t("⚠️ Please login or register first using the button at the top.",
                             "⚠️ يرجى تسجيل الدخول أو إنشاء حساب أولاً عبر الزر في الأعلى"))
            elif not lang_chosen:
                st.warning(t("⚠️ Please select your preferred language first.",
                             "⚠️ يرجى اختيار اللغة المفضلة أولاً"))
            elif fr==opts[0]:
                st.warning(t("⚠️ Please select your role.",
                             "⚠️ يرجى اختيار دورك الوظيفي"))
            elif not diff_chosen:
                st.warning(t("⚠️ Please select a difficulty level.",
                             "⚠️ يرجى اختيار مستوى الصعوبة"))
            else:
                if "scenario_order" in st.session_state: del st.session_state["scenario_order"]
                # EN: fresh "Start Training" click → fresh scenario/recipient
                # order too, so re-running training mid-session also gets new
                # variety (not just a full logout/retake).
                # AR: ضغطة جديدة على "بدء التدريب" → ترتيب سيناريو/مستلم جديد
                # كذلك، حتى إعادة تشغيل التدريب بنفس الجلسة يطلع له تنوع جديد.
                for k in list(st.session_state.keys()):
                    if k.startswith(("scenario_order_", "recipient_order_", "category_order_", "used_topics_")):
                        st.session_state.pop(k, None)
                go_to_learning(fr); st.rerun()
        st.markdown('</div>',unsafe_allow_html=True)

    ft = t("Together, let's build a stronger, phishing-resistant healthcare environment.","معًا نبني بيئة صحية أكثر مقاومة للتصيد الإلكتروني")
    fs = t("Stay aware, Stay secure, Save lives.","كن واعيًا، ابق آمنًا، وساهم في حماية الأرواح")
    if is_arabic:
        f1=f'<div class="footer-side" style="direction:ltr;"><span style="direction:rtl;">{fs}</span>&nbsp;{ECG_SVG}</div>'
        f2=f'<div class="footer-side" style="direction:ltr;justify-content:flex-end;"><span style="direction:rtl;">{ft}</span>{SHFOOT}</div>'
    else:
        f1=f'<div class="footer-side">{SHFOOT}<span>{ft}</span></div>'
        f2=f'<div class="footer-side">{ECG_SVG}&nbsp;{fs}</div>'
    st.markdown(f'<div class="footer-bar">{f1}{f2}</div>',unsafe_allow_html=True)


def page_learning():
    is_arabic  = st.session_state["language"]=="Arabic"
    dir_attr   = 'rtl' if is_arabic else 'ltr'
    text_align = 'right' if is_arabic else 'left'
    TOTAL      = 6
    idx       = st.session_state["example_index"]

    # Inject JS listener for phishing click events from email buttons
    import streamlit.components.v1 as _comp
    _comp.html("""
    <script>
    window.addEventListener('message', function(e) {
        if (e.data && e.data.type === 'phishing_click') {
            // Send to Streamlit via query param change to trigger rerun
            const url = new URL(window.parent.location.href);
            url.searchParams.set('phishing_click', '1');
            window.parent.history.replaceState({}, '', url.toString());
        }
    });
    </script>
    """, height=0, width=0)

    if st.query_params.get("phishing_click") == "1":
        st.query_params.clear()
        st.session_state["page"] = "phishing_caught"
        st.rerun()

    if st.session_state.get("cache_version",0) < 20:
        st.session_state["emails"]={}
        st.session_state.pop("assess_emails", None)
        st.session_state["cache_version"]=20

    st.markdown(f"""<style>
#MainMenu,header,footer{{visibility:hidden;}}
.stApp{{background:radial-gradient(circle at top left,#0B2E68 0%,#020617 35%,#020617 100%);color:white;}}
.block-container{{max-width:1200px;padding-top:2rem;}}
.tutor-panel{{background:rgba(2,6,23,.7);border:1px solid rgba(37,99,235,.45);border-radius:16px;padding:1.4rem 1.5rem;direction:{dir_attr};text-align:{text_align};}}
.tutor-section{{font-size:1rem;font-weight:800;color:#F1F5F9;margin:1rem 0 .4rem;direction:{dir_attr};text-align:{text_align};}}
.tutor-text{{color:#94A3B8;font-size:.92rem;line-height:1.65;direction:{dir_attr};text-align:{text_align};}}
.tip-box{{background:rgba(16,185,129,.1);border:1px solid rgba(16,185,129,.35);border-radius:10px;padding:.8rem 1rem;color:#6EE7B7;font-size:.9rem;line-height:1.6;margin-top:.8rem;direction:{dir_attr};text-align:{text_align};}}
.stButton>button{{min-height:52px;background:linear-gradient(90deg,#0B4FA8,#0284C7) !important;color:white !important;border:none !important;font-weight:800 !important;border-radius:12px !important;font-size:1rem !important;}}
</style>""", unsafe_allow_html=True)

    if idx not in st.session_state["emails"]:
        with st.spinner(t("🤖 Generating phishing example...","🤖 جارٍ توليد مثال التصيد...")):
            st.session_state["emails"][idx] = generate_email(st.session_state["role"], idx, st.session_state["language"], st.session_state.get("difficulty", "medium"))
            st.rerun()

    email = st.session_state["emails"].get(idx,{})
    pct   = int((idx/TOTAL)*100)

    st.markdown(f"""
<div style="margin-bottom:1.5rem;direction:{dir_attr};">
  <div style="font-size:2.2rem;font-weight:900;color:#F8FAFC;margin-bottom:.4rem;text-align:{text_align};">
    {t("AI Tutor-Guided Learning Phase","مرحلة التعلم بتوجيه الذكاء الاصطناعي")}
  </div>
  <div style="width:100%;height:6px;background:rgba(37,99,235,.25);border-radius:99px;margin:.8rem 0;">
    <div style="height:6px;border-radius:99px;background:linear-gradient(90deg,#1EA7FF,#2563EB);width:{pct}%;transition:width .4s ease;"></div>
  </div>
  <div style="color:#7DD3FC;font-size:.95rem;font-weight:600;">
    {t(f"Example {idx+1} of {TOTAL}",f"مثال {idx+1} من {TOTAL}")}
  </div>
</div>""", unsafe_allow_html=True)

    if "error" in email:
        st.error(f"**{t('Error','خطأ')}:** " + _safe_error_text(email['error'], st.session_state['language']))
        if st.button(t("🔄 Try Again","🔄 حاول مرة أخرى"),key="retry_btn"):
            del st.session_state["emails"][idx]; st.rerun()
        return

    if is_arabic:
        col_tutor, col_email = st.columns([1,1.1],gap="large")
    else:
        col_email, col_tutor = st.columns([1.1,1],gap="large")

    with col_email:
        render_email_window(email, is_arabic, show_badges=True)

    with col_tutor:
        indicators    = email.get("indicators",[])
        indicators_html = ""
        for pos, ind in enumerate(indicators, 1):
            # Backward compatibility: providers/fallbacks may occasionally
            # return a plain string instead of the expected dictionary.
            if isinstance(ind, dict):
                ind_number = ind.get("number") or pos
                ind_title = ind.get("title") or ind.get("name") or ""
                ind_description = ind.get("description") or ind.get("detail") or ""
            else:
                ind_number = pos
                ind_title = str(ind or "")
                ind_description = ""
            row_dir = 'rtl' if is_arabic else 'ltr'
            pad     = 'padding-right:2rem;' if is_arabic else 'padding-left:2rem;'
            ta2     = 'right' if is_arabic else 'left'
            indicators_html += f"""
<div style="margin-bottom:1rem;">
  <div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.3rem;direction:{row_dir};">
    <span style="display:inline-flex;align-items:center;justify-content:center;width:24px;height:24px;border-radius:50%;background:#DC2626;color:white;font-size:.75rem;font-weight:800;flex-shrink:0;">{ind_number}</span>
    <span style="font-weight:700;color:#E2E8F0;font-size:.95rem;">{html_lib.escape(str(ind_title))}</span>
  </div>
  <div style="color:#94A3B8;font-size:.9rem;line-height:1.65;{pad};direction:{row_dir};text-align:{ta2};">{html_lib.escape(str(ind_description))}</div>
</div>"""

        st.markdown(f"""
<div class="tutor-panel">
  <div style="font-size:1.3rem;font-weight:900;color:#F8FAFC;margin-bottom:.2rem;">🎯 {t("AI Tutor Analysis","تحليل المعلم الذكي")}</div>
  <div style="color:#64748B;font-size:.85rem;margin-bottom:1.2rem;">{t("AI-guided phishing awareness","شرح توعوي بالتصيد")}</div>
  <div class="tutor-section">{t("What is suspicious?","ما هو المشبوه؟")}</div>
  {indicators_html}
  <div class="tutor-section">{t("Why is it risky?","لماذا هو خطير؟")}</div>
  <div class="tutor-text">{(email.get("why_risky") or "")}</div>
  <div class="tutor-section">💡 {t("Learning Tip","نصيحة تعليمية")}</div>
  <div class="tip-box">{(email.get("learning_tip") or "")}</div>
</div>""", unsafe_allow_html=True)

    st.markdown('<div style="height:1.5rem"></div>',unsafe_allow_html=True)
    bc,_ = st.columns([1,3])
    with bc:
        if idx<TOTAL-1:
            if st.button(t("Next Example →","← المثال التالي"),key="next_btn"):
                st.session_state["example_index"]+=1; st.rerun()
        else:
            if st.button(t("Complete Learning Phase →","← إتمام مرحلة التعلم"),key="complete_btn"):
                st.session_state["page"]="complete"; st.rerun()


def page_complete():
    is_arabic=st.session_state["language"]=="Arabic"; da='rtl' if is_arabic else 'ltr'
    def tc(e,a): return a if is_arabic else e
    st.markdown("""<style>#MainMenu,header,footer{visibility:hidden;}.stApp{background:radial-gradient(circle at top left,#0B2E68 0%,#020617 35%,#020617 100%);color:white;}.block-container{max-width:700px;padding-top:4rem;}.stButton>button{min-height:52px;background:linear-gradient(90deg,#0B4FA8,#0284C7) !important;color:white !important;border:none !important;font-weight:800 !important;border-radius:12px !important;font-size:1rem !important;width:100%;}</style>""",unsafe_allow_html=True)
    msg=tc("You have completed all 6 learning examples.\nYou are now ready to test your phishing awareness skills.","لقد أكملت جميع الأمثلة التعليمية الـ 6.\nأنت الآن جاهز لاختبار مستوى وعيك بالتصيد الإلكتروني.")
    st.markdown(f'<div style="text-align:center;padding:3rem 2rem;border:1px solid rgba(37,99,235,.45);border-radius:24px;background:linear-gradient(135deg,rgba(2,6,23,.96),rgba(8,47,73,.88));direction:{da};"><div style="font-size:4rem;margin-bottom:1rem;">🎓</div><div style="font-size:2rem;font-weight:900;color:#F8FAFC;margin-bottom:1rem;">{tc("Great job","ممتاز")}</div><div style="font-size:1.05rem;color:#DCEBFF;line-height:1.8;margin-bottom:2rem;white-space:pre-line;">{msg}</div></div>',unsafe_allow_html=True)
    st.markdown('<div style="height:1.5rem"></div>',unsafe_allow_html=True)
    if st.button(tc("Start Assessment →","← ابدأ الاختبار"),key="go_assessment"):
        if "assess_scenario_order" in st.session_state: del st.session_state["assess_scenario_order"]
        for k in list(st.session_state.keys()):
            if k.startswith(("assess_recipient_order_", "assess_scenario_order_", "assess_category_order_", "used_topics_")):
                st.session_state.pop(k, None)
        st.session_state.update({"page":"assessment","assess_index":0,"assess_emails":{},"assess_answers":{}})
        st.rerun()


def page_assessment():
    is_arabic=st.session_state["language"]=="Arabic"; da='rtl' if is_arabic else 'ltr'
    TOTAL=10; idx=st.session_state.get("assess_index",0)
    def ta(e,a): return a if is_arabic else e

    if "assess_pattern" not in st.session_state:
        p=[True]*5+[False]*5; random.shuffle(p); st.session_state["assess_pattern"]=p
    pattern=st.session_state["assess_pattern"]

    st.markdown(f"""<style>
#MainMenu,header,footer{{visibility:hidden;}}
.stApp{{background:radial-gradient(circle at top left,#0B2E68 0%,#020617 35%,#020617 100%);color:white;}}
.block-container{{max-width:960px;padding-top:2rem;}}
.stButton>button{{min-height:52px;font-weight:800 !important;border-radius:12px !important;font-size:1rem !important;background:rgba(15,23,42,.88) !important;color:white !important;border:1px solid rgba(37,99,235,.55) !important;width:100% !important;transition:.2s ease;}}
.stButton>button:hover{{background:linear-gradient(90deg,#0B4FA8,#0284C7) !important;border-color:#1EA7FF !important;}}
</style>""",unsafe_allow_html=True)

    if idx not in st.session_state["assess_emails"]:
        with st.spinner(ta("🤖 Generating scenario...","🤖 جارٍ توليد السيناريو...")):
            st.session_state["assess_emails"][idx]=generate_assess_email(st.session_state["role"], idx, pattern[idx], st.session_state["language"], st.session_state.get("difficulty", "medium"))
            st.rerun()

    email=st.session_state["assess_emails"].get(idx,{})
    answered_count = len(st.session_state.get("assess_answers", {}))
    pct=int((answered_count/TOTAL)*100)
    st.markdown(f"""
<div style="margin-bottom:1.5rem;direction:{da};">
  <div style="font-size:2rem;font-weight:900;color:#F8FAFC;margin-bottom:.4rem;">{ta("AI-Generated Assessment","مرحلة الاختبار")}</div>
  <div style="display:flex;justify-content:space-between;align-items:center;margin:.8rem 0 .3rem;">
    <div style="color:#7DD3FC;font-size:.95rem;font-weight:600;">{ta(f"Question {idx+1} of {TOTAL}",f"السؤال {idx+1} من {TOTAL}")}</div>
    <div style="color:#F59E0B;font-size:.9rem;font-weight:700;">{pct}%</div>
  </div>
  <div style="width:100%;height:8px;background:rgba(37,99,235,.2);border-radius:99px;overflow:hidden;">
    <div style="height:8px;border-radius:99px;background:linear-gradient(90deg,#F59E0B,#EF4444);width:{pct}%;transition:width .5s ease;"></div>
  </div>
  <div style="display:flex;justify-content:space-between;margin-top:.4rem;">
    {"".join(f'<div style="width:{100//TOTAL}%;height:4px;border-radius:2px;background:{"#F59E0B" if i < answered_count else "rgba(255,255,255,.1)"};margin:0 1px;"></div>' for i in range(TOTAL))}
  </div>
</div>""",unsafe_allow_html=True)

    if "error" in email:
        st.error(f"{t('Error','خطأ')}: " + _safe_error_text(email['error'], st.session_state['language']))
        if st.button(ta("🔄 Try Again","🔄 حاول مرة أخرى"),key="assess_retry"):
            del st.session_state["assess_emails"][idx]; st.rerun()
        return

    if is_arabic: col_action,col_email=st.columns([1,1.2],gap="large")
    else:         col_email,col_action=st.columns([1.2,1],gap="large")

    with col_email: render_email_window(email,is_arabic,show_badges=False)

    with col_action:
        q=ta("Is this email phishing or legitimate?","هل هذه الرسالة تصيد إلكتروني أم شرعية؟")
        st.markdown(f'<div style="background:rgba(2,6,23,.7);border:1px solid rgba(37,99,235,.45);border-radius:16px;padding:1.5rem;text-align:center;direction:{da};margin-bottom:1rem;"><div style="font-size:1.1rem;font-weight:800;color:#F1F5F9;">{q}</div></div>',unsafe_allow_html=True)

        answered=idx in st.session_state["assess_answers"]
        if not answered:
            c1,c2=st.columns(2)
            with c1:
                if st.button(f"🚨 {ta('Phishing','تصيد إلكتروني')}",key=f"ph_{idx}", use_container_width=True):
                    st.session_state["assess_answers"][idx]="phishing"; st.rerun()
            with c2:
                if st.button(f"✅ {ta('Legitimate','شرعية')}",key=f"lg_{idx}", use_container_width=True):
                    st.session_state["assess_answers"][idx]="legitimate"; st.rerun()
        else:
            ua=st.session_state["assess_answers"][idx]; ca2="phishing" if pattern[idx] else "legitimate"; ok=ua==ca2
            c="#6EE7B7" if ok else "#FCA5A5"; bg="rgba(16,185,129,.15)" if ok else "rgba(239,68,68,.15)"; br="rgba(16,185,129,.5)" if ok else "rgba(239,68,68,.5)"
            ic="✅" if ok else "❌"; lb=ta("Correct!","إجابة صحيحة!") if ok else ta("Incorrect!","إجابة خاطئة!")
            st.markdown(f'<div style="background:{bg};border:2px solid {br};border-radius:12px;padding:1rem;text-align:center;color:{c};font-weight:800;font-size:1.1rem;margin-bottom:1rem;">{ic} {lb}</div>',unsafe_allow_html=True)
            if idx<TOTAL-1:
                if st.button(ta("Next Question →","← السؤال التالي"),key=f"na_{idx}", use_container_width=True):
                    st.session_state["assess_index"]+=1; st.rerun()
            else:
                if st.button(ta("View Results →","← عرض النتائج"),key="vr", use_container_width=True):
                    st.session_state["page"]="results"; st.rerun()


def page_results():
    is_arabic=st.session_state["language"]=="Arabic"; da='rtl' if is_arabic else 'ltr'; TOTAL=10
    def tr(e,a): return a if is_arabic else e
    st.markdown("""<style>#MainMenu,header,footer{visibility:hidden;}.stApp{background:radial-gradient(circle at top left,#0B2E68 0%,#020617 35%,#020617 100%);color:white;}.block-container{max-width:900px;padding-top:2rem;}.stButton>button{min-height:52px;background:linear-gradient(90deg,#0B4FA8,#0284C7) !important;color:white !important;border:none !important;font-weight:800 !important;border-radius:12px !important;}div[style*="direction:rtl"]{text-align:right;}</style>""",unsafe_allow_html=True)
    answers=st.session_state.get("assess_answers",{}); pattern=st.session_state.get("assess_pattern",[True]*5+[False]*5); emails=st.session_state.get("assess_emails",{})
    score=sum(1 for i in range(TOTAL) if answers.get(i)==("phishing" if pattern[i] else "legitimate"))
    pct=int((score/TOTAL)*100)
    sc="#10B981" if pct>=80 else "#F59E0B" if pct>=60 else "#EF4444"
    sl=tr("Excellent 🎉","ممتاز 🎉") if pct>=80 else tr("Good job 👍","جيد 👍") if pct>=60 else tr("Keep practicing 💪","استمر في التدريب 💪")
    st.markdown(f'<div style="text-align:center;padding:2rem;border:1px solid rgba(37,99,235,.45);border-radius:24px;background:linear-gradient(135deg,rgba(2,6,23,.96),rgba(8,47,73,.88));margin-bottom:2rem;direction:{da};"><div style="font-size:1.5rem;font-weight:900;color:#F8FAFC;margin-bottom:1rem;">{tr("Your Results","نتائجك")}</div><div style="font-size:4rem;font-weight:900;color:{sc};">{score}/{TOTAL}</div><div style="font-size:1.2rem;color:{sc};font-weight:700;">{sl}</div><div style="color:#94A3B8;margin-top:.5rem;">{tr(f"You answered {score} of {TOTAL} correctly ({pct}%)",f"أجبت على {score} من {TOTAL} بشكل صحيح ({pct}٪)")}</div></div>',unsafe_allow_html=True)
    st.markdown(f'<div style="font-size:1.3rem;font-weight:900;color:#F8FAFC;margin-bottom:1rem;direction:{da};">📋 {tr("Review","مراجعة الإجابات")}</div>',unsafe_allow_html=True)
    for i in range(TOTAL):
        em=emails.get(i,{})
        if not em or "error" in em: continue
        ua=answers.get(i,""); ca2="phishing" if pattern[i] else "legitimate"; ok=ua==ca2
        bc2="rgba(16,185,129,.5)" if ok else "rgba(239,68,68,.5)"; bg2="rgba(16,185,129,.05)" if ok else "rgba(239,68,68,.05)"
        ri="✅" if ok else "❌"; tl=tr("Phishing","تصيد") if pattern[i] else tr("Legitimate","شرعية"); ic="🚨" if pattern[i] else "✅"
        exp=re.sub(r'<[^>]+>','',(em.get("explanation") or ""))
        # FIX: bidi issue — Arabic explanations often embed Latin/domain
        # substrings (e.g. "hosp1tal-clinic.org") in the middle of a
        # sentence, sometimes inside parentheses with Arabic words on
        # both sides. The browser's bidi algorithm can let that embedded
        # LTR run "leak" and flip the visual order of the surrounding
        # Arabic words, especially at a line wrap. Wrapping each such run
        # in a <bdi dir="ltr"> tag creates a true isolated bidi run, which
        # is the standards-correct fix and keeps the rest of the Arabic
        # sentence in correct right-to-left reading order regardless of
        # where the line wraps.
        if is_arabic:
            exp = re.sub(
                r'[A-Za-z][A-Za-z0-9\.\-_/:@]*[A-Za-z0-9]',
                lambda m: f'<bdi dir="ltr">{m.group(0)}</bdi>',
                exp
            )
        st.markdown(f'<div style="border:1px solid {bc2};border-radius:14px;padding:1.2rem 1.5rem;background:{bg2};margin-bottom:1rem;direction:{da};"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:.5rem;flex-wrap:wrap;gap:.5rem;"><span style="font-weight:800;color:#E2E8F0;">{ri} {tr(f"Q{i+1}",f"س{i+1}")} — {html_lib.escape((em.get("subject") or ""))}</span><span style="background:{"rgba(239,68,68,.2)" if pattern[i] else "rgba(16,185,129,.2)"};color:{"#FCA5A5" if pattern[i] else "#6EE7B7"};padding:.2rem .8rem;border-radius:99px;font-size:.85rem;font-weight:700;">{ic} {tl}</span></div><div dir="{da}" style="color:#94A3B8;font-size:.9rem;line-height:1.6;direction:{da};text-align:{"right" if is_arabic else "left"};unicode-bidi:embed;">{exp}</div></div>',unsafe_allow_html=True)
    st.markdown('<div style="height:1rem"></div>',unsafe_allow_html=True)
    if st.button(tr("Go to Report →","← الانتقال للتقرير"),key="go_report"):
        st.session_state["page"]="report"; st.rerun()


def page_report():
    is_arabic=st.session_state["language"]=="Arabic"; da='rtl' if is_arabic else 'ltr'; TOTAL=10
    def tp(e,a): return a if is_arabic else e
    st.markdown(f"""<style>#MainMenu,header,footer{{visibility:hidden;}}.stApp{{background:radial-gradient(circle at top left,#0B2E68 0%,#020617 35%,#020617 100%);color:white;}}.block-container{{max-width:900px;padding-top:2rem;}}.stButton>button{{min-height:52px !important;font-weight:800 !important;border-radius:12px !important;background:rgba(15,23,42,.88) !important;color:white !important;border:1px solid rgba(37,99,235,.55) !important;width:100% !important;}}.stButton>button:hover{{background:linear-gradient(90deg,#0B4FA8,#0284C7) !important;border-color:#1EA7FF !important;}}div[style*="direction:rtl"]{{text-align:right;}}</style>""",unsafe_allow_html=True)
    answers=st.session_state.get("assess_answers",{}); pattern=st.session_state.get("assess_pattern",[True]*5+[False]*5)
    role=(st.session_state.get("role") or ""); lang=st.session_state.get("language","English")
    score=pc=lc=0; pt=sum(1 for p in pattern if p); lt=TOTAL-pt
    for i in range(TOTAL):
        ca2="phishing" if pattern[i] else "legitimate"
        if answers.get(i)==ca2:
            score+=1
            if pattern[i]: pc+=1
            else: lc+=1
    pct=int((score/TOTAL)*100); pp=int(pc/pt*100) if pt else 0; lp=int(lc/lt*100) if lt else 0
    aw="🥇" if pct>=80 else "🥈" if pct>=60 else "🥉"
    sc2="#10B981" if pct>=80 else "#F59E0B" if pct>=60 else "#EF4444"
    awl=tp("High","عالي") if pct>=80 else tp("Moderate","متوسط") if pct>=60 else tp("Needs Improvement","يحتاج تحسين")
    strengths=[]; areas=[]
    if pp>=70: strengths.append(tp("Good at identifying phishing emails","جيد في تحديد رسائل التصيد"))
    else: areas.append(tp("Review phishing indicators more carefully","راجع مؤشرات التصيد بعناية أكبر"))
    if lp>=70: strengths.append(tp("Good at identifying legitimate emails","جيد في تمييز الرسائل الشرعية"))
    else: areas.append(tp("Be cautious not to flag legitimate emails","احذر من تصنيف الرسائل الشرعية كتصيد"))
    # Personalized recommendations based on the trainee's actual error pattern.
    recs=[]
    if lp < 70:
        recs.append(tp(
            "Do not flag an email only because its subject sounds urgent. First check the official domain, whether it requests credentials, and whether it uses an external link.",
            "لا تصنّف الرسالة كتصيد لمجرد أن عنوانها يبدو عاجلاً. افحص أولاً النطاق الرسمي، وهل تطلب بيانات دخول، وهل تستخدم رابطاً خارجياً."
        ))
    if pp < 70:
        recs.append(tp(
            "Pause before acting on credential, PIN, OTP, or verification requests—especially when they arrive through an unexpected link.",
            "توقّف قبل تنفيذ أي طلب لكلمة مرور أو رقم موظف أو رمز تحقق، خصوصاً عندما يصل عبر رابط غير متوقع."
        ))
    if pp >= 70 and lp >= 70:
        recs.append(tp(
            "Keep using multiple clues together: sender domain, requested action, channel, link destination, and message context.",
            "استمر في جمع أكثر من دليل معاً: نطاق المرسل، الإجراء المطلوب، القناة، وجهة الرابط، وسياق الرسالة."
        ))
    recs.extend([
        tp("Verify unusual requests through the hospital's official system or a trusted internal contact.","تحقق من الطلبات غير المعتادة عبر نظام المستشفى الرسمي أو جهة داخلية موثوقة."),
        tp("When uncertain, do not reply with sensitive information and do not use the email link.","عند الشك، لا ترسل معلومات حساسة ولا تستخدم الرابط الموجود في البريد.")
    ])
    user_name  = (st.session_state.get("user_name") or "")
    user_email = (st.session_state.get("user_email") or "")
    name_line  = f'<div style="color:#F8FAFC;font-size:1rem;font-weight:700;margin-bottom:.2rem;">{html_lib.escape(user_name)}</div>' if user_name else ""
    email_line = f'<div style="color:#64748B;font-size:.8rem;margin-bottom:.3rem;">{html_lib.escape(user_email)}</div>' if user_email else ""
    st.markdown(f'<div style="text-align:center;padding:2rem;border:1px solid rgba(37,99,235,.45);border-radius:24px;background:linear-gradient(135deg,rgba(2,6,23,.96),rgba(8,47,73,.88));margin-bottom:1.5rem;direction:{da};">'
                f'<div style="font-size:1.6rem;font-weight:900;color:#F8FAFC;margin-bottom:.5rem;">📊 {tp("Your Performance Report","تقرير أدائك")}</div>'
                f'{name_line}{email_line}'
                f'<div style="color:#7DD3FC;font-size:.95rem;">{tp(f"Role: {role}","الدور: "+role)}</div></div>',
                unsafe_allow_html=True)
    c1,c2,c3=st.columns(3)
    card = "border:1px solid rgba(37,99,235,.45);border-radius:16px;padding:1.5rem;text-align:center;background:rgba(2,6,23,.6);direction:{da};height:200px;display:flex;flex-direction:column;justify-content:center;align-items:center;box-sizing:border-box;"
    with c1: st.markdown(f'<div style="{card.format(da=da)}"><div style="color:#94A3B8;font-size:.85rem;margin-bottom:.4rem;">{tp("Overall Score","النتيجة الإجمالية")}</div><div style="font-size:2.5rem;font-weight:900;color:{sc2};">{score}/{TOTAL}</div><div style="color:{sc2};font-size:.9rem;">{pct}%</div></div>',unsafe_allow_html=True)
    with c2: st.markdown(f'<div style="{card.format(da=da)}"><div style="color:#94A3B8;font-size:.85rem;margin-bottom:.4rem;">{tp("Awareness Level","مستوى الوعي")}</div><div style="font-size:2.5rem;">{aw}</div><div style="color:{sc2};font-weight:700;font-size:.95rem;">{awl}</div></div>',unsafe_allow_html=True)
    with c3: st.markdown(f'<div style="{card.format(da=da)}"><div style="color:#94A3B8;font-size:.85rem;margin-bottom:.3rem;">{tp("Detection Rate","معدل الاكتشاف")}</div><div style="font-size:1rem;font-weight:700;color:#FCA5A5;margin-bottom:.3rem;">🚨 {tp("Phishing detected","التصيد المكتشف")}: {pp}%</div><div style="font-size:1rem;font-weight:700;color:#6EE7B7;">✅ {tp("Legitimate identified","الشرعية المميزة")}: {lp}%</div></div>',unsafe_allow_html=True)
    st.markdown('<div style="height:1rem"></div>',unsafe_allow_html=True)
    s1,s2=st.columns(2)
    with s1:
        si="".join([f'<div style="color:#6EE7B7;margin-bottom:.4rem;text-align:{"right" if is_arabic else "left"};">✅ {s}</div>' for s in strengths]) or f'<div style="color:#94A3B8;">{tp("Keep practicing","استمر في التدريب")}</div>'
        st.markdown(f'<div style="border:1px solid rgba(16,185,129,.35);border-radius:14px;padding:1.2rem;background:rgba(16,185,129,.05);direction:{da};text-align:{"right" if is_arabic else "left"};min-height:160px;"><div style="font-weight:800;color:#F1F5F9;margin-bottom:.8rem;text-align:{"right" if is_arabic else "left"};">💪 {tp("Strengths","نقاط القوة")}</div>{si}</div>',unsafe_allow_html=True)
    with s2:
        ai2="".join([f'<div style="color:#FCA5A5;margin-bottom:.4rem;text-align:{"right" if is_arabic else "left"};">⚠️ {a}</div>' for a in areas]) or f'<div style="color:#94A3B8;">{tp("Great work!","عمل رائع!")}</div>'
        st.markdown(f'<div style="border:1px solid rgba(239,68,68,.35);border-radius:14px;padding:1.2rem;background:rgba(239,68,68,.05);direction:{da};text-align:{"right" if is_arabic else "left"};min-height:160px;"><div style="font-weight:800;color:#F1F5F9;margin-bottom:.8rem;text-align:{"right" if is_arabic else "left"};">📈 {tp("Areas to Improve","مجالات التحسين")}</div>{ai2}</div>',unsafe_allow_html=True)
    st.markdown('<div style="height:1rem"></div>',unsafe_allow_html=True)
    ri="".join([f'<div style="color:#DCEBFF;margin-bottom:.5rem;text-align:{"right" if is_arabic else "left"};">📌 {r}</div>' for r in recs])
    st.markdown(f'<div style="border:1px solid rgba(37,99,235,.45);border-radius:14px;padding:1.2rem 1.5rem;background:rgba(2,6,23,.6);margin-bottom:1.5rem;direction:{da};text-align:{"right" if is_arabic else "left"};"><div style="font-weight:800;color:#F1F5F9;margin-bottom:.8rem;text-align:{"right" if is_arabic else "left"};">💡 {tp("Recommendations","التوصيات")}</div>{ri}</div>',unsafe_allow_html=True)
    st.markdown(f'<div style="text-align:center;padding:.8rem;border:1px solid rgba(37,99,235,.3);border-radius:10px;background:rgba(37,99,235,.08);color:#7DD3FC;margin-bottom:1.5rem;">⭐ {tp("Your awareness helps keep your organization safe","وعيك يساهم في حماية مؤسستك")}</div>',unsafe_allow_html=True)
    if st.button(tp("Retake Training","إعادة التدريب من البداية"),key="retake", use_container_width=True):
        # FIX 6: مسح كامل لكل session data لضمان تنوع المحتوى
        keys_to_clear = [
            "page","example_index","emails","assess_index",
            "assess_emails","assess_answers","assess_pattern",
            "cache_version","role","scenario_order",
            "assess_scenario_order","difficulty",
            "user_name","user_email",
            "lang_explicitly_chosen","diff_explicitly_chosen",
            "login_mode","assess_index",
        ]
        for k in keys_to_clear:
            st.session_state.pop(k, None)
        # EN: also drop the dynamic per-role scenario/recipient random-order
        # picks (scenario_order_*, recipient_order_*, assess_*_order_*) so
        # "Retake Training" actually reshuffles them instead of reusing the
        # same session pick — otherwise picking the same role again would
        # silently replay the same scenario/recipient order as before.
        # AR: نمسح كمان اختيارات الترتيب العشوائي الديناميكية لكل دور
        # (scenario_order_*, recipient_order_*, assess_*_order_*) حتى زر
        # "إعادة التدريب" يعيد خلطها فعليًا بدل إعادة استخدام نفس اختيار
        # الجلسة القديم — لو اخترتِ نفس الدور بعد إعادة المحاولة.
        for k in list(st.session_state.keys()):
            if k.startswith(("scenario_order_", "recipient_order_", "assess_recipient_order_", "assess_scenario_order_",
                              "category_order_", "assess_category_order_", "used_topics_")):
                st.session_state.pop(k, None)
        # تجديد الـ cache_version لإجبار النموذج على توليد محتوى جديد
        st.session_state["cache_version"] = int(__import__("time").time()) % 99999 + 20
        st.rerun()


def page_login():
    is_arabic = st.session_state["language"] == "Arabic"
    da = 'rtl' if is_arabic else 'ltr'
    mode = st.session_state.get("login_mode","login")
    def tl(e,a): return a if is_arabic else e

    is_reg = mode == "register"
    page_title = tl("Create Account","إنشاء حساب") if is_reg else tl("Welcome Back","مرحباً بك")
    page_sub   = tl("Enter your details to get started","أدخل بياناتك للبدء") if is_reg else tl("Enter your details to personalise your experience","أدخل بياناتك لتخصيص تجربتك التدريبية")
    page_icon  = "✨" if is_reg else "👤"

    st.markdown(f"""<style>
#MainMenu,header,footer{{visibility:hidden;}}
.stApp{{background:radial-gradient(circle at top left,#0B2E68 0%,#020617 35%,#020617 100%);color:white;}}
.block-container{{max-width:480px;padding-top:4rem;}}
.stTextInput>div>div>input{{background:rgba(15,23,42,.88) !important;color:white !important;border:1px solid rgba(37,99,235,.55) !important;border-radius:12px !important;min-height:48px;direction:{da};font-size:.95rem !important;}}
.stTextInput label{{color:#94A3B8 !important;font-size:.85rem !important;}}
.stButton>button{{width:100% !important;min-height:48px !important;max-height:48px !important;font-weight:700 !important;border-radius:12px !important;font-size:.9rem !important;padding:0 16px !important;line-height:48px !important;}}
div[data-testid="stHorizontalBlock"] > div:first-child .stButton>button{{background:rgba(15,23,42,.88) !important;color:#EAF4FF !important;border:1px solid rgba(37,99,235,.55) !important;}}
div[data-testid="stHorizontalBlock"] > div:last-child .stButton>button{{background:rgba(15,23,42,.88) !important;color:#EAF4FF !important;border:1px solid rgba(37,99,235,.55) !important;}}
</style>""", unsafe_allow_html=True)

    st.markdown(f"""
<div style="text-align:center;padding:2.5rem 2rem 2rem;border:1px solid rgba(37,99,235,.45);border-radius:24px;background:linear-gradient(135deg,rgba(2,6,23,.96),rgba(8,47,73,.88));direction:{da};margin-bottom:1.5rem;">
  <div style="font-size:2.8rem;margin-bottom:.8rem;">{page_icon}</div>
  <div style="font-size:1.4rem;font-weight:900;color:#F8FAFC;margin-bottom:.4rem;">{page_title}</div>
  <div style="font-size:.9rem;color:#94A3B8;">{page_sub}</div>
</div>""", unsafe_allow_html=True)

    if is_arabic:
        st.markdown('<style>.stTextInput label{direction:rtl;text-align:right;display:block;}.stTextInput input{text-align:right;direction:rtl;}</style>', unsafe_allow_html=True)

    user_name  = st.text_input(tl("Full name","الاسم الكامل"), value=(st.session_state.get("user_name") or ""), placeholder=tl("e.g. Dr. Sarah Al-Mutairi","مثال: د. سارة المطيري"))
    user_email = st.text_input(tl("Email address","البريد الإلكتروني"), value=(st.session_state.get("user_email") or ""), placeholder="name@hospital.org")

    st.markdown('<div style="height:.8rem;"></div>', unsafe_allow_html=True)
    st.markdown("""<style>div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button{height:48px !important;min-height:48px !important;max-height:48px !important;padding-top:0 !important;padding-bottom:0 !important;display:flex !important;align-items:center !important;justify-content:center !important;box-sizing:border-box !important;}</style>""", unsafe_allow_html=True)

    c1, c2 = st.columns([1,1])
    with c1:
        if st.button(tl("← Back","← رجوع"), key="login_back", use_container_width=True):
            st.session_state["page"] = "home"; st.rerun()
    with c2:
        if st.button(tl("Continue","متابعة"), key="login_continue", use_container_width=True):
            email_pattern = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')
            if not user_name.strip():
                st.warning(tl("⚠️ Please enter your full name.","⚠️ يرجى إدخال اسمك الكامل"))
            elif not user_email.strip():
                st.warning(tl("⚠️ Please enter your email address.","⚠️ يرجى إدخال بريدك الإلكتروني"))
            elif not email_pattern.match(user_email.strip()):
                st.warning(tl("⚠️ Please enter a valid email address (e.g. name@hospital.org).","⚠️ يرجى إدخال بريد إلكتروني صحيح مثل: name@hospital.org"))
            else:
                st.session_state["user_name"]  = user_name.strip()
                st.session_state["user_email"] = user_email.strip()
                st.session_state["page"] = "home"; st.rerun()


# ══════════════════════════════════════════════════════════════
# ADMIN PANEL — مخفي، يظهر فقط عبر ?admin=true
# ══════════════════════════════════════════════════════════════
def page_admin():
    def get_secret(key):
        try:
            return st.secrets[key]
        except Exception:
            return os.environ.get(key, "")

    ADMIN_PASSWORD = get_secret("ADMIN_PASSWORD") or "admin2025"

    # ── اللغة المختارة من المستخدم في الصفحة الرئيسية ──
    _lang = st.session_state.get("language", "English")
    _is_ar = (_lang == "Arabic")

    _T = {
        "admin_title":      {"en": "Admin Panel",                          "ar": "لوحة التحكم"},
        "admin_subtitle":   {"en": "Researcher Access Only",               "ar": "للباحثين فقط"},
        "pwd_placeholder":  {"en": "Enter admin password",                 "ar": "أدخل كلمة سر الإدارة"},
        "login_btn":        {"en": "🔓 Login",                             "ar": "🔓 دخول"},
        "back_btn":         {"en": "← Back to App",                       "ar": "→ رجوع للتطبيق"},
        "wrong_pwd":        {"en": "❌ Incorrect password",                "ar": "❌ كلمة السر غير صحيحة"},
        "panel_title":      {"en": "🔬 Admin Research Panel",              "ar": "🔬 لوحة البحث الإدارية"},
        "authenticated":    {"en": "● Authenticated",                      "ar": "● تم تسجيل الدخول"},
        "tab_provider":     {"en": "⚙️ Provider Control",                  "ar": "⚙️ التحكم بالمزوّد"},
        "tab_score":        {"en": "📊 Score Card",                       "ar": "📊 بطاقة التقييم"},
        "tab_manual":       {"en": "👍 Manual Ratings",                   "ar": "👍 التقييم اليدوي"},
        "select_provider":  {"en": "Select Active AI Provider",            "ar": "اختر مزوّد الذكاء الاصطناعي النشط"},
        "active":           {"en": "ACTIVE",                               "ar": "نشط"},
        "activate_btn":     {"en": "Activate",                             "ar": "تفعيل"},
        "api_status":       {"en": "API Keys Status",                      "ar": "حالة مفاتيح API"},
        "not_set":          {"en": "Pending setup",                          "ar": "بانتظار الإضافة"},
        "difficulty_lvl":   {"en": "Difficulty Level",                     "ar": "مستوى الصعوبة"},
        "language_lbl":     {"en": "Language",                             "ar": "اللغة"},
        "set_btn":          {"en": "Set",                                  "ar": "تطبيق"},
        "easy":             {"en": "Easy",                                 "ar": "سهل"},
        "medium":           {"en": "Medium",                               "ar": "متوسط"},
        "hard":             {"en": "Hard",                                 "ar": "صعب"},
        "clear_cache":      {"en": "🔄 Clear All Cached Emails (Force Regenerate)", "ar": "🔄 مسح كل الرسائل المخزّنة (توليد من جديد)"},
        "cache_cleared":    {"en": "✅ Cache cleared — next generation will produce fresh content", "ar": "✅ تم مسح الذاكرة المؤقتة — التوليد القادم سينتج محتوى جديد"},
        "logout_btn":       {"en": "🚪 Logout",                            "ar": "🚪 تسجيل خروج"},
        "score_title":      {"en": "📊 Comparison Score Card",             "ar": "📊 بطاقة المقارنة"},
        "metric_col":       {"en": "Metric",                               "ar": "المعيار"},
        "speed_metric":     {"en": "⚡ Avg Speed (s)",                     "ar": "⚡ متوسط السرعة (ث)"},
        "json_metric":      {"en": "✅ JSON Success Rate",                 "ar": "✅ نسبة نجاح JSON"},
        "error_metric":     {"en": "🚫 Error Rate",                        "ar": "🚫 نسبة الأخطاء"},
        "diversity_metric": {"en": "🔄 Unique Responses",                  "ar": "🔄 الردود الفريدة"},
        "quality_metric":   {"en": "🎯 Quality",                          "ar": "🎯 الجودة"},
        "difficulty_metric":{"en": "📊 Difficulty Level",                 "ar": "📊 مستوى الصعوبة"},
        "arabic_metric":    {"en": "🌐 Arabic Quality",                   "ar": "🌐 جودة العربية"},
        "medical_metric":   {"en": "🏥 Medical Realism",                  "ar": "🏥 الواقعية الطبية"},
        "auto_manual_note": {"en": "Auto-tracked: Speed, JSON, Errors, Diversity | Manual: Quality, Difficulty, Arabic, Medical",
                              "ar": "تلقائي: السرعة، JSON، الأخطاء، التنوع | يدوي: الجودة، الصعوبة، العربية، الطبي"},
        "reset_metrics":    {"en": "🗑️ Reset All Metrics",                "ar": "🗑️ إعادة ضبط كل المعايير"},
        "metrics_reset":    {"en": "✅ All metrics reset",                 "ar": "✅ تم إعادة ضبط كل المعايير"},
        "rate_title":       {"en": "Rate the last generated content",      "ar": "قيّم آخر محتوى تم توليده"},
        "active_provider":  {"en": "Active provider",                     "ar": "المزوّد النشط"},
        "saved_permanently": {"en": "saved permanently",                  "ar": "محفوظ بشكل دائم"},
        "quality_label":    {"en": "🎯 Model Quality",                    "ar": "🎯 جودة النموذج"},
        "quality_desc":     {"en": "How accurate and realistic is the phishing email?",
                              "ar": "ما مدى دقة وواقعية رسالة التصيد؟"},
        "diff_acc_label":   {"en": "📊 Difficulty Accuracy",              "ar": "📊 دقة الصعوبة"},
        "diff_acc_desc":    {"en": "Does the difficulty level (Easy/Medium/Hard) feel right?",
                              "ar": "هل مستوى الصعوبة (سهل/متوسط/صعب) مناسب؟"},
        "arabic_label":     {"en": "🌐 Arabic Quality",                   "ar": "🌐 جودة اللغة العربية"},
        "arabic_desc":      {"en": "How good is the Arabic language quality? (skip if English)",
                              "ar": "ما مدى جودة اللغة العربية؟ (تخطّى إذا إنجليزي)"},
        "medical_label":    {"en": "🏥 Medical Realism",                  "ar": "🏥 الواقعية الطبية"},
        "medical_desc":     {"en": "How realistic is the healthcare/hospital context?",
                              "ar": "ما مدى واقعية السياق الطبي/الصحي؟"},
        "note_label":       {"en": "📝 Note (optional)",                  "ar": "📝 ملاحظة (اختياري)"},
        "note_placeholder": {"en": "Any observation about this provider...",
                              "ar": "أي ملاحظة حول هذا المزوّد..."},
        "save_btn":         {"en": "💾 Save Ratings",                     "ar": "💾 حفظ التقييم"},
        "ratings_saved":    {"en": "✅ Ratings saved for",                 "ar": "✅ تم حفظ التقييم لـ"},
        "quick_rating":     {"en": "Quick Rating",                        "ar": "تقييم سريع"},
        "good_btn":         {"en": "👍 Good (4/5)",                   "ar": "👍 جيد (4/5)"},
        "avg_btn":          {"en": "😐 Average (3/5)",                   "ar": "😐 متوسط (3/5)"},
        "poor_btn":         {"en": "👎 Poor (2/5)",                   "ar": "👎 ضعيف (2/5)"},
        "saved_45":         {"en": "✅ Saved this cycle — Overall 4/5",         "ar": "✅ تم حفظ هذي الدورة — الانطباع العام 4/5"},
        "saved_35":         {"en": "✅ Saved this cycle — Overall 3/5",         "ar": "✅ تم حفظ هذي الدورة — الانطباع العام 3/5"},
        "saved_25":         {"en": "✅ Saved this cycle — Overall 2/5",         "ar": "✅ تم حفظ هذي الدورة — الانطباع العام 2/5"},
        "rating_history":   {"en": "Rating History for",                  "ar": "سجل التقييم لـ"},
        "avg_label":        {"en": "avg",                                 "ar": "المتوسط"},
        "ratings_label":    {"en": "ratings",                             "ar": "تقييم"},
    }

    def T(key):
        return _T.get(key, {}).get("ar" if _is_ar else "en", key)

    _dir = 'rtl' if _is_ar else 'ltr'
    _align = 'right' if _is_ar else 'left'
    _align_opp = 'left' if _is_ar else 'right'
    _flex_dir = 'row-reverse' if _is_ar else 'row'

    st.markdown(f"""<style>
#MainMenu,header,footer{{visibility:hidden;}}
.stApp{{background:radial-gradient(circle at top left,#0B1A0B 0%,#020617 40%,#020617 100%);color:white;direction:{_dir};}}
.block-container{{max-width:1200px;padding-top:1.5rem;direction:{_dir};text-align:{_align};}}
[data-testid="stMarkdownContainer"]{{direction:{_dir};text-align:{_align};}}
[data-testid="stMarkdownContainer"] > div{{direction:{_dir};text-align:{_align};}}
.block-container div,.block-container span,.block-container p,.block-container label{{
    direction:{_dir};
}}
[data-baseweb="tab-list"]{{direction:{_dir};}}
[data-baseweb="tab-border"]{{direction:{_dir};}}
.stTabs [data-baseweb="tab"]{{direction:{_dir};}}
div[style*="justify-content:space-between"]{{flex-direction:{_flex_dir}!important;}}
.stButton>button{{
    min-height:44px;
    font-weight:700!important;
    border-radius:10px!important;
    background:rgba(15,23,42,.8)!important;
    color:#94A3B8!important;
    border:1px solid rgba(37,99,235,.35)!important;
    direction:{_dir};
}}
.stButton>button:hover{{
    background:rgba(11,79,168,.3)!important;
    color:#E2E8F0!important;
    border-color:#1EA7FF!important;
}}
a[download],
a[download]:link,
a[download]:visited,
div[data-testid="stDownloadButton"] button,
div[data-testid="stDownloadButton"] a{{
    font-weight:700!important;
    border-radius:10px!important;
    background:rgba(15,23,42,.8)!important;
    color:#94A3B8!important;
    border:1px solid rgba(37,99,235,.35)!important;
    text-decoration:none!important;
}}
a[download]:hover,
a[download]:focus,
a[download]:active,
div[data-testid="stDownloadButton"] button:hover,
div[data-testid="stDownloadButton"] button:focus,
div[data-testid="stDownloadButton"] button:active,
div[data-testid="stDownloadButton"] a:hover,
div[data-testid="stDownloadButton"] a:focus,
div[data-testid="stDownloadButton"] a:active{{
    background:rgba(11,79,168,.3)!important;
    color:#E2E8F0!important;
    border-color:#1EA7FF!important;
    text-decoration:none!important;
}}
button[kind="primary"]{{
    background:rgba(37,99,235,.18)!important;
    color:#93C5FD!important;
    border:1px solid rgba(37,99,235,.5)!important;
}}
button[kind="primary"]:hover{{
    background:rgba(37,99,235,.28)!important;
    color:#BFDBFE!important;
}}
.stButton>button:disabled,.stButton>button[disabled]{{
    background:rgba(37,99,235,.12)!important;
    color:#60A5FA!important;
    border:1px solid rgba(37,99,235,.4)!important;
    opacity:1!important;
    cursor:default!important;
}}
.stTextInput input{{
    direction:{_dir};
    text-align:{_align};
    background:transparent!important;
    color:#E2E8F0!important;
    border:none!important;
    box-shadow:none!important;
}}
.stTextInput input::placeholder{{color:#6B7280!important;}}
.stTextInput>div>div,
.stTextInput div[data-baseweb="base-input"],
.stTextInput div[data-baseweb="input"]{{
    background:rgba(15,23,42,.80)!important;
    border:2px solid rgba(148,163,184,.75)!important;
    border-radius:12px!important;
    box-shadow:0 0 0 3px rgba(148,163,184,.08)!important;
}}
.stTextInput>div>div:focus-within,
.stTextInput div[data-baseweb="base-input"]:focus-within,
.stTextInput div[data-baseweb="input"]:focus-within{{
    border:2px solid #60A5FA!important;
    box-shadow:0 0 0 3px rgba(96,165,250,.20)!important;
}}
div[data-baseweb="select"] *{{color:#EAF4FF!important;}}
div[data-baseweb="select"] > div{{background:rgba(15,23,42,.78)!important;border:1px solid rgba(37,99,235,.55)!important;}}
.stSelectbox label{{color:#EAF4FF!important;}}
</style>""", unsafe_allow_html=True)

    def _div(content, extra=""):
        """صندوق نصي يحترم اتجاه اللغة الحالية"""
        return f'<div dir="{_dir}" style="text-align:{_align};{extra}">{content}</div>'

    # ── Authentication ──────────────────────────────────────────
    if not st.session_state.get("admin_authenticated", False):
        st.markdown(f"""
<div style="max-width:420px;margin:6rem auto;padding:2.5rem;
     border:1px solid rgba(34,197,94,.4);border-radius:20px;
     background:linear-gradient(135deg,rgba(2,6,23,.97),rgba(4,20,4,.9));
     text-align:center;">
  <div style="font-size:2.5rem;margin-bottom:.5rem;">🔐</div>
  <div style="font-size:1.3rem;font-weight:900;color:#F0FDF4;margin-bottom:.3rem;">{T('admin_title')}</div>
  <div style="font-size:.85rem;color:#6B7280;margin-bottom:1.5rem;">{T('admin_subtitle')}</div>
</div>""", unsafe_allow_html=True)
        pwd = st.text_input(T('pwd_placeholder'), type="password", placeholder=T('pwd_placeholder'),
                            label_visibility="collapsed")
        col1, col2 = st.columns([1,1])
        with col1:
            if st.button(T('login_btn'), use_container_width=True):
                if pwd == ADMIN_PASSWORD:
                    st.session_state["admin_authenticated"] = True
                    st.rerun()
                else:
                    st.error(T('wrong_pwd'))
        with col2:
            if st.button(T('back_btn'), use_container_width=True):
                st.query_params.clear()
                st.session_state["page"] = "home"
                st.rerun()
        return

    # ══════════════════════════════════════════════════════════
    # MAIN ADMIN PANEL (after login)
    # ══════════════════════════════════════════════════════════
    st.markdown(f"""
<div dir="{_dir}" style="display:flex;justify-content:space-between;align-items:center;
     padding:.8rem 1.2rem;border:1px solid rgba(34,197,94,.35);border-radius:14px;
     background:rgba(4,20,4,.6);margin-bottom:1.5rem;">
  <div style="font-size:1.3rem;font-weight:900;color:#F0FDF4;">{T('panel_title')}</div>
  <div style="font-size:.8rem;color:#4ADE80;">{T('authenticated')}</div>
</div>""", unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs([T('tab_provider'), T('tab_score'), T('tab_manual'), "🐞 Debug Log"])

    _persist_pk = st.session_state.get("ai_provider", load_persistent_provider("openai"))
    _persist_labels = {
        "groq":      "🟠 Groq (LLaMA 3.3-70b)",
        "anthropic": "🟣 Claude (claude-sonnet-4-6)",
        "openai":    "🟢 OpenAI (GPT-4o)",
        "gemini":    "🔵 Gemini",
    }
    st.markdown(f"""
<div dir="{_dir}" style="display:flex;justify-content:space-between;align-items:center;
     padding:.6rem 1.2rem;border:1px solid rgba(245,158,11,.5);border-radius:12px;
     background:rgba(40,30,4,.5);margin-bottom:1rem;">
  <div style="font-size:.85rem;color:#FCD34D;">💾 {T('active_provider')} ({T('saved_permanently')}):</div>
  <div style="font-size:.95rem;font-weight:900;color:#FBBF24;">{_persist_labels.get(_persist_pk, _persist_pk)}</div>
</div>""", unsafe_allow_html=True)

    # ──────────────────────────────────────────────────────────
    # TAB 1 — Provider Control
    # ──────────────────────────────────────────────────────────
    with tab1:
        st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)

        provider_info = {
            "groq":      {"label": "🟠 Groq — LLaMA 3.3-70b",       "secret": "GROQ_API_KEY",      "color": "#F97316"},
            "anthropic": {"label": "🟣 Claude — claude-sonnet-4-6",  "secret": "ANTHROPIC_API_KEY", "color": "#A855F7"},
            "openai":    {"label": "🟢 OpenAI — GPT-4o",             "secret": "OPENAI_API_KEY",    "color": "#22C55E"},
            "gemini":    {"label": "🔵 Gemini — 2.5 Flash",          "secret": "GEMINI_API_KEY",    "color": "#3B82F6"},
        }

        cur = st.session_state.get("ai_provider", "openai")

        st.markdown(f'<div dir="{_dir}" style="font-weight:800;color:#D1FAE5;margin-bottom:.8rem;">{T("select_provider")}</div>', unsafe_allow_html=True)
        cols = st.columns(4)
        for i, (pk, pv) in enumerate(provider_info.items()):
            with cols[i]:
                is_sel = cur == pk
                status_line = f'<div style="font-size:.7rem;color:#9CA3AF;margin-top:.3rem;">● {T("active")}</div>' if is_sel else '<div style="font-size:.7rem;margin-top:.3rem;">&nbsp;</div>'
                st.markdown(f'<div dir="{_dir}" style="background:rgba(0,0,0,.2);border:1px solid rgba(255,255,255,.15);border-radius:12px;padding:.8rem;text-align:center;margin-bottom:.5rem;height:84px;display:flex;flex-direction:column;justify-content:center;align-items:center;overflow:hidden;">'
                            f'<div style="font-size:.85rem;font-weight:700;color:#9CA3AF;line-height:1.25;display:flex;align-items:center;justify-content:center;height:2.5em;">{pv["label"]}</div>'
                            f'{status_line}'
                            f'</div>', unsafe_allow_html=True)
                if not is_sel:
                    if st.button(T('activate_btn'), key=f"adm_prov_{pk}", use_container_width=True):
                        set_active_provider(pk)
                        # Clear cached emails to regenerate with new provider
                        st.session_state["emails"] = {}
                        st.session_state.pop("assess_emails", None)
                        st.session_state["cache_version"] = int(__import__("time").time()) % 99999 + 20
                        st.rerun()
                else:
                    st.button(T('active'), key=f"adm_prov_{pk}_disabled", use_container_width=True, disabled=True)

        st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)
        st.markdown(f'<div dir="{_dir}" style="font-weight:800;color:#D1FAE5;margin-bottom:.8rem;">{T("api_status")}</div>', unsafe_allow_html=True)
        key_cols = st.columns(4)
        for i, (pk, pv) in enumerate(provider_info.items()):
            with key_cols[i]:
                val = get_secret(pv["secret"])
                has_key = bool(val and len(val) > 10)
                dot = "🟢" if has_key else "🔴"
                preview = f"...{val[-4:]}" if has_key else T('not_set')
                st.markdown(f'<div style="border:1px solid rgba(255,255,255,.1);border-radius:10px;padding:.7rem;text-align:center;">'
                            f'<div style="font-size:.75rem;color:#9CA3AF;margin-bottom:.3rem;">{pv["label"].split("—")[0].strip()}</div>'
                            f'<div style="font-size:.85rem;font-weight:700;color:white;">{dot} {preview}</div>'
                            f'</div>', unsafe_allow_html=True)

        st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

        st.markdown(f'<div dir="{_dir}" style="font-weight:800;color:#D1FAE5;margin-bottom:.5rem;">{T("language_lbl")}</div>', unsafe_allow_html=True)
        cur_lang = st.session_state.get("language", "English")
        lang_display = {"English": "English", "Arabic": "العربية"}
        col_lang1, col_lang2 = st.columns(2)
        for lk, lcol in zip(["English", "Arabic"], [col_lang1, col_lang2]):
            with lcol:
                is_l = cur_lang == lk
                label = f"{lang_display[lk]} ✓" if is_l else lang_display[lk]
                if st.button(label, key=f"adm_lang_{lk}", use_container_width=True,
                             type="primary" if is_l else "secondary"):
                    if not is_l:
                        st.session_state["language"] = lk
                        st.session_state["emails"] = {}
                        st.rerun()

        st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)
        _gs_ok, _gs_msg = gsheet_setup_status()
        _gs_id = _get_gsheet_id()
        _gs_border = "rgba(34,197,94,.4)" if _gs_ok else "rgba(245,158,11,.4)"
        _gs_bg = "rgba(4,20,4,.5)" if _gs_ok else "rgba(40,30,4,.5)"
        _gs_title_color = "#86EFAC" if _gs_ok else "#FCD34D"
        _gs_title = "📊 " + ("نسخة Google Sheets الدائمة" if _is_ar else "Durable Google Sheets backup")
        _gs_link_html = ""
        if _gs_ok and _gs_id:
            _gs_link_text = "فتح الشيت ↗" if _is_ar else "Open Sheet ↗"
            _gs_link_html = (
                '<a href="https://docs.google.com/spreadsheets/d/' + _gs_id + '/edit" '
                'target="_blank" style="color:#60A5FA;font-weight:700;text-decoration:none;">'
                + _gs_link_text + '</a>'
            )
        _gs_html = (
            '<div dir="' + _dir + '" style="padding:.8rem 1rem;border-radius:10px;'
            'border:1px solid ' + _gs_border + ';background:' + _gs_bg + ';'
            'display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem;">'
            '<div>'
            '<div style="font-weight:800;color:' + _gs_title_color + ';">' + _gs_title + '</div>'
            '<div style="font-size:.8rem;color:#9CA3AF;">' + _gs_msg + '</div>'
            '</div>'
            + _gs_link_html +
            '</div>'
        )
        st.markdown(_gs_html, unsafe_allow_html=True)

        st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)
        if st.button(T('clear_cache'), use_container_width=True):
            st.session_state["emails"] = {}
            st.session_state.pop("assess_emails", None)
            st.session_state["cache_version"] = int(__import__("time").time()) % 99999 + 20
            st.success(T('cache_cleared'))
            st.rerun()

        if st.button(T('logout_btn'), use_container_width=True):
            st.session_state["admin_authenticated"] = False
            st.query_params.clear()
            st.rerun()

    # ──────────────────────────────────────────────────────────
    # TAB 2 — Score Card: one independent card per provider
    # ──────────────────────────────────────────────────────────
    with tab2:
        st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)
        metrics = st.session_state.get("metrics", {})
        runs = load_runs()

        PROV_ORDER = ["groq", "anthropic", "openai", "gemini"]
        PROV_META = {
            "groq":      {"label": "🟠 Groq — LLaMA 3.3-70b",      "color": "#F97316"},
            "anthropic": {"label": "🟣 Claude — claude-sonnet-4-6", "color": "#A855F7"},
            "openai":    {"label": "🟢 OpenAI — GPT-4o",            "color": "#22C55E"},
            "gemini":    {"label": "🔵 Gemini — 2.5 Flash",         "color": "#3B82F6"},
        }
        TARGET_PER_LANG = 5

        def avg(lst):
            return round(sum(lst)/len(lst), 1) if lst else None

        def get_m(p): return metrics.get(p, {})

        # ── Rotation plan reference table (informational only) ──
        with st.expander("📋 " + ("جدول التدوير المنظّم (10 دورات)" if _is_ar else "Systematic Rotation Plan (10 cycles)")):
            rcols = st.columns([1,2,2,1.5])
            for ci, hdr in enumerate([("#" if _is_ar else "#"), ("الوظيفة" if _is_ar else "Role"), ("المستوى" if _is_ar else "Difficulty"), ("اللغة" if _is_ar else "Language")]):
                with rcols[ci]:
                    st.markdown(f'<div style="font-weight:800;color:#9CA3AF;font-size:.78rem;">{hdr}</div>', unsafe_allow_html=True)
            for plan in ROTATION_PLAN:
                rcols = st.columns([1,2,2,1.5])
                role_show = plan["role_ar"] if _is_ar else plan["role_en"]
                diff_show = {"easy": ("سهل" if _is_ar else "Easy"), "medium": ("متوسط" if _is_ar else "Medium"), "hard": ("صعب" if _is_ar else "Hard")}[plan["difficulty"]]
                lang_show = ("🇸🇦 عربي" if plan["language"]=="Arabic" else "🇬🇧 EN") if _is_ar else ("Arabic" if plan["language"]=="Arabic" else "English")
                for ci, val in enumerate([str(plan["cycle"]), role_show, diff_show, lang_show]):
                    with rcols[ci]:
                        st.markdown(f'<div style="color:#E2E8F0;font-size:.82rem;padding:.15rem 0;">{val}</div>', unsafe_allow_html=True)

        st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

        # ── One independent card per provider ──
        _gsheet_auto_latest = pull_latest_auto_metrics_from_gsheet()
        for p in PROV_ORDER:
            meta = PROV_META[p]
            m = get_m(p)
            speeds = m.get("speed", [])
            total_j = m.get("json_ok",0) + m.get("json_fail",0)
            json_rate = int(m.get("json_ok",0)/total_j*100) if total_j > 0 else None
            calls = m.get("calls",0)
            err_rate = int(m.get("errors",0)/calls*100) if calls > 0 else None
            hashes = m.get("hashes",[])

            # No local activity recorded for this provider this session
            # (e.g. right after a container restart) — fall back to the
            # last snapshot synced to Google Sheets instead of showing
            # blank boxes for data that genuinely exists.
            if calls == 0:
                _snap = _gsheet_auto_latest.get(p)
                if _snap:
                    try:
                        _sp = _snap.get("avg_speed_s")
                        speeds = [float(_sp)] if _sp not in ("", None) else []
                    except (ValueError, TypeError):
                        speeds = []
                    try:
                        json_rate = int(float(_snap.get("json_success_rate_pct"))) if _snap.get("json_success_rate_pct") not in ("", None) else None
                    except (ValueError, TypeError):
                        json_rate = None
                    try:
                        err_rate = int(float(_snap.get("error_rate_pct"))) if _snap.get("error_rate_pct") not in ("", None) else None
                    except (ValueError, TypeError):
                        err_rate = None
                    try:
                        calls = int(float(_snap.get("calls"))) if _snap.get("calls") not in ("", None) else 0
                    except (ValueError, TypeError):
                        calls = 0
                    try:
                        _uniq = int(float(_snap.get("unique_responses"))) if _snap.get("unique_responses") not in ("", None) else 0
                    except (ValueError, TypeError):
                        _uniq = 0
                    hashes = list(range(_uniq))  # only its length is used below

            p_runs_en = [r for r in runs if r.get("provider")==p and r.get("language")=="English"]
            p_runs_ar = [r for r in runs if r.get("provider")==p and r.get("language")=="Arabic"]
            ordered_runs = p_runs_en + p_runs_ar

            def _avg_field(field):
                vals = [r.get(field) for r in ordered_runs if r.get(field) is not None]
                return round(sum(vals)/len(vals), 1) if vals else None

            avg_diff    = _avg_field("auto_difficulty")
            avg_arabic  = _avg_field("auto_arabic")
            avg_quality = _avg_field("auto_quality")
            avg_medical = _avg_field("auto_medical")
            avg_overall = _avg_field("overall")

            st.markdown(f"""
<div dir="{_dir}" style="border:1px solid {meta['color']}55;border-radius:14px;padding:1rem 1.2rem;margin-bottom:1.2rem;background:rgba(255,255,255,.02);">
  <div style="font-weight:900;font-size:1.05rem;color:{meta['color']};margin-bottom:.6rem;">{meta['label']}</div>
</div>""", unsafe_allow_html=True)

            # 9 uniform boxes — 4 auto-performance + 4 auto-content + 1 manual overall
            box_items = [
                (T('speed_metric'),      f"{avg(speeds):.1f}s" if speeds else "—"),
                (T('json_metric'),       f"{json_rate}%" if json_rate is not None else "—"),
                (T('error_metric'),      f"{err_rate}%" if err_rate is not None else "—"),
                (T('diversity_metric'),  f"{len(hashes)}/{calls}" if calls > 0 else "—"),
                (("صعوبة%" if _is_ar else "Difficulty%"), f"{avg_diff}%" if avg_diff is not None else "—"),
                (("عربي%" if _is_ar else "Arabic%"),      f"{avg_arabic}%" if avg_arabic is not None else "—"),
                (("جودة%" if _is_ar else "Quality%"),     f"{avg_quality}%" if avg_quality is not None else "—"),
                (("طبي%" if _is_ar else "Medical%"),      f"{avg_medical}%" if avg_medical is not None else "—"),
                (("الانطباع⭐" if _is_ar else "Overall⭐"), f"{avg_overall}/5" if avg_overall is not None else "—"),
            ]
            box_rows = [box_items[i:i+3] for i in range(0, 9, 3)]
            for row in box_rows:
                bc = st.columns(3)
                for ci, (lbl, val) in enumerate(row):
                    with bc[ci]:
                        st.markdown(f'<div style="text-align:center;border:1px solid rgba(255,255,255,.08);border-radius:10px;padding:.5rem;margin-bottom:.5rem;">'
                                    f'<div style="font-size:.7rem;color:#9CA3AF;">{lbl}</div>'
                                    f'<div style="font-size:1rem;font-weight:800;color:#E2E8F0;">{val}</div></div>', unsafe_allow_html=True)

            n_en, n_ar = len(p_runs_en), len(p_runs_ar)
            tag_en = "✅" if n_en >= TARGET_PER_LANG else "🟡"
            tag_ar = "✅" if n_ar >= TARGET_PER_LANG else "🟡"
            st.markdown(f'<div style="color:#9CA3AF;font-size:.78rem;margin:.2rem 0 .6rem;">'
                        f'{tag_en} English: {n_en}/{TARGET_PER_LANG} &nbsp;&nbsp; {tag_ar} Arabic: {n_ar}/{TARGET_PER_LANG}</div>',
                        unsafe_allow_html=True)

            # Detailed per-cycle table, tucked away in an expander so the
            # 9 boxes above stay the clean at-a-glance summary.
            with st.expander("📋 " + ("عرض تفاصيل الـ10 دورات" if _is_ar else "View detailed 10-cycle breakdown")):
                col_widths = [0.5,0.7,0.8,0.7,0.7,0.8,0.9,0.9,0.9,0.9,0.9]
                cols_hdr = st.columns(col_widths)
                headers = [
                    "#", ("لغة" if _is_ar else "Lang"),
                    ("سرعة" if _is_ar else "Speed"), ("JSON%"), ("أخطاء%" if _is_ar else "Err%"),
                    ("تنوع" if _is_ar else "Divers."),
                    ("صعوبة%" if _is_ar else "Diff%"), ("عربي%" if _is_ar else "Arabic%"),
                    ("جودة%" if _is_ar else "Quality%"), ("طبي%" if _is_ar else "Medical%"),
                    ("انطباع" if _is_ar else "Overall"),
                ]
                for ci, hdr in enumerate(headers):
                    with cols_hdr[ci]:
                        st.markdown(f'<div style="font-weight:800;color:#9CA3AF;font-size:.68rem;border-bottom:1px solid rgba(255,255,255,.1);padding:.2rem 0;">{hdr}</div>', unsafe_allow_html=True)

                if ordered_runs:
                    for i, r in enumerate(ordered_runs, 1):
                        rc = st.columns(col_widths)
                        lang_short = "EN" if r.get("language")=="English" else "AR"
                        vals = [
                            str(i),
                            lang_short,
                            f"{r.get('avg_speed')}s" if r.get('avg_speed') is not None else "—",
                            f"{r.get('json_rate')}%" if r.get('json_rate') is not None else "—",
                            f"{r.get('error_rate')}%" if r.get('error_rate') is not None else "—",
                            r.get('diversity') or "—",
                            f"{r.get('auto_difficulty')}%" if r.get('auto_difficulty') is not None else "—",
                            f"{r.get('auto_arabic')}%" if r.get('auto_arabic') is not None else "—",
                            f"{r.get('auto_quality')}%" if r.get('auto_quality') is not None else "—",
                            f"{r.get('auto_medical')}%" if r.get('auto_medical') is not None else "—",
                            f"{r.get('overall')}/5" if r.get('overall') is not None else "—",
                        ]
                        for ci, val in enumerate(vals):
                            with rc[ci]:
                                st.markdown(f'<div style="color:#E2E8F0;font-size:.74rem;padding:.3rem .2rem;border-radius:6px;'
                                            f'background:{"rgba(255,255,255,.03)" if i%2==0 else "transparent"};">{val}</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div style="color:#6B7280;font-size:.8rem;padding:.5rem 0;">{("لا توجد دورات محفوظة لهذا المزوّد بعد" if _is_ar else "No cycles saved for this provider yet")}</div>', unsafe_allow_html=True)

            st.markdown('<div style="height:.8rem"></div>', unsafe_allow_html=True)

        st.markdown(f'<div dir="{_dir}" style="font-size:.75rem;color:#6B7280;">{T("auto_manual_note")}</div>', unsafe_allow_html=True)
        st.markdown('<div style="height:.8rem"></div>', unsafe_allow_html=True)

        if not runs:
            st.markdown(f'<div dir="{_dir}" style="text-align:{_align};font-size:.8rem;color:#6B7280;">⚠️ ' + ("احفظي تقييمًا واحدًا على الأقل من تبويب Manual Ratings لتفعيل التصدير." if _is_ar else "Save at least one rating from the Manual Ratings tab to enable export.") + '</div>', unsafe_allow_html=True)
        col_exp, col_reset = st.columns(2)
        with col_exp:
            if runs:
                st.download_button(
                    "⬇️ " + ("تصدير كل النتائج Excel" if _is_ar else "Export full results (Excel)"),
                    data=build_excel_export(),
                    file_name="phishing_research_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="export_excel_scorecard",
                )
            else:
                st.button("⬇️ " + ("تصدير كل النتائج Excel" if _is_ar else "Export full results (Excel)"), use_container_width=True, disabled=True, key="export_excel_scorecard_disabled")
        with col_reset:
            if st.button(T('reset_metrics'), use_container_width=True):
                st.session_state["metrics"] = {}
                save_metrics_file({})
                delete_all_runs()
                clear_gsheet_data()
                _save_json_list(_AUTO_EVAL_FILE_PATH, [])
                _save_pending_buckets({})
                _save_json_dict(_PENDING_PERF_FILE_PATH, {})
                st.success(T('metrics_reset'))
                st.rerun()

    # ──────────────────────────────────────────────────────────
    # TAB 3 — Manual Ratings (👍 اليدوية)
    # ──────────────────────────────────────────────────────────
    with tab3:
        st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)
        cur_prov = st.session_state.get("ai_provider", "groq")
        prov_label = provider_info.get(cur_prov, {}).get("label", cur_prov)

        st.markdown(
            f'<div style="font-weight:900;color:#D1FAE5;margin-bottom:.3rem;">'
            f'{"قيّمي الدورة الكاملة (٦ تعلّم + ١٠ اختبار) بعد ما تخلّصينها" if _is_ar else "Rate the FULL cycle (6 learning + 10 assessment) after finishing it"}'
            f'</div>'
            f'<div style="color:#9CA3AF;font-size:.85rem;margin-bottom:1rem;">{T("active_provider")}: {prov_label}</div>',
            unsafe_allow_html=True)

        # Compute which cycle # comes next (based on saved cycles for this
        # provider) BEFORE showing the language radio, so the radio can
        # default to whatever the rotation plan says for that cycle.
        _runs_now_pre = load_runs()
        _n_done_total = len([r for r in _runs_now_pre if r.get("provider")==cur_prov])
        _cycle_no = min(_n_done_total + 1, 10)
        _plan = ROTATION_PLAN[_cycle_no - 1]
        _plan_role = _plan["role_ar"] if _is_ar else _plan["role_en"]
        _plan_diff = {"easy": ("سهل" if _is_ar else "Easy"), "medium": ("متوسط" if _is_ar else "Medium"), "hard": ("صعب" if _is_ar else "Hard")}[_plan["difficulty"]]
        _plan_lang = _plan["language"]

        # Which language was this cycle run in? Defaults to whatever the
        # rotation plan says for this cycle number (can be overridden).
        lang_for_run = st.radio(
            ("🏷️ صنّفي بيانات هذي الدورة: المحتوى المولّد كان بـ" if _is_ar else "🏷️ Tag this cycle's data: generated content was in"),
            options=["English", "Arabic"],
            index=0 if _plan_lang=="English" else 1,
            horizontal=True,
            key="run_lang_selector",
        )

        # Progress counter for this provider+language combo (target 5+5=10)
        _runs_now = load_runs()
        _n_done_lang = len([r for r in _runs_now if r.get("provider")==cur_prov and r.get("language")==lang_for_run])
        _target_lang = 5
        _pct = min(_n_done_lang / _target_lang, 1.0)

        st.markdown(
            f'<div style="margin:.3rem 0 .6rem;">'
            f'<div style="color:#93C5FD;font-size:.9rem;font-weight:700;margin-bottom:.3rem;">'
            f'{("اللغة:" if _is_ar else "Language:")} {lang_for_run} — {_n_done_lang}/{_target_lang}'
            f'</div>'
            f'<div style="background:rgba(255,255,255,.1);border-radius:6px;height:8px;overflow:hidden;">'
            f'<div style="background:#22C55E;height:100%;width:{_pct*100:.0f}%;"></div>'
            f'</div></div>',
            unsafe_allow_html=True)

        _plan_lang_show = ("🇸🇦 عربي" if _plan_lang=="Arabic" else "🇬🇧 إنجليزي") if _is_ar else _plan_lang
        st.markdown(
            f'<div dir="{_dir}" style="border:1px solid rgba(245,158,11,.4);border-radius:10px;padding:.6rem .9rem;'
            f'background:rgba(40,30,4,.4);margin-bottom:1rem;font-size:.85rem;color:#FCD34D;">'
            f'📋 {("الدورة التالية حسب جدول التدوير رقم" if _is_ar else "Next cycle per rotation plan, #")} {_cycle_no}/10 — '
            f'{("الوظيفة" if _is_ar else "Role")}: <b>{_plan_role}</b> | {("المستوى" if _is_ar else "Difficulty")}: <b>{_plan_diff}</b> | {("اللغة" if _is_ar else "Language")}: <b>{_plan_lang_show}</b>'
            f'</div>',
            unsafe_allow_html=True)

        # Snapshot of the automatic scores collected since the last saved cycle
        _pending = _load_pending_buckets().get(f"{cur_prov}__{lang_for_run}", [])
        _pending_perf = _load_json_dict(_PENDING_PERF_FILE_PATH).get(f"{cur_prov}__{lang_for_run}", [])
        if _pending or _pending_perf:
            def _pavg(field):
                vals = [it[field] for it in _pending if it.get(field) is not None]
                return round(sum(vals)/len(vals), 1) if vals else None
            _perf_speeds = [it["speed"] for it in _pending_perf if it.get("speed") is not None]
            _perf_speed_avg = round(sum(_perf_speeds)/len(_perf_speeds), 1) if _perf_speeds else None
            _perf_calls = len(_pending_perf)
            _perf_errors = sum(1 for it in _pending_perf if it.get("is_error"))
            _perf_err_rate = round(_perf_errors/_perf_calls*100) if _perf_calls else None
            st.markdown(
                f'<div style="border:1px solid rgba(34,197,94,.35);border-radius:10px;padding:.6rem .9rem;'
                f'background:rgba(4,30,10,.4);margin-bottom:1rem;font-size:.82rem;color:#86EFAC;">'
                f'⚙️ {("نتائج آلية لهذي الدورة لحد الآن" if _is_ar else "Automatic scores collected so far this cycle")} '
                f'({len(_pending)} {("إيميل" if _is_ar else "emails")}): '
                f'{("صعوبة" if _is_ar else "Difficulty")} {_pavg("difficulty_score")}% · '
                f'{("عربي" if _is_ar else "Arabic")} {_pavg("arabic_score")}% · '
                f'{("جودة" if _is_ar else "Quality")} {_pavg("quality_score")}% · '
                f'{("طبي" if _is_ar else "Medical")} {_pavg("medical_score")}% · '
                f'{("سرعة" if _is_ar else "Speed")} {_perf_speed_avg if _perf_speed_avg is not None else "—"}s · '
                f'{("أخطاء" if _is_ar else "Errors")} {_perf_err_rate if _perf_err_rate is not None else "—"}%'
                f'</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div style="color:#6B7280;font-size:.78rem;margin-bottom:1rem;">'
                f'{("لسا ما ولّدتي أي إيميل بهذي الدورة — النتائج الآلية بتظهر هنا أوتوماتيك" if _is_ar else "No emails generated yet this cycle — automatic scores will appear here")}'
                f'</div>', unsafe_allow_html=True)

        # تنسيق حقل الملاحظات بشكل شفاف مع إطار مثل باقي العناصر
        st.markdown("""<style>
.stTextInput input{
    background:transparent!important;
    color:#E2E8F0!important;
    border:none!important;
    box-shadow:none!important;
}
.stTextInput input::placeholder{color:#6B7280!important;}
.stTextInput>div>div,
.stTextInput div[data-baseweb="base-input"]{
    background:transparent!important;
    background-color:transparent!important;
    border:none!important;
    box-shadow:none!important;
}
.stTextInput div[data-baseweb="input"]{
    background:rgba(15,23,42,.5)!important;
    border:1px solid rgba(255,255,255,.15)!important;
    border-radius:8px!important;
    box-shadow:none!important;
}
</style>""", unsafe_allow_html=True)

        st.markdown(f'<div style="margin-bottom:.2rem;"><span style="font-weight:700;color:#E2E8F0;">⭐ '
                    f'{("الانطباع العام" if _is_ar else "Overall Impression")}</span>'
                    f'<span style="color:#6B7280;font-size:.8rem;margin-right:.5rem;"> — '
                    f'{("حكمك الشامل عن جودة/واقعية/لغة هذي الدورة بالكامل" if _is_ar else "Your holistic judgement of quality/realism/language for this whole cycle")}</span></div>',
                    unsafe_allow_html=True)
        # NEW: a per-(provider, language) "form version" counter. Bumping it
        # after every successful save changes the slider/note widget keys,
        # which forces Streamlit to reset them to their defaults instead of
        # silently keeping the just-saved values — this was the root cause
        # of accidental double-saves (clicking Save twice resubmitted the
        # same note/rating because the widgets never visibly reset).
        _form_ver_key = f"cycle_form_version_{cur_prov}_{lang_for_run}"
        _form_ver = st.session_state.get(_form_ver_key, 0)

        overall_rating = st.select_slider(
            label="overall",
            options=[1, 2, 3, 4, 5],
            value=3,
            format_func=lambda x: f"{'⭐'*x}{'☆'*(5-x)} ({x}/5)",
            key=f"rating_overall_{cur_prov}_{lang_for_run}_{_form_ver}",
            label_visibility="collapsed"
        )
        st.markdown('<div style="height:.4rem"></div>', unsafe_allow_html=True)

        col_note, _ = st.columns([2,1])
        with col_note:
            note = st.text_input(T('note_label'), placeholder=T('note_placeholder'),
                                 key=f"note_{cur_prov}_{lang_for_run}_{_form_ver}")

        def _save_cycle_rating(overall_val, note_text):
            snap = snapshot_and_clear_pending_cycle(cur_prov, lang_for_run)
            perf_snap = snapshot_and_clear_pending_perf(cur_prov, lang_for_run)
            record = {
                "timestamp": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
                "provider": cur_prov,
                "language": lang_for_run,
                "overall": overall_val,
                "auto_difficulty": snap["difficulty_score"],
                "auto_arabic": snap["arabic_score"],
                "auto_quality": snap["quality_score"],
                "auto_medical": snap["medical_score"],
                "n_auto_emails": snap["n_emails"],
                "avg_speed": perf_snap["avg_speed"],
                "json_rate": perf_snap["json_rate"],
                "error_rate": perf_snap["error_rate"],
                "diversity": perf_snap["diversity"],
                "note": note_text or "",
            }
            save_run(record)
            # Bump the form version so the slider/note reset to defaults on
            # the next render, making it visually obvious the save went
            # through and preventing a second click from resubmitting it.
            st.session_state[_form_ver_key] = _form_ver + 1

        if st.button(T('save_btn'), use_container_width=True):
            _save_cycle_rating(overall_rating, note)
            st.success(f"{T('ratings_saved')} {prov_label} ({lang_for_run}) — {_n_done_lang+1}/{_target_lang}")
            st.rerun()

        # Quick thumbs up/down shortcut — still one record per FULL cycle
        st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)
        st.markdown(f'<div style="color:#9CA3AF;font-size:.85rem;margin-bottom:.4rem;">{T("quick_rating")}</div>', unsafe_allow_html=True)
        qc1, qc2, qc3 = st.columns(3)
        with qc1:
            if st.button(T('good_btn'), use_container_width=True, key="quick_good"):
                _save_cycle_rating(4, note)
                st.success(T('saved_45'))
                st.rerun()
        with qc2:
            if st.button(T('avg_btn'), use_container_width=True, key="quick_avg"):
                _save_cycle_rating(3, note)
                st.success(T('saved_35'))
                st.rerun()
        with qc3:
            if st.button(T('poor_btn'), use_container_width=True, key="quick_bad"):
                _save_cycle_rating(2, note)
                st.success(T('saved_25'))
                st.rerun()

        # Undo last entry for this provider+language, in case of a mis-click
        # (also restores the snapshot back into the pending bucket so no
        # automatic data is lost).
        st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)
        if st.button("↩️ " + ("تراجع عن آخر دورة محفوظة لهذا المزوّد/اللغة" if _is_ar else "Undo last saved cycle for this provider/language"),
                     use_container_width=True, key="undo_last_run"):
            all_runs = load_runs()
            for i in range(len(all_runs) - 1, -1, -1):
                if all_runs[i].get("provider") == cur_prov and all_runs[i].get("language") == lang_for_run:
                    removed_record = all_runs.pop(i)
                    try:
                        with open(_RUNS_FILE_PATH, "w", encoding="utf-8") as f:
                            json.dump(all_runs, f, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                    delete_run_from_gsheet(removed_record)
                    st.success("✅ " + ("تم حذف آخر دورة" if _is_ar else "Last cycle removed"))
                    st.rerun()
                    break

        # History summary for this provider+language
        st.markdown('<div style="height:.8rem"></div>', unsafe_allow_html=True)
        my_runs = [r for r in load_runs() if r.get("provider")==cur_prov and r.get("language")==lang_for_run]
        if my_runs:
            overall_vals = [r.get("overall") for r in my_runs if r.get("overall") is not None]
            if overall_vals:
                a = round(sum(overall_vals)/len(overall_vals), 1)
                st.markdown(f'<div style="font-weight:700;color:#D1FAE5;margin-bottom:.2rem;">{T("rating_history")} {prov_label} ({lang_for_run})</div>', unsafe_allow_html=True)
                st.markdown(f'<div style="color:#9CA3AF;font-size:.82rem;">{("الانطباع العام" if _is_ar else "Overall")}: {T("avg_label")} {a}/5 ({len(overall_vals)} {("دورة" if _is_ar else "cycles")}) {"⭐"*round(a)}</div>', unsafe_allow_html=True)

        # Same full Excel export, available here too so the researcher
        # doesn't need to switch tabs just to download her results.
        st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)
        _all_runs_now = load_runs()
        if _all_runs_now:
            st.download_button(
                "⬇️ " + ("تصدير كل النتائج Excel" if _is_ar else "Export full results (Excel)"),
                data=build_excel_export(),
                file_name="phishing_research_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="export_excel_manual",
            )
        else:
            st.markdown(f'<div dir="{_dir}" style="text-align:{_align};font-size:.8rem;color:#6B7280;">⚠️ ' + ("احفظي تقييمًا واحدًا على الأقل أعلاه لتفعيل التصدير." if _is_ar else "Save at least one rating above to enable export.") + '</div>', unsafe_allow_html=True)
            st.button("⬇️ " + ("تصدير كل النتائج Excel" if _is_ar else "Export full results (Excel)"), use_container_width=True, disabled=True, key="export_excel_manual_disabled")

    with tab4:
        st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)
        st.markdown(
            f'<div style="color:#9CA3AF;font-size:.85rem;margin-bottom:1rem;">'
            f'{"آخر 20 خطأ تقني حصل أثناء التوليد (محفوظ على القرص، يبقى حتى بعد تحديث الصفحة)" if _is_ar else "Last 20 technical generation errors (saved to disk — survives page reloads)"}'
            f'</div>', unsafe_allow_html=True)
        disk_log = _load_debug_log()
        session_log = st.session_state.get("_debug_log", [])
        # Merge + dedupe by (stage, ts) while preserving order, disk is the
        # source of truth since it survives reloads; session entries cover
        # the (rare) case a write to disk failed on a read-only host.
        seen = set()
        merged = []
        for entry in disk_log + session_log:
            key = (entry.get("stage"), str(entry.get("ts")), str(entry.get("error"))[:50])
            if key not in seen:
                seen.add(key)
                merged.append(entry)
        debug_log = list(reversed(merged))[:20]
        if not debug_log:
            st.info("لا توجد أخطاء مسجّلة حتى الآن." if _is_ar else "No errors logged yet.")
        else:
            if st.button("🗑️ " + ("تفريغ السجل" if _is_ar else "Clear log"), key="clear_debug_log"):
                st.session_state["_debug_log"] = []
                try:
                    with open(_DEBUG_LOG_PATH, "w", encoding="utf-8") as f:
                        json.dump([], f)
                except Exception:
                    pass
                st.rerun()
            for i, entry in enumerate(debug_log):
                with st.expander(f"#{len(debug_log)-i} — {entry.get('stage','?')}"):
                    st.json(entry.get("error"))






# =============================================================
# FINAL DIVERSITY + DIFFICULTY OVERRIDE — Study 3
# -------------------------------------------------------------
# This block intentionally overrides the earlier prompt/quality
# functions before the Streamlit router runs. It keeps the UI and
# provider code unchanged, but strengthens:
# 1) unbounded learning generation,
# 2) unbounded assessment generation,
# 3) the 9-criterion difficulty contract,
# 4) Arabic/English parity,
# 5) analysis quality and attack-type diversity.
# =============================================================

ADVANCED_BANNED_DOMAIN_WORDS = {
    "secure", "update", "verify", "verification", "login", "reset",
    "password", "urgent", "alert", "account", "emr", "phish", "fake"
}

ATTACK_PLAYBOOK = {
    "clinical": [
        {"attack":"Credential Harvesting", "vector":"external portal link", "persuasion":"routine clinical workflow", "en":"patient-records access review", "ar":"مراجعة وصول السجلات الطبية"},
        {"attack":"Attachment Malware", "vector":"PDF/DOCX/XLSM attachment", "persuasion":"patient safety", "en":"urgent patient report attachment", "ar":"مرفق تقرير مريض عاجل"},
        {"attack":"QR Phishing", "vector":"QR code instruction", "persuasion":"mobile workflow", "en":"mobile clinical checklist QR", "ar":"رمز QR لقائمة تحقق سريرية"},
        {"attack":"Authority Impersonation", "vector":"reply with sensitive data", "persuasion":"medical director authority", "en":"consultant/director request", "ar":"طلب من مدير أو استشاري"},
        {"attack":"MFA / OTP Abuse", "vector":"OTP or MFA confirmation", "persuasion":"access continuity", "en":"MFA confirmation for clinical system", "ar":"تأكيد MFA لنظام سريري"},
        {"attack":"Spear Phishing", "vector":"contextual link or reply", "persuasion":"personalized prior meeting", "en":"follow-up to department meeting", "ar":"متابعة لاجتماع القسم"},
        {"attack":"Legitimate-Looking Compliance Phish", "vector":"policy acknowledgement", "persuasion":"compliance", "en":"infection-control acknowledgement", "ar":"إقرار مكافحة العدوى"},
        {"attack":"Cloud Document Scam", "vector":"shared document", "persuasion":"collaboration", "en":"shared clinical handover document", "ar":"مستند تسليم سريري مشترك"},
    ],
    "admin": [
        {"attack":"Invoice Fraud", "vector":"invoice attachment or portal", "persuasion":"payment deadline", "en":"supplier invoice approval", "ar":"اعتماد فاتورة مورد"},
        {"attack":"BEC / Executive Impersonation", "vector":"reply or bank-transfer instruction", "persuasion":"executive authority", "en":"director payment request", "ar":"طلب دفع من مدير"},
        {"attack":"Payroll Phishing", "vector":"bank-detail update", "persuasion":"salary continuity", "en":"payroll account confirmation", "ar":"تأكيد حساب الرواتب"},
        {"attack":"Benefits Enrollment Scam", "vector":"benefits portal", "persuasion":"reward/opportunity", "en":"employee benefits enrollment", "ar":"التسجيل في مزايا الموظفين"},
        {"attack":"Vendor Portal Scam", "vector":"supplier portal", "persuasion":"contract renewal", "en":"vendor contract renewal", "ar":"تجديد عقد مورد"},
        {"attack":"Shared Document Scam", "vector":"document share", "persuasion":"administrative process", "en":"shared patient-files report", "ar":"تقرير ملفات مرضى مشترك"},
        {"attack":"Insurance Portal Phish", "vector":"coverage verification", "persuasion":"claim processing", "en":"insurance claim validation", "ar":"التحقق من مطالبة تأمين"},
        {"attack":"Meeting Invitation Scam", "vector":"calendar attachment/link", "persuasion":"routine meeting", "en":"accreditation review meeting", "ar":"اجتماع مراجعة الاعتماد"},
    ],
    "it": [
        {"attack":"Credential Harvesting", "vector":"admin portal", "persuasion":"service continuity", "en":"network access review", "ar":"مراجعة وصول الشبكة"},
        {"attack":"MFA Fatigue / OTP Abuse", "vector":"MFA approval request", "persuasion":"security validation", "en":"MFA push validation", "ar":"اعتماد إشعار MFA"},
        {"attack":"Attachment Malware", "vector":"script/XLSM/DOCM attachment", "persuasion":"system repair", "en":"backup verification script", "ar":"سكربت تحقق النسخ الاحتياطي"},
        {"attack":"CIO/CISO Impersonation", "vector":"reply with configuration/data", "persuasion":"executive pressure", "en":"security exception request", "ar":"طلب استثناء أمني"},
        {"attack":"License Renewal Scam", "vector":"renewal portal", "persuasion":"service deadline", "en":"software license renewal", "ar":"تجديد ترخيص برنامج"},
        {"attack":"Cloud Backup Scam", "vector":"cloud console link", "persuasion":"data protection", "en":"cloud backup access review", "ar":"مراجعة وصول النسخ السحابي"},
        {"attack":"QR Phishing", "vector":"QR enrollment", "persuasion":"device enrollment", "en":"device compliance QR", "ar":"رمز امتثال الجهاز"},
        {"attack":"Spear Phishing", "vector":"contextual reply/link", "persuasion":"known incident follow-up", "en":"follow-up to outage ticket", "ar":"متابعة تذكرة عطل"},
    ],
    "other": [
        {"attack":"Mixed Department Phishing", "vector":"fresh mixed vector", "persuasion":"workplace routine", "en":"new hospital workflow", "ar":"إجراء مستشفى جديد"},
        {"attack":"Vendor / HR / IT Scam", "vector":"link, attachment, reply, or QR", "persuasion":"authority or compliance", "en":"cross-department request", "ar":"طلب بين الأقسام"},
        {"attack":"Legitimate-Looking Internal Process", "vector":"contextual workflow", "persuasion":"routine", "en":"ordinary internal process abused", "ar":"استغلال إجراء داخلي عادي"},
    ]
}

LEGITIMATE_PLAYBOOK = {
    "clinical": [
        {"en":"shift schedule update", "ar":"تحديث جدول المناوبات", "sender":"Head Nurse / Clinical Operations"},
        {"en":"infection-control guideline notice", "ar":"إشعار إرشادات مكافحة العدوى", "sender":"Infection Control Team"},
        {"en":"CME workshop invitation", "ar":"دعوة ورشة تعليم طبي مستمر", "sender":"Medical Education"},
        {"en":"lab downtime notice", "ar":"إشعار توقف مؤقت للمختبر", "sender":"Laboratory Department"},
        {"en":"radiology protocol reminder", "ar":"تذكير ببروتوكول الأشعة", "sender":"Radiology Department"},
    ],
    "admin": [
        {"en":"policy acknowledgement with no link", "ar":"إقرار سياسة بدون رابط", "sender":"HR Department"},
        {"en":"procurement order confirmation", "ar":"تأكيد طلب مشتريات", "sender":"Procurement Department"},
        {"en":"insurance training session", "ar":"جلسة تدريب التأمين", "sender":"Training Department"},
        {"en":"meeting room change", "ar":"تغيير قاعة اجتماع", "sender":"Admin Office"},
        {"en":"accreditation visit reminder", "ar":"تذكير بزيارة الاعتماد", "sender":"Quality Department"},
    ],
    "it": [
        {"en":"scheduled maintenance notice", "ar":"إشعار صيانة مجدولة", "sender":"IT Department"},
        {"en":"helpdesk ticket closure", "ar":"إغلاق تذكرة دعم", "sender":"IT Helpdesk"},
        {"en":"approved patching window", "ar":"نافذة تحديث معتمدة", "sender":"Systems Team"},
        {"en":"VPN maintenance reminder", "ar":"تذكير صيانة VPN", "sender":"Network Team"},
        {"en":"security awareness training", "ar":"تدريب توعية أمنية", "sender":"Cybersecurity Office"},
    ],
    "other": [
        {"en":"general hospital announcement", "ar":"إعلان مستشفى عام", "sender":"Hospital Administration"},
        {"en":"training reminder", "ar":"تذكير تدريب", "sender":"Training Department"},
        {"en":"safe operational update", "ar":"تحديث تشغيلي آمن", "sender":"Operations Office"},
    ]
}

# =============================================================
# DIVERSITY EXPANSION — non-classic lure themes
# -------------------------------------------------------------
# EN: Added to EVERY role so learning/assessment examples are not
# always "IT/security"-flavoured. These are realistic lure types
# hospital staff actually receive: MOH-style public offers,
# restaurant/cafe staff discounts, prize/ad campaigns, and
# bank/telecom partner promos — used both as phishing pretexts and
# as safe legitimate look-alikes.
# AR: تُضاف لكل الأدوار حتى لا تكون الأمثلة دومًا بنكهة "تقنية/أمنية".
# هذه أنواع طُعم واقعية يستقبلها الموظفون فعليًا: عروض شبيهة بوزارة
# الصحة، خصومات موظفين بمطاعم/كوفيهات، حملات جوائز/إعلانات، وعروض
# شراكة بنكية/اتصالات — تُستخدم كذريعة تصيد وكذلك كرسائل شرعية آمنة.
# =============================================================
_EXTRA_ATTACK_THEMES = [
    {"attack": "Fake Government/MOH Offer", "vector": "benefits or wellness program link",
     "persuasion": "limited-time government benefit",
     "en": "MOH wellness program enrollment", "ar": "التسجيل في برنامج العافية بوزارة الصحة"},
    {"attack": "Restaurant/Retail Voucher Scam", "vector": "discount voucher link or QR",
     "persuasion": "exclusive staff discount",
     "en": "exclusive staff discount at a partner restaurant/cafe", "ar": "خصم حصري للموظفين في مطعم/كوفي شريك"},
    {"attack": "Prize/Reward Ad Scam", "vector": "promotional ad link or attachment",
     "persuasion": "limited-time prize or reward",
     "en": "hospital anniversary prize draw", "ar": "سحب جوائز بمناسبة ذكرى تأسيس المستشفى"},
    {"attack": "Telecom/Bank Promo Scam", "vector": "promotional offer link",
     "persuasion": "too-good-to-be-true financial offer",
     "en": "telecom/bank partnership offer for staff", "ar": "عرض شراكة بنكية/اتصالات للموظفين"},
]
_EXTRA_LEGIT_THEMES = [
    {"en": "official MOH public-health awareness bulletin", "ar": "نشرة توعوية صحية رسمية من وزارة الصحة",
     "sender": "Ministry of Health Communications"},
    {"en": "hospital cafeteria new menu / Ramadan timing notice", "ar": "إشعار قائمة الكافيتيريا الجديدة / مواعيد رمضان",
     "sender": "Hospital Facilities"},
    {"en": "staff wellness day announcement", "ar": "إعلان يوم العافية للموظفين",
     "sender": "HR Wellness Office"},
    {"en": "National Day / public holiday schedule notice", "ar": "إشعار جدول اليوم الوطني / العطلة الرسمية",
     "sender": "Hospital Administration"},
]
for _rt in list(ATTACK_PLAYBOOK.keys()):
    ATTACK_PLAYBOOK[_rt] = ATTACK_PLAYBOOK[_rt] + _EXTRA_ATTACK_THEMES
for _rt in list(LEGITIMATE_PLAYBOOK.keys()):
    LEGITIMATE_PLAYBOOK[_rt] = LEGITIMATE_PLAYBOOK[_rt] + _EXTRA_LEGIT_THEMES




def get_role_unbounded_context(role_type, is_ar=False):
    if is_ar:
        return {
            "clinical": "الدور سريري داخل مستشفى سعودي. يمكن أن تشمل البيئة: طبيب، تمريض، صيدلة، مختبر، أشعة، عيادات، طوارئ، عناية مركزة، سجلات مرضى، PACS، LIS، EMR، تسليم مناوبات، بروتوكولات وزارة الصحة.",
            "admin": "الدور إداري داخل مستشفى سعودي. يمكن أن تشمل البيئة: استقبال، سكرتارية طبية، ملفات مرضى، تأمين، فوترة، مشتريات، موارد بشرية، اعتماد، عقود موردين، رواتب، اجتماعات إدارية.",
            "it": "الدور تقني داخل مستشفى سعودي. يمكن أن تشمل البيئة: VPN، شبكات، خوادم، EMR، نسخ احتياطي، Active Directory، MFA، شهادات، جدار ناري، تراخيص، مكتب مساعدة، أمن سيبراني.",
            "other": "الدور موظف عام داخل مستشفى سعودي. اختر قسمًا منطقيًا جديدًا في كل مرة: سريري، إداري، تقني، تشغيلي، دعم، تدريب، جودة.",
        }.get(role_type, "موظف في مستشفى سعودي.")
    return {
        "clinical": "Clinical role in a Saudi hospital. Possible context: doctor, nurse, pharmacy, laboratory, radiology, clinics, ER, ICU, patient records, PACS, LIS, EMR, handover, MOH clinical protocols.",
        "admin": "Administrative role in a Saudi hospital. Possible context: reception, medical secretary, patient records, insurance, billing, procurement, HR, accreditation, vendor contracts, payroll, administrative meetings.",
        "it": "IT/Informatics role in a Saudi hospital. Possible context: VPN, networks, servers, EMR, backups, Active Directory, MFA, certificates, firewall, licenses, helpdesk, cybersecurity.",
        "other": "General Saudi hospital employee. Choose a fresh logical department each time: clinical, administrative, technical, operations, support, training, quality.",
    }.get(role_type, "Saudi hospital employee.")







# =============================================================
# STRICT ROLE + DIFFICULTY GUARDRAILS (patched)
# -------------------------------------------------------------
# These checks reject AI outputs that drift away from the selected
# job role (clinical/admin/IT) or violate the documented difficulty
# framework before the email is shown to the trainee.
# =============================================================




def get_generation_quality_issues(result, difficulty, is_phishing=True):
    if not isinstance(result, dict):
        return ["result is not a JSON object"]
    body = (result.get("body") or "") or ""
    subject = (result.get("subject") or "") or ""
    sender = (result.get("from") or "") or ""
    link = (result.get("suspicious_link") or "") or ""
    attach = (result.get("attachment") or "") or ""
    attack_type = (result.get("attack_type") or "") or (result.get("email_type") or "") or ""
    combined = " ".join([body, subject, sender, link, attach, attack_type])
    domains = extract_domains_from_result(result)
    non_official = [d for d in domains if _domain_root(d) not in {"hospital.org", "moh.gov.sa"}]
    issues = []

    if is_phishing:
        has_vector = bool(non_official or attach or re.search(r'\bQR\b|رمز\s*QR|reply|رد\s|OTP|MFA|اتصل|call\s', combined, re.I))
        if not has_vector:
            issues.append("phishing item needs a clear vector: fake domain/link, attachment, QR, reply request, MFA/OTP, or phone request")
        if difficulty == "easy":
            if not _has_generic_greeting(body):
                issues.append("Beginner must use a generic greeting")
            if not re.search(r'password|credential|login|verify|account|bank|IBAN|OTP|MFA|كلمة مرور|بيانات الدخول|تحقق|حساب|آيبان|رمز', combined, re.I):
                issues.append("Beginner needs an obvious sensitive request")
            if not re.search(r'urgent|immediately|today|suspended|terminated|locked|عاجل|فورًا|اليوم|تعليق|إيقاف', combined, re.I):
                issues.append("Beginner needs obvious urgency/threat")
        elif difficulty == "medium":
            if _contains_long_all_caps(combined):
                issues.append("Intermediate must not use aggressive all-caps")
            if re.search(r'permanent termination|within 1 hour|act now|account closed|إنهاء دائم|خلال ساعة|تصرف الآن', combined, re.I):
                issues.append("Intermediate threat is too aggressive")
        elif difficulty == "hard":
            if _has_generic_greeting(body):
                issues.append("Advanced must use a personalized greeting")
            if _contains_long_all_caps(combined):
                issues.append("Advanced must not contain all-caps")
            if re.search(r'act now|failure to comply|account will be closed|enter your password|full credentials|تصرف الآن|سيتم إغلاق|أدخل كلمة المرور|بيانات الدخول كاملة', combined, re.I):
                issues.append("Advanced contains beginner-style direct threat or password request")
            if any(_domain_has_obvious_advanced_word(d) for d in non_official):
                issues.append("Advanced fake domain is too obvious")
            # Advanced should not be only an obvious link-verification email unless the context is spear-phishing.
            if re.search(r'verify your account|account verification|تحقق من حسابك|تأكيد الحساب', combined, re.I) and not re.search(r'meeting|ticket|case|shift|review|اعتماد|اجتماع|تذكرة|مناوبة|مراجعة', combined, re.I):
                issues.append("Advanced needs contextual spear-phishing, not generic account verification")
    else:
        bad = [d for d in domains if _domain_root(d) not in {"hospital.org", "moh.gov.sa"}]
        if bad:
            issues.append("Legitimate item must not contain external or fake domains")
        if re.search(r'password|credential|verify your account|enter your login|OTP|MFA|كلمة مرور|بيانات الدخول|تحقق من حسابك|رمز تحقق', combined, re.I):
            issues.append("Legitimate item must not ask for credentials/MFA/account verification")
        if re.search(r'suspended|terminated|locked|account closed|تعليق|إيقاف|إغلاق الحساب', combined, re.I):
            issues.append("Legitimate item must not threaten account suspension")
    return issues




# =============================================================
# STUDY 3 — FINAL 12-NOTE CONTENT OVERRIDE
# -------------------------------------------------------------
# Implements the final review notes collected from the five stages:
# 1) Tie the AI analysis directly to the generated attack type.
# 2) Keep Beginner / Intermediate / Advanced visibly different.
# 3) Keep assessment content aligned with the selected difficulty.
# 4) Improve Arabic and English parity.
# 5) Avoid repeated fixed templates and repeated example slots.
# 6) Diversify attack vectors: link, attachment, QR, reply, MFA/OTP, phone, shared document.
# 7) Diversify senders, recipients, departments, domains, and scenarios.
# 8) Do not over-focus analysis on Domain/Urgency/Spelling.
# 9) Prioritize sensitive requests and role-context risk in the analysis.
# 10) Keep legitimate assessment emails safe and clearly official.
# 11) Keep advanced phishing realistic and less obvious.
# 12) Reduce provider friction by using strict JSON, concise outputs, retries, and quality checks.
# =============================================================

ROLE_ATTACK_HINTS = {
    "clinical": "patient safety, EMR access, clinical workflow, handover, lab/radiology/pharmacy systems, or patient-data confidentiality",
    "admin": "billing, insurance, HR, procurement, patient-file administration, supplier payments, or executive workflow",
    "it": "network access, server administration, MFA, VPN, backups, security policy, identity management, or system availability",
    "other": "a realistic hospital workflow matching the selected mixed-role scenario",
}

ROLE_ATTACK_HINTS_AR = {
    "clinical": "سلامة المرضى، الوصول للسجلات الطبية، سير العمل السريري، التسليم، المختبر/الأشعة/الصيدلية، أو سرية بيانات المرضى",
    "admin": "الفوترة، التأمين، الموارد البشرية، المشتريات، ملفات المرضى، مدفوعات الموردين، أو سير العمل الإداري",
    "it": "وصول الشبكة، إدارة الخوادم، MFA، VPN، النسخ الاحتياطي، سياسات الأمن، إدارة الهوية، أو استمرارية الأنظمة",
    "other": "سير عمل مستشفى واقعي مطابق للسيناريو المختلط المختار",
}

ATTACK_ANALYSIS_PRIORITIES = [
    (r"OTP|MFA|رمز|مصادقة", "MFA / OTP Abuse", "طلب رمز تحقق أو اعتماد MFA"),
    (r"password|credential|login|username|IBAN|bank|كلمة مرور|بيانات الدخول|اسم المستخدم|آيبان|حساب بنكي", "Credential / Sensitive Data Request", "طلب بيانات دخول أو بيانات حساسة"),
    (r"invoice|payment|supplier|SAR|فاتورة|دفع|مورد|ريال", "Invoice / Payment Fraud", "احتيال فاتورة أو دفع"),
    (r"QR|رمز QR|scan", "QR Phishing", "تصيد عبر رمز QR"),
    (r"\.pdf|\.docx|\.xlsx|attachment|attached|مرفق|ملف", "Attachment-Based Attack", "هجوم عبر مرفق"),
    (r"reply|respond|send me|رد|أرسل", "Reply-Based Social Engineering", "هندسة اجتماعية عبر الرد"),
    (r"call|phone|extension|اتصل|هاتف|تحويلة", "Phone / Callback Phishing", "تصيد عبر اتصال أو رقم بديل"),
    (r"shared|document|drive|portal|مستند|مشترك|بوابة", "Cloud / Portal Phishing", "تصيد عبر مستند أو بوابة"),
]

def infer_attack_type_from_content(result, is_ar=False):
    combined = " ".join(str(result.get(k, "")) for k in ["email_type", "attack_type", "subject", "attachment", "body", "suspicious_text", "suspicious_link"])
    for pattern, en, ar in ATTACK_ANALYSIS_PRIORITIES:
        if re.search(pattern, combined, re.I):
            return ar if is_ar else en
    existing = result.get("attack_type") or result.get("email_type")
    return existing or ("تصيد موجه" if is_ar else "Spear Phishing")

def _place_link_in_body(body, link):
    """
    GUARANTEED link placement — removes link from anywhere it appears
    (including after the signature) and re-inserts it either:
      (a) immediately after the last cue sentence ('click the link below', etc.)
          that appears before the signature, OR
      (b) immediately before the closing signature block.
    This means the link can NEVER end up after 'Best regards / Regards / Sincerely'.
    """
    body = (body or "").strip()
    link = (link or "").strip()
    if not link or not body:
        return body

    # ── 1. Remove all occurrences of the link from the body ─────────────────
    lines = body.split("\n")
    cleaned = []
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        if stripped == link or (link in stripped and len(stripped) <= len(link) + 8):
            # eat a trailing blank line too
            if i + 1 < len(lines) and not lines[i + 1].strip():
                i += 1
        else:
            cleaned.append(lines[i])
        i += 1
    body = "\n".join(cleaned).strip()

    # ── 2. Find the signature block start ───────────────────────────────────
    _SIG_RE = re.compile(
        r"^(regards|best regards|kind regards|warm regards|thank you|sincerely|"
        r"respectfully|yours truly|yours sincerely|best wishes|"
        r"شكرا|شكراً|مع تحياتي|تحياتي|أطيب التحيات|وتفضلوا)",
        re.I,
    )
    lines = body.split("\n")
    sig_start = None
    for idx, line in enumerate(lines):
        if _SIG_RE.match(line.strip()):
            sig_start = idx
            break

    body_above = lines[:sig_start] if sig_start is not None else lines[:]
    sig_block  = lines[sig_start:]  if sig_start is not None else []

    # ── Early-exit: link is already present above the signature ─────────────
    # This happens when the AI writes the link inline in a sentence
    # (e.g. "visit this link: http://..."). Step 1 only removes standalone
    # link lines, so inline occurrences survive. If the link is already
    # above the signature, we must NOT insert a second copy.
    if link in "\n".join(body_above):
        merged = body_above[:]
        if sig_block:
            if merged and merged[-1].strip():
                merged.append("")
            merged.extend(sig_block)
        return "\n".join(merged).strip()
    _CUE_RE = re.compile(
        r"(link below|following link|the link|click here|click the|visit the|"
        r"access.*link|download.*link|open.*link|survey.*link|"
        r"link:|here:|click:|visit:|download:|scan:|submit:|complete:|"
        r"الرابط|اضغط|افتح|امسح|انقر|حمّل|زيارة|هنا:|أكمل)",
        re.I,
    )
    cue_idx = None
    for idx in range(len(body_above) - 1, -1, -1):
        if _CUE_RE.search(body_above[idx]):
            cue_idx = idx
            break

    # ── 4. Insert the link ───────────────────────────────────────────────────
    if cue_idx is not None:
        insert_at = cue_idx + 1
        # Skip any blank lines that immediately follow the cue
        while insert_at < len(body_above) and not body_above[insert_at].strip():
            insert_at += 1
        body_above.insert(insert_at, link)
    else:
        # No cue sentence — append just before the signature
        body_above.append(link)

    # ── 5. Reassemble with exactly one blank line before signature ───────────
    merged = body_above[:]
    if sig_block:
        if merged and merged[-1].strip():
            merged.append("")
        merged.extend(sig_block)

    return "\n".join(merged).strip()


# Backward-compatible aliases so every existing call site works unchanged.
def _insert_before_signature(body, marker):
    return _place_link_in_body(body, marker)


def _reposition_trailing_lone_link(body, link):
    return _place_link_in_body(body, link)

def _enforce_attack_vector(result, vector):
    """Many providers (Groq/Llama especially) acknowledge the requested
    attack vector in their reasoning but then default to inventing an
    attachment filename regardless, or skip writing the required
    [QR: ...] / [Label](url) marker in the body. This made every example
    look the same (always an attachment) even when the diversity plan
    asked for a link, QR code, or button. We post-process the raw result
    against the vector that was actually requested for this generation:
      - If the vector is NOT attachment-related, strip any attachment
        the model invented anyway.
      - If the vector IS attachment-related but the model gave a link
        instead, drop the stray link/QR markers and keep it attachment-only.
      - If the vector calls for a QR code and no [QR: ...] marker exists
        in the body, synthesize one.
      - If the vector calls for a link/button/portal and neither a QR nor
        a [Label](url) marker exists, synthesize a button.
    """
    if not isinstance(result, dict) or not vector:
        return result
    v = vector.lower()
    body = result.get("body")
    body = body if isinstance(body, str) else ""
    has_qr_marker  = bool(re.search(r'\[\s*QR', body, re.I))
    has_btn_marker = bool(re.search(r'\[[^\]]{1,80}\]\s*\(\s*https?://', body))

    wants_attachment = any(k in v for k in ["attachment", "pdf", "docx", "xlsm", "docm", "script"])
    wants_qr         = "qr" in v
    wants_link       = (not wants_attachment and not wants_qr and
                        any(k in v for k in ["link", "portal", "document", "url", "console", "enrollment", "share"]))

    def _as_str(v):
        return v if isinstance(v, str) else ("" if v is None else str(v))

    if not wants_attachment and _as_str(result.get("attachment")).strip():
        result["attachment"] = ""

    if wants_attachment and (has_qr_marker or has_btn_marker):
        # Vector says attachment but model produced a link/QR marker instead —
        # strip the markers (keep the plain sentence) so the email stays
        # attachment-only and doesn't show two vectors at once.
        body = re.sub(r'\[\s*QR(?:\s*Code)?\s*:?\s*[^\]]*\]', '', body, flags=re.I)
        body = re.sub(r'\[([^\]]{1,80})\]\s*\(\s*https?://[^\)\s]+\s*\)', r'\1', body)
        result["body"] = re.sub(r'[ \t]*\n[ \t]*\n[ \t]*\n+', '\n\n', body).strip()
        if not _as_str(result.get("attachment")).strip():
            if "docx" in v or "docm" in v:
                ext = ".docm" if "docm" in v else ".docx"
                stem = random.choice(["Patient_Report", "Handover_Notes", "Policy_Update", "Meeting_Minutes"])
            elif "xlsm" in v:
                ext, stem = ".xlsm", random.choice(["Backup_Verification", "Audit_Sheet", "Schedule_Update"])
            elif "script" in v:
                ext, stem = ".zip", random.choice(["Verification_Tool", "System_Update", "Backup_Script"])
            else:
                ext, stem = ".pdf", random.choice(["Patient_Report", "Invoice", "Policy_Notice", "Schedule_Update"])
            result["attachment"] = f"{stem}{ext}"

    elif wants_qr and not has_qr_marker:
        link = _as_str(result.get("suspicious_link")).strip() or "https://example-training-only.invalid/verify"
        result["suspicious_link"] = link
        result["body"] = _insert_before_signature(body, "[QR: Scan to continue]")

    elif wants_link and not has_qr_marker and not has_btn_marker:
        link = _as_str(result.get("suspicious_link")).strip() or "https://example-training-only.invalid/verify"
        result["suspicious_link"] = link
        result["body"] = _insert_before_signature(body, "[Open Link](" + link + ")")

    return result

def normalize_learning_analysis(result, role_type, difficulty, is_ar=False):
    """Post-processes model output so analysis is tied to the attack type and role context."""
    if not isinstance(result, dict):
        return result
    result = _recover_from_nested_email_blob(result)
    result = _enforce_attack_vector(result, st.session_state.get("_last_learn_vector", ""))
    # Defensive: trust what the model actually wrote over the requested
    # language flag. If the body/subject came back in Arabic script despite
    # English being requested (or vice versa), all of OUR appended sentences
    # below must match the model's actual language — otherwise we end up
    # gluing an English clause onto an Arabic sentence (or the reverse).
    is_ar = _detect_is_ar(result, is_ar)
    attack_type = result.get("attack_type") or infer_attack_type_from_content(result, is_ar)
    result["attack_type"] = attack_type
    if not result.get("email_type"):
        result["email_type"] = attack_type
    risk = result.get("risk_level") or ("Medium" if difficulty == "medium" else ("High" if difficulty == "hard" else "Low"))
    result["risk_level"] = risk
    role_hint = (ROLE_ATTACK_HINTS_AR if is_ar else ROLE_ATTACK_HINTS).get(role_type, ROLE_ATTACK_HINTS["other"])

    indicators = result.get("indicators")
    if not isinstance(indicators, list):
        indicators = []
    # Defensive: some providers occasionally return indicator items as plain
    # strings instead of {"number","title","description"} dicts, which
    # crashes any .get() call on them downstream. Normalize every item to a
    # dict shape before anything else touches it.
    indicators = [
        it if isinstance(it, dict) else {"number": i+1, "title": str(it or ""), "description": ""}
        for i, it in enumerate(indicators)
    ]
    while len(indicators) < 3:
        indicators.append({"number": len(indicators)+1, "title": "", "description": ""})

    # Indicator 1 must connect the attack type to the risky action — but we
    # only OVERWRITE the model's own wording when it's missing/empty or is a
    # generic placeholder. Otherwise we keep the model's real (varied,
    # non-repeating) analysis and merely make sure attack_type is mentioned
    # somewhere in it. Forcing the exact same template sentence every time
    # (regardless of provider/example) was the root cause of every "AI Tutor
    # Analysis" panel reading identically across different examples.
    def _is_placeholder(desc):
        d = (desc or "").strip()
        return (not d) or len(d) < 25

    if is_ar:
        d0 = (indicators[0].get("description") or "").strip()
        if _is_placeholder(d0):
            indicators[0] = {
                "number": 1,
                "title": indicators[0].get("title") or f"نوع الهجوم: {attack_type}",
                "description": f"الخطر الأساسي هنا مرتبط بـ {attack_type} داخل سياق {role_hint}."
            }
        elif attack_type not in d0:
            indicators[0]["description"] = f"{d0} (هذا يرتبط مباشرة بنوع هجوم {attack_type}.)"
        if _is_placeholder(indicators[1].get("description")) or re.search(r"^النطاق$|^Domain$", indicators[1].get("title", "").strip(), re.I):
            indicators[1] = {"number": 2, "title": indicators[1].get("title") or "طلب أو سلوك غير معتاد", "description": indicators[1].get("description") or "الرسالة تطلب إجراءً لا يتم عادة عبر بريد عادي في بيئة المستشفى."}
        if _is_placeholder(indicators[2].get("description")) or re.search(r"^إملاء$|^Spelling$", indicators[2].get("title", "").strip(), re.I):
            indicators[2] = {"number": 3, "title": indicators[2].get("title") or "عدم توافق السياق أو القناة", "description": indicators[2].get("description") or "القناة أو المرسل لا يطابقان طريقة التعامل الرسمية مع هذا النوع من الطلبات."}
        wr = (result.get("why_risky") or "").strip()
        if _is_placeholder(wr):
            wr = f"هذه رسالة {attack_type} بمستوى خطورة {risk}. قد تؤثر على {role_hint} إذا تم تنفيذ الطلب دون تحقق."
        elif attack_type not in wr:
            wr = f"{wr} (هذه رسالة {attack_type}.)"
        result["why_risky"] = wr
        tip = (result.get("learning_tip") or "").strip()
        if not tip:
            tip = "تحقق من الطلب عبر قناة المستشفى الرسمية قبل فتح رابط أو مرفق أو مشاركة أي بيانات."
        result["learning_tip"] = tip
    else:
        d0 = (indicators[0].get("description") or "").strip()
        if _is_placeholder(d0):
            indicators[0] = {
                "number": 1,
                "title": indicators[0].get("title") or f"Attack Type: {attack_type}",
                "description": f"The main risk is {attack_type} in a hospital role involving {role_hint}."
            }
        elif attack_type not in d0:
            indicators[0]["description"] = f"{d0} (This directly ties to the {attack_type} attack pattern.)"
        if _is_placeholder(indicators[1].get("description")) or re.search(r"^Domain$", indicators[1].get("title", "").strip(), re.I):
            indicators[1] = {"number": 2, "title": indicators[1].get("title") or "Unusual request or workflow", "description": indicators[1].get("description") or "The message asks for an action that should normally use an official hospital channel."}
        if _is_placeholder(indicators[2].get("description")) or re.search(r"^Spelling$", indicators[2].get("title", "").strip(), re.I):
            indicators[2] = {"number": 3, "title": indicators[2].get("title") or "Role-context or channel mismatch", "description": indicators[2].get("description") or "The sender or channel does not match how this workplace process should be handled."}
        wr = (result.get("why_risky") or "").strip()
        if _is_placeholder(wr):
            wr = f"This is a {attack_type} email with a {risk} risk level. It can affect {role_hint} if the recipient acts without verification."
        elif attack_type not in wr:
            wr = f"{wr} (This is a {attack_type} email.)"
        result["why_risky"] = wr
        tip = (result.get("learning_tip") or "").strip()
        if not tip:
            tip = "Verify the request through an official hospital channel before opening links, attachments, QR codes, or sharing data."
        result["learning_tip"] = tip

    # Keep exactly three indicators and correct numbering.
    result["indicators"] = [{**indicators[i], "number": i+1} for i in range(3)]
    return result

def clear_gsheet_data():
    """Wipe all data rows (keeping headers) from both synced tabs. Used by
    'Reset All Metrics' — now that load_runs() merges in the durable
    Google Sheet copy, a local-only reset would otherwise be silently
    undone on the next page load as old rows get pulled back in."""
    client = _get_gsheet_client()
    sheet_id = _get_gsheet_id()
    if not client or not sheet_id:
        return
    for tab_name in ["Cycle Ratings", "Auto Metrics"]:
        try:
            sheet = client.open_by_key(sheet_id)
            ws = sheet.worksheet(tab_name)
            n_rows = ws.row_count
            if n_rows > 1:
                ws.delete_rows(2, n_rows)
        except Exception:
            pass
    st.session_state.pop("_gsheet_runs_cache", None)

def delete_run_from_gsheet(record):
    """Remove the matching row (by timestamp+provider+language) from the
    'Cycle Ratings' tab. Used by the Undo button so a removed cycle
    doesn't silently reappear on the next load_runs() merge — without
    this, Undo would only ever be temporary since Google Sheets is the
    durable source of truth and gets re-merged in on every page load."""
    client = _get_gsheet_client()
    sheet_id = _get_gsheet_id()
    if not client or not sheet_id or not record:
        return
    try:
        sheet = client.open_by_key(sheet_id)
        ws = sheet.worksheet("Cycle Ratings")
        all_vals = ws.get_all_values()
        if not all_vals:
            return
        headers = all_vals[0]
        try:
            ts_i = headers.index("timestamp")
            prov_i = headers.index("provider")
            lang_i = headers.index("language")
        except ValueError:
            return
        target = (str(record.get("timestamp", "")), str(record.get("provider", "")), str(record.get("language", "")))
        for row_idx in range(len(all_vals) - 1, 0, -1):
            row = all_vals[row_idx]
            if len(row) > max(ts_i, prov_i, lang_i):
                if (row[ts_i], row[prov_i], row[lang_i]) == target:
                    ws.delete_rows(row_idx + 1)  # +1: sheet rows are 1-indexed
                    break
        # Invalidate the cached pull so the next load_runs() reflects the deletion.
        st.session_state.pop("_gsheet_runs_cache", None)
    except Exception:
        pass

def _is_nonempty_str(v):
    return isinstance(v, str) and bool(v.strip())

def _detect_is_ar(result, fallback=False):
    """Decide the real language of a generated email from its own prose,
    ignoring URLs and our own [QR:...]/[Label](url) markup — those always
    contain Latin characters (and English words like "Button"/"QR") even
    inside a fully Arabic email, which previously made the script-based
    language check always see "mixed" content and silently fall back to
    the (unreliable) caller-provided flag. Stripping markup first means a
    genuinely Arabic email is correctly detected as Arabic even though its
    button/link markup is in Latin script."""
    subject = result.get("subject") or ""
    body = result.get("body") or ""
    sample = f"{subject} {body}"
    sample = re.sub(r"https?://\S+", " ", sample)
    sample = re.sub(r"\[[^\]]{1,80}\](?:\s*\([^)]*\))?", " ", sample)
    sample = re.sub(r"\b(Button|QR|Open|Link|Document|pdf|docx|xlsx|xlsm|docm)\b", " ", sample, flags=re.I)
    has_ar = bool(re.search(r"[\u0600-\u06FF]", sample))
    has_lat = bool(re.search(r"[A-Za-z]{4,}", sample))
    if has_ar and not has_lat:
        return True
    if has_lat and not has_ar:
        return False
    return fallback

def _is_substantial_str(v, min_len=15):
    return isinstance(v, str) and len(v.strip()) >= min_len

_GENERIC_HOSPITAL_NAME = {"en": "Riyadh Specialist Hospital", "ar": "مستشفى الرياض التخصصي"}

def _resolve_leftover_placeholders(result, is_ar=False):
    """Some providers occasionally leave an unfilled template placeholder
    inside the body instead of inventing real content, e.g. literally
    writing "[Hospital Name]" or "[اسم المستشفى]" instead of a name. These
    survive because they aren't QR/link markers, so the renderer doesn't
    touch them. Replace any leftover bracket placeholder that looks like a
    hospital-name slot with a real, consistent generic name; for any other
    leftover bracket placeholder, just drop the brackets and keep the
    (likely still-readable) text inside, since a literal bracket label
    looks far more obviously broken than blending it into the sentence."""
    body = result.get("body")
    if not isinstance(body, str) or "[" not in body:
        return result
    hospital_name = _GENERIC_HOSPITAL_NAME["ar" if is_ar else "en"]
    body = re.sub(r"\[\s*(?:اسم\s*المستشفى|hospital\s*name)\s*\]", hospital_name, body, flags=re.I)
    # Any other still-unresolved bracket placeholder that isn't one of our
    # own QR/link/button markers (those are handled separately at render
    # time) — drop the brackets but keep the label so it doesn't look like
    # raw template syntax leaked into a "real" email.
    body = re.sub(r"\[([^\]]{1,40})\](?!\s*\()", lambda m: m.group(1) if not re.match(r"^\s*QR", m.group(1), re.I) else m.group(0), body)
    result["body"] = body
    return result


# Stronger difficulty contract: short, provider-friendly, and visibly different.

def build_prompt(role, index, language):
    base = _OLD_BUILD_PROMPT_STUDY3(role, index, language)
    is_ar = (language == "Arabic")
    extra = """

قاعدة نهائية مهمة:
- يجب أن يرتبط التحليل مباشرة بنوع الهجوم المكتوب في attack_type.
- أول مؤشر في indicators يجب أن يبدأ بنوع الهجوم، وليس النطاق دائمًا.
- اربط كل سبب بسياق الدور الوظيفي والمستشفى.
- لا تجعل التحليل أطول من اللازم.
""" if is_ar else """

Final important rule:
- The AI analysis must directly connect to the attack_type field.
- The first indicator must start from the attack type, not always the domain.
- Link each reason to the role context and hospital workflow.
- Keep the analysis concise.
"""
    return base + extra

def build_assess_prompt(role, index, is_phishing, language):
    base = _OLD_BUILD_ASSESS_PROMPT_STUDY3(role, index, is_phishing, language)
    is_ar = (language == "Arabic")
    extra = """

قاعدة الاختبار النهائية:
- ركّز فقط على محتوى البريد وتصنيفه وصعوبته.
- لا تضف تحليل تعليمي طويل داخل سؤال الاختبار.
- يجب أن يكون البريد مناسبًا تمامًا لمستوى الصعوبة المختار.
""" if is_ar else """

Final assessment rule:
- Focus only on the email content, correct label, and selected difficulty.
- Do not add long learning analysis inside assessment questions.
- The email must clearly fit the selected difficulty level.
"""
    return base + extra

# Override generators to apply the final normalization after each provider response.

_DEBUG_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "debug_log.json")

def _store_debug(stage, err):
    """Keep technical diagnostics out of the user-facing UI; persist them to
    disk (not just session_state) so the admin/debug panel still shows the
    real error even if the page was reloaded or the admin panel was opened
    in a fresh session — which was making the log look empty right when it
    was needed most."""
    entry = {"stage": stage, "error": err, "ts": __import__("time").time()}
    try:
        log = st.session_state.setdefault("_debug_log", [])
        log.append(entry)
        st.session_state["_debug_log"] = log[-20:]
    except Exception:
        pass
    try:
        with open(_DEBUG_LOG_PATH, "r", encoding="utf-8") as f:
            disk_log = json.load(f)
        if not isinstance(disk_log, list):
            disk_log = []
    except Exception:
        disk_log = []
    disk_log.append(entry)
    disk_log = disk_log[-20:]
    try:
        with open(_DEBUG_LOG_PATH, "w", encoding="utf-8") as f:
            json.dump(disk_log, f, ensure_ascii=False, default=str)
    except Exception:
        pass

def _load_debug_log():
    try:
        with open(_DEBUG_LOG_PATH, "r", encoding="utf-8") as f:
            disk_log = json.load(f)
        if isinstance(disk_log, list):
            return disk_log
    except Exception:
        pass
    return []











# Make the final system prompt shorter and more explicit for all providers.

# =============================================================
# END FINAL DIVERSITY + DIFFICULTY OVERRIDE
# =============================================================

# ══════════════════════════════════════════════════════════════


# =============================================================
# FINAL PATCH — Stability, diversity, and analysis-quality fixes
# =============================================================
# Notes covered:
# - Stronger provider retries for Gemini 503/high-demand and slow providers.
# - Claude/Gemini get shorter prompts and larger timeout/token safety.
# - Learning analysis always links attack_type to the explanation.
# - Placeholder indicator titles are replaced with meaningful titles.
# - Beginner/Intermediate/Advanced are kept visibly different.
# - Assessment stays focused on email difficulty and classification only.

import time as _time_patch

_PROVIDER_RETRYABLE_PATTERNS = re.compile(
    r"503|UNAVAILABLE|high demand|temporar|try again|rate limit|overloaded|timeout|timed out|deadline",
    re.I,
)

# Keep the original network caller, then wrap it with provider-aware retry.
_BASE_CALL_AI_FINAL = call_ai

def _is_retryable_ai_error(data):
    if not isinstance(data, dict) or "error" not in data:
        return False
    err = data.get("error")
    msg = err.get("message", str(err)) if isinstance(err, dict) else str(err)
    return bool(_PROVIDER_RETRYABLE_PATTERNS.search(msg))

def _compact_prompt_for_slow_provider(prompt):
    """Shorten long prompts for Claude/Gemini without removing the JSON contract."""
    if not isinstance(prompt, str) or len(prompt) < 3500:
        return prompt
    # Preserve the most important parts: top task + JSON contract + final rules.
    head = prompt[:2200]
    tail = prompt[-1800:]
    return head + "\n\n[Prompt shortened: keep all rules above; avoid repetition; return valid JSON only.]\n\n" + tail

def call_ai(prompt, max_tokens=1600):
    provider = st.session_state.get("ai_provider", "groq")
    attempts = 3 if provider in {"gemini", "anthropic", "groq"} else 2
    last = None
    for attempt in range(attempts):
        use_prompt = _compact_prompt_for_slow_provider(prompt) if provider in {"gemini", "anthropic"} else prompt
        data = _BASE_CALL_AI_FINAL(use_prompt, max_tokens=max(max_tokens, 3500 if provider in {"gemini", "anthropic"} else max_tokens))
        if not _is_retryable_ai_error(data):
            return data
        last = data
        _time_patch.sleep(1.2 * (attempt + 1))
    return last or {"error": {"message": "AI provider failed after retries."}}


# Stronger prompt wording that avoids placeholder-like indicator titles.
_BASE_BUILD_PROMPT_FINAL = build_prompt
_BASE_BUILD_ASSESS_PROMPT_FINAL = build_assess_prompt

def build_prompt(role, index, language):
    base = _BASE_BUILD_PROMPT_FINAL(role, index, language)
    if language == "Arabic":
        return base + """

تعليمات جودة نهائية:
- لا تكتب عناوين عامة مثل "مؤشر ثالث" أو "مؤشر مختلف".
- اكتب أسماء مؤشرات واضحة مثل: طلب بيانات حساسة، إساءة MFA/OTP، مرفق مشبوه، QR غير موثوق، انتحال سلطة، عدم توافق سير العمل.
- يجب أن يوضح أول مؤشر: نوع الهجوم + الإجراء الخطر داخل البريد.
- يجب أن تكون أمثلة المبتدئ واضحة، والمتوسط مقنعة جزئياً، والمتقدم قريب من الشرعي لكن لا يزال خطراً.
- اجعل كل مثال جديد مختلفاً في: المرسل، النطاق، القسم، الطلب، وناقل الهجوم.
"""
    return base + """

Final quality instructions:
- Do not write generic placeholder titles like "different clue" or "third different clue".
- Use concrete indicator titles such as: Sensitive data request, MFA/OTP abuse, Suspicious attachment, Untrusted QR code, Authority impersonation, Workflow mismatch.
- Indicator 1 must state the attack type plus the risky action inside the email.
- Beginner must be obvious, Intermediate partly convincing, Advanced near-legitimate but still risky.
- Make every example differ in sender, domain, department, request, and attack vector.
"""

def build_assess_prompt(role, index, is_phishing, language):
    base = _BASE_BUILD_ASSESS_PROMPT_FINAL(role, index, is_phishing, language)
    if language == "Arabic":
        return base + """

تعليمات اختبار نهائية:
- لا تضف تحليل المعلم داخل الاختبار.
- البريد نفسه فقط يجب أن يثبت مستوى الصعوبة المختار.
- إذا كان شرعياً: لا روابط خارجية، لا تهديد، لا بيانات حساسة، ولا نطاق غير رسمي.
- إذا كان تصيداً: اجعل العلامات مناسبة للصعوبة، وليس كلها واضحة في المتقدم.
"""
    return base + """

Final assessment instructions:
- Do not include tutor-analysis sections in assessment questions.
- The email content itself must prove the selected difficulty level.
- If legitimate: no external links, no threats, no sensitive requests, and no unofficial domain.
- If phishing: red flags must match the difficulty; Advanced must not look Beginner-level obvious.
"""

# =============================================================
# UX REALISM PATCH — QR/link rendering contract + professional tone
# -------------------------------------------------------------
# EN: Adds two things on top of every previous prompt layer:
#  1) A strict, render-friendly contract for how a QR code and a
#     clickable link must be written inside `body`, so the UI can
#     turn them into a REAL scannable QR image / a REAL clickable
#     button instead of literal bracket text like "[QR Code: ...]".
#  2) An instruction to make the wording read like a genuinely
#     professional, polished email — not generic training copy.
# AR: تضيف طبقتين فوق كل طبقات التعليمات السابقة:
#  1) صيغة ثابتة وصارمة لكتابة رمز QR والرابط داخل body، حتى تقدر
#     الواجهة تحوّلها لصورة QR فعلية قابلة للمسح / زر فعلي قابل
#     للنقر، بدل نص بين أقواس لا يظهر كباركود حقيقي.
#  2) تعليمة لجعل صياغة الرسالة تقرأ كبريد احترافي حقيقي، وليس
#     نصًا تدريبيًا عامًا.
# =============================================================
_BASE_BUILD_PROMPT_UX = build_prompt
_BASE_BUILD_ASSESS_PROMPT_UX = build_assess_prompt

_UX_CONTRACT_AR = """

تنسيق إلزامي: QR → اكتب مرة واحدة فقط [QR: نص قصير]، والرابط الفعلي في suspicious_link فقط لا تكرره بالنص أو التوقيع.
زر → اكتب مرة واحدة فقط [نص الزر](الرابط)، ولا تكرر الرابط بعدها.
اجعل الصياغة كاملة تقرأ كبريد احترافي حقيقي، لا حشو تدريبي.
مهم جدًا: أعد القيم بالحقول المباشرة بالمستوى الأول فقط كما هي بالمخطط تمامًا، بدون أي حقول إضافية أو متداخلة غير المطلوبة. استخدم نفس أسماء الحقول المحددة بالمخطط حرفيًا، بدون أي تسمية بديلة.
اكتب المحتوى بلغة واحدة فقط هي اللغة المطلوبة بهذا الطلب. لا تضف نسخة ثانية بلغة أخرى بأي مكان من الرد.
"""
_UX_CONTRACT_EN = """

Mandatory format: QR -> write ONCE [QR: short label]; real URL only in suspicious_link, never repeated in body/signature.
Button -> write ONCE [Button label](url); never print that URL again afterward.
Write the whole message like a real professional email, no generic training filler.
CRITICAL: Return values directly as the top-level fields defined in the schema only — no extra or nested fields beyond what is shown. Use the exact field names given in the schema, character for character — do not substitute alternate names.
Write the content in ONLY the single language requested for this request. Do not include a second-language version anywhere in the response.
"""

def build_prompt(role, index, language):
    base = _BASE_BUILD_PROMPT_UX(role, index, language)
    base = base + (_UX_CONTRACT_AR if language == "Arabic" else _UX_CONTRACT_EN)
    if language == "Arabic":
        return base + "\n\nتعليمة لغة نهائية صارمة: كل حقل نصي (subject, body, indicators, why_risky, learning_tip) يجب يكون بالعربية الفصحى بالكامل. ممنوع أي كلمة أو جملة إنجليزية إلا أسماء العلامات التجارية أو الاختصارات التقنية الشائعة (مثل MFA، OTP، PACS). أي رد يحتوي فقرة كاملة بالإنجليزية يعتبر خطأ ويجب رفضه.\n"
    return base + "\n\nFINAL strict language directive: every text field (subject, body, indicators, why_risky, learning_tip) must be written entirely in English. Do not include any Arabic word or sentence anywhere in the response.\n"

def build_assess_prompt(role, index, is_phishing, language):
    base = _BASE_BUILD_ASSESS_PROMPT_UX(role, index, is_phishing, language)
    base = base + (_UX_CONTRACT_AR if language == "Arabic" else _UX_CONTRACT_EN)
    if language == "Arabic":
        return base + "\n\nتعليمة لغة نهائية صارمة: كل حقل نصي يجب يكون بالعربية الفصحى بالكامل. ممنوع أي جملة إنجليزية كاملة إلا الاختصارات التقنية الشائعة.\n"
    return base + "\n\nFINAL strict language directive: every text field must be written entirely in English. Do not include any Arabic word or sentence anywhere in the response.\n"
# =============================================================
# END UX REALISM PATCH
# =============================================================

# =============================================================
# DIFFICULTY ENFORCEMENT PATCH
# Enforces the 4-axis framework strictly at prompt level:
# - EASY: generic greeting, plain text URL, NO QR, NO button, 2 spelling errors
# - MEDIUM: semi-personal greeting, simple button, NO QR, 1 spelling error
# - HARD: full name+title, QR mandatory, professional button, zero errors
# Also post-processes the AI output to remove QR from easy/medium
# =============================================================
_BASE_BUILD_PROMPT_DIFF = build_prompt
_BASE_BUILD_ASSESS_PROMPT_DIFF = build_assess_prompt

_DIFF_ADDON_EASY_AR = """

⚠️ تعليمات صارمة جداً لمستوى السهل — يجب الالتزام بها حرفياً:
1. التحية: "عزيزي الموظف" أو "عزيزي الزميل" فقط — ممنوع منعاً باتاً أي اسم شخصي.
2. هوية المرسل (from): يجب أن تكون جهة عامة أو اسم قسم فقط (مثل "فريق الدعم الفني" أو "قسم الموارد البشرية") — ممنوع تماماً أن يكون المرسل شخصاً باسمه الشخصي أو بلقب وظيفي (ممنوع كتابة "د." أو "Dr." أو اسم كامل + منصب) حتى في التوقيع بنهاية الرسالة.
3. النطاق: يجب أن يكون واضح التزوير تماماً ولا يشبه أي جهة حقيقية إطلاقاً — ممنوع استخدام كلمات مثل gov أو moh أو board أو ministry أو أي تركيبة تحاكي moh.gov.sa أو hospital.org.
4. الأخطاء: ضع بالضبط خطأين إملائيين أو نحويين واضحين في جسم الرسالة — هذا إلزامي ومطلوب.
5. الرابط: ضع الرابط كنص خام مرئي فقط في جسم الرسالة (مثل: http://fake-hospital.com/update) — ممنوع استخدام زر أو markdown.
6. QR: محظور تماماً — لا تكتب [QR:...] أبداً.
7. المرفق: محظور تماماً — الحقل attachment يجب أن يبقى فارغاً تماماً.
8. الإلحاح: صريح ومباشر ("الآن فوراً" أو "سيُغلق حسابك اليوم").
"""

_DIFF_ADDON_EASY_EN = """

⚠️ STRICT EASY LEVEL RULES — follow these literally or the output is invalid:
1. Greeting: MUST be "Dear Employee" or "Dear Staff" — ANY personal name is FORBIDDEN.
2. Sender identity (from): MUST be a generic team or department only (e.g. "IT Support Team" or "HR Department") — a named individual with a personal title (e.g. "Dr.", a full first+last name, a job title in the signature) is STRICTLY FORBIDDEN, including in the closing signature.
3. Domain: must be completely and obviously fake, with NO resemblance to any real organization — do NOT use words like gov, moh, board, ministry, or any pattern that mimics moh.gov.sa or hospital.org.
4. Errors: place EXACTLY TWO obvious spelling/grammar mistakes in the body — this is REQUIRED.
5. Link: place the URL as RAW VISIBLE PLAIN TEXT in the body (e.g. http://fake-hospital.com/update) — NO button, NO markdown link.
6. QR: COMPLETELY FORBIDDEN — do NOT write [QR:...] anywhere.
7. Attachment: FORBIDDEN — the "attachment" field MUST be left completely empty.
8. Urgency: direct and explicit ("Act NOW", "your account closes TODAY").
"""

_DIFF_ADDON_MEDIUM_AR = """

⚠️ تعليمات صارمة جداً لمستوى المتوسط — يجب الالتزام بها حرفياً:
1. التحية: استخدم الاسم الأول أو المسمى الوظيفي فقط (مثل "عزيزي د. أحمد") — ممنوع التحية العامة الكاملة وممنوع الاسم الكامل + اللقب الدقيق.
2. النطاق: مشابه للرسمي مع فرق بسيط يمكن ملاحظته عند التدقيق (مثل hospital-it.net) — ليس واضح التزوير تماماً كمستوى السهل، وليس شبه رسمي بذكاء كمستوى الصعب.
3. الأخطاء: ضع بالضبط خطأً إملائياً واحداً خفيفاً في جسم الرسالة — هذا إلزامي.
4. الرابط: استخدم زراً بسيطاً [نص](رابط) — مسموح. مرفق PDF بسيط وعام مسموح أيضاً (اختياري).
5. QR: محظور تماماً — لا تكتب [QR:...] أبداً.
6. الارتباط الوظيفي: يجب ذكر اسم القسم الحقيقي (سريري/إداري/تقني) داخل الرسالة.
"""

_DIFF_ADDON_MEDIUM_EN = """

⚠️ STRICT INTERMEDIATE LEVEL RULES — follow these literally:
1. Greeting: use first name or job title ONLY (e.g. "Dear Dr. Ahmed") — no fully generic greeting and no full name + precise title.
2. Domain: look-alike to the real one with a small detectable difference (e.g. hospital-it.net) — not obviously fake like Easy, and not near-official like Advanced.
3. Errors: place EXACTLY ONE subtle spelling/grammar mistake in the body — this is REQUIRED.
4. Link: a simple button [label](url) is allowed. A simple generic PDF attachment is also allowed (optional).
5. QR: COMPLETELY FORBIDDEN — do NOT write [QR:...] anywhere.
6. Role alignment: mention the real department name (Clinical/Administrative/IT) inside the email.
"""

_DIFF_ADDON_HARD_AR = """

⚠️ تعليمات صارمة جداً لمستوى الصعب — يجب الالتزام بها حرفياً (الكل إلزامي معاً):
1. التحية: الاسم الكامل + اللقب الوظيفي الدقيق (مثل "عزيزتي د. نورة العتيبي، استشارية الأمراض الداخلية").
2. النطاق: شبه رسمي بذكاء (مثل moh-staff.net) — ممنوع كلمات: secure, update, verify, login, reset.
3. المرفق: إلزامي ومطلوب دائماً — يجب أن يحتوي الحقل attachment على اسم مستند رسمي واقعي (مثل Compliance_Protocol_2024.pdf).
4. QR: اختياري ومتنوّع — لا تضعه في كل إيميل. ضعه فقط إذا كان يخدم السيناريو فعلاً (مثل تحقق عبر جهاز موبايل أو تسجيل جهاز)، واكتبه [QR: نص قصير وصفي]. إذا لم يكن مناسباً للسيناريو، لا تكتب أي رمز QR إطلاقاً واعتمد على الزر أو الرابط فقط.
5. الزر: استخدم زراً رسمياً باسم وصفي واضح (ليس "Open Link" أو "اضغط هنا").
6. الأخطاء: صفر أخطاء — لغة احترافية كاملة.
7. الإلحاح: خفيف ومهذب فقط ("إجراء روتيني") — ممنوع أي تهديد.
8. التوقيع: اسم كامل + المنصب + القسم + رقم تحويلة داخلية حقيقي (ليس XX).
9. الارتباط الوظيفي: يجب ربط الرسالة بمهمة يومية محددة جداً للدور المختار، وذكر اسم زميل أو قسم داخلي محدد (وليس عاماً).
"""

_DIFF_ADDON_HARD_EN = """

⚠️ STRICT ADVANCED LEVEL RULES — follow these literally (ALL are mandatory together):
1. Greeting: FULL NAME + precise job title (e.g. "Dear Dr. Noura Al-Otaibi, Internal Medicine Consultant").
2. Domain: near-official but not matching (e.g. moh-staff.net) — FORBIDDEN words: secure, update, verify, login, reset.
3. Attachment: MANDATORY AND REQUIRED — the "attachment" field MUST contain a realistic, officially named document (e.g. Compliance_Protocol_2024.pdf).
4. QR: OPTIONAL and VARIED — do NOT include it in every email. Only include [QR: short descriptive label] when it genuinely fits the scenario (e.g. mobile device check-in, device enrollment). If it doesn't fit naturally, omit it completely and rely on the button/link instead.
5. Button: use a professionally descriptive label (NOT "Open Link" or "Click Here").
6. Errors: ZERO spelling or grammar errors.
7. Urgency: polite and subtle ONLY ("routine procedure") — NO threats.
8. Signature: full name + title + department + real internal extension (no XX placeholders).
9. Role alignment: tie the email to a very specific daily task of the selected role, and name a specific colleague or internal department (not generic).
"""

def build_prompt(role, index, language):
    base = _BASE_BUILD_PROMPT_DIFF(role, index, language)
    difficulty = st.session_state.get("difficulty", "medium")
    is_ar = (language == "Arabic")
    if difficulty == "easy":
        base += _DIFF_ADDON_EASY_AR if is_ar else _DIFF_ADDON_EASY_EN
    elif difficulty == "medium":
        base += _DIFF_ADDON_MEDIUM_AR if is_ar else _DIFF_ADDON_MEDIUM_EN
    elif difficulty in ("hard", "advanced"):
        base += _DIFF_ADDON_HARD_AR if is_ar else _DIFF_ADDON_HARD_EN
    return base

def build_assess_prompt(role, index, is_phishing, language):
    base = _BASE_BUILD_ASSESS_PROMPT_DIFF(role, index, is_phishing, language)
    difficulty = st.session_state.get("difficulty", "medium")
    is_ar = (language == "Arabic")
    if difficulty == "easy":
        base += _DIFF_ADDON_EASY_AR if is_ar else _DIFF_ADDON_EASY_EN
    elif difficulty == "medium":
        base += _DIFF_ADDON_MEDIUM_AR if is_ar else _DIFF_ADDON_MEDIUM_EN
    elif difficulty in ("hard", "advanced"):
        base += _DIFF_ADDON_HARD_AR if is_ar else _DIFF_ADDON_HARD_EN
    return base

# =============================================================
# END DIFFICULTY ENFORCEMENT PATCH
# =============================================================

_BAD_INDICATOR_TITLES = re.compile(
    r"^(different behavioral or technical clue|third different clue|indicator\s*\d+|مؤشر\s*\d+|مؤشر ثالث|مؤشر سلوكي أو تقني مختلف)$",
    re.I,
)

def _clean_indicator_title(title, attack_type, n, is_ar=False):
    t0 = (title or "").strip()
    if t0 and not _BAD_INDICATOR_TITLES.match(t0):
        return t0
    if is_ar:
        defaults = [f"نوع الهجوم: {attack_type}", "طلب أو سير عمل غير معتاد", "عدم توافق القناة أو الدور"]
    else:
        defaults = [f"Attack Type: {attack_type}", "Unusual request or workflow", "Role-context or channel mismatch"]
    return defaults[min(max(n-1,0),2)]

_BASE_NORMALIZE_LEARNING_FINAL = normalize_learning_analysis

_FIELD_SYNONYMS = {
    "from":             ["from", "sender", "sender_email", "sender_address", "from_email", "from_address", "sender_display"],
    "to":               ["to", "recipient", "recipient_email", "to_email", "to_address"],
    "subject":          ["subject", "email_subject", "title"],
    "body":             ["body", "body_html", "body_text", "message", "message_body", "content", "email_body"],
    "suspicious_link":  ["suspicious_link", "link", "malicious_link", "phishing_link", "url"],
    "attachment":       ["attachment", "attachment_name", "file_attachment"],
}
_CANONICAL_FIELDS = list(_FIELD_SYNONYMS.keys())

def _strip_html(text):
    """Quick plain-text fallback for a body field that came back as
    body_html (e.g. '<p>Dear Doctor</p>') instead of plain text."""
    if not isinstance(text, str):
        return text
    text = re.sub(r"</p>|</div>|<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()

def _try_parse_blob(value):
    """If `value` is a string that LOOKS like a stringified dict (Python-
    repr style, single-quoted — not valid JSON), parse it into a real
    dict. Tries the exact/clean parse first, then a couple of common
    real-world breakages (stray backslashes) before giving up."""
    if not isinstance(value, str):
        return None
    s = value.strip()
    if not (s.startswith("{") and s.endswith("}")):
        return None
    for candidate in (s, re.sub(r'\\+(?=["\'])', '', s)):
        try:
            parsed = ast.literal_eval(candidate)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            continue
    return None

def _deep_flatten(value, pool, depth=0):
    """Walk `value` (dict, or a string that might itself be a stringified
    dict) and collect EVERY scalar key/value pair it contains into `pool`,
    at any nesting depth, under any key name the model invented — so it
    does not matter whether a provider nests content under 'email',
    'email_content', 'analysis', or anything else not seen before.
    Shallower keys win: once a key is in the pool, deeper duplicates of
    the same key name are ignored, so top-level fields are never
    overwritten by something buried inside a sub-blob."""
    if depth > 6:
        return
    if isinstance(value, dict):
        # First pass at this level: capture plain scalars so this level's
        # own keys take priority over anything nested deeper inside it.
        # An EMPTY placeholder (e.g. top-level "subject": "") must NOT
        # block a real, non-empty value found deeper from winning.
        for k, v in value.items():
            if isinstance(v, (str, int, float, bool)):
                if k not in pool or (not pool[k] and v):
                    pool[k] = v
        # Second pass: recurse into nested dicts / stringified blobs.
        for k, v in value.items():
            if isinstance(v, dict):
                _deep_flatten(v, pool, depth + 1)
            elif isinstance(v, str):
                parsed = _try_parse_blob(v)
                if parsed is not None:
                    _deep_flatten(parsed, pool, depth + 1)
            elif isinstance(v, list):
                for item in v:
                    _deep_flatten(item, pool, depth + 1)

def _recover_from_nested_email_blob(result):
    """Universal recovery: regardless of WHICH unexpected shape a provider
    used (wrong field names, content nested one or more levels deep under
    an arbitrary wrapper key, HTML body instead of plain text, etc.), this
    rebuilds the canonical top-level fields (from/to/subject/body/
    suspicious_link/attachment) from whatever the model actually produced,
    anywhere in the response. This intentionally does NOT special-case any
    specific wrapper key name (like 'email' or 'analysis') — new naming
    variants are handled automatically without needing a code change.
    Wrapped in a try/except as a whole: this is glue code patching up
    inconsistent provider output, so a bug in it must never crash the whole
    generation pipeline — worst case, it just leaves the original result
    untouched and the existing core_missing check handles it normally."""
    if not isinstance(result, dict):
        return result
    try:
        pool = {}
        _deep_flatten(result, pool)

        for canonical in _CANONICAL_FIELDS:
            cur = result.get(canonical)
            if isinstance(cur, str) and cur.strip():
                continue  # already has a real value, leave it alone
            for alt in _FIELD_SYNONYMS[canonical]:
                val = pool.get(alt)
                if isinstance(val, str) and val.strip():
                    result[canonical] = val
                    break

        # body_html (or any HTML-flavoured body) needs tag-stripping to be
        # readable plain text in the email window.
        body = result.get("body")
        if isinstance(body, str) and re.search(r"<[a-zA-Z]+[^>]*>", body):
            result["body"] = _strip_html(body)

        # Final type safety net: nothing downstream expects these fields to
        # ever be anything but a plain string. If recovery still leaves a
        # non-string (e.g. a dict slipped through), drop it to "" rather
        # than let it crash re.sub()/.strip() elsewhere.
        for canonical in _CANONICAL_FIELDS:
            if canonical in result and not isinstance(result[canonical], str):
                result[canonical] = "" if result[canonical] is None else str(result[canonical])
    except Exception:
        pass
    return result

def normalize_learning_analysis(result, role_type, difficulty, is_ar=False):
    result = _BASE_NORMALIZE_LEARNING_FINAL(result, role_type, difficulty, is_ar)
    if not isinstance(result, dict):
        return result
    result = _recover_from_nested_email_blob(result)
    is_ar = _detect_is_ar(result, is_ar)
    # SAFETY NET: if the core fields are still empty after every parsing /
    # repair step, this generation effectively failed (most likely a
    # response that got cut off before the JSON closed). Surface this as a
    # clear error + retry button instead of silently rendering a blank
    # email window, which was confusing and hard to diagnose.
    if "error" not in result:
        core_missing = not (_is_nonempty_str(result.get("from")) and
                             _is_nonempty_str(result.get("subject")) and
                             _is_substantial_str(result.get("body")))
        if core_missing:
            debug_keys = {k: (str(v)[:400] if v else v) for k, v in result.items()}
            return {"error": {"code": "incomplete_content", "message": "incomplete_generation", "debug": debug_keys}}
    result = _resolve_leftover_placeholders(result, is_ar)
    attack_type = result.get("attack_type") or infer_attack_type_from_content(result, is_ar)
    result["attack_type"] = attack_type
    indicators = result.get("indicators") if isinstance(result.get("indicators"), list) else []
    while len(indicators) < 3:
        indicators.append({"number": len(indicators)+1, "title": "", "description": ""})
    for i in range(3):
        indicators[i]["number"] = i + 1
        indicators[i]["title"] = _clean_indicator_title(indicators[i].get("title"), attack_type, i+1, is_ar)
        desc = (indicators[i].get("description") or "").strip()
        if i == 0 and attack_type not in desc:
            if is_ar:
                desc = f"هذا المؤشر يوضح {attack_type} لأنه يرتبط مباشرة بالإجراء الخطر المطلوب في البريد. " + desc
            else:
                desc = f"This shows {attack_type} because it directly matches the risky action requested in the email. " + desc
        indicators[i]["description"] = desc.strip()
    result["indicators"] = indicators[:3]
    wr = (result.get("why_risky") or "").strip()
    if attack_type and attack_type not in wr:
        role_hint = (ROLE_ATTACK_HINTS_AR if is_ar else ROLE_ATTACK_HINTS).get(role_type, ROLE_ATTACK_HINTS["other"])
        prefix = f"هذه رسالة {attack_type} وتؤثر على {role_hint}. " if is_ar else f"This is a {attack_type} email that affects {role_hint}. "
        result["why_risky"] = prefix + wr
    return result

_BASE_GENERATION_ISSUES_FINAL = get_generation_quality_issues



# =============================================================
# STRICT ROLE + DIFFICULTY GUARDRAILS (patched)
# -------------------------------------------------------------
# These checks reject AI outputs that drift away from the selected
# job role (clinical/admin/IT) or violate the documented difficulty
# framework before the email is shown to the trainee.
# =============================================================




def get_generation_quality_issues(result, difficulty, is_phishing=True):
    issues = _BASE_GENERATION_ISSUES_FINAL(result, difficulty, is_phishing)
    if not isinstance(result, dict):
        return issues
    if is_phishing:
        # Prevent model-copy artifacts that appeared during testing.
        titles = " ".join(str((x.get("title") or "")) for x in result.get("indicators", []) if isinstance(x, dict))
        if _BAD_INDICATOR_TITLES.search(titles):
            issues.append("indicator titles must be concrete, not placeholders")
        # The analysis must mention the attack type in learning outputs.
        attack_type = result.get("attack_type") or result.get("email_type") or ""
        analysis_text = " ".join(str(result.get(k, "")) for k in ["why_risky", "learning_tip"])
        analysis_text += " " + " ".join(str((x.get("description") or "")) for x in result.get("indicators", []) if isinstance(x, dict))
        if attack_type and attack_type not in analysis_text:
            issues.append("analysis must connect to attack_type")
    return issues

# =============================================================
# FRAMEWORK COMPLIANCE PATCH — Difficulty_Framework.docx enforcement
# Adds deterministic (non-LLM) gating on top of every previous prompt
# layer. Covers, per the accumulated findings across Easy/Medium/Hard
# testing rounds:
#   1) Verifiable spelling-error count (Easy=2, Medium=1, Hard=0) using
#      the model-reported injected_errors field, with a regex fallback
#      list of common misspellings when that field is missing.
#   2) Sender identity per Axis 1: no personal name/title at Easy,
#      no FULL name+title (Advanced-style) leaking into Medium, and a
#      required personal name+title at Hard. Domain sophistication:
#      Easy must be obviously fake, not government-like.
#   3) Axis 4 role alignment, deepened: reject generic commercial/
#      prize/marketing-themed phishing for role-specific roles even if
#      a role keyword is superficially mentioned once.
#   4) Axis 3: NO attachment at Easy; MANDATORY named attachment at
#      Hard (QR is now OPTIONAL/varied at Hard, not mandatory); and a
#      consistency check so the body never claims an attachment exists
#      when the attachment field is actually empty.
#   5) Recipient/greeting name consistency: the greeting must not name
#      someone entirely unrelated to the "to" address.
# This wraps the same shared function used by build_prompt's retry
# loop, so it automatically applies across all 4 AI providers and
# both languages (Arabic/English) without touching each provider path.
# =============================================================
_BASE_GENERATION_ISSUES_FRAMEWORK = get_generation_quality_issues

_PERSONAL_TITLE_RE = re.compile(
    r'\b(Dr\.?|Prof\.?|Professor|Doctor)\s+[A-Z][a-zA-Z\-]+(\s+[A-Z][a-zA-Z\-]+)?'
    r'|(د\.|دكتور|دكتورة|الدكتور|الدكتورة|أ\.د)\s*[\u0600-\u06FF]{2,}',
    re.I,
)
# Stricter: title + TWO capitalized name words (full first+last name) —
# this is the Advanced-level sender signature and must not leak into
# Easy or Medium (Medium may only use a single first name/title).
_FULL_NAME_TITLE_RE = re.compile(
    r'\b(Dr\.?|Prof\.?|Professor|Doctor)\s+[A-Z][a-zA-Z\-]+\s+[A-Z][a-zA-Z\-]+'
    r'|(د\.|دكتور|دكتورة|الدكتور|الدكتورة|أ\.د)\s*[\u0600-\u06FF]{2,}\s+[\u0600-\u06FF]{2,}',
    re.I,
)
_LOOKALIKE_GOV_DOMAIN_RE = re.compile(r'\b(gov|moh|ministry|board)\b', re.I)

_COMMON_MISSPELLINGS = [
    "acess", "informatin", "comunity", "recieve", "seperate", "occured", "untill",
    "goverment", "priviledge", "enviroment", "maintainance", "noticable",
    "reccommend", "adress", "begining", "calender", "definately", "embarass",
    "harrass", "independant", "occassion", "reccomend", "succesful", "tommorow",
    "wich", "teh", "hte", "loosing", "patint", "guidlines", "aknowledge",
    "acknowlegde", "requiered", "immediatly", "urgant", "pleaes", "thier",
]
_MISSPELLING_RE = re.compile(r'\b(' + '|'.join(_COMMON_MISSPELLINGS) + r')\b', re.I)

_COMMERCIAL_THEME_RE = re.compile(
    r'\bprize\b|\breward\b|\bcash\s*back\b|\bdiscount\b|\bpromotion\b|\banniversary\b|'
    r'\bwellness program\b|\btelecom\b|\bbank offer\b|\bloan\b|\bvoucher\b|\bgift\b|'
    r'\bwin(?:ner|ning)?\b|\braffle\b|\blucky draw\b|'
    r'جائزة|خصم|عرض تجاري|كاش باك|مكافأة|قرض|هدية|رابح|سحب',
    re.I,
)

_ROLE_KEYWORD_RE = {
    "clinical": re.compile(
        r'\b(patient|emr|lab result|medication|ward round|nurse|physician|diagnosis|radiology|icu)\w*'
        r'|مريض|عيادة|مختبر|دواء|تمريض|تشخيص|أشعة|طوارئ|عناية\s*مركزة',
        re.I,
    ),
    "admin": re.compile(
        r'\b(invoice|contract|billing|insurance|procurement|hr|schedule|vendor|accreditation)\w*'
        r'|فاتورة|عقد|تأمين|مشتريات|موارد\s*بشرية|جدولة|مورد|اعتماد',
        re.I,
    ),
    "it": re.compile(
        r'\b(network|vpn|server|firewall|backup|active directory|certificate|license|helpdesk|cybersecurity)\w*'
        r'|شبكة|خادم|جدار\s*ناري|نسخ\s*احتياطي|شهادة|ترخيص|الدعم\s*الفني',
        re.I,
    ),
}

# Generic commercial/prize themes are never role-specific regardless of any
# incidental keyword match (handled separately via _COMMERCIAL_THEME_RE).
# NOTE: MFA/OTP is intentionally NOT treated as an automatic violation here
# anymore — "MFA / OTP Abuse" tied to clinical-system access is one of the
# legitimately pre-approved Attack Playbook vectors for the clinical role
# (see ATTACK_PLAYBOOK["clinical"]), so rejecting it outright just caused
# repeated failed retries on a perfectly valid scenario. We only flag a
# login/credential theme as a role-mismatch when it is otherwise completely
# generic (no role keyword at all) — which the primary keyword-missing
# check below already covers.
_GENERIC_IT_HR_THEME_RE = re.compile(
    r'document sharing|log in now',
    re.I,
)

_ATTACHMENT_MENTION_RE = re.compile(r'\battach(?:ed|ment)?\b|مرفق|مرفقة|المرفق', re.I)

def _extract_name_tokens_from_email(addr):
    local = (addr or "").split("@")[0]
    parts = re.split(r'[.\-_]', local)
    return [p for p in parts if len(p) >= 3 and not p.isdigit()]



# =============================================================
# STRICT ROLE + DIFFICULTY GUARDRAILS (patched)
# -------------------------------------------------------------
# These checks reject AI outputs that drift away from the selected
# job role (clinical/admin/IT) or violate the documented difficulty
# framework before the email is shown to the trainee.
# =============================================================




def get_generation_quality_issues(result, difficulty, is_phishing=True):
    issues = _BASE_GENERATION_ISSUES_FRAMEWORK(result, difficulty, is_phishing)
    if not isinstance(result, dict) or not is_phishing:
        return issues

    body = str(result.get("body") or "")
    subject = str(result.get("subject") or "")
    sender = str(result.get("from") or "")
    to_addr = str(result.get("to") or "")
    attachment = str(result.get("attachment") or "").strip()
    combined_lower = f"{subject} {body}".lower()

    domain_match = re.search(r"@([\w.-]+)>?", sender)
    domain = (domain_match.group(1) if domain_match else "")

    # --- Axis 2 (verifiable spelling-error count) ---
    injected = result.get("injected_errors")
    if isinstance(injected, list):
        err_count = len([e for e in injected if str(e).strip()])
    else:
        err_count = len(_MISSPELLING_RE.findall(body))
    expected = {"easy": 2, "medium": 1, "hard": 0}.get(difficulty)
    if expected is not None and err_count != expected:
        issues.append(f"{difficulty} must contain exactly {expected} spelling/grammar mistake(s) in body (found {err_count})")

    # --- Axis 1 (Sender Identity) + Axis 3 (Technical Elements) ---
    if difficulty == "easy":
        if attachment:
            issues.append("Easy must have NO attachment (attachment field must be empty)")
        if _PERSONAL_TITLE_RE.search(sender):
            issues.append("Easy sender must be generic (team/department) — no personal name or title in 'from'")
        if _LOOKALIKE_GOV_DOMAIN_RE.search(domain):
            issues.append("Easy domain must not resemble a real government/hospital domain (avoid gov/moh/board/ministry)")
    elif difficulty == "medium":
        if _has_generic_greeting(body):
            issues.append("Intermediate greeting must be semi-personal (first name or title), not fully generic")
        if _FULL_NAME_TITLE_RE.search(sender) or _FULL_NAME_TITLE_RE.search(body[:200]):
            issues.append("Intermediate sender must NOT use a full name + job title (that is an Advanced-level signature) — first name/title only")
    elif difficulty == "hard":
        if not attachment:
            issues.append("Advanced must include a mandatory named attachment (attachment field cannot be empty)")
        if not _PERSONAL_TITLE_RE.search(sender) and not _PERSONAL_TITLE_RE.search(body[:200]):
            issues.append("Advanced sender must use a full personal name and job title")
        # QR is intentionally OPTIONAL/varied at Hard now — no check here.

    # --- Axis 3 consistency: body must not claim an attachment that doesn't exist ---
    if _ATTACHMENT_MENTION_RE.search(body) and not attachment:
        issues.append("body mentions an attachment but the attachment field is empty — keep them consistent")

    # --- Axis 4 (Role & Healthcare Context alignment), deepened ---
    role = st.session_state.get("role", "Clinical")
    role_info = ROLE_MAP.get(role, ROLE_MAP.get("Clinical"))
    role_type = role_info[2] if role_info else "clinical"
    role_kw_re = _ROLE_KEYWORD_RE.get(role_type)
    if role_kw_re:
        has_role_keyword = bool(role_kw_re.search(combined_lower))
        if not has_role_keyword:
            issues.append(f"email content must clearly reflect the '{role_type}' role context (missing role-specific keywords)")
        elif _COMMERCIAL_THEME_RE.search(combined_lower):
            issues.append(f"generic commercial/prize/marketing-themed phishing is not allowed for the '{role_type}' role — the scenario must revolve around an actual {role_type} work task, not a superficial keyword mention")
        elif role_type in ("clinical", "admin") and _GENERIC_IT_HR_THEME_RE.search(combined_lower):
            issues.append(f"generic MFA/login/document-sharing phishing is not allowed for the '{role_type}' role — that is an IT-themed scenario; the email must revolve around an actual {role_type} task instead")

    # --- Recipient/greeting name consistency (English only — Arabic
    # transliteration vs Latin email-derived tokens can't be reliably
    # matched, and skip for intentionally generic Easy greetings) ---
    if not _has_generic_greeting(body) and not re.search(r'[\u0600-\u06FF]', body[:120]):
        name_tokens = _extract_name_tokens_from_email(to_addr)
        if name_tokens:
            greeting_zone = body[:120].lower()
            if not any(tok.lower() in greeting_zone for tok in name_tokens):
                issues.append("greeting name does not match the 'to' recipient address — keep them consistent")

    return issues
# =============================================================
# END FRAMEWORK COMPLIANCE PATCH
# =============================================================

# Regeneration helper: when Try Again is clicked, also clear the used topic/domain
# memory for the current phase enough to allow a truly fresh attempt.

# More explicit system prompt for all providers.
def get_system_prompt():
    difficulty = st.session_state.get("difficulty", "medium")
    role = st.session_state.get("role", "Clinical")
    role_info = ROLE_MAP.get(role, ROLE_MAP.get("Clinical"))
    _, _, role_type = role_info
    role_context = get_role_unbounded_context(role_type, False)
    return f"""
You are a safe cybersecurity training content generator for a Saudi healthcare phishing-awareness study.
Role context: {role_context}
Difficulty: {difficulty}.
Return valid JSON only. No markdown, no comments, no prose outside JSON.
Do not use fixed templates, repeated domains, or placeholder indicator titles.
Learning email JSON must include attack_type, risk_level, indicators, why_risky, and learning_tip.
Assessment email JSON must stay concise and contain assessment fields only.
Phishing content must be simulated and use invented non-real domains.
Legitimate content must use hospital.org or moh.gov.sa only and must not request credentials, payment, MFA, OTP, bank details, or urgent account verification.
Beginner = obvious; Intermediate = partly convincing; Advanced = near-legitimate and subtle.
Analysis must link attack_type to role-context risk and should not rely only on Domain/Urgency/Spelling.
Arabic and English must have equal depth and quality.
""".strip()

# =============================================================
# END FINAL PATCH
# =============================================================

# =============================================================
# NOTE: Superseded generation-engine iterations removed (2026-07 cleanup)
# -------------------------------------------------------------
# The following abandoned engine iterations were deleted here because
# they were 100% dead code: every name they defined (generate_email,
# generate_assess_email, generate_other_email, generate_other_assess_email,
# get_generation_quality_issues, etc.) was re-defined again later in the
# file, so Python only ever executed the LAST definition (the v18 engine
# below). Removed blocks, in order: ROOT-CONTROLLED GENERATION ENGINE,
# ENHANCED HYBRID SCENARIO ENGINE v2, SCENARIO ENGINE v3, SCENARIO ENGINE v4,
# SCENARIO CONTENT ENGINE v5, EMAIL GENERATION ENGINE v6,
# API-FIRST DYNAMIC GENERATION ENGINE v10, FINAL REVIEW PATCH v11,
# FINAL STABLE PATCH v12, FINAL RESEARCH PATCH v15 (~3,850 lines total).
# None of this removal changes app behaviour: it was unreachable before
# the cleanup too.
# =============================================================
# =============================================================
# FINAL RESEARCH ENGINE v18 — ROLE × DIFFICULTY × HIGH-DIVERSITY
# -------------------------------------------------------------
# This engine is intentionally defined BEFORE Streamlit routing.
# It replaces the previous fixed-template patches with:
#   1) strict role alignment (clinical/admin/it/other mix),
#   2) strict progressive difficulty contracts,
#   3) combinatorial scenario generation (hundreds of blueprints),
#   4) multi-structure email composition,
#   5) cross-session anti-repeat history,
#   6) API wording generation with deterministic safe fallback.
# =============================================================

_V18_HISTORY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scenario_history_v18.json")

V18_ROLE_BANKS = {
    "clinical": {
        "easy": [
            ("Clinical Staff Portal", "basic account activation"), ("Shift Portal", "shift access check"),
            ("Hospital App", "mobile app access update"), ("Training Portal", "mandatory training login"),
            ("Clinic Account", "basic profile verification"), ("EMR Access", "simple account reset"),
            ("Staff Badge", "badge access confirmation"), ("Patient Portal", "basic sign-in verification"),
            ("Ward Schedule", "schedule access confirmation"), ("Clinical Email", "mailbox storage warning"),
            ("On-call Portal", "on-call access update"), ("Nursing Dashboard", "dashboard password reset"),
        ],
        "medium": [
            ("Emergency Department", "triage dashboard review"), ("Radiology", "PACS access verification"),
            ("Laboratory", "LIS result-routing update"), ("Pharmacy", "dispensing workflow review"),
            ("Outpatient Clinics", "appointment workflow revision"), ("ICU", "clinical handover update"),
            ("Infection Control", "policy acknowledgement"), ("Patient Safety", "incident portal review"),
            ("Blood Bank", "inventory access confirmation"), ("Operating Theatre", "theatre list review"),
            ("Dialysis Unit", "treatment schedule update"), ("Medical Records", "record access review"),
            ("Cardiology", "ECG workflow update"), ("Oncology", "treatment-plan portal review"),
            ("Respiratory Therapy", "device allocation review"), ("Endoscopy", "procedure-list confirmation"),
        ],
        "hard": [
            ("Blood Bank Transfusion Committee", "blood product release approval"),
            ("Radiology Informatics", "PACS reconciliation exception"),
            ("Clinical Governance", "sentinel event evidence review"),
            ("Antimicrobial Stewardship", "restricted-antibiotic approval"),
            ("Mortality and Morbidity Committee", "case review pack acknowledgement"),
            ("Research Ethics Office", "protocol deviation response"),
            ("Medication Safety Committee", "high-alert medication audit"),
            ("Operating Room Governance", "surgical safety checklist variance"),
            ("Pathology Quality Unit", "critical-result escalation audit"),
            ("Accreditation Office", "JCI tracer evidence submission"),
            ("Clinical Coding", "DRG case validation request"),
            ("Medical Affairs", "privileging document review"),
            ("Radiation Safety", "dosimetry compliance attestation"),
            ("Infection Prevention", "outbreak line-list review"),
            ("Transplant Coordination", "donor-case document verification"),
            ("Patient Experience", "formal complaint case response"),
        ],
    },
    "admin": {
        "easy": [
            ("Employee Portal", "profile update"), ("Payroll Portal", "salary account confirmation"),
            ("Leave System", "leave balance verification"), ("HR Account", "password reset"),
            ("Benefits Portal", "benefit enrollment check"), ("Training System", "course login update"),
            ("Attendance Portal", "timesheet verification"), ("Staff Directory", "contact details update"),
            ("Expense Portal", "basic reimbursement login"), ("Recruitment Portal", "candidate account update"),
            ("Policy Portal", "policy account activation"), ("Employee App", "mobile access confirmation"),
        ],
        "medium": [
            ("Human Resources", "annual appraisal workflow"), ("Finance", "expense claim review"),
            ("Procurement", "supplier onboarding verification"), ("Insurance Office", "coverage data review"),
            ("Medical Records Administration", "records retention acknowledgement"),
            ("Quality Office", "department audit submission"), ("Training and Development", "CME registration review"),
            ("Facilities", "work-order portal update"), ("Executive Office", "meeting action tracker"),
            ("Patient Relations", "complaint routing update"), ("Corporate Communications", "staff survey review"),
            ("Legal Affairs", "contract document acknowledgement"), ("Revenue Cycle", "billing exception review"),
            ("Admissions", "patient registration workflow"), ("Scheduling Office", "clinic capacity review"),
            ("Compliance", "annual declaration submission"),
        ],
        "hard": [
            ("Chief Executive Office", "confidential board pack review"),
            ("Internal Audit", "evidence request for control testing"),
            ("Procurement Governance", "single-source justification approval"),
            ("Finance Operations", "high-value payment exception"),
            ("Legal Affairs", "litigation hold acknowledgement"),
            ("Accreditation Program Office", "survey evidence submission"),
            ("Corporate Compliance", "conflict-of-interest attestation"),
            ("Revenue Integrity", "coding variance investigation"),
            ("Vendor Management", "banking detail change validation"),
            ("Board Secretariat", "committee resolution signature"),
            ("Risk Management", "enterprise risk register update"),
            ("Insurance Contracting", "payer agreement amendment"),
            ("Human Capital", "executive succession file review"),
            ("Strategic Planning", "restricted KPI dashboard access"),
            ("Data Protection Office", "privacy incident response pack"),
            ("Supply Chain", "critical shortage allocation approval"),
        ],
    },
    "it": {
        "easy": [
            ("IT Helpdesk", "password reset"), ("Email Support", "mailbox reactivation"),
            ("VPN Portal", "VPN sign-in verification"), ("Wi-Fi Portal", "wireless access update"),
            ("MFA Service", "MFA enrollment"), ("Account Support", "basic account reactivation"),
            ("Device Portal", "device login update"), ("Remote Access", "remote access confirmation"),
            ("Software Center", "basic software account update"), ("Service Desk", "ticket portal login"),
            ("Cloud Storage", "storage account verification"), ("Network Account", "password expiry notice"),
        ],
        "medium": [
            ("Enterprise Applications", "application access recertification"), ("Network Operations", "VPN profile migration"),
            ("Identity Management", "MFA token re-registration"), ("Microsoft 365 Support", "shared mailbox review"),
            ("Service Desk", "remote-support session approval"), ("Cybersecurity", "suspicious sign-in review"),
            ("Cloud Operations", "backup console verification"), ("Database Services", "maintenance-window acknowledgement"),
            ("Endpoint Management", "device compliance review"), ("Telephony", "softphone profile update"),
            ("PACS Support", "clinical imaging account review"), ("Integration Team", "interface monitoring update"),
            ("Data Center", "server access recertification"), ("Application Support", "license renewal review"),
            ("Digital Health", "portal migration confirmation"), ("Information Security", "policy attestation"),
        ],
        "hard": [
            ("Security Operations Center", "privileged access anomaly review"),
            ("Identity and Access Management", "emergency role-elevation approval"),
            ("Network Security", "firewall exception validation"),
            ("Cloud Security", "conditional-access policy exception"),
            ("Database Administration", "production credential rotation"),
            ("Infrastructure Architecture", "certificate chain remediation"),
            ("Incident Response", "forensic evidence handoff"),
            ("Enterprise Integration", "HL7 interface certificate renewal"),
            ("PACS Infrastructure", "DICOM routing exception"),
            ("Business Continuity", "disaster recovery failover approval"),
            ("Privileged Access Management", "vault checkout reconciliation"),
            ("Security Governance", "third-party risk exception"),
            ("Endpoint Detection", "quarantine release authorization"),
            ("Cloud Platform", "tenant federation change review"),
            ("Data Protection", "encryption key escrow verification"),
            ("Change Advisory Board", "emergency production change approval"),
        ],
    },
}

V18_ATTACKS = {
    "easy": ["credential_request", "password_reset", "pin_verification", "account_reactivation", "fake_login"],
    "medium": ["credential_request", "fake_survey", "shared_document", "reply_information", "pdf_review", "login_portal"],
    "hard": ["sharepoint", "microsoft365", "qr", "pdf", "excel", "reply_chain", "docusign", "calendar_invite"],
}

V18_STRUCTURES = {
    "easy": ["problem_action_link", "warning_link_consequence", "notice_action_deadline", "short_alert", "support_request", "account_status"],
    "medium": ["context_request_deadline", "department_notice", "workflow_exception", "review_and_confirm", "service_change", "audit_followup", "two_paragraph_professional"],
    "hard": ["internal_memo", "reply_thread", "committee_request", "document_share", "calendar_followup", "audit_evidence", "executive_request", "system_notification"],
}

V18_SUBJECT_PREFIX = {
    "easy": ["Urgent:", "Important:", "Action required:", "Account alert:", "Final notice:", "Immediate action:"],
    "medium": ["Reminder:", "Workflow notice:", "Review required:", "Department update:", "Access review:", "Follow-up:"],
    "hard": ["Re:", "Fwd:", "For review:", "Action item:", "Committee follow-up:", "Document shared:", "Meeting follow-up:", "Exception review:"],
}

V18_TIME_OPTIONS = {
    "easy": ["Today, 8:15 AM", "Today, 10:42 AM", "Today, 2:31 PM", "Yesterday, 4:05 PM", "Monday, 7:03 AM", "Thursday, 4:07 PM"],
    "medium": ["Monday, 9:18 AM", "Tuesday, 1:26 PM", "Wednesday, 11:04 AM", "Thursday, 8:47 AM", "Friday, 2:12 PM", "Yesterday, 3:39 PM"],
    "hard": ["Mon 8:13 AM", "Tue 3:47 PM", "Wed 10:06 AM", "Thu 4:22 PM", "Fri 7:56 AM", "Yesterday 5:14 PM", "Last Friday 2:08 PM"],
}

V18_DOMAINS = {
    "easy": ["secure-staff-verify.net", "hospital-access-check.com", "portal-verify-now.net", "account-update-care.org", "staff-login-update.co", "secure-login-alert.net"],
    "medium": ["hospital-services.org.co", "staffportal-health.net", "clinical-workflow.org", "healthsystems-support.com", "hospital-share.net", "care-portal-services.com"],
    "hard": ["hospitalorg-support.com", "m365-hospital.org", "sharepoint-hospital.net", "hospital-docs.co", "secure-hospitaloffice.com", "hospital-cloud.org"],
}

# =============================================================
# LIBRARY-DRIVEN DIVERSITY SYSTEM v19
# -------------------------------------------------------------
# Design agreed with the researcher: instead of handing the API a
# finished seed sentence and asking it to "reword it" (which kept
# collapsing onto the same handful of shapes), each generated email
# is now assembled from FIVE independent seed libraries — subject,
# opening angle, content angle (V18_ROLE_BANKS, already existed),
# urgency reason, and the existing sender/recipient/signature pools.
# Every library is split by role_type x difficulty (a strict grid —
# picking for "clinical/easy" can never reach into "it/hard"), and
# every pick is drawn through the session-scoped anti-repeat helper
# (_v18_no_repeat_choice) so the SAME user never sees the same seed
# twice in one session. The API receives the seeds as creative
# DIRECTION (not literal text) and is instructed to write the actual
# subject/body wording itself, constrained by V19_DIFFICULTY_CONTRACT
# below — so the AI still does the real writing (per the project's
# core idea), it just starts from a much richer, role-and-difficulty
# -correct starting point instead of one fixed template per cell.
# =============================================================

# Full difficulty contract (matches the researcher's own table):
# link visibility, greeting type, language register, how directly
# credentials are requested, the urgency style, indicator-count
# range, attachment policy, QR policy, MS365/Outlook usage, and how
# close the fake login page is to the real one. Used both to build
# the API prompt's rules section and to validate/repair the result.
V19_DIFFICULTY_CONTRACT = {
    "easy": {
        "link_style": "very obvious fake link (unrelated or misspelled domain)",
        "greeting": "generic (\"Dear Staff\" style, no name)",
        "language": "direct and simple",
        "credential_request": "direct — asks for the password/PIN/OTP outright",
        "urgency_style": "blunt words like \"Immediately\" or \"Today\"",
        "indicator_range": (4, 5),
        "attachment": "never",
        "qr": "never",
        "ms365_outlook": "never",
        "login_page_realism": "simple, does not resemble the real system",
    },
    "medium": {
        "link_style": "looks fairly official, close to the real domain",
        "greeting": "role/department based (e.g. \"Dear Radiology Team\")",
        "language": "more professional",
        "credential_request": "asks the user to sign in / verify — not a bare password request",
        "urgency_style": "a reasonable deadline, not a bare command",
        "indicator_range": (3, 4),
        "attachment": "sometimes a plain PDF",
        "qr": "never",
        "ms365_outlook": "never",
        "login_page_realism": "resembles the real system's look",
    },
    "hard": {
        "link_style": "very close to the official one, may use a shortener",
        "greeting": "the real person's name",
        "language": "very natural, reads like an internal message",
        "credential_request": "review/approve a document instead of a direct login",
        "urgency_style": "a logical business reason (audit, policy, patient safety), not a threat",
        "indicator_range": (1, 2),
        "attachment": "PDF, Excel, or SharePoint depending on the scenario",
        "qr": "sometimes, when the scenario calls for it",
        "ms365_outlook": "sometimes",
        "login_page_realism": "near-identical copy of the real one",
    },
}

# --- Opening-angle seeds: the DIRECTION for the first 1-2 sentences,
# tailored per role (clinical/admin/it) AND per difficulty. These are
# not final sentences — the API rewrites them into real prose in the
# requested language; the deterministic fallback path turns them into
# a simple sentence directly when the API is unavailable.
V19_OPENING_SEEDS = {
    "clinical": {
        "easy": [
            "a routine system check flagged a problem with the account",
            "access was automatically paused pending a quick check",
            "the mobile app logged an unusual sign-in attempt",
            "a scheduled maintenance step needs the user's confirmation",
            "the account was flagged for a missing verification step",
        ],
        "medium": [
            "a scheduled review of clinical system access identified an item needing confirmation",
            "the department's monthly access audit found an outstanding item",
            "a recent system upgrade requires staff to reconfirm their access",
            "clinical governance requested a routine credential check for this unit",
            "the on-call roster system flagged an inconsistency that needs review",
        ],
        "hard": [
            "following up on the item discussed at this week's unit meeting",
            "the accreditation team asked for confirmation on this file before the tracer visit",
            "sharing the file we discussed for your sign-off before the committee meets",
            "a minor exception was logged during the last audit cycle and needs your input",
            "the case reference from the last handover still needs a confirmation from your side",
        ],
    },
    "admin": {
        "easy": [
            "the payroll system flagged an incomplete profile step",
            "a routine HR check found a missing confirmation on file",
            "the leave-management app needs a quick account check",
            "the directory sync job flagged an outdated profile entry",
            "a scheduled account cleanup needs the user to confirm details",
        ],
        "medium": [
            "the finance team's quarterly access review found an item needing confirmation",
            "a recent policy update requires staff to reconfirm their account details",
            "the procurement portal migration needs a short confirmation from your side",
            "HR's routine audit flagged an outstanding item on this account",
            "the department's compliance check found a pending action item",
        ],
        "hard": [
            "following up on the file shared during this week's management meeting",
            "internal audit asked for a quick confirmation before closing this item",
            "sharing the document we discussed for your review before the board pack is finalised",
            "a minor exception surfaced during the procurement review and needs your input",
            "the finance exception noted last week still needs a confirmation from your office",
        ],
    },
    "it": {
        "easy": [
            "the security system flagged an unusual sign-in on the account",
            "a routine password-policy check found an expired credential",
            "the VPN service logged a failed authentication attempt",
            "the helpdesk queue has a pending action item on this account",
            "the mailbox quota system flagged the account for review",
        ],
        "medium": [
            "the quarterly access-recertification review found an item needing confirmation",
            "a scheduled identity-management migration requires staff to reconfirm access",
            "the security team's routine review flagged an outstanding item on this account",
            "a recent policy change requires re-authentication for privileged accounts",
            "the monitoring system logged an anomaly that needs a quick confirmation",
        ],
        "hard": [
            "following up on the change request discussed in this week's CAB meeting",
            "security operations asked for confirmation before closing this ticket",
            "sharing the access-review file we discussed for your sign-off",
            "a minor exception was logged during the last security audit and needs your input",
            "the incident reference from last week still needs a confirmation from your side",
        ],
    },
}

# --- Urgency-reason seeds: mirrors the researcher's own contract —
# Easy is a blunt command word, Medium is a believable deadline,
# Hard is a logical business reason (never a bare threat).
V19_URGENCY_SEEDS = {
    "clinical": {
        "easy": ["immediately", "today", "within the hour", "before your next shift"],
        "medium": ["within 24 hours", "before Thursday's shift handover", "by the end of the business day", "before the next audit cycle"],
        "hard": ["ahead of the accreditation tracer visit", "per the patient-safety review timeline", "before the committee finalises this cycle's file", "as part of the routine governance schedule"],
    },
    "admin": {
        "easy": ["immediately", "today", "within the hour", "before close of business"],
        "medium": ["within 24 hours", "before the next payroll run", "by the end of the business day", "before the quarterly close"],
        "hard": ["ahead of the internal audit deadline", "per this quarter's compliance timeline", "before the board pack is finalised", "as part of the routine procurement schedule"],
    },
    "it": {
        "easy": ["immediately", "today", "within the hour", "before your account locks"],
        "medium": ["within 24 hours", "before the scheduled migration window", "by the end of the business day", "before the next recertification cycle"],
        "hard": ["ahead of the change-advisory-board deadline", "per this quarter's security review timeline", "before the ticket is auto-closed", "as part of the routine access-audit schedule"],
    },
}

# --- Subject-line templates: kept difficulty-specific rather than a
# 3x3x3 grid, since the {area}/{topic} substitution (already fully
# role-specific via V18_ROLE_BANKS) is what actually varies the
# wording per role; the TEMPLATE SHAPE only needs to vary by
# difficulty (easy = blunt, medium = professional, hard = internal).
V19_SUBJECT_TEMPLATES = {
    "easy": [
        "Immediate action needed: {topic}",
        "Your {topic} is on hold",
        "We could not verify your {topic}",
        "{topic} — action required today",
        "Reminder: {topic} pending",
        "Access alert: {topic}",
    ],
    "medium": [
        "{area}: {topic} requires your review",
        "Update on {topic} — {area}",
        "Follow-up needed: {topic}",
        "{area} notice regarding {topic}",
        "Please confirm: {topic}",
        "Workflow change — {topic}",
    ],
    "hard": [
        "Re: {topic}",
        "{topic} — evidence requested",
        "Follow-up on our {area} discussion",
        "{area}: outstanding item — {topic}",
        "For your review: {topic}",
        "{topic} — sign-off needed",
    ],
}

# Arabic urgency phrases for the deterministic fallback path (used
# when the API is unavailable and the language is Arabic). Kept
# role-shared to control scope, but still difficulty-correct: Easy is
# a blunt command, Medium a believable deadline, Hard a logical
# business reason rather than a threat.
V19_URGENCY_SEEDS_AR = {
    "easy": ["فورًا", "اليوم", "خلال الساعة القادمة", "قبل نهاية الدوام"],
    "medium": ["خلال 24 ساعة", "قبل نهاية يوم العمل", "قبل نهاية الأسبوع", "قبل دورة المراجعة القادمة"],
    "hard": ["قبل زيارة الاعتماد القادمة", "ضمن الجدول الزمني لمراجعة السلامة", "قبل إغلاق الملف من اللجنة", "ضمن الجدول الدوري المعتاد للحوكمة"],
}

# Arabic subject-line templates (fallback path). NOTE: {topic}/{area}
# still come from V18_ROLE_BANKS, which is English-only, so the topic
# text itself stays in English inside the Arabic sentence — the same
# accepted trade-off already used throughout the Arabic body templates
# below. Fully bilingual role banks would remove this, but that is a
# separate, larger content task.
V19_SUBJECT_TEMPLATES_AR = {
    "easy": [
        "إجراء فوري مطلوب: {topic}",
        "حسابك بخصوص {topic} معلّق",
        "تعذر التحقق من {topic}",
        "{topic} — مطلوب إجراء اليوم",
        "تذكير: {topic} بانتظار الإجراء",
    ],
    "medium": [
        "{area}: {topic} بحاجة لمراجعتكم",
        "تحديث بخصوص {topic} — {area}",
        "متابعة مطلوبة: {topic}",
        "إشعار من {area} بخصوص {topic}",
        "يرجى التأكيد: {topic}",
    ],
    "hard": [
        "رد: {topic}",
        "{topic} — مطلوب دليل",
        "متابعة لنقاشنا في {area}",
        "{area}: بند معلّق — {topic}",
        "للمراجعة: {topic}",
    ],
}

# FIXED: why_risky / learning_tip used to be ONE hardcoded sentence per
# language, reused verbatim on every single fallback email — exactly
# the repetition the researcher flagged in the "AI Tutor Analysis"
# panel across otherwise-different emails. Each difficulty now has its
# own small pool, picked with the same anti-repeat helper as everything
# else.


def _v18_load_history():
    try:
        with open(_V18_HISTORY_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _v18_save_history(items):
    try:
        with open(_V18_HISTORY_PATH, "w", encoding="utf-8") as f:
            json.dump(items[-2000:], f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _v18_role_type(role):
    return ROLE_MAP.get(role, ROLE_MAP.get("Clinical"))[2]


def _v18_effective_role(role, index, phase):
    rt = _v18_role_type(role)
    if rt != "other":
        return rt
    key = f"v18_other_roles_{phase}_{st.session_state.get('v18_cycle_id', 0)}"
    if key not in st.session_state:
        seq = ["clinical", "admin", "it", "clinical", "admin", "it"]
        random.shuffle(seq)
        st.session_state[key] = seq
    seq = st.session_state[key]
    return seq[index % len(seq)]


def _v18_rng(index, role_type, difficulty, phase):
    nonce = st.session_state.get("v18_cycle_id")
    if nonce is None:
        nonce = random.randint(100000, 999999)
        st.session_state["v18_cycle_id"] = nonce
    seed = f"{nonce}|{index}|{role_type}|{difficulty}|{phase}|{random.random()}"
    return random.Random(seed)


def _v18_no_repeat_choice(items, memory_key, rng=None):
    """FIXED: shared, session-guaranteed anti-repeat picker. Streamlit's
    st.session_state always works within a browser session (unlike writing
    to a file next to the script, which silently fails on read-only/ephemeral
    deployments such as Streamlit Community Cloud — see the note on
    _v18_pick_blueprint below). Avoids repeating a value seen recently in
    THIS session before falling back to allowing repeats once every option
    has been used.
    """
    recent = st.session_state.get(memory_key, [])
    pool = [x for x in items if x not in recent]
    if not pool:
        pool = list(items)
        recent = []
    choice = (rng or random).choice(pool)
    st.session_state[memory_key] = (recent + [choice])[-(max(1, len(items) - 1)):]
    return choice


def _v18_pick_blueprint(role, index, language, difficulty, phase, is_phishing=True):
    difficulty = str(difficulty or "medium").lower()
    if difficulty not in ("easy", "medium", "hard"):
        difficulty = "medium"
    role_type = _v18_effective_role(role, index, phase)
    rng = _v18_rng(index, role_type, difficulty, phase)
    bank = V18_ROLE_BANKS[role_type][difficulty]

    # FIXED (root cause of repeated scenarios/domains across a 6-example
    # batch): this used to rely ONLY on a JSON file written next to the
    # script (scenario_history_v18.json) to avoid repeats. On Streamlit
    # Community Cloud that directory is frequently read-only or reset on
    # restart, so _v18_save_history() was failing silently every time
    # (wrapped in try/except: pass) and _v18_load_history() always came
    # back empty — meaning the "used" set was ALWAYS empty and every pick
    # was fully random with no memory at all. That is exactly why the same
    # domain and the same scenario could appear two or three times in one
    # six-example session. We now track "used" primarily in
    # st.session_state (guaranteed to work, scoped to this cycle), and keep
    # the on-disk file only as a secondary best-effort layer for repeat
    # visits across sessions on setups where the disk is actually writable.
    cycle_key = f"v18_used_signatures_{st.session_state.get('v18_cycle_id', 0)}"
    used_session = set(st.session_state.get(cycle_key, []))
    history = _v18_load_history()
    used_file = {x.get("signature") for x in history[-600:] if isinstance(x, dict)}
    used = used_session | used_file

    candidates = []
    for area, topic in bank:
        for attack in V18_ATTACKS[difficulty]:
            for structure in V18_STRUCTURES[difficulty]:
                sig = f"{role_type}|{difficulty}|{area}|{topic}|{attack}|{structure}"
                candidates.append((sig, area, topic, attack, structure))
    rng.shuffle(candidates)
    chosen = next((c for c in candidates if c[0] not in used), candidates[0])
    sig, area, topic, attack, structure = chosen

    st.session_state[cycle_key] = list(used_session | {sig})
    history.append({"signature": sig, "role": role_type, "difficulty": difficulty, "phase": phase})
    _v18_save_history(history)

    # FIXED: domain / subject prefix / display time previously had NO
    # anti-repeat protection at all (plain rng.choice every time) — with
    # pools of only 6 options and 6 examples per session, repeats were
    # statistically expected, not a fluke.
    domain = _v18_no_repeat_choice(V18_DOMAINS[difficulty], f"v18_recent_domain_{difficulty}_{st.session_state.get('v18_cycle_id', 0)}", rng)
    prefix = _v18_no_repeat_choice(V18_SUBJECT_PREFIX[difficulty], f"v18_recent_prefix_{difficulty}_{st.session_state.get('v18_cycle_id', 0)}", rng)
    display_time = _v18_no_repeat_choice(V18_TIME_OPTIONS[difficulty], f"v18_recent_time_{difficulty}_{st.session_state.get('v18_cycle_id', 0)}", rng)

    # v19 library-driven seeds — each pick is strictly scoped to
    # role_type x difficulty (a separate list per cell, never mixed
    # across roles) and anti-repeat protected within this session.
    cid = st.session_state.get('v18_cycle_id', 0)
    subject_template = _v18_no_repeat_choice(
        V19_SUBJECT_TEMPLATES[difficulty], f"v19_recent_subject_{difficulty}_{cid}", rng)
    subject_template_ar = _v18_no_repeat_choice(
        V19_SUBJECT_TEMPLATES_AR[difficulty], f"v19_recent_subject_ar_{difficulty}_{cid}", rng)
    opening_seed = _v18_no_repeat_choice(
        V19_OPENING_SEEDS[role_type][difficulty], f"v19_recent_opening_{role_type}_{difficulty}_{cid}", rng)
    urgency_seed = _v18_no_repeat_choice(
        V19_URGENCY_SEEDS[role_type][difficulty], f"v19_recent_urgency_{role_type}_{difficulty}_{cid}", rng)
    urgency_seed_ar = _v18_no_repeat_choice(
        V19_URGENCY_SEEDS_AR[difficulty], f"v19_recent_urgency_ar_{difficulty}_{cid}", rng)

    return {
        "signature": sig, "role_type": role_type, "difficulty": difficulty,
        "area": area, "topic": topic, "attack": attack, "structure": structure,
        "domain": domain, "subject_prefix": prefix, "display_time": display_time,
        "subject_template": subject_template, "subject_template_ar": subject_template_ar,
        "opening_seed": opening_seed,
        "urgency_seed": urgency_seed, "urgency_seed_ar": urgency_seed_ar,
        "language": language, "is_phishing": bool(is_phishing), "phase": phase,
    }








# FIXED: the easy-difficulty greeting pool used to have only 3 fixed
# strings, AND (see _v18_enforce below) the enforcement step forced the
# body's first line to be exactly the first of those 3 strings almost every
# time — which is why every screenshot showed "Dear Staff," no matter what.
# The pool is now bigger, and both the deterministic path (_v18_greeting)
# and the AI-enforcement path pull from the SAME pool using the SAME
# session-scoped anti-repeat memory, so they can no longer disagree or
# collapse onto one value.
V18_EASY_GREETINGS_EN = ["Dear Staff,", "Dear Healthcare Team,", "Dear Employee,", "Dear Colleague,", "Dear Team Member,", "Hello Staff,"]
V18_EASY_GREETINGS_AR = ["عزيزي الموظف،", "فريق العمل العزيز،", "الزملاء الأعزاء،", "عزيزي الزميل،", "أعضاء الفريق الأعزاء،", "مرحباً بكم،"]




def _v18_indicator_objects(bp, link, is_phishing=True):
    is_ar = bp["language"] == "Arabic"
    if not is_phishing:
        return []
    # Indicator COUNT now varies within the contracted range (easy 4-5,
    # medium 3-4, hard 1-2) instead of always being the same fixed
    # number every time, per the researcher's difficulty table. The
    # coin flip is seeded off this email's own signature so it's
    # reproducible for a given seed but varies across different emails.
    extra = random.Random(bp.get("signature", "") + "_indicator_count").random() < 0.5
    if bp["difficulty"] == "easy":
        rows = [
            ("نطاق مرسل مزيف وواضح", "المرسل لا يستخدم نطاق المستشفى الرسمي."),
            ("طلب مباشر لبيانات الدخول", "تطلب الرسالة كلمة المرور أو الرمز أو بيانات الحساب مباشرة."),
            ("إلحاح أو تهديد واضح", "تضغط الرسالة للتصرف اليوم أو خلال ساعات."),
            ("رابط خارجي ظاهر", f"الرابط يقود إلى نطاق غير رسمي: {link}"),
            ("تحية عامة غير شخصية", "الرسالة لا تخاطبك باسمك، بل بصيغة عامة تدل على إرسال جماعي."),
        ] if is_ar else [
            ("Obvious fake sender domain", "The sender does not use the hospital's official domain."),
            ("Direct credential request", "The message directly asks for a password, PIN, OTP, or login details."),
            ("Strong urgency or threat", "The email pressures the user to act today or within hours."),
            ("Visible external link", f"The link points to a non-official domain: {link}"),
            ("Generic, impersonal greeting", "The message does not address you by name, suggesting a mass mailing."),
        ]
        rows = rows[:4] + (rows[4:5] if extra else [])
    elif bp["difficulty"] == "medium":
        rows = [
            ("نطاق يبدو رسميًا لكنه غير مطابق", "النطاق قريب من نطاق المستشفى لكنه ليس النطاق المعتمد."),
            ("طلب غير معتاد ضمن سير العمل", "الطلب يبدو مهنيًا لكنه يطلب إجراءً لا يتم عادة عبر البريد."),
            ("مهلة زمنية مقنعة", "تستخدم الرسالة موعدًا معقولًا للضغط دون تهديد مباشر."),
            ("رابط لا يطابق الوجهة المتوقعة", "الرابط الظاهر لا يقود لنفس النظام الذي تتحدث عنه الرسالة."),
        ] if is_ar else [
            ("Plausible but incorrect domain", "The domain looks professional but does not match the approved hospital domain."),
            ("Unusual workflow request", "The request sounds legitimate but asks for an action not normally completed by email."),
            ("Plausible deadline pressure", "The message uses a believable deadline rather than an obvious threat."),
            ("Link destination mismatch", "The visible link does not point to the same system the email refers to."),
        ]
        rows = rows[:3] + (rows[3:4] if extra else [])
    else:
        rows = [
            ("اختلاف دقيق في القناة", "الطلب مهني جدًا لكن القناة أو الرابط لا يطابقان الإجراء الداخلي المعتاد."),
            ("طلب حساس بصياغة مقنعة", "تستخدم الرسالة سياقًا داخليًا لإقناع المستلم بتنفيذ إجراء حساس."),
        ] if is_ar else [
            ("Subtle channel mismatch", "The message is highly realistic, but the channel or destination differs from the normal internal process."),
            ("Sensitive request in a convincing context", "The email uses credible internal context to encourage a sensitive action."),
        ]
        rows = rows[:1] if not extra else rows[:2]
    return [{"number": i+1, "title": t, "description": d} for i, (t, d) in enumerate(rows)]






def _v18_api_prompt(seed, bp):
    is_ar = bp["language"] == "Arabic"
    language_rule = "Write every natural-language field in Arabic." if is_ar else "Write every natural-language field in English."
    contract = V19_DIFFICULTY_CONTRACT[bp["difficulty"]]
    lo, hi = contract["indicator_range"]
    indicator_desc = f"exactly {lo}" if lo == hi else f"between {lo} and {hi}"
    subject_hint = bp["subject_template"].format(area=bp["area"], topic=bp["topic"])
    return f"""
You are generating one email for a PhD phishing-awareness experiment in a Saudi hospital.
{language_rule}
Return valid JSON only. Preserve the exact scenario, role, difficulty, recipient, sender domain, attachment, and display_time from the seed.
Do not copy the seed wording. Write the subject and body yourself, from scratch, using the creative direction below — do not just reword the seed's fallback sentences.

CRITICAL LINK RULE — read carefully:
- Do NOT write out any URL or link text yourself, anywhere in the body. Do NOT invent, shorten, paraphrase, or repeat a link.
- Instead, put the exact placeholder token {{{{LINK}}}} (literally these characters, once) at the single point in the body where a link would naturally appear.
- The system will replace {{{{LINK}}}} with the real controlled link after you respond. If the email is not phishing (legitimate), omit {{{{LINK}}}} entirely.
- A body containing any raw "http" text, or more than one {{{{LINK}}}} token, is invalid.

Role: {bp['role_type']}
Difficulty: {bp['difficulty']}
Department/context: {bp['area']} — {bp['topic']}
Attack channel: {bp['attack']}
Required structure style: {bp['structure']}

Creative direction (use these as inspiration, put them in your own words — do not quote them verbatim):
- Subject line idea: "{subject_hint}" — you may adjust the wording, but keep it about {bp['topic']}.
- Opening idea: the email should open around this situation — {bp['opening_seed']}.
- Urgency/deadline idea: frame the time pressure around — {bp['urgency_seed']}.

Difficulty contract for "{bp['difficulty']}" (follow every point):
- Link appearance: {contract['link_style']}.
- Greeting: {contract['greeting']}.
- Language register: {contract['language']}.
- How credentials are requested: {contract['credential_request']}.
- Urgency style: {contract['urgency_style']}.
- Number of red-flag indicators: {indicator_desc}.
- Attachment: {contract['attachment']}.
- QR code: {contract['qr']}.
- Microsoft 365 / Outlook branding: {contract['ms365_outlook']}.
- Fake login page realism: {contract['login_page_realism']}.

Seed JSON (for protected fields only — do not copy its wording):
{json.dumps(seed, ensure_ascii=False)}
"""




_V18_URL_RE = re.compile(r"(?:https?://|www\.)\S+", re.I)


def _v18_strip_and_place_link(body, link, is_phishing):
    """FIXED (root cause of the duplicate/conflicting-link bug): the model
    was asked to 'preserve the seed link' but LLMs routinely ignore that and
    write out their own plausible-looking URL inline instead. _v18_enforce
    then compared the seed link against the body with an exact substring
    check; since the AI's invented link never matched byte-for-byte, the
    real link got appended a second time — leaving TWO different fake links
    in the same email (one AI-invented and unstyled, one correct and
    highlighted). This function guarantees exactly one link, always the
    correct controlled one:
      1) Remove every raw http(s)/www URL the model wrote on its own.
      2) Fill in the {{LINK}} placeholder if the model used it.
      3) Otherwise insert the single controlled link once, at a position
         that varies (see _v18_place_link_fallback) instead of always
         landing in the exact same spot right before the signature.
    """
    text = str(body or "")
    text = _V18_URL_RE.sub("", text)
    text = re.sub(r"[ \t]+\n", "\n", text)          # trailing spaces left by a removed URL
    text = re.sub(r"\n{3,}", "\n\n", text).strip()  # collapse blank-line gaps left behind

    if not is_phishing or not link:
        return text.replace("{{LINK}}", "").replace("{LINK}", "").strip()

    if "{{LINK}}" in text:
        return text.replace("{{LINK}}", link)
    if "{LINK}" in text:
        return text.replace("{LINK}", link)

    return _v18_place_link_fallback(text, link)


_V18_LINK_ACTION_WORDS = (
    "verify", "confirm", "click", "provide", "update", "reactivate", "sign in",
    "log in", "review", "complete", "restore", "reset",
    "تحقق", "أكّد", "أكد", "قدّم", "حدّث", "أعد تفعيل", "سجّل الدخول", "راجع", "أكمل",
)


def _v18_place_link_fallback(text, link):
    """FIXED: previously the link, whenever the model skipped the
    {{LINK}} placeholder, always landed in the exact same spot (right
    before the signature) in every single email — a monotony that stood
    out just as much as an outright duplicate. It now rotates, once per
    session, between three natural positions so consecutive examples don't
    all look identical."""
    lines = text.split("\n")
    sig_markers = ("regards", "sincerely", "thank you", "best,", "support",
                   "مع التحية", "تحياتي", "مع الشكر", "شكرًا", "وتفضلوا", "مع التقدير")
    sig_idx = next((i for i, ln in enumerate(lines)
                     if ln.strip() and any(m in ln.strip().lower() for m in sig_markers)), None)
    action_idx = next((i for i, ln in enumerate(lines)
                        if ln.strip() and any(w in ln.lower() for w in _V18_LINK_ACTION_WORDS)), None)

    strategies = ["before_signature", "own_line_end"]
    if action_idx is not None:
        strategies.append("after_action_sentence")

    key = f"v18_recent_link_position_{st.session_state.get('v18_cycle_id', 0)}"
    strategy = _v18_no_repeat_choice(strategies, key)

    if strategy == "after_action_sentence" and action_idx is not None:
        lines[action_idx:action_idx + 1] = [lines[action_idx].rstrip(), "", link]
        return "\n".join(lines).strip()
    if strategy == "before_signature" and sig_idx is not None:
        lines[sig_idx:sig_idx] = [link, ""]
        return "\n".join(lines).strip()
    return text.rstrip() + "\n\n" + link


def _v18_enforce(result, seed, bp):
    if not isinstance(result, dict):
        return seed
    protected = ["from", "to", "attachment", "suspicious_link", "display_time", "is_phishing", "scenario_id", "scenario_meta", "risk_level"]
    for k in protected:
        result[k] = seed.get(k)
    result.setdefault("email_type", seed.get("email_type"))
    result.setdefault("attack_type", seed.get("attack_type"))
    result.setdefault("subject", seed.get("subject"))
    result.setdefault("body", seed.get("body"))
    result.setdefault("why_risky", seed.get("why_risky"))
    result.setdefault("learning_tip", seed.get("learning_tip"))
    result["indicators"] = _v18_indicator_objects(bp, seed.get("suspicious_link", ""), bool(seed.get("is_phishing")))
    if bp["difficulty"] == "easy":
        result["attachment"] = ""
        result["body"] = re.sub(r"\[QR[^\]]*\]", "", str(result.get("body", "")), flags=re.I)
        lines = str(result.get("body", "")).splitlines()
        if lines:
            # FIXED (root cause of every email showing "Dear Staff,"): this
            # used to compare against only 3 fixed strings and, since an
            # AI paraphrase essentially never matches one of them verbatim,
            # it silently overwrote line 1 with allowed[0] on almost every
            # single call — collapsing all greeting diversity to one value.
            # We now draw from the same expanded, session-deduped pool used
            # by the deterministic path (_v18_greeting), so a forced
            # rewrite still lands on a fresh, varied greeting instead of
            # always the same one.
            pool = V18_EASY_GREETINGS_AR if bp["language"] == "Arabic" else V18_EASY_GREETINGS_EN
            if not any(lines[0].strip().lower() == x.lower() for x in pool):
                key = f"v18_recent_greeting_easy_{bp['language']}_{st.session_state.get('v18_cycle_id', 0)}"
                lines[0] = _v18_no_repeat_choice(pool, key)
            result["body"] = "\n".join(lines)
    if bp["difficulty"] == "medium":
        result["body"] = re.sub(r"\[QR[^\]]*\]", "", str(result.get("body", "")), flags=re.I)
        if str(result.get("attachment", "")).lower().endswith((".xlsx", ".xls", ".docx")):
            result["attachment"] = ""
    # FIXED: exactly one link, always the correct controlled one — see
    # _v18_strip_and_place_link docstring above for why this replaced the
    # old naive "append if missing" substring check.
    link = seed.get("suspicious_link", "")
    result["body"] = _v18_strip_and_place_link(result.get("body", ""), link, bool(seed.get("is_phishing")))
    try:
        result = clean_result(result, bp["language"] == "Arabic")
    except Exception:
        result = seed
    return result








def generate_other_email(index, language, difficulty):
    role = "أخرى" if language == "Arabic" else "Other"
    return generate_email(role, index, language, difficulty)


def generate_other_assess_email(index, is_phishing, language, difficulty):
    role = "أخرى" if language == "Arabic" else "Other"
    return generate_assess_email(role, index, is_phishing, language, difficulty)


def go_to_learning(role):
    st.session_state["role"] = role
    st.session_state["page"] = "learning"
    st.session_state["example_index"] = 0
    st.session_state["emails"] = {}
    st.session_state["v18_cycle_id"] = random.randint(100000, 999999)
    for k in list(st.session_state.keys()):
        if str(k).startswith("v18_") and k != "v18_cycle_id":
            del st.session_state[k]

# =============================================================
# END FINAL RESEARCH ENGINE v18
# =============================================================

# =============================================================
# RESEARCH ENGINE v20 — semantic diversity + grounded analysis
# Added without changing the UI, providers, admin panel, exports,
# scoring, bilingual flow, or existing data structures.
# =============================================================

V20_SCENARIO_DIMENSIONS = {
    "clinical": {
        "units": ["Emergency Department", "Outpatient Clinic", "ICU", "Cardiology", "Oncology", "Pediatrics", "Radiology", "Laboratory", "Pharmacy", "Infection Control", "Operating Theatre", "Blood Bank", "Dialysis Unit", "Patient Safety", "Medical Affairs", "Clinical Education"],
        "events": [
            ("schedule_change", "a revised clinic or duty schedule requiring acknowledgement"),
            ("lab_followup", "a patient laboratory result requiring a documented follow-up"),
            ("referral_review", "a referral or transfer request awaiting clinical review"),
            ("medication_review", "a medication reconciliation or restricted-drug approval"),
            ("infection_notice", "an infection-control exposure or isolation update"),
            ("patient_safety", "a patient-safety incident or near-miss follow-up"),
            ("imaging_review", "an imaging report or PACS exception requiring review"),
            ("procedure_list", "a procedure list or theatre booking amendment"),
            ("handover_issue", "a handover discrepancy requiring confirmation"),
            ("equipment_alert", "a clinical-device allocation or maintenance notice"),
            ("training_update", "a mandatory clinical competency or simulation session"),
            ("policy_update", "a clinical protocol or guideline acknowledgement"),
            ("blood_product", "a blood-product release or transfusion documentation request"),
            ("audit_case", "a case audit, mortality review, or tracer evidence request"),
            ("telemedicine", "a remote consultation schedule or platform notification"),
            ("credentialing", "a clinical privilege or professional credential review"),
        ],
        "actions": ["review the case", "confirm receipt", "acknowledge the update", "validate the record", "approve the pending item", "open the referenced file", "respond with the case reference", "complete the requested review"],
    },
    "admin": {
        "units": ["Human Resources", "Finance", "Procurement", "Patient Access", "Medical Records", "Insurance Office", "Revenue Cycle", "Quality Office", "Facilities", "Executive Office", "Training and Development", "Legal Affairs", "Patient Relations", "Compliance", "Supply Chain", "Scheduling Office"],
        "events": [
            ("payroll_exception", "a payroll or allowance exception requiring confirmation"),
            ("leave_update", "a leave balance, roster, or attendance correction"),
            ("invoice_review", "a supplier invoice or payment exception"),
            ("vendor_change", "a vendor profile or banking-detail change"),
            ("insurance_claim", "an insurance claim or coverage discrepancy"),
            ("appointment_capacity", "a clinic-capacity or appointment scheduling update"),
            ("records_request", "a medical-records retention or release request"),
            ("audit_evidence", "an audit evidence or accreditation submission"),
            ("contract_review", "a contract amendment or renewal document"),
            ("benefits_enrollment", "an employee benefits or allowance enrollment notice"),
            ("training_registration", "a mandatory course or professional-development registration"),
            ("complaint_case", "a patient complaint or service-recovery case"),
            ("procurement_shortage", "a critical supply shortage or allocation approval"),
            ("board_document", "a confidential committee or board document"),
            ("privacy_case", "a privacy incident or data-protection response"),
            ("facility_workorder", "a facility access or work-order update"),
        ],
        "actions": ["review the request", "confirm the record", "approve the item", "validate the change", "acknowledge receipt", "open the supporting document", "submit the requested information", "respond with the reference number"],
    },
    "it": {
        "units": ["IT Service Desk", "Identity Management", "Network Operations", "Cybersecurity", "Cloud Operations", "Database Services", "Enterprise Applications", "PACS Support", "Digital Health", "Endpoint Management", "Data Center", "Integration Team", "Information Security", "Business Continuity", "Telephony", "Change Advisory Board"],
        "events": [
            ("identity_review", "an account-access or identity recertification request"),
            ("mfa_change", "an MFA token enrollment or re-registration notice"),
            ("vpn_migration", "a VPN profile migration or remote-access update"),
            ("security_alert", "a suspicious sign-in or endpoint security alert"),
            ("certificate_expiry", "a certificate renewal or trust-chain exception"),
            ("backup_failure", "a backup verification or recovery test exception"),
            ("database_change", "a database maintenance or credential-rotation request"),
            ("firewall_exception", "a firewall or conditional-access exception"),
            ("software_license", "a software license or application entitlement review"),
            ("device_compliance", "a managed-device compliance or quarantine notice"),
            ("interface_issue", "an HL7 or system-interface monitoring exception"),
            ("pacs_routing", "a PACS or DICOM routing exception"),
            ("incident_handoff", "an incident-response evidence or forensic handoff"),
            ("dr_test", "a disaster-recovery or failover approval"),
            ("change_request", "an emergency production change request"),
            ("cloud_federation", "a cloud tenant or federation configuration review"),
        ],
        "actions": ["review the alert", "confirm the change", "validate the exception", "approve the request", "acknowledge the maintenance window", "open the technical record", "respond with the ticket number", "complete the access review"],
    },
}

_V20_OLD_PICK_BLUEPRINT = _v18_pick_blueprint
_V20_OLD_API_PROMPT = _v18_api_prompt


def _v18_pick_blueprint(role, index, language, difficulty, phase, is_phishing=True):
    """Create a role-locked, semantically unique scenario blueprint.
    Uniqueness is based on the event family, not merely wording."""
    bp = _V20_OLD_PICK_BLUEPRINT(role, index, language, difficulty, phase, is_phishing)
    rt = bp["role_type"]
    dims = V20_SCENARIO_DIMENSIONS[rt]
    rng = _v18_rng(index, rt, bp["difficulty"], phase + "_v20")
    cid = st.session_state.get("v18_cycle_id", 0)

    # Learning and assessment share this memory, so test items do not reuse
    # the same semantic family already shown during learning.
    family_key = f"v20_used_families_{rt}_{cid}"
    used_families = set(st.session_state.get(family_key, []))
    available_events = [e for e in dims["events"] if e[0] not in used_families]
    if not available_events:
        available_events = list(dims["events"])
        used_families = set()
    family, event_text = rng.choice(available_events)
    used_families.add(family)
    st.session_state[family_key] = list(used_families)

    unit = _v18_no_repeat_choice(dims["units"], f"v20_units_{rt}_{cid}", rng)
    action = _v18_no_repeat_choice(dims["actions"], f"v20_actions_{rt}_{cid}", rng)
    bp["area"] = unit
    bp["scenario_family"] = family
    bp["event_text"] = event_text
    bp["action_text"] = action
    bp["topic"] = f"{event_text}; requested action: {action}"
    bp["signature"] = f"v20|{rt}|{bp['difficulty']}|{family}|{unit}|{action}|{bp['attack']}|{bp['structure']}"
    return bp


def _v20_find_phrase(body, patterns):
    for pattern in patterns:
        m = re.search(pattern, body, re.I)
        if m:
            return m.group(0).strip(" .,:;،؛")
    return ""


def _v20_grounded_analysis(result, bp):
    """Build tutor indicators from evidence that is visible in the final email.
    Each indicator also carries a render target so badge numbers and tutor text
    always point to the same element."""
    if not result.get("is_phishing"):
        result["indicators"] = []
        result["suspicious_text"] = ""
        return result

    is_ar = bp["language"] == "Arabic"
    body = str(result.get("body", ""))
    subject = str(result.get("subject", ""))
    sender = str(result.get("from", ""))
    link = str(result.get("suspicious_link", "")).strip()
    candidates = []

    def add(key, title_en, title_ar, desc_en, desc_ar, evidence, target):
        if evidence:
            candidates.append({
                "key": key,
                "title": title_ar if is_ar else title_en,
                "description": desc_ar if is_ar else desc_en,
                "evidence": evidence,
                "target": target,
            })

    mdom = re.search(r"@([A-Za-z0-9.-]+)", sender)
    sender_domain = mdom.group(1) if mdom else ""
    if sender_domain and sender_domain.lower() != "hospital.org":
        add("domain", "Non-official sender domain", "نطاق مرسل غير رسمي",
            f"The sender uses {sender_domain}, not the hospital's official domain.",
            f"عنوان المرسل يستخدم النطاق {sender_domain} وليس نطاق المستشفى الرسمي.",
            sender_domain, "from")

    urgency = _v20_find_phrase(subject + "\n" + body, [
        r"(?:immediate(?:ly)?|urgent(?:ly)?|within (?:the )?(?:hour|\d+ hours?)|today|before your next shift|as soon as possible|failure to act)",
        r"(?:فوراً|فورًا|عاجل|بشكل عاجل|خلال ساعة|خلال \d+ ساعات?|اليوم|قبل المناوبة القادمة|عدم الاستجابة)",
    ])
    if urgency:
        target = "subject" if urgency.lower() in subject.lower() else "body"
        add("urgency", "Time pressure", "ضغط زمني",
            "The message uses urgency or a deadline to push the recipient into acting quickly.",
            "تستخدم الرسالة مهلة أو استعجالاً لدفع المستلم للتصرف بسرعة.",
            urgency, target)

    if link and link in body:
        add("link", "Unapproved external link", "رابط خارجي غير معتمد",
            f"The link points to a non-official destination: {link}",
            f"الرابط يقود إلى نطاق غير رسمي: {link}",
            link, "link")

    first = next((ln.strip() for ln in body.splitlines() if ln.strip()), "")
    if re.match(r"(?:Dear (?:Staff|Team|Employee|Colleague|Healthcare Team)|Hello Staff|عزيزي الموظف|فريق العمل العزيز|الزملاء الأعزاء|عزيزي الزميل)", first, re.I):
        add("greeting", "Generic greeting", "تحية عامة",
            "The email does not address the recipient by name, which can indicate bulk targeting.",
            "لا تخاطب الرسالة المستلم باسمه، ما قد يدل على إرسال جماعي.",
            first.rstrip("،,"), "greeting")

    credential = _v20_find_phrase(body, [
        r"(?:enter|confirm|verify|provide|reset|submit)\s+(?:your\s+)?(?:password|PIN|OTP|login credentials|credentials|account details)",
        r"(?:أدخل|أكد|تحقق من|زوّدنا بـ|أرسل)\s+(?:كلمة المرور|الرقم السري|رمز التحقق|بيانات الدخول)",
    ])
    if credential:
        add("credential", "Credential request", "طلب بيانات دخول",
            "The email asks for sensitive sign-in information that should not be provided by email.",
            "تطلب الرسالة بيانات حساسة لا ينبغي إرسالها عبر البريد.",
            credential, "body")

    attachment = str(result.get("attachment", "")).strip()
    if attachment:
        add("attachment", "Unexpected attachment", "مرفق غير متوقع",
            f"The message asks the recipient to open {attachment} as part of a sensitive request.",
            f"تطلب الرسالة فتح المرفق {attachment} ضمن طلب حساس.",
            attachment, "attachment")

    desired = {"easy": 4, "medium": 3, "hard": 2}[bp["difficulty"]]
    # Prefer distinct visible UI targets. This prevents two tutor numbers from
    # pointing to the same URL or a badge appearing with the wrong explanation.
    priority = {
        "easy": ["domain", "urgency", "link", "greeting", "credential", "attachment"],
        "medium": ["domain", "link", "credential", "urgency", "attachment", "greeting"],
        "hard": ["domain", "link", "attachment", "credential", "urgency", "greeting"],
    }[bp["difficulty"]]
    by_key = {c["key"]: c for c in candidates}
    chosen = [by_key[k] for k in priority if k in by_key][:desired]

    result["indicators"] = []
    for i, item in enumerate(chosen, 1):
        result["indicators"].append({
            "number": i,
            "title": item["title"],
            "description": item["description"],
            "evidence": item["evidence"],
            "target": item["target"],
            "key": item["key"],
        })

    # Legacy renderer compatibility; the updated renderer uses indicator targets.
    body_item = next((x for x in result["indicators"] if x.get("target") == "body"), None)
    result["suspicious_text"] = body_item.get("evidence", "") if body_item else ""
    return result



def _v18_api_prompt(seed, bp):
    base = _V20_OLD_API_PROMPT(seed, bp)
    length_rule = {"easy": "65-120", "medium": "85-150", "hard": "100-180"}[bp["difficulty"]]
    return base + f"""

V20 QUALITY AND DIVERSITY REQUIREMENTS:
- This email belongs ONLY to the {bp['role_type']} role. Do not introduce unrelated IT, administrative, or clinical scenarios.
- Semantic scenario family: {bp.get('scenario_family')}. Event: {bp.get('event_text')}.
- Requested action: {bp.get('action_text')}.
- Target body length: {length_rule} words (Arabic may be moderately shorter).
- Use 2-4 coherent paragraphs plus a natural closing and a role-appropriate signature.
- Do not repeat the subject verbatim in the body.
- Do not use the generic account-update/password-reset story unless the selected event explicitly requires it.
- Make sender name, subject, opening, call-to-action, closing, and signature fit this exact workplace event.
- Keep exactly one natural location for {{{{LINK}}}}. Never put it after the signature.
- Return indicators only if they are directly supported by visible evidence in the final email. The application will verify them.
"""


_V20_OLD_ENFORCE = _v18_enforce


def _v18_enforce(result, seed, bp):
    result = _V20_OLD_ENFORCE(result, seed, bp)
    return _v20_grounded_analysis(result, bp)




# Public entry points remain unchanged for the rest of the application.



# =============================================================
# END RESEARCH ENGINE v20
# =============================================================


# =============================================================
# RULE-GUIDED PHISHING SCENARIO ENGINE v30
# Knowledge base -> planner -> difficulty contract -> composer -> validator
# This block intentionally overrides the legacy generate_email entry points.
# =============================================================
import hashlib as _v30_hashlib
import time as _v30_time

_V30_RNG = random.SystemRandom()

V30_DIFFICULTY = {
    "easy": {
        "indicator_count": 5, "generic_greeting": True, "personal_greeting": False,
        "direct_credentials": True, "urgency": "strong", "domain_style": "obvious",
        "allowed_channels": ("link",), "attachments": False, "qr": False,
        "body_words": (70, 125),
    },
    "medium": {
        "indicator_count": 3, "generic_greeting": False, "personal_greeting": False,
        "direct_credentials": False, "urgency": "moderate", "domain_style": "similar",
        "allowed_channels": ("button", "link", "pdf"), "attachments": "sometimes", "qr": False,
        "body_words": (85, 145),
    },
    "hard": {
        "indicator_count": 2, "generic_greeting": False, "personal_greeting": True,
        "direct_credentials": False, "urgency": "logical", "domain_style": "near",
        "allowed_channels": ("button", "pdf", "xlsx", "docx", "sharepoint", "m365", "qr"),
        "attachments": "allowed", "qr": "allowed", "body_words": (95, 165),
    },
}

# Each family is a coherent workplace workflow. Sender identities and actions are
# tied to the family so Pharmacy can never sign an Infection Control message, etc.
V30_KNOWLEDGE = {
    "clinical": [
        {"id":"lab_critical","area":"Laboratory Services","area_ar":"خدمات المختبر","event":"a critical laboratory result requiring documented follow-up","event_ar":"نتيجة مخبرية حرجة تتطلب متابعة موثقة","objects":["potassium result","blood culture result","coagulation result","troponin result","specimen rejection notice"],"objects_ar":["نتيجة البوتاسيوم","نتيجة مزرعة الدم","نتيجة التخثر","نتيجة التروبونين","إشعار رفض العينة"],"actions":["review the result","acknowledge the finding","open the case record"],"actions_ar":["مراجعة النتيجة","إقرار النتيجة","فتح سجل الحالة"],"senders":["Laboratory Results Desk","Clinical Laboratory Services"],"senders_ar":["مكتب نتائج المختبر","خدمات المختبر السريري"],"signatures":["Laboratory Results Team","Clinical Laboratory Services"],"signatures_ar":["فريق نتائج المختبر","خدمات المختبر السريري"]},
        {"id":"referral","area":"Referral Coordination","area_ar":"تنسيق الإحالات","event":"a specialist referral awaiting clinical review","event_ar":"إحالة تخصصية بانتظار المراجعة السريرية","objects":["cardiology referral","oncology referral","neurology referral","urgent outpatient referral","inter-facility transfer request"],"objects_ar":["إحالة قلبية","إحالة أورام","إحالة أعصاب","إحالة عيادات خارجية عاجلة","طلب تحويل بين المنشآت"],"actions":["review the referral","confirm clinical acceptance","open the referral record"],"actions_ar":["مراجعة الإحالة","تأكيد القبول السريري","فتح سجل الإحالة"],"senders":["Referral Coordination Unit","Patient Flow Office"],"senders_ar":["وحدة تنسيق الإحالات","مكتب تدفق المرضى"],"signatures":["Referral Coordination Team","Patient Flow Office"],"signatures_ar":["فريق تنسيق الإحالات","مكتب تدفق المرضى"]},
        {"id":"radiology","area":"Radiology","area_ar":"الأشعة","event":"a diagnostic imaging report awaiting review","event_ar":"تقرير تصوير تشخيصي بانتظار المراجعة","objects":["CT report","MRI report","ultrasound report","critical imaging addendum","contrast safety checklist"],"objects_ar":["تقرير الأشعة المقطعية","تقرير الرنين المغناطيسي","تقرير الموجات فوق الصوتية","ملحق تصوير حرج","قائمة تدقيق سلامة الصبغة"],"actions":["review the report","acknowledge the addendum","open the imaging record"],"actions_ar":["مراجعة التقرير","إقرار الملحق","فتح سجل التصوير"],"senders":["Radiology Reporting Office","Diagnostic Imaging Services"],"senders_ar":["مكتب تقارير الأشعة","خدمات التصوير التشخيصي"],"signatures":["Radiology Reporting Team","Diagnostic Imaging Services"],"signatures_ar":["فريق تقارير الأشعة","خدمات التصوير التشخيصي"]},
        {"id":"medication","area":"Pharmacy","area_ar":"الصيدلية","event":"a medication-safety item requiring clinical review","event_ar":"بند متعلق بسلامة الأدوية يتطلب مراجعة سريرية","objects":["high-alert medication review","medication reconciliation exception","formulary substitution notice","antimicrobial approval","dose clarification request"],"objects_ar":["مراجعة دواء عالي الخطورة","استثناء مطابقة الأدوية","إشعار بديل بالتشكيلة الدوائية","اعتماد مضاد حيوي","طلب توضيح جرعة"],"actions":["review the medication item","acknowledge the safety notice","open the medication record"],"actions_ar":["مراجعة بند الدواء","إقرار إشعار السلامة","فتح سجل الدواء"],"senders":["Medication Safety Office","Clinical Pharmacy Services"],"senders_ar":["مكتب سلامة الدواء","خدمات الصيدلية السريرية"],"signatures":["Medication Safety Team","Clinical Pharmacy Services"],"signatures_ar":["فريق سلامة الدواء","خدمات الصيدلية السريرية"]},
        {"id":"infection","area":"Infection Prevention","area_ar":"مكافحة العدوى","event":"an infection-control notification requiring acknowledgement","event_ar":"إشعار مكافحة عدوى يتطلب إقراراً","objects":["isolation precaution update","exposure follow-up","outbreak advisory","hand-hygiene audit finding","screening protocol update"],"objects_ar":["تحديث احتياطات العزل","متابعة تعرض","تنبيه تفشي","نتيجة تدقيق نظافة اليدين","تحديث بروتوكول الفحص"],"actions":["review the notice","acknowledge the update","open the exposure record"],"actions_ar":["مراجعة الإشعار","إقرار التحديث","فتح سجل التعرض"],"senders":["Infection Prevention and Control","Occupational Health Infection Desk"],"senders_ar":["مكافحة العدوى والوقاية","مكتب الصحة المهنية للعدوى"],"signatures":["Infection Prevention Team","Occupational Health"],"signatures_ar":["فريق مكافحة العدوى","الصحة المهنية"]},
        {"id":"patient_safety","area":"Patient Safety","area_ar":"سلامة المرضى","event":"a patient-safety case awaiting documented review","event_ar":"حالة سلامة مرضى بانتظار مراجعة موثقة","objects":["near-miss report","medication variance","fall-risk incident","handover concern","clinical escalation record"],"objects_ar":["تقرير حالة كادت تقع","تباين دوائي","حادثة خطر سقوط","ملاحظة تسليم","سجل تصعيد سريري"],"actions":["review the case","acknowledge the incident","open the safety record"],"actions_ar":["مراجعة الحالة","إقرار الحادثة","فتح سجل السلامة"],"senders":["Patient Safety Office","Clinical Governance Unit"],"senders_ar":["مكتب سلامة المرضى","وحدة الحوكمة السريرية"],"signatures":["Patient Safety Team","Clinical Governance"],"signatures_ar":["فريق سلامة المرضى","الحوكمة السريرية"]},
        {"id":"schedule","area":"Clinical Operations","area_ar":"العمليات السريرية","event":"a clinical schedule change requiring confirmation","event_ar":"تغيير بجدول سريري يتطلب تأكيداً","objects":["operating theatre list","on-call rota","clinic coverage schedule","weekend duty roster","procedure-room allocation"],"objects_ar":["قائمة غرفة العمليات","جدول المناوبة","جدول تغطية العيادة","جدول نوبات نهاية الأسبوع","تخصيص غرفة الإجراءات"],"actions":["review the revised schedule","confirm availability","open the duty roster"],"actions_ar":["مراجعة الجدول المعدّل","تأكيد التوفر","فتح جدول النوبات"],"senders":["Clinical Operations Office","Medical Workforce Scheduling"],"senders_ar":["مكتب العمليات السريرية","جدولة القوى العاملة الطبية"],"signatures":["Clinical Operations","Medical Workforce Team"],"signatures_ar":["العمليات السريرية","فريق القوى العاملة الطبية"]},
        {"id":"blood_bank","area":"Blood Bank","area_ar":"بنك الدم","event":"a transfusion workflow item requiring clinical action","event_ar":"بند بمسار نقل الدم يتطلب إجراءً سريرياً","objects":["crossmatch request","blood-product release","transfusion reaction follow-up","massive transfusion checklist","blood availability notice"],"objects_ar":["طلب مطابقة الدم","إفراج منتج دموي","متابعة تفاعل نقل دم","قائمة تدقيق نقل دم ضخم","إشعار توفر الدم"],"actions":["review the transfusion item","acknowledge the request","open the blood-bank record"],"actions_ar":["مراجعة بند نقل الدم","إقرار الطلب","فتح سجل بنك الدم"],"senders":["Transfusion Services","Blood Bank Coordination"],"senders_ar":["خدمات نقل الدم","تنسيق بنك الدم"],"signatures":["Transfusion Services Team","Blood Bank"],"signatures_ar":["فريق خدمات نقل الدم","بنك الدم"]},
        {"id":"discharge","area":"Health Information Management","area_ar":"إدارة المعلومات الصحية","event":"a discharge-documentation item awaiting completion","event_ar":"بند توثيق خروج بانتظار الإكمال","objects":["discharge summary","clinical coding query","unsigned progress note","incomplete medication list","pending follow-up plan"],"objects_ar":["ملخص الخروج","استفسار ترميز سريري","ملاحظة تقدم غير موقعة","قائمة أدوية غير مكتملة","خطة متابعة معلقة"],"actions":["review the documentation","complete the acknowledgement","open the patient record"],"actions_ar":["مراجعة التوثيق","إكمال الإقرار","فتح سجل المريض"],"senders":["Clinical Documentation Office","Health Information Management"],"senders_ar":["مكتب التوثيق السريري","إدارة المعلومات الصحية"],"signatures":["Clinical Documentation Team","Health Information Management"],"signatures_ar":["فريق التوثيق السريري","إدارة المعلومات الصحية"]},
        {"id":"education","area":"Clinical Education","area_ar":"التعليم السريري","event":"a mandatory clinical learning item requiring completion","event_ar":"بند تعلّم سريري إلزامي يتطلب إكمالاً","objects":["annual competency assessment","resuscitation update","medication-safety module","infection-control module","patient-identification refresher"],"objects_ar":["التقييم السنوي للكفاءة","تحديث الإنعاش","وحدة سلامة الأدوية","وحدة مكافحة العدوى","تذكير تعريف هوية المريض"],"actions":["open the learning item","confirm completion","review the assigned module"],"actions_ar":["فتح بند التعلم","تأكيد الإكمال","مراجعة الوحدة المخصصة"],"senders":["Clinical Education Department","Professional Development Unit"],"senders_ar":["قسم التعليم السريري","وحدة التطوير المهني"],"signatures":["Clinical Education Team","Professional Development"],"signatures_ar":["فريق التعليم السريري","التطوير المهني"]},
        {"id":"equipment","area":"Biomedical Engineering","area_ar":"الهندسة الطبية الحيوية","event":"a medical-device notice requiring acknowledgement","event_ar":"إشعار جهاز طبي يتطلب إقراراً","objects":["infusion pump advisory","patient monitor update","ventilator safety notice","defibrillator inspection","glucose meter recall"],"objects_ar":["إشعار سحب مضخة تسريب","تحديث جهاز مراقبة المريض","إشعار سلامة جهاز التنفس","فحص جهاز الصدمات الكهربائية","سحب جهاز قياس الجلوكوز"],"actions":["review the device notice","acknowledge the advisory","open the equipment record"],"actions_ar":["مراجعة إشعار الجهاز","إقرار الإشعار","فتح سجل الجهاز"],"senders":["Biomedical Engineering","Medical Device Safety Desk"],"senders_ar":["الهندسة الطبية الحيوية","مكتب سلامة الأجهزة الطبية"],"signatures":["Biomedical Engineering Team","Medical Device Safety"],"signatures_ar":["فريق الهندسة الطبية الحيوية","سلامة الأجهزة الطبية"]},
        {"id":"telehealth","area":"Virtual Care","area_ar":"الرعاية الافتراضية","event":"a remote-consultation item awaiting clinician review","event_ar":"بند استشارة عن بعد بانتظار مراجعة الطبيب","objects":["telemedicine appointment","remote monitoring alert","virtual clinic handover","video consultation note","home-care escalation"],"objects_ar":["موعد التطبيب عن بعد","تنبيه مراقبة عن بعد","تسليم عيادة افتراضية","ملاحظة استشارة فيديو","تصعيد رعاية منزلية"],"actions":["review the virtual-care item","open the consultation record","confirm the handover"],"actions_ar":["مراجعة بند الرعاية الافتراضية","فتح سجل الاستشارة","تأكيد التسليم"],"senders":["Virtual Care Coordination","Telehealth Operations"],"senders_ar":["تنسيق الرعاية الافتراضية","عمليات التطبيب عن بعد"],"signatures":["Virtual Care Team","Telehealth Operations"],"signatures_ar":["فريق الرعاية الافتراضية","عمليات التطبيب عن بعد"]},
    ],
    "admin": [
        {"id":"insurance","area":"Insurance Coordination","area_ar":"تنسيق التأمين","event":"an insurance case awaiting administrative review","event_ar":"حالة تأمين بانتظار المراجعة الإدارية","objects":["coverage exception","claim rejection","pre-authorization request","policy eligibility update","reimbursement query"],"objects_ar":["استثناء تغطية","رفض مطالبة","طلب تصريح مسبق","تحديث أهلية بوليصة","استفسار استرداد"],"actions":["review the case","open the claim record","confirm the administrative response"],"actions_ar":["مراجعة الحالة","فتح سجل المطالبة","تأكيد الرد الإداري"],"senders":["Insurance Coordination Unit","Revenue Cycle Office"],"senders_ar":["وحدة تنسيق التأمين","مكتب دورة الإيرادات"],"signatures":["Insurance Coordination","Revenue Cycle Team"],"signatures_ar":["تنسيق التأمين","فريق دورة الإيرادات"]},
        {"id":"billing","area":"Patient Billing","area_ar":"فوترة المرضى","event":"a billing item requiring reconciliation","event_ar":"بند فوترة يتطلب مطابقة","objects":["invoice discrepancy","unposted payment","billing adjustment","patient account exception","coding-related charge query"],"objects_ar":["تباين فاتورة","دفعة غير مرحّلة","تعديل فوترة","استثناء حساب مريض","استفسار رسوم متعلق بالترميز"],"actions":["review the billing item","open the account record","acknowledge the discrepancy"],"actions_ar":["مراجعة بند الفوترة","فتح سجل الحساب","إقرار التباين"],"senders":["Patient Billing Office","Revenue Integrity Unit"],"senders_ar":["مكتب فوترة المرضى","وحدة نزاهة الإيرادات"],"signatures":["Patient Billing Team","Revenue Integrity"],"signatures_ar":["فريق فوترة المرضى","نزاهة الإيرادات"]},
        {"id":"procurement","area":"Procurement","area_ar":"المشتريات","event":"a supplier transaction awaiting review","event_ar":"معاملة مورد بانتظار المراجعة","objects":["purchase order amendment","supplier invoice","contract renewal","delivery discrepancy","quotation approval"],"objects_ar":["تعديل أمر شراء","فاتورة مورد","تجديد عقد","تباين تسليم","اعتماد عرض سعر"],"actions":["review the transaction","open the procurement record","confirm receipt"],"actions_ar":["مراجعة المعاملة","فتح سجل المشتريات","تأكيد الاستلام"],"senders":["Medical Procurement Office","Supply Chain Management"],"senders_ar":["مكتب المشتريات الطبية","إدارة سلسلة الإمداد"],"signatures":["Procurement Team","Supply Chain Management"],"signatures_ar":["فريق المشتريات","إدارة سلسلة الإمداد"]},
        {"id":"appointments","area":"Patient Access","area_ar":"وصول المرضى","event":"an appointment workflow item requiring action","event_ar":"بند بمسار المواعيد يتطلب إجراءً","objects":["clinic overbooking notice","appointment rescheduling batch","waiting-list release","referral booking exception","patient registration correction"],"objects_ar":["إشعار حجز زائد بالعيادة","دفعة إعادة جدولة مواعيد","إفراج قائمة انتظار","استثناء حجز إحالة","تصحيح تسجيل مريض"],"actions":["review the booking item","open the scheduling record","acknowledge the change"],"actions_ar":["مراجعة بند الحجز","فتح سجل الجدولة","إقرار التغيير"],"senders":["Patient Access Services","Appointment Coordination"],"senders_ar":["خدمات وصول المرضى","تنسيق المواعيد"],"signatures":["Patient Access Team","Appointment Coordination"],"signatures_ar":["فريق وصول المرضى","تنسيق المواعيد"]},
        {"id":"hr","area":"Human Resources","area_ar":"الموارد البشرية","event":"an employee administration item requiring review","event_ar":"بند إداري لموظف يتطلب مراجعة","objects":["leave balance correction","benefits enrollment","contract detail update","attendance exception","staff credential renewal"],"objects_ar":["تصحيح رصيد إجازة","تسجيل مزايا","تحديث تفاصيل عقد","استثناء حضور","تجديد اعتماد موظف"],"actions":["review the employee item","open the HR record","confirm the update"],"actions_ar":["مراجعة بند الموظف","فتح سجل الموارد البشرية","تأكيد التحديث"],"senders":["Human Resources Services","Employee Relations Office"],"senders_ar":["خدمات الموارد البشرية","مكتب علاقات الموظفين"],"signatures":["Human Resources","Employee Relations"],"signatures_ar":["الموارد البشرية","علاقات الموظفين"]},
        {"id":"payroll","area":"Payroll","area_ar":"الرواتب","event":"a payroll exception requiring confirmation","event_ar":"استثناء رواتب يتطلب تأكيداً","objects":["salary payment exception","IBAN verification case","allowance adjustment","overtime reconciliation","end-of-service calculation"],"objects_ar":["استثناء دفع راتب","حالة تحقق آيبان","تعديل بدل","مطابقة وقت إضافي","احتساب نهاية خدمة"],"actions":["review the payroll item","open the payment record","confirm the correction"],"actions_ar":["مراجعة بند الراتب","فتح سجل الدفع","تأكيد التصحيح"],"senders":["Payroll Services","Compensation and Benefits"],"senders_ar":["خدمات الرواتب","التعويضات والمزايا"],"signatures":["Payroll Team","Compensation and Benefits"],"signatures_ar":["فريق الرواتب","التعويضات والمزايا"]},
        {"id":"records","area":"Medical Records Administration","area_ar":"إدارة السجلات الطبية","event":"an administrative patient-record item awaiting review","event_ar":"بند إداري بسجل مريض بانتظار المراجعة","objects":["release-of-information request","record merge exception","demographic correction","archiving notice","document indexing query"],"objects_ar":["طلب الإفراج عن معلومات","استثناء دمج سجل","تصحيح بيانات ديموغرافية","إشعار أرشفة","استفسار فهرسة مستند"],"actions":["review the records item","open the administrative record","acknowledge the request"],"actions_ar":["مراجعة بند السجلات","فتح السجل الإداري","إقرار الطلب"],"senders":["Medical Records Administration","Health Information Services"],"senders_ar":["إدارة السجلات الطبية","خدمات المعلومات الصحية"],"signatures":["Medical Records Team","Health Information Services"],"signatures_ar":["فريق السجلات الطبية","خدمات المعلومات الصحية"]},
        {"id":"facilities","area":"Facilities Management","area_ar":"إدارة المرافق","event":"a facilities service item requiring coordination","event_ar":"بند خدمة مرافق يتطلب تنسيقاً","objects":["access badge review","maintenance closure","office relocation","parking permit update","building access exception"],"objects_ar":["مراجعة بطاقة الدخول","إغلاق صيانة","نقل مكتب","تحديث تصريح موقف","استثناء دخول مبنى"],"actions":["review the service notice","open the facilities request","confirm the arrangement"],"actions_ar":["مراجعة إشعار الخدمة","فتح طلب المرافق","تأكيد الترتيب"],"senders":["Facilities Management","Workplace Services"],"senders_ar":["إدارة المرافق","خدمات مكان العمل"],"signatures":["Facilities Team","Workplace Services"],"signatures_ar":["فريق المرافق","خدمات مكان العمل"]},
    ],
    "it": [
        {"id":"vpn","area":"Network Services","area_ar":"خدمات الشبكة","event":"a remote-access service item requiring technical review","event_ar":"بند خدمة وصول عن بعد يتطلب مراجعة تقنية","objects":["VPN certificate update","remote access exception","gateway maintenance","privileged access review","secure tunnel migration"],"objects_ar":["تحديث شهادة VPN","استثناء وصول عن بعد","صيانة بوابة","مراجعة وصول مميز","ترحيل نفق آمن"],"actions":["review the service item","open the network ticket","acknowledge the change"],"actions_ar":["مراجعة بند الخدمة","فتح تذكرة الشبكة","إقرار التغيير"],"senders":["Network Operations Centre","Infrastructure Services"],"senders_ar":["مركز عمليات الشبكة","خدمات البنية التحتية"],"signatures":["Network Operations","Infrastructure Services"],"signatures_ar":["عمليات الشبكة","خدمات البنية التحتية"]},
        {"id":"identity","area":"Identity and Access Management","area_ar":"إدارة الهوية والوصول","event":"an identity-governance item awaiting review","event_ar":"بند حوكمة هوية بانتظار المراجعة","objects":["privileged role recertification","inactive account review","MFA registration exception","service-account ownership","access entitlement review"],"objects_ar":["إعادة اعتماد دور مميز","مراجعة حساب غير نشط","استثناء تسجيل تحقق ثنائي","ملكية حساب خدمة","مراجعة صلاحية وصول"],"actions":["review the access item","open the IAM case","acknowledge the entitlement"],"actions_ar":["مراجعة بند الوصول","فتح حالة إدارة الهوية","إقرار الصلاحية"],"senders":["Identity and Access Management","Cybersecurity Operations"],"senders_ar":["إدارة الهوية والوصول","عمليات الأمن السيبراني"],"signatures":["IAM Team","Cybersecurity Operations"],"signatures_ar":["فريق إدارة الهوية","عمليات الأمن السيبراني"]},
        {"id":"server","area":"Infrastructure Operations","area_ar":"عمليات البنية التحتية","event":"a server operation item requiring action","event_ar":"بند عملية خادم يتطلب إجراءً","objects":["storage threshold alert","backup verification","patch maintenance window","cluster failover review","virtual machine ownership"],"objects_ar":["تنبيه حد التخزين","تحقق نسخ احتياطي","نافذة صيانة تحديثات","مراجعة تبديل تلقائي للعنقود","ملكية جهاز افتراضي"],"actions":["review the infrastructure item","open the operations ticket","acknowledge the maintenance"],"actions_ar":["مراجعة بند البنية التحتية","فتح تذكرة العمليات","إقرار الصيانة"],"senders":["Infrastructure Operations","Data Centre Services"],"senders_ar":["عمليات البنية التحتية","خدمات مركز البيانات"],"signatures":["Infrastructure Operations","Data Centre Services"],"signatures_ar":["عمليات البنية التحتية","خدمات مركز البيانات"]},
        {"id":"security","area":"Cybersecurity","area_ar":"الأمن السيبراني","event":"a security case awaiting technical review","event_ar":"حالة أمنية بانتظار مراجعة تقنية","objects":["endpoint detection alert","firewall rule exception","vulnerability remediation","phishing investigation","certificate anomaly"],"objects_ar":["تنبيه كشف نقطة نهاية","استثناء قاعدة جدار حماية","معالجة ثغرة","تحقيق تصيد","شذوذ شهادة"],"actions":["review the security case","open the incident ticket","acknowledge the finding"],"actions_ar":["مراجعة الحالة الأمنية","فتح تذكرة الحادثة","إقرار النتيجة"],"senders":["Security Operations Centre","Cybersecurity Governance"],"senders_ar":["مركز عمليات الأمن","حوكمة الأمن السيبراني"],"signatures":["Security Operations","Cybersecurity Governance"],"signatures_ar":["عمليات الأمن","حوكمة الأمن السيبراني"]},
        {"id":"database","area":"Database Services","area_ar":"خدمات قواعد البيانات","event":"a database service item requiring confirmation","event_ar":"بند خدمة قاعدة بيانات يتطلب تأكيداً","objects":["backup integrity check","replication lag alert","schema change request","database account review","restore-test result"],"objects_ar":["فحص سلامة نسخة احتياطية","تنبيه تأخر تكرار","طلب تغيير مخطط","مراجعة حساب قاعدة بيانات","نتيجة اختبار استعادة"],"actions":["review the database item","open the service ticket","acknowledge the change"],"actions_ar":["مراجعة بند قاعدة البيانات","فتح تذكرة الخدمة","إقرار التغيير"],"senders":["Database Administration","Enterprise Applications"],"senders_ar":["إدارة قواعد البيانات","التطبيقات المؤسسية"],"signatures":["Database Services","Enterprise Applications"],"signatures_ar":["خدمات قواعد البيانات","التطبيقات المؤسسية"]},
        {"id":"application","area":"Clinical Applications","area_ar":"التطبيقات السريرية","event":"an enterprise application item awaiting review","event_ar":"بند تطبيق مؤسسي بانتظار المراجعة","objects":["EMR interface error","application release note","license allocation","integration exception","production change request"],"objects_ar":["خطأ واجهة النظام الطبي","ملاحظة إصدار تطبيق","تخصيص ترخيص","استثناء تكامل","طلب تغيير إنتاج"],"actions":["review the application item","open the change record","acknowledge the release"],"actions_ar":["مراجعة بند التطبيق","فتح سجل التغيير","إقرار الإصدار"],"senders":["Clinical Applications Support","Enterprise Systems"],"senders_ar":["دعم التطبيقات السريرية","الأنظمة المؤسسية"],"signatures":["Clinical Applications","Enterprise Systems"],"signatures_ar":["التطبيقات السريرية","الأنظمة المؤسسية"]},
        {"id":"cloud","area":"Cloud and Backup Services","area_ar":"الخدمات السحابية والنسخ الاحتياطي","event":"a cloud service item requiring technical review","event_ar":"بند خدمة سحابية يتطلب مراجعة تقنية","objects":["backup retention exception","cloud access review","storage policy update","recovery test","subscription renewal"],"objects_ar":["استثناء الاحتفاظ بنسخ احتياطي","مراجعة وصول سحابي","تحديث سياسة تخزين","اختبار استعادة","تجديد اشتراك"],"actions":["review the cloud item","open the service record","acknowledge the policy"],"actions_ar":["مراجعة البند السحابي","فتح سجل الخدمة","إقرار السياسة"],"senders":["Cloud Platform Operations","Backup and Recovery Services"],"senders_ar":["عمليات المنصة السحابية","خدمات النسخ الاحتياطي والاستعادة"],"signatures":["Cloud Operations","Backup and Recovery"],"signatures_ar":["عمليات السحابة","النسخ الاحتياطي والاستعادة"]},
        {"id":"helpdesk","area":"IT Service Desk","area_ar":"مكتب خدمات تقنية المعلومات","event":"a support case requiring technician action","event_ar":"حالة دعم تتطلب إجراء فني","objects":["escalated service ticket","remote support request","device enrollment case","software deployment issue","asset handover"],"objects_ar":["تذكرة خدمة مصعّدة","طلب دعم عن بعد","حالة تسجيل جهاز","مشكلة نشر برنامج","تسليم أصل"],"actions":["review the support case","open the ticket","acknowledge the assignment"],"actions_ar":["مراجعة حالة الدعم","فتح التذكرة","إقرار التكليف"],"senders":["IT Service Desk","End User Computing"],"senders_ar":["مكتب خدمات تقنية المعلومات","حوسبة المستخدم النهائي"],"signatures":["IT Service Desk","End User Computing"],"signatures_ar":["مكتب الخدمات","حوسبة المستخدم النهائي"]},
    ],
    "other": [
        {"id":"ops_policy","area":"Hospital Operations","area_ar":"عمليات المستشفى","event":"a hospital operations notice requiring acknowledgement","event_ar":"إشعار عمليات مستشفى يتطلب إقراراً","objects":["visitor policy update","parking permit renewal","staff directory correction","emergency contact list","facility access notice"],"objects_ar":["تحديث سياسة الزوار","تجديد تصريح موقف سيارات","تصحيح دليل الموظفين","قائمة اتصال الطوارئ","إشعار دخول المرفق"],"actions":["review the notice","acknowledge the update","open the operations record"],"actions_ar":["مراجعة الإشعار","إقرار التحديث","فتح سجل العمليات"],"senders":["Hospital Operations","Corporate Services Office"],"senders_ar":["عمليات المستشفى","مكتب الخدمات المؤسسية"],"signatures":["Hospital Operations Team","Corporate Services"],"signatures_ar":["فريق عمليات المستشفى","الخدمات المؤسسية"]},
        {"id":"staff_training","area":"Staff Development","area_ar":"تطوير الموظفين","event":"a mandatory training item requiring completion","event_ar":"بند تدريب إلزامي يتطلب إكمالاً","objects":["mandatory induction module","cybersecurity awareness course","fire safety training","patient privacy module","annual competency record"],"objects_ar":["وحدة التهيئة الإلزامية","دورة توعية الأمن السيبراني","تدريب السلامة من الحريق","وحدة خصوصية المريض","سجل الكفاءة السنوي"],"actions":["open the learning item","confirm completion","review the assigned module"],"actions_ar":["فتح بند التعلم","تأكيد الإكمال","مراجعة الوحدة المخصصة"],"senders":["Staff Development Office","Learning and Training Centre"],"senders_ar":["مكتب تطوير الموظفين","مركز التعلم والتدريب"],"signatures":["Staff Development Team","Learning and Training"],"signatures_ar":["فريق تطوير الموظفين","التعلم والتدريب"]},
    ],
}


# Extra workflow families ensure that 6 learning + 10 assessment items can use
# different semantic families in one full cycle, not merely different wording.
V30_KNOWLEDGE["clinical"].extend([
    {"id":"emergency","area":"Emergency Medicine","area_ar":"طب الطوارئ","event":"an emergency-care workflow item requiring review","event_ar":"بند بمسار رعاية طارئة يتطلب مراجعة","objects":["triage escalation","observation-unit handover","trauma pathway exception","emergency discharge follow-up","critical-care transfer note"],"objects_ar":["تصعيد فرز","تسليم وحدة الملاحظة","استثناء مسار الإصابات","متابعة خروج طارئ","ملاحظة تحويل رعاية حرجة"],"actions":["review the emergency record","acknowledge the handover","open the escalation note"],"actions_ar":["مراجعة سجل الطوارئ","إقرار التسليم","فتح ملاحظة التصعيد"],"senders":["Emergency Operations Desk","Acute Care Coordination"],"senders_ar":["مكتب عمليات الطوارئ","تنسيق الرعاية الحرجة"],"signatures":["Emergency Operations","Acute Care Coordination"],"signatures_ar":["عمليات الطوارئ","تنسيق الرعاية الحرجة"]},
    {"id":"surgery","area":"Surgical Services","area_ar":"الخدمات الجراحية","event":"a perioperative item awaiting clinical review","event_ar":"بند محيط بالجراحة بانتظار المراجعة السريرية","objects":["surgical consent exception","theatre list amendment","pre-operative checklist","post-operative handover","implant traceability record"],"objects_ar":["استثناء موافقة جراحية","تعديل قائمة غرفة العمليات","قائمة تدقيق ما قبل الجراحة","تسليم ما بعد الجراحة","سجل تتبع غرسة"],"actions":["review the surgical item","acknowledge the amendment","open the perioperative record"],"actions_ar":["مراجعة البند الجراحي","إقرار التعديل","فتح سجل ما حول الجراحة"],"senders":["Surgical Services Office","Perioperative Coordination"],"senders_ar":["مكتب الخدمات الجراحية","تنسيق ما حول الجراحة"],"signatures":["Surgical Services","Perioperative Coordination"],"signatures_ar":["الخدمات الجراحية","تنسيق ما حول الجراحة"]},
    {"id":"maternity","area":"Maternity Services","area_ar":"خدمات الأمومة","event":"a maternal-care item requiring follow-up","event_ar":"بند رعاية أمومة يتطلب متابعة","objects":["antenatal risk review","labour ward handover","newborn screening result","postnatal follow-up","fetal monitoring addendum"],"objects_ar":["مراجعة مخاطر ما قبل الولادة","تسليم جناح الولادة","نتيجة فحص المولود","متابعة ما بعد الولادة","ملحق مراقبة الجنين"],"actions":["review the maternity item","acknowledge the handover","open the maternal record"],"actions_ar":["مراجعة بند الأمومة","إقرار التسليم","فتح سجل الأمومة"],"senders":["Maternity Coordination","Women’s Health Services"],"senders_ar":["تنسيق الأمومة","خدمات صحة المرأة"],"signatures":["Maternity Services","Women’s Health Services"],"signatures_ar":["خدمات الأمومة","خدمات صحة المرأة"]},
    {"id":"dialysis","area":"Renal Services","area_ar":"خدمات الكلى","event":"a renal-care workflow item requiring review","event_ar":"بند رعاية كلوية يتطلب مراجعة","objects":["dialysis schedule change","vascular access note","renal medication review","dialysis adequacy result","transplant clinic follow-up"],"objects_ar":["تغيير جدول الغسيل الكلوي","ملاحظة وصول وعائي","مراجعة دواء كلوي","نتيجة كفاءة الغسيل الكلوي","متابعة عيادة الزراعة"],"actions":["review the renal item","acknowledge the update","open the dialysis record"],"actions_ar":["مراجعة البند الكلوي","إقرار التحديث","فتح سجل الغسيل الكلوي"],"senders":["Renal Services Coordination","Dialysis Unit Office"],"senders_ar":["تنسيق خدمات الكلى","مكتب وحدة الغسيل الكلوي"],"signatures":["Renal Services","Dialysis Unit"],"signatures_ar":["خدمات الكلى","وحدة الغسيل الكلوي"]},
    {"id":"respiratory","area":"Respiratory Care","area_ar":"الرعاية التنفسية","event":"a respiratory-care item awaiting acknowledgement","event_ar":"بند رعاية تنفسية بانتظار الإقرار","objects":["ventilator setting review","oxygen therapy order","pulmonary function report","respiratory isolation note","home oxygen assessment"],"objects_ar":["مراجعة إعدادات جهاز التنفس","طلب علاج بالأكسجين","تقرير وظائف الرئة","ملاحظة عزل تنفسي","تقييم أكسجين منزلي"],"actions":["review the respiratory item","acknowledge the order","open the care record"],"actions_ar":["مراجعة البند التنفسي","إقرار الطلب","فتح سجل الرعاية"],"senders":["Respiratory Care Services","Pulmonary Coordination"],"senders_ar":["خدمات الرعاية التنفسية","تنسيق أمراض الرئة"],"signatures":["Respiratory Care","Pulmonary Services"],"signatures_ar":["الرعاية التنفسية","خدمات الرئة"]},
    {"id":"rehabilitation","area":"Rehabilitation Services","area_ar":"خدمات التأهيل","event":"a rehabilitation plan requiring clinical review","event_ar":"خطة تأهيل تتطلب مراجعة سريرية","objects":["physiotherapy plan","occupational therapy assessment","mobility risk update","speech therapy note","discharge rehabilitation plan"],"objects_ar":["خطة العلاج الطبيعي","تقييم العلاج الوظيفي","تحديث مخاطر الحركة","ملاحظة علاج النطق","خطة تأهيل الخروج"],"actions":["review the rehabilitation plan","acknowledge the assessment","open the therapy record"],"actions_ar":["مراجعة خطة التأهيل","إقرار التقييم","فتح سجل العلاج"],"senders":["Rehabilitation Coordination","Allied Health Services"],"senders_ar":["تنسيق التأهيل","خدمات المهن الصحية المساندة"],"signatures":["Rehabilitation Services","Allied Health Services"],"signatures_ar":["خدمات التأهيل","المهن الصحية المساندة"]},
    {"id":"vaccination","area":"Occupational Health","area_ar":"الصحة المهنية","event":"a workforce-health item requiring follow-up","event_ar":"بند صحة القوى العاملة يتطلب متابعة","objects":["vaccination record","fitness-to-work review","exposure assessment","screening result","immunity status update"],"objects_ar":["سجل التطعيم","مراجعة اللياقة للعمل","تقييم التعرض","نتيجة الفحص","تحديث حالة المناعة"],"actions":["review the occupational-health item","acknowledge the notice","open the staff-health record"],"actions_ar":["مراجعة بند الصحة المهنية","إقرار الإشعار","فتح سجل صحة الموظف"],"senders":["Occupational Health Services","Employee Health Clinic"],"senders_ar":["خدمات الصحة المهنية","عيادة صحة الموظفين"],"signatures":["Occupational Health","Employee Health Clinic"],"signatures_ar":["الصحة المهنية","عيادة صحة الموظفين"]},
    {"id":"pathology","area":"Anatomic Pathology","area_ar":"علم الأمراض التشريحي","event":"a pathology workflow item awaiting review","event_ar":"بند مسار علم الأمراض بانتظار المراجعة","objects":["histopathology addendum","specimen discrepancy","cytology report","tumour board pathology note","biopsy tracking exception"],"objects_ar":["ملحق الأنسجة المرضية","تباين عينة","تقرير خلوي","ملاحظة علم أمراض لمجلس الأورام","استثناء تتبع خزعة"],"actions":["review the pathology item","acknowledge the addendum","open the specimen record"],"actions_ar":["مراجعة بند علم الأمراض","إقرار الملحق","فتح سجل العينة"],"senders":["Anatomic Pathology Services","Specimen Coordination Desk"],"senders_ar":["خدمات علم الأمراض التشريحي","مكتب تنسيق العينات"],"signatures":["Anatomic Pathology","Specimen Coordination"],"signatures_ar":["علم الأمراض التشريحي","تنسيق العينات"]},
    {"id":"nutrition","area":"Clinical Nutrition","area_ar":"التغذية السريرية","event":"a nutrition-care item requiring review","event_ar":"بند رعاية تغذوية يتطلب مراجعة","objects":["enteral feeding plan","nutrition risk assessment","diet order exception","parenteral nutrition review","allergy-related meal update"],"objects_ar":["خطة التغذية الأنبوبية","تقييم مخاطر التغذية","استثناء طلب حمية","مراجعة تغذية وريدية","تحديث وجبة متعلقة بالحساسية"],"actions":["review the nutrition item","acknowledge the plan","open the nutrition record"],"actions_ar":["مراجعة بند التغذية","إقرار الخطة","فتح سجل التغذية"],"senders":["Clinical Nutrition Services","Dietetic Coordination"],"senders_ar":["خدمات التغذية السريرية","تنسيق التغذية"],"signatures":["Clinical Nutrition","Dietetic Services"],"signatures_ar":["التغذية السريرية","خدمات التغذية"]},
    {"id":"oncology","area":"Oncology Services","area_ar":"خدمات الأورام","event":"an oncology-care item requiring clinical follow-up","event_ar":"بند رعاية أورام يتطلب متابعة سريرية","objects":["chemotherapy protocol review","tumour board decision","infusion schedule change","treatment consent note","oncology follow-up plan"],"objects_ar":["مراجعة بروتوكول العلاج الكيميائي","قرار مجلس الأورام","تغيير جدول التسريب","ملاحظة موافقة العلاج","خطة متابعة الأورام"],"actions":["review the oncology item","acknowledge the plan","open the treatment record"],"actions_ar":["مراجعة بند الأورام","إقرار الخطة","فتح سجل العلاج"],"senders":["Oncology Care Coordination","Cancer Centre Operations"],"senders_ar":["تنسيق رعاية الأورام","عمليات مركز السرطان"],"signatures":["Oncology Services","Cancer Centre Operations"],"signatures_ar":["خدمات الأورام","عمليات مركز السرطان"]},
])

V30_KNOWLEDGE["admin"].extend([
    {"id":"quality_admin","area":"Quality Administration","area_ar":"إدارة الجودة","event":"a quality-governance item requiring administrative follow-up","event_ar":"بند حوكمة جودة يتطلب متابعة إدارية","objects":["accreditation evidence request","policy publication record","committee action log","audit document index","quality dashboard exception"],"objects_ar":["طلب دليل اعتماد","سجل نشر سياسة","سجل إجراءات لجنة","فهرس مستندات تدقيق","استثناء لوحة معلومات الجودة"],"actions":["review the quality item","open the evidence record","acknowledge the request"],"actions_ar":["مراجعة بند الجودة","فتح سجل الدليل","إقرار الطلب"],"senders":["Quality Management Office","Accreditation Coordination"],"senders_ar":["مكتب إدارة الجودة","تنسيق الاعتماد"],"signatures":["Quality Management","Accreditation Coordination"],"signatures_ar":["إدارة الجودة","تنسيق الاعتماد"]},
    {"id":"legal","area":"Legal Affairs","area_ar":"الشؤون القانونية","event":"a legal-administration item awaiting review","event_ar":"بند إداري قانوني بانتظار المراجعة","objects":["contract clarification","records preservation notice","consent-form revision","legal correspondence log","regulatory response draft"],"objects_ar":["توضيح عقد","إشعار حفظ سجلات","تعديل نموذج موافقة","سجل مراسلات قانونية","مسودة رد تنظيمي"],"actions":["review the legal item","open the case record","acknowledge the notice"],"actions_ar":["مراجعة البند القانوني","فتح سجل الحالة","إقرار الإشعار"],"senders":["Legal Affairs Office","Compliance Administration"],"senders_ar":["مكتب الشؤون القانونية","إدارة الامتثال"],"signatures":["Legal Affairs","Compliance Administration"],"signatures_ar":["الشؤون القانونية","إدارة الامتثال"]},
    {"id":"training_admin","area":"Learning and Development","area_ar":"التعلم والتطوير","event":"an employee-learning administration item requiring action","event_ar":"بند إداري لتعلم الموظف يتطلب إجراءً","objects":["course enrollment","attendance correction","certificate record","training nomination","orientation schedule"],"objects_ar":["تسجيل دورة","تصحيح حضور","سجل شهادة","ترشيح تدريب","جدول تهيئة"],"actions":["review the training item","open the learning record","confirm the nomination"],"actions_ar":["مراجعة بند التدريب","فتح سجل التعلم","تأكيد الترشيح"],"senders":["Learning and Development","Workforce Development"],"senders_ar":["التعلم والتطوير","تطوير القوى العاملة"],"signatures":["Learning and Development","Workforce Development"],"signatures_ar":["التعلم والتطوير","تطوير القوى العاملة"]},
    {"id":"executive","area":"Executive Office","area_ar":"المكتب التنفيذي","event":"an executive-administration item requiring coordination","event_ar":"بند إداري تنفيذي يتطلب تنسيقاً","objects":["committee briefing pack","leadership meeting action","executive correspondence","board document review","departmental status request"],"objects_ar":["حزمة إحاطة اللجنة","إجراء اجتماع قيادي","مراسلات تنفيذية","مراجعة مستند مجلس الإدارة","طلب حالة قسم"],"actions":["review the executive item","open the briefing record","acknowledge the action"],"actions_ar":["مراجعة البند التنفيذي","فتح سجل الإحاطة","إقرار الإجراء"],"senders":["Executive Office","Corporate Affairs"],"senders_ar":["المكتب التنفيذي","الشؤون المؤسسية"],"signatures":["Executive Office","Corporate Affairs"],"signatures_ar":["المكتب التنفيذي","الشؤون المؤسسية"]},
    {"id":"communications","area":"Corporate Communications","area_ar":"الاتصال المؤسسي","event":"a communications item awaiting administrative review","event_ar":"بند اتصال بانتظار المراجعة الإدارية","objects":["public notice draft","staff announcement","media enquiry log","campaign approval","website content update"],"objects_ar":["مسودة إشعار عام","إعلان للموظفين","سجل استفسار إعلامي","اعتماد حملة","تحديث محتوى الموقع"],"actions":["review the communication","open the approval record","acknowledge the draft"],"actions_ar":["مراجعة الاتصال","فتح سجل الاعتماد","إقرار المسودة"],"senders":["Corporate Communications","Public Affairs Office"],"senders_ar":["الاتصال المؤسسي","مكتب الشؤون العامة"],"signatures":["Corporate Communications","Public Affairs"],"signatures_ar":["الاتصال المؤسسي","الشؤون العامة"]},
    {"id":"transport","area":"Transport Services","area_ar":"خدمات النقل","event":"a logistics item requiring coordination","event_ar":"بند لوجستي يتطلب تنسيقاً","objects":["patient transport schedule","fleet booking exception","courier route change","ambulance transfer log","vehicle access permit"],"objects_ar":["جدول نقل المرضى","استثناء حجز الأسطول","تغيير مسار المندوب","سجل نقل إسعاف","تصريح دخول مركبة"],"actions":["review the logistics item","open the transport record","confirm the arrangement"],"actions_ar":["مراجعة البند اللوجستي","فتح سجل النقل","تأكيد الترتيب"],"senders":["Transport Services","Logistics Coordination"],"senders_ar":["خدمات النقل","تنسيق اللوجستيات"],"signatures":["Transport Services","Logistics Coordination"],"signatures_ar":["خدمات النقل","تنسيق اللوجستيات"]},
    {"id":"housing","area":"Staff Services","area_ar":"خدمات الموظفين","event":"a staff-services item requiring administrative review","event_ar":"بند خدمات موظفين يتطلب مراجعة إدارية","objects":["housing allocation","staff shuttle request","uniform issue","employee ID replacement","staff accommodation maintenance"],"objects_ar":["تخصيص سكن","طلب حافلة الموظفين","صرف زي موحد","استبدال هوية موظف","صيانة سكن الموظفين"],"actions":["review the staff-service item","open the request record","confirm the arrangement"],"actions_ar":["مراجعة بند خدمة الموظفين","فتح سجل الطلب","تأكيد الترتيب"],"senders":["Staff Services Office","Employee Experience"],"senders_ar":["مكتب خدمات الموظفين","تجربة الموظف"],"signatures":["Staff Services","Employee Experience"],"signatures_ar":["خدمات الموظفين","تجربة الموظف"]},
    {"id":"inventory","area":"Warehouse Operations","area_ar":"عمليات المستودع","event":"an inventory item requiring reconciliation","event_ar":"بند مخزون يتطلب مطابقة","objects":["stock discrepancy","expiry review","goods receipt exception","inventory transfer","consignment count"],"objects_ar":["تباين مخزون","مراجعة انتهاء صلاحية","استثناء استلام بضائع","تحويل مخزون","جرد شحنة أمانة"],"actions":["review the inventory item","open the stock record","acknowledge the discrepancy"],"actions_ar":["مراجعة بند المخزون","فتح سجل المخزون","إقرار التباين"],"senders":["Warehouse Operations","Materials Management"],"senders_ar":["عمليات المستودع","إدارة المواد"],"signatures":["Warehouse Operations","Materials Management"],"signatures_ar":["عمليات المستودع","إدارة المواد"]},
])

V30_KNOWLEDGE["it"].extend([
    {"id":"integration","area":"Integration Services","area_ar":"خدمات التكامل","event":"an interface workflow item requiring technical review","event_ar":"بند مسار واجهة يتطلب مراجعة تقنية","objects":["HL7 queue exception","API certificate update","interface mapping change","message-routing alert","integration endpoint review"],"objects_ar":["استثناء قائمة انتظار HL7","تحديث شهادة API","تغيير تخطيط الواجهة","تنبيه توجيه الرسائل","مراجعة نقطة نهاية التكامل"],"actions":["review the integration item","open the interface ticket","acknowledge the change"],"actions_ar":["مراجعة بند التكامل","فتح تذكرة الواجهة","إقرار التغيير"],"senders":["Integration Services","Interoperability Team"],"senders_ar":["خدمات التكامل","فريق التشغيل البيني"],"signatures":["Integration Services","Interoperability Team"],"signatures_ar":["خدمات التكامل","فريق التشغيل البيني"]},
    {"id":"endpoint","area":"Endpoint Management","area_ar":"إدارة الأجهزة الطرفية","event":"a managed-device item awaiting review","event_ar":"بند جهاز مُدار بانتظار المراجعة","objects":["device compliance exception","encryption status alert","software deployment","laptop ownership review","mobile device enrollment"],"objects_ar":["استثناء امتثال جهاز","تنبيه حالة التشفير","نشر برنامج","مراجعة ملكية جهاز محمول","تسجيل جهاز جوال"],"actions":["review the endpoint item","open the device record","acknowledge the deployment"],"actions_ar":["مراجعة بند الجهاز الطرفي","فتح سجل الجهاز","إقرار النشر"],"senders":["Endpoint Management","Digital Workplace Operations"],"senders_ar":["إدارة الأجهزة الطرفية","عمليات مكان العمل الرقمي"],"signatures":["Endpoint Management","Digital Workplace"],"signatures_ar":["إدارة الأجهزة الطرفية","مكان العمل الرقمي"]},
    {"id":"telecom","area":"Unified Communications","area_ar":"الاتصالات الموحدة","event":"a communications-platform item requiring technical review","event_ar":"بند منصة اتصالات يتطلب مراجعة تقنية","objects":["call routing change","contact centre queue","Teams voice migration","extension ownership review","video conference gateway"],"objects_ar":["تغيير توجيه المكالمات","قائمة انتظار مركز الاتصال","ترحيل صوت Teams","مراجعة ملكية تحويلة","بوابة مؤتمر فيديو"],"actions":["review the communications item","open the service request","acknowledge the change"],"actions_ar":["مراجعة بند الاتصالات","فتح طلب الخدمة","إقرار التغيير"],"senders":["Unified Communications","Telecommunications Services"],"senders_ar":["الاتصالات الموحدة","خدمات الاتصالات"],"signatures":["Unified Communications","Telecommunications Services"],"signatures_ar":["الاتصالات الموحدة","خدمات الاتصالات"]},
    {"id":"pacs","area":"Imaging Informatics","area_ar":"معلوماتية التصوير","event":"an imaging-system item requiring technical review","event_ar":"بند نظام تصوير يتطلب مراجعة تقنية","objects":["PACS archive alert","modality worklist exception","DICOM routing change","image retention policy","viewer access review"],"objects_ar":["تنبيه أرشيف نظام الصور","استثناء قائمة عمل الجهاز","تغيير توجيه DICOM","سياسة الاحتفاظ بالصور","مراجعة وصول العارض"],"actions":["review the imaging-system item","open the PACS ticket","acknowledge the change"],"actions_ar":["مراجعة بند نظام التصوير","فتح تذكرة نظام الصور","إقرار التغيير"],"senders":["Imaging Informatics","PACS Support"],"senders_ar":["معلوماتية التصوير","دعم نظام الصور"],"signatures":["Imaging Informatics","PACS Support"],"signatures_ar":["معلوماتية التصوير","دعم نظام الصور"]},
    {"id":"analytics","area":"Data and Analytics","area_ar":"البيانات والتحليلات","event":"a data-platform item awaiting review","event_ar":"بند منصة بيانات بانتظار المراجعة","objects":["dashboard refresh failure","data-quality exception","report ownership review","analytics workspace access","scheduled extract alert"],"objects_ar":["فشل تحديث لوحة المعلومات","استثناء جودة البيانات","مراجعة ملكية تقرير","وصول مساحة عمل التحليلات","تنبيه استخراج مجدول"],"actions":["review the data item","open the analytics ticket","acknowledge the exception"],"actions_ar":["مراجعة بند البيانات","فتح تذكرة التحليلات","إقرار الاستثناء"],"senders":["Data and Analytics","Business Intelligence Services"],"senders_ar":["البيانات والتحليلات","خدمات ذكاء الأعمال"],"signatures":["Data and Analytics","Business Intelligence"],"signatures_ar":["البيانات والتحليلات","ذكاء الأعمال"]},
    {"id":"change","area":"IT Change Management","area_ar":"إدارة التغيير التقني","event":"a production-change item requiring technical approval","event_ar":"بند تغيير إنتاج يتطلب اعتماداً تقنياً","objects":["emergency change request","maintenance window","rollback plan","change collision alert","post-implementation review"],"objects_ar":["طلب تغيير طارئ","نافذة صيانة","خطة تراجع","تنبيه تعارض تغيير","مراجعة ما بعد التنفيذ"],"actions":["review the change item","open the change record","acknowledge the schedule"],"actions_ar":["مراجعة بند التغيير","فتح سجل التغيير","إقرار الجدول"],"senders":["IT Change Management","Technology Governance"],"senders_ar":["إدارة التغيير التقني","حوكمة التقنية"],"signatures":["IT Change Management","Technology Governance"],"signatures_ar":["إدارة التغيير التقني","حوكمة التقنية"]},
    {"id":"vendor_it","area":"Technology Vendor Management","area_ar":"إدارة موردي التقنية","event":"a vendor-support item requiring technical coordination","event_ar":"بند دعم مورد يتطلب تنسيقاً تقنياً","objects":["support entitlement renewal","vendor remote session","license true-up","maintenance agreement","technical escalation"],"objects_ar":["تجديد استحقاق الدعم","جلسة عن بعد للمورد","تسوية ترخيص","اتفاقية صيانة","تصعيد تقني"],"actions":["review the vendor item","open the support record","acknowledge the request"],"actions_ar":["مراجعة بند المورد","فتح سجل الدعم","إقرار الطلب"],"senders":["Technology Vendor Management","IT Commercial Services"],"senders_ar":["إدارة موردي التقنية","الخدمات التجارية لتقنية المعلومات"],"signatures":["Technology Vendor Management","IT Commercial Services"],"signatures_ar":["إدارة موردي التقنية","الخدمات التجارية"]},
    {"id":"continuity","area":"IT Service Continuity","area_ar":"استمرارية خدمات تقنية المعلومات","event":"a resilience item requiring technical review","event_ar":"بند مرونة يتطلب مراجعة تقنية","objects":["disaster recovery exercise","business continuity dependency","recovery time validation","failover test","critical service inventory"],"objects_ar":["تمرين التعافي من الكوارث","اعتمادية استمرارية الأعمال","تحقق زمن الاستعادة","اختبار التبديل التلقائي","جرد الخدمات الحرجة"],"actions":["review the continuity item","open the resilience record","acknowledge the test"],"actions_ar":["مراجعة بند الاستمرارية","فتح سجل المرونة","إقرار الاختبار"],"senders":["IT Service Continuity","Technology Resilience Office"],"senders_ar":["استمرارية خدمات تقنية المعلومات","مكتب مرونة التقنية"],"signatures":["IT Service Continuity","Technology Resilience"],"signatures_ar":["استمرارية الخدمات","مرونة التقنية"]},
])
V30_GENERIC_GREETINGS = {"English":["Dear Staff Member","Hello Team","Dear Healthcare Employee","Attention Staff","Dear Colleague"], "Arabic":["عزيزي الموظف","فريق العمل العزيز","عزيزي الممارس الصحي","إلى الزملاء الأعزاء","عزيزي الزميل"]}
V30_DEPARTMENT_GREETINGS = {"English":["Dear {area} Team","Hello {area} Colleague","To the {area} Team"], "Arabic":["فريق {area} العزيز","الزميل في {area}","إلى فريق {area}"]}
V30_FIRST_NAMES = ["Sarah Almutairi","Ahmed Alotaibi","Noura Alshamri","Fahad Aldosari","Mona Alharbi","Khalid Alanazi","Lama Alqahtani","Faisal Alzahrani"]

V30_OBVIOUS_DOMAINS = ["secure-staff-check.net","hospital-access-alert.co","portal-confirm-now.net","clinical-update-login.org","staff-verify-center.com"]
V30_SIMILAR_DOMAINS = ["hospital-portal.org","hospital-services.net","moh-clinical.org","hospital-workflow.com","staff-hospital.org"]
V30_NEAR_DOMAINS = ["hospitaI.org","hospital-sso.org","hospital365.org","hospital-sharepoint.org","hospital-sa.org"]






def _v30_role_type(role):
    return ROLE_MAP.get(role, ROLE_MAP["Clinical"])[2]








def _v30_recipient(role, index, language, phase):
    try: return get_recipient(role, index, language, phase="assess" if phase=="assess" else "learn")
    except Exception: return "employee@hospital.org"


def _v30_display_name(email):
    local = (email or "employee").split("@")[0].replace("."," ").replace("_"," ")
    parts = [p for p in local.split() if len(p)>1 and p.lower() not in {"dr","n","m","t","ph","s"}]
    return " ".join(p.capitalize() for p in parts[-2:]) or _V30_RNG.choice(V30_FIRST_NAMES)


def _v30_title_for(email, language):
    """Hard-level greetings use the real person's name — for physicians that
    should read naturally too (e.g. "Dear Dr. Ahmed Alotaibi"), since a real
    internal message would never drop a colleague's title."""
    local = (email or "").split("@")[0].lower()
    prefix = local.split(".")[0] if "." in local else local
    if prefix == "dr":
        return "Dr. " if language != "Arabic" else "د. "
    return ""


def _v30_domain(difficulty):
    return _V30_RNG.choice({"easy":V30_OBVIOUS_DOMAINS,"medium":V30_SIMILAR_DOMAINS,"hard":V30_NEAR_DOMAINS}[difficulty])


def _v30_subject(plan, urgency_phrase):
    is_ar = plan.get("language") == "Arabic"
    disp = plan.get("object_disp", plan["object"])
    obj = disp if is_ar else disp.title()
    if is_ar:
        easy = [f"عاجل: إجراء مطلوب اليوم بشأن {obj}", f"مراجعة فورية مطلوبة: {obj}", f"إشعار أخير: متابعة {obj}", f"إجراء مطلوب اليوم — {obj}"]
        medium = [f"متابعة مطلوبة: {obj}", f"إشعار سير عمل — {obj}", f"فترة مراجعة {obj}", f"بانتظار التأكيد: {obj}"]
        hard = [f"{obj} — تحديث المرجع", f"متابعة بشأن {obj}", f"ملاحظة مراجعة: {obj}", f"بخصوص {obj}"]
    else:
        easy = [f"Urgent: {obj} Requires Action Today", f"Immediate Review Required: {obj}", f"Final Notice: {obj} Access Pending", f"Action Needed Today — {obj}"]
        medium = [f"Follow-up Required: {obj}", f"Workflow Notice — {obj}", f"Review Window for {obj}", f"Pending Confirmation: {obj}"]
        hard = [f"{obj} — Reference Update", f"Follow-up on {obj}", f"Review note: {obj}", f"Regarding {obj}"]
    return _V30_RNG.choice({"easy":easy,"medium":medium,"hard":hard}[plan["difficulty"]])


def _v30_link(plan, domain):
    short = re.sub(r"[^a-z0-9]+","-",plan["family_id"].lower()).strip("-")
    token = _V30_RNG.randrange(1000,9999)
    return f"http://{domain}/{short}/{token}"


def _v30_indicator(number, key, title, description, evidence, target):
    return {"number":number,"key":key,"title":title,"description":description,"evidence":evidence,"target":target}


def _v30_compose_phishing(plan, role, index):
    lang = plan["language"]; ar = lang == "Arabic"; diff = plan["difficulty"]
    recipient = _v30_recipient(role,index,lang,plan["phase"]); person = _v30_display_name(recipient)
    domain = _v30_domain(diff); link = _v30_link(plan,domain)
    mailbox = _V30_RNG.choice(["alerts","workflow","notifications","case-review","service-desk"])
    sender = f'{plan["sender"]} <{mailbox}@{domain}>'
    urgency = _V30_RNG.choice(["today","immediately","within the next hour","before your next shift"])
    subject = _v30_subject(plan, urgency)

    if diff == "easy":
        greeting = _V30_RNG.choice(V30_GENERIC_GREETINGS[lang])
        credential = _V30_RNG.choice(["enter your password","confirm your staff PIN","submit your login credentials","verify your account details"])
        if ar:
            cred_map={"enter your password":"أدخل كلمة المرور","confirm your staff PIN":"أكد الرقم السري للموظف","submit your login credentials":"أرسل بيانات الدخول","verify your account details":"تحقق من بيانات حسابك"}; credential=cred_map[credential]
            body = f"{greeting}،\n\nورد إلى {plan['area']} إشعار بخصوص {plan['object']}، ويتطلب الإجراء متابعة سريعة حتى لا يتوقف سير العمل. يرجى {credential} {('فورًا' if urgency=='immediately' else 'اليوم')} لإتاحة {plan['action']} وإغلاق التنبيه.\n\nاستخدم الرابط التالي لإكمال الإجراء: {link}\n\nعدم الاستجابة {('خلال الساعة القادمة' if 'hour' in urgency else 'قبل المناوبة القادمة')} قد يؤدي إلى تعليق الوصول مؤقتًا.\n\nشكرًا لتعاونك،\n{plan['signature']}"
            indicators=[
                _v30_indicator(1,"domain","نطاق مرسل غير رسمي",f"عنوان المرسل يستخدم النطاق {domain} وليس نطاق المستشفى الرسمي.",domain,"from"),
                _v30_indicator(2,"urgency","استعجال أو تهديد واضح","يستخدم العنوان والنص مهلة قصيرة لدفع المستلم إلى التصرف دون تحقق.",next(x for x in ["فورًا","اليوم","خلال الساعة القادمة","قبل المناوبة القادمة"] if x in subject+body),"subject" if any(x in subject for x in ["عاجل","فوري","اليوم"]) else "body"),
                _v30_indicator(3,"credential","طلب مباشر لبيانات الدخول","تطلب الرسالة كلمة مرور أو رقمًا سريًا أو بيانات حساب مباشرة.",credential,"body"),
                _v30_indicator(4,"link","رابط خارجي واضح",f"الرابط يقود إلى النطاق غير الرسمي {domain}.",link,"link"),
                _v30_indicator(5,"greeting","تحية عامة","لا تخاطب الرسالة المستلم باسمه، ما قد يدل على إرسال جماعي.",greeting,"greeting"),
            ]
            why="تجمع الرسالة بين نطاق غير رسمي وطلب مباشر لبيانات الدخول ورابط خارجي مع ضغط زمني واضح، وهي مؤشرات أساسية على التصيد."
            tip="لا تُدخل كلمة المرور أو الرقم السري عبر رابط وارد في بريد مفاجئ. افتح النظام الرسمي مباشرة أو تواصل مع القسم عبر قناة موثوقة."
        else:
            body = f"{greeting},\n\n{plan['area']} has received a notification concerning {plan['object']}. The item requires prompt follow-up to prevent disruption to the related clinical workflow. Please {credential} {urgency} so you can {plan['action']} and close the notification.\n\nUse the following link to complete the action: {link}\n\nFailure to respond {('within the next hour' if 'hour' in urgency else 'before your next shift')} may result in temporary access suspension.\n\nThank you for your cooperation,\n{plan['signature']}"
            urgency_ev = next(x for x in ["Urgent","Immediate","Today","today","immediately","within the next hour","before your next shift"] if x in subject+body)
            indicators=[
                _v30_indicator(1,"domain","Non-official sender domain",f"The sender uses {domain}, not the hospital's official domain.",domain,"from"),
                _v30_indicator(2,"urgency","Strong urgency or threat","The message uses an immediate deadline or access consequence to pressure the recipient.",urgency_ev,"subject" if urgency_ev in subject else "body"),
                _v30_indicator(3,"credential","Direct credential request","The email directly asks for a password, staff PIN, or login credentials.",credential,"body"),
                _v30_indicator(4,"link","Visible external link",f"The link points to the non-official domain {domain}.",link,"link"),
                _v30_indicator(5,"greeting","Generic greeting","The message does not address the recipient by name, which may indicate bulk targeting.",greeting,"greeting"),
            ]
            why="The email combines a non-official domain, a direct credential request, an external link, and strong time pressure—clear beginner-level phishing indicators."
            tip="Never enter a password or staff PIN through an unexpected email link. Open the official system directly or verify the request through a trusted channel."
        attachment=""
    elif diff == "medium":
        greeting = _V30_RNG.choice(V30_DEPARTMENT_GREETINGS[lang]).format(area=plan["area"])
        deadline = _V30_RNG.choice(["within 24 hours","by tomorrow afternoon","before the current review window closes"])
        if ar:
            dmap={"within 24 hours":"خلال 24 ساعة","by tomorrow afternoon":"قبل ظهر الغد","before the current review window closes":"قبل إغلاق فترة المراجعة الحالية"}; deadline=dmap[deadline]
            cta=f"سجّل الدخول إلى بوابة المتابعة لمراجعة {plan['object']}"; label="فتح سجل المتابعة"
            body=f"{greeting}،\n\nتجري {plan['area']} مراجعة دورية للطلبات المفتوحة، وقد ظهر {plan['object']} ضمن العناصر التي تحتاج إلى متابعة. يرجى {plan['action']} {deadline} حتى يبقى السجل ضمن مسار العمل الحالي.\n\n{cta}: [{label}]({link})\n\nإذا لم تكن مسؤولًا عن هذا العنصر، يرجى إعادة توجيه الإشعار إلى المنسق المعتمد.\n\nمع التحية،\n{plan['signature']}"
            indicators=[
                _v30_indicator(1,"domain","نطاق مشابه لكنه غير رسمي",f"النطاق {domain} يبدو قريبًا من جهة رسمية لكنه ليس hospital.org.",domain,"from"),
                _v30_indicator(2,"workflow","طلب تسجيل دخول خارج مسار العمل المعتاد","تطلب الرسالة فتح بوابة خارجية بدل استخدام النظام الداخلي المعروف.",cta,"body"),
                _v30_indicator(3,"link","رابط بوابة غير معتمدة",f"زر المتابعة يقود إلى {domain}، وهو نطاق غير رسمي.",link,"link"),
            ]; why="تبدو الرسالة مهنية ومرتبطة بالعمل، لكن نطاق البوابة غير رسمي وطريقة تسجيل الدخول لا تتبع المسار المعتاد."; tip="افتح النظام من الاختصار الرسمي أو المفضلة، ولا تعتمد على زر تسجيل دخول داخل بريد غير متوقع."
        else:
            cta=f"Sign in to the review portal to {plan['action']}"; label="Open review record"
            body=f"{greeting},\n\n{plan['area']} is completing a routine review of open workflow items. {plan['object'].title()} has been listed for follow-up. Please {plan['action']} {deadline} so the record remains in the current workflow queue.\n\n{cta}: [{label}]({link})\n\nIf this item is not assigned to you, forward the notice to the approved coordinator rather than entering case details by email.\n\nRegards,\n{plan['signature']}"
            indicators=[
                _v30_indicator(1,"domain","Look-alike sender domain",f"The domain {domain} resembles a workplace service but is not hospital.org.",domain,"from"),
                _v30_indicator(2,"workflow","Unusual external sign-in workflow","The email directs the recipient to sign in through an emailed portal instead of the known internal system.",cta,"body"),
                _v30_indicator(3,"link","Unapproved portal link",f"The review button resolves to the non-official domain {domain}.",link,"link"),
            ]; why="The message is professionally written and job-relevant, but the look-alike domain and external sign-in workflow are inconsistent with approved hospital practice."; tip="Navigate to the official system independently. Do not use an emailed sign-in button unless the destination has been verified."
        attachment=""
    else:
        title = _v30_title_for(recipient, lang)
        greeting = (f"Dear {title}{person}" if not ar else f"عزيزي {title}{person}")
        # hard uses subtle channel/identity issue; no direct credentials or artificial deadline
        channel=plan["channel"]
        attachment=""; qr_marker=""
        disp_obj = plan.get("object_disp", plan["object"])
        disp_area = plan.get("area_disp", plan["area"])
        disp_action = plan.get("action_disp", plan["action"])
        disp_signature = plan.get("signature_disp", plan["signature"])
        if channel in ("pdf","xlsx","docx"):
            ext = {"pdf":".pdf","xlsx":".xlsx","docx":".docx"}[channel]
            attachment=re.sub(r"[^a-z0-9]+","_",plan["family_id"])+f"_{_V30_RNG.randrange(10,99)}{ext}"
        elif channel == "qr":
            qr_label = "Verification code" if not ar else "رمز التحقق"
            qr_marker = f"[QR: {qr_label}]"
        elif channel == "m365":
            domain = "hospital365.org" if "365" not in domain else domain
        if ar:
            area_ar_disp = disp_area
            intro = _V30_RNG.choice([
                f"أرسل لك متابعة بخصوص {disp_obj} التي نوقشت ضمن {area_ar_disp}.",
                f"بالإشارة إلى ما ناقشناه مع {area_ar_disp}، حبيت أشاركك تفاصيل {disp_obj}.",
                f"طلب مني فريق {area_ar_disp} إرسال تحديث {disp_obj} لك مباشرة.",
                f"لدي تفاصيل {disp_obj} جاهزة — تطرقنا لها بإيجاز باجتماع {area_ar_disp}.",
                f"أتابع معك بخصوص {disp_obj} من {area_ar_disp} — حبيت أوصلها لك قبل نهاية اليوم.",
                f"كما اتفقنا مع {area_ar_disp}، هذا مرجع {disp_obj} اللي تحتاجه.",
            ])
            bridge = _V30_RNG.choice([
                f"أضفت مرجع الحالة حتى تتمكن من {disp_action} عند توفر الوقت، ثم تدوين الملاحظة في السجل المعتاد.",
                f"تقدر {disp_action} مباشرة من هذا، وتسجل النتيجة بالسجل المعتاد بعدها.",
                f"خذ وقتك بالإطلاع، وبعدها {disp_action} ودوّن الملاحظة بالنظام كالعادة.",
                f"ما فيه استعجال — بس {disp_action} متى ما ناسبك ودوّنها بالنظام كالمعتاد.",
            ])
            body=f"{greeting}،\n\n{intro} {bridge}\n\n"
            if attachment: body += _V30_RNG.choice([
                f"ستجد التفاصيل في المرفق {attachment}. يرجى مطابقته مع رقم الحالة في النظام قبل اعتماده.\n\n",
                f"أرفقت {attachment} وفيه التفاصيل كاملة — يفضل تتأكد من رقم الحالة قبل الاعتماد عليه.\n\n",
                f"{attachment} فيه كل التفاصيل. طابقه مع رقم الحالة بالنظام قبل ما تتصرف بناءً عليه.\n\n",
                f"التفاصيل موجودة بالمرفق {attachment}؛ تأكد إن رقم الحالة يطابق قبل ما تكمل.\n\n",
            ])
            elif qr_marker: body += _V30_RNG.choice([
                f"للوصول السريع من جوالك، امسح الرمز التالي: {qr_marker}\n\n",
                f"إذا أسهل لك، تقدر تمسح الرمز التالي من جوالك: {qr_marker}\n\n",
                f"مسح الرمز أدناه بياخذك للصفحة مباشرة: {qr_marker}\n\n",
            ])
            elif channel == "m365": body += _V30_RNG.choice([
                f"سجّل الدخول عبر [تسجيل الدخول بحساب Microsoft 365]({link}) لعرض المرجع.\n\n",
                f"بتحتاج [تسجيل الدخول بحساب Microsoft 365]({link}) عشان تفتح هذا.\n\n",
                f"الوصول يتطلب [تسجيل دخول سريع بحساب Microsoft 365]({link}).\n\n",
            ])
            elif channel == "sharepoint": body += _V30_RNG.choice([
                f"المستند متاح على SharePoint: [فتح في SharePoint]({link})\n\n",
                f"حطيته على SharePoint — [فتح في SharePoint]({link}) متى ما جهزت.\n\n",
                f"بتلقاه على SharePoint حقنا: [فتح في SharePoint]({link})\n\n",
            ])
            else: body += _V30_RNG.choice([
                f"مرجع المتابعة متاح هنا: [عرض مرجع الحالة]({link})\n\n",
                f"هذا رابط مرجع الحالة: [عرض مرجع الحالة]({link})\n\n",
                f"تقدر تفتحه من هنا: [عرض مرجع الحالة]({link})\n\n",
            ])
            closing = _V30_RNG.choice([
                "لا توجد حاجة لإرسال أي معلومات حساسة عبر البريد.",
                "كالعادة، الرجاء عدم تضمين أي تفاصيل حساسة بالبريد.",
                "ما فيه شي حساس يحتاج يرسل بالبريد بخصوص هذا.",
                "بس تذكير — خلي أي معلومات حساسة برا البريد الإلكتروني.",
            ])
            sign_off = _V30_RNG.choice(["مع التقدير","مع خالص الشكر","تحياتي"])
            body += f"{closing}\n\n{sign_off}،\n{disp_signature}"
            indicators=[_v30_indicator(1,"domain","اختلاف دقيق في هوية النطاق",f"النطاق {domain} قريب بصريًا من النطاق الرسمي لكنه مختلف.",domain,"from")]
            if attachment: indicators.append(_v30_indicator(2,"attachment","مرفق غير متوقع","يجب التحقق من المرفق عبر النظام الداخلي قبل فتحه، حتى لو بدا السياق واقعيًا.",attachment,"attachment"))
            elif qr_marker: indicators.append(_v30_indicator(2,"qr","رمز QR غير متوقع","لا يُفترض مسح رمز غير معروف للوصول إلى نظام داخلي؛ افتح النظام مباشرة بدل المسح.",qr_label,"body"))
            elif channel == "m365": indicators.append(_v30_indicator(2,"m365","انتحال تسجيل دخول Microsoft 365","يطلب زر تسجيل الدخول بحساب Microsoft عبر رابط بريدي بدل بوابة المستشفى الرسمية.",domain,"link"))
            else: indicators.append(_v30_indicator(2,"link","رابط خارجي خفي",f"نص الرابط يبدو مهنيًا لكن وجهته هي {domain}.",link,"link"))
            why=_V30_RNG.choice([
                "السياق واللغة طبيعيان جدًا، لكن هناك اختلافًا دقيقًا في هوية النطاق وقناة خارجية غير متوقعة. هذه مؤشرات متقدمة تتطلب التحقق من التفاصيل.",
                "الرسالة مكتوبة بأسلوب داخلي مقنع تمامًا، إلا أن دقة النطاق والقناة المستخدمة لا تطابقان الإجراءات المعتمدة — وهذا نمط تصيد متقدم يصعب رصده بالنظر السريع.",
                "لا يوجد استعجال أو لغة تهديد هنا، وهذا بالضبط ما يجعل هذا النوع خطيرًا: الاعتماد فقط على واقعية السياق دون التحقق من النطاق والقناة قد يؤدي لخطأ.",
                "أسلوب الرسالة وسياقها الوظيفي يجعلانها تبدو صادرة من زميل حقيقي، لكن الفارق الدقيق بنطاق المرسل هو المؤشر الوحيد الفعلي هنا — ولهذا يصعب اكتشافه بدون تدقيق متأنٍ.",
                "غياب أي ضغط زمني أو صياغة تهديدية لا يعني أن الرسالة آمنة؛ التصيد المتقدم يعتمد على الثقة والسياق الواقعي بدل الاستعجال.",
                "المعلومات المذكورة بالرسالة (الاسم، القسم، طبيعة المهمة) دقيقة وواقعية جدًا، وهذا تحديدًا ما يجعل المستلم يقل حذره تجاه تفاصيل النطاق والقناة المستخدمة.",
                "كل شي بالرسالة يوحي بأنها من داخل المستشفى، لكن القناة المستخدمة (سواء رابط أو مرفق أو رمز) لا تتبع المسار الرسمي المعتاد لهذا النوع من الطلبات.",
                "هذا النمط من الرسائل مصمم خصيصًا ليتجاوز الفحص السريع؛ يحتاج المستلم فحصًا دقيقًا للنطاق والقناة، لا الاعتماد على انطباعه العام عن الرسالة.",
            ])
            tip=_V30_RNG.choice([
                "في الرسائل الواقعية، ركّز على النطاق الفعلي ومسار العمل، وليس على جودة اللغة أو معرفة المرسل بالسياق.",
                "لا تحكم على الرسالة من طبيعتها أو معرفتها بتفاصيل عملك — تحقق دائمًا من النطاق الدقيق للمرسل قبل أي إجراء.",
                "اجعل عادتك التحقق من مسار العمل نفسه: هل هذا القناة اللي تُستخدم فعليًا لهذا النوع من الطلبات، أم قناة جديدة غير معتادة؟",
                "قارن نطاق المرسل حرفًا بحرف مع النطاق الرسمي المعروف — الفروقات الدقيقة (حرف مشابه، امتداد إضافي) هي المؤشر الأهم بهذا المستوى.",
                "لو الرسالة تطلب فتح رابط أو مرفق أو مسح رمز خارج النظام المعتاد، تحقق أولًا مباشرة مع الجهة المعنية قبل أي إجراء، حتى لو بدت الرسالة مألوفة.",
                "الثقة بالسياق وحدها غير كافية — اسأل نفسك دائمًا: هل هذه بالضبط القناة الرسمية اللي يفترض تصل منها هذا النوع من الطلبات؟",
                "خصص لحظة للتحقق من عنوان المرسل الكامل قبل أي تفاعل، خصوصًا إذا كانت الرسالة تبدو مألوفة أو من زميل تعرفه.",
            ])
        else:
            intro = _V30_RNG.choice([
                f"I am following up on {disp_obj}, which was discussed through {disp_area}.",
                f"Following our conversation with {disp_area}, I wanted to send over the {disp_obj} details.",
                f"{disp_area} asked me to share the {disp_obj} update with you directly.",
                f"I have the {disp_obj} item ready — we touched on this briefly during the {disp_area} meeting.",
                f"Circling back on {disp_obj} from {disp_area} — wanted to get this to you before end of day.",
                f"As discussed with {disp_area}, here is the {disp_obj} reference you needed.",
            ])
            bridge = _V30_RNG.choice([
                f"I included the case reference so you can {disp_action} when convenient and record the outcome in the usual system.",
                f"You should be able to {disp_action} directly from this and log it in the usual system afterward.",
                f"Take a look when you get a chance and {disp_action}; the usual system will pick up the update.",
                f"No rush — just {disp_action} whenever it suits you and note it in the system as usual.",
            ])
            body=f"{greeting},\n\n{intro} {bridge}\n\n"
            if attachment: body += _V30_RNG.choice([
                f"The supporting detail is in {attachment}. Please match it against the case number in the internal system before relying on it.\n\n",
                f"I've attached {attachment} with the full breakdown — worth checking it against the case number first.\n\n",
                f"{attachment} has the details. Cross-check it with the case number in the system before acting on it.\n\n",
                f"Details are in the attached {attachment}; confirm the case number matches before you proceed.\n\n",
            ])
            elif qr_marker: body += _V30_RNG.choice([
                f"For quick access from your phone, scan the code below: {qr_marker}\n\n",
                f"If it's easier, you can scan this from your phone instead: {qr_marker}\n\n",
                f"Scanning the code below will take you straight there: {qr_marker}\n\n",
            ])
            elif channel == "m365": body += _V30_RNG.choice([
                f"Please [Sign in with Microsoft 365]({link}) to view the reference.\n\n",
                f"You'll need to [sign in with your Microsoft 365 account]({link}) to open this.\n\n",
                f"Access requires a quick [Microsoft 365 sign-in]({link}).\n\n",
            ])
            elif channel == "sharepoint": body += _V30_RNG.choice([
                f"The document is available on SharePoint: [Open in SharePoint]({link})\n\n",
                f"I've put this on SharePoint — [Open in SharePoint]({link}) when ready.\n\n",
                f"You'll find it on our SharePoint: [Open in SharePoint]({link})\n\n",
            ])
            else: body += _V30_RNG.choice([
                f"The reference is available here: [View case reference]({link})\n\n",
                f"Here's the link to the case reference: [View case reference]({link})\n\n",
                f"You can pull it up here: [View case reference]({link})\n\n",
            ])
            closing = _V30_RNG.choice([
                "No sensitive information needs to be sent by email.",
                "As always, please don't include any sensitive details over email.",
                "Nothing sensitive needs to go over email for this one.",
                "Just a reminder — keep any sensitive information out of email.",
            ])
            sign_off = _V30_RNG.choice(["Kind regards","Best regards","Many thanks"])
            body += f"{closing}\n\n{sign_off},\n{disp_signature}"
            indicators=[_v30_indicator(1,"domain","Subtle domain discrepancy",f"The domain {domain} is visually close to the official domain but is not identical.",domain,"from")]
            if attachment: indicators.append(_v30_indicator(2,"attachment","Unexpected contextual attachment","The attachment should be verified in the internal system before it is opened, despite the realistic context.",attachment,"attachment"))
            elif qr_marker: indicators.append(_v30_indicator(2,"qr","Unexpected QR code","An unfamiliar QR code should never be scanned to reach an internal system; open the system directly instead.","Verification code","body"))
            elif channel == "m365": indicators.append(_v30_indicator(2,"m365","Impersonated Microsoft 365 sign-in","The button asks for a Microsoft 365 sign-in through an emailed link instead of the hospital's own portal.",domain,"link"))
            else: indicators.append(_v30_indicator(2,"link","Hidden external destination",f"The professional-looking link text resolves to {domain}.",link,"link"))
            why=_V30_RNG.choice([
                "The message is natural and context-aware, but a subtle domain discrepancy and an unexpected external channel indicate a sophisticated phishing attempt.",
                "The email reads like a genuinely internal message, yet the exact domain and the delivery channel don't match approved practice — an advanced pattern that's easy to miss at a glance.",
                "There is no urgency or threatening language here, which is exactly what makes this dangerous: relying on how realistic the context feels, without checking the domain and channel, can lead to a costly mistake.",
                "The tone and workplace context make this feel like it's from a real colleague, but the one genuine tell is the subtle sender-domain mismatch — easy to miss without a careful check.",
                "The absence of urgency or threatening wording doesn't mean this is safe; advanced phishing relies on trust and plausible context rather than pressure.",
                "The specific details in the message (name, department, task) are accurate and realistic, which is exactly what lowers a reader's guard toward the domain and delivery channel.",
                "Everything about this message suggests it came from inside the hospital, but the channel used — whether a link, attachment, or code — doesn't follow the normal approved path for this kind of request.",
                "This pattern is deliberately built to slide past a quick glance; it takes a careful look at the domain and channel, not a general impression of the message, to catch it.",
            ])
            tip=_V30_RNG.choice([
                "For realistic emails, inspect the exact domain and verify the workflow independently; polished language is not proof of legitimacy.",
                "Don't judge a message by how natural it sounds or how well it knows your work — always verify the sender's exact domain before acting.",
                "Make it a habit to check the workflow itself: is this really the channel normally used for this type of request, or an unfamiliar one?",
                "Compare the sender's domain character by character against the known official one — subtle differences (a lookalike letter, an extra word) are the key signal at this level.",
                "If a message asks you to open a link or attachment, or scan a code, outside the normal system, verify directly with the sender first — even if the message feels familiar.",
                "Trusting the context alone isn't enough — always ask: is this really the official channel this type of request is supposed to come through?",
                "Take a moment to check the full sender address before engaging, especially when the message feels familiar or appears to come from someone you know.",
            ])

    return {"from":sender,"to":recipient,"subject":subject,"body":body,"attachment":attachment,
            "suspicious_link":link if not attachment else (link if diff!="hard" else ""),"suspicious_text":next((i["evidence"] for i in indicators if i["target"]=="body"),""),
            "indicators":indicators,"why_risky":why,"learning_tip":tip,"is_phishing":True,
            "email_type":"Phishing","attack_type":{"easy":"Credential harvesting","medium":"Look-alike portal","hard":"Contextual spear phishing"}[diff],
            "risk_level":diff,"scenario_id":plan["fingerprint"],"scenario_meta":plan,"display_time":_V30_RNG.choice(["Today, 8:15 AM","Today, 10:42 AM","Monday, 2:31 PM","Yesterday, 4:05 PM","Thursday, 9:18 AM"])}











# =============================================================
# END RULE-GUIDED ENGINE v30
# =============================================================


# =============================================================
# RULE-GUIDED PHISHING SCENARIO ENGINE v31
# Diversity writer: varied rhetorical structures, natural legitimate
# messages, and strict evidence-linked tutor analysis.
# =============================================================

V31_PHISHING_STYLES = [
    "brief_alert", "case_followup", "manager_escalation", "service_notice",
    "deadline_reminder", "workflow_exception", "handover_request", "record_release"
]
V31_BANNED_PHRASES = [
    "has received a notification concerning",
    "prevent disruption to the related clinical workflow",
    "close the notification",
    "use the following link to complete the action",
    "this message does not request a password",
    "contains no external sign-in link",
]


def _v31_pick(items, salt=0):
    # SystemRandom keeps sessions fresh; salt avoids repeated adjacent choices.
    if not items:
        return None
    return items[(_V30_RNG.randrange(len(items)) + int(salt or 0)) % len(items)]


def _v31_subject(plan, urgency, style):
    obj = plan["object"].title()
    area = plan["area"]
    easy = {
        "brief_alert": [f"Action Required Today: {obj}", f"Urgent Review: {obj}"],
        "case_followup": [f"Final Reminder — {obj}", f"{obj}: Response Needed Today"],
        "manager_escalation": [f"Escalated: {obj} Awaiting Your Response", f"Immediate Attention Requested — {obj}"],
        "service_notice": [f"Access Notice for {obj}", f"Service Alert: {obj}"],
        "deadline_reminder": [f"Deadline Today: {obj}", f"{obj} Will Expire {urgency.title()}"],
        "workflow_exception": [f"Workflow Exception: {obj}", f"Unresolved {obj} Requires Action"],
        "handover_request": [f"Shift Handover Item: {obj}", f"Before Your Next Shift — {obj}"],
        "record_release": [f"Release Pending: {obj}", f"Confirm Receipt of {obj}"],
    }
    medium = [f"{obj} Review Window", f"Follow-up Required: {obj}", f"{area} Update — {obj}"]
    hard = [f"{obj} Follow-up", f"Re: {obj}", f"{area} — {obj}"]
    if plan["difficulty"] == "easy":
        return _v31_pick(easy.get(style, easy["brief_alert"]))
    return _v31_pick(medium if plan["difficulty"] == "medium" else hard)


def _v31_easy_phishing(plan, role, index, style):
    lang = plan["language"]; ar = lang == "Arabic"
    recipient = _v30_recipient(role, index, lang, plan["phase"])
    domain = _v30_domain("easy")
    link = _v30_link(plan, domain)
    mailbox = _v31_pick(["alerts", "workflow", "review", "updates", "coordination"], index)
    sender = f'{plan["sender"]} <{mailbox}@{domain}>'
    urgency = _v31_pick(["today", "immediately", "within the next hour", "before your next shift"], index)
    subject = _v31_subject(plan, urgency, style)
    greeting = _v31_pick(V30_GENERIC_GREETINGS[lang], index)
    credential = _v31_pick([
        "enter your password", "confirm your staff PIN", "submit your login credentials",
        "verify your account details", "provide your verification code"
    ], index)
    obj = plan["object"]; action = plan["action"]; area = plan["area"]

    if ar:
        templates = {
            "brief_alert": f"{greeting}،\n\nتعذر إكمال متابعة {obj} في {area}. يلزم {credential} {urgency} لفتح السجل ومراجعة التفاصيل.\n\nرابط المتابعة: {link}\n\nسيؤدي عدم الإجراء في الوقت المحدد إلى تعليق الوصول مؤقتًا.\n\nمع التحية،\n{plan['signature']}",
            "case_followup": f"{greeting}،\n\nهذه متابعة أخيرة بشأن {obj}. ما زالت الخطوة المطلوبة غير مكتملة، ويجب {credential} {urgency} حتى تتمكن من {action}.\n\nافتح الحالة من هنا: {link}\n\nإذا لم يتم الرد، ستُعاد الحالة تلقائيًا وقد يتوقف الوصول.\n\n{plan['signature']}",
            "manager_escalation": f"{greeting}،\n\nتم تصعيد {obj} إلى {plan['sender']} لعدم وجود رد. أكمل التحقق عبر {credential} {urgency}.\n\nمتابعة الطلب: {link}\n\nالتأخير قد يؤدي إلى تعليق حسابك عن هذا المسار.\n\nشكرًا،\n{plan['signature']}",
            "service_notice": f"{greeting}،\n\nتم إيقاف الوصول إلى {obj} مؤقتًا أثناء تحديث الخدمة. لاستعادة الوصول، يرجى {credential} {urgency}.\n\nاستعادة الخدمة: {link}\n\nلن تتمكن من فتح السجل حتى تكتمل الخطوة.\n\n{plan['signature']}",
            "deadline_reminder": f"{greeting}،\n\nتنتهي مهلة {obj} {urgency}. استخدم {credential} لإبقاء الطلب فعالًا ثم {action}.\n\nإكمال الطلب: {link}\n\nسيتم إلغاء الوصول بعد انتهاء المهلة.\n\nمع التقدير،\n{plan['signature']}",
            "workflow_exception": f"{greeting}،\n\nرصد النظام استثناءً في {obj}. لتصحيح الحالة، يجب {credential} {urgency} ثم إعادة فتح العنصر.\n\nبوابة التصحيح: {link}\n\nقد يؤدي تجاهل الاستثناء إلى تعليق صلاحياتك.\n\n{plan['signature']}",
            "handover_request": f"{greeting}،\n\nيوجد عنصر تسليم متعلق بـ {obj} قبل مناوبتك القادمة. يرجى {credential} لعرضه وتأكيد الاستلام.\n\nعرض عنصر التسليم: {link}\n\nعدم التأكيد قد يمنع الوصول أثناء المناوبة.\n\n{plan['signature']}",
            "record_release": f"{greeting}،\n\nأصبح {obj} جاهزًا للإصدار، لكن يلزم تأكيد هويتك أولًا. يرجى {credential} {urgency}.\n\nفتح السجل: {link}\n\nسيتم سحب السجل إذا لم يُستكمل الإجراء.\n\n{plan['signature']}",
        }
        body = templates[style]
        urgency_ev = next((x for x in ["فورًا", "اليوم", "خلال الساعة القادمة", "قبل مناوبتك القادمة"] if x in body or x in subject), urgency)
        indicators = [
            _v30_indicator(1,"domain","نطاق مرسل غير رسمي",f"عنوان المرسل يستخدم النطاق {domain} وليس نطاق المستشفى الرسمي.",domain,"from"),
            _v30_indicator(2,"urgency","استعجال أو تهديد واضح","تستخدم الرسالة مهلة قصيرة أو عاقبة على الوصول لدفع المستلم إلى التصرف بسرعة.",urgency_ev,"subject" if urgency_ev in subject else "body"),
            _v30_indicator(3,"credential","طلب مباشر لبيانات الدخول","تطلب الرسالة كلمة مرور أو رقمًا سريًا أو رمز تحقق مباشرة.",credential,"body"),
            _v30_indicator(4,"link","رابط خارجي ظاهر",f"الرابط يقود إلى النطاق غير الرسمي {domain}.",link,"link"),
            _v30_indicator(5,"greeting","تحية عامة","لا تستخدم الرسالة اسم المستلم، ما قد يشير إلى استهداف جماعي.",greeting,"greeting"),
        ]
        why = "تجمع الرسالة بين نطاق غير رسمي وطلب مباشر لبيانات الدخول ورابط خارجي وضغط زمني؛ وهي مؤشرات واضحة على التصيد للمستوى المبتدئ."
        tip = "لا تدخل كلمة مرور أو رمز موظف من رابط وصل عبر البريد. افتح النظام الرسمي مباشرة وتحقق من الطلب عبر قناة موثوقة."
    else:
        templates = {
            "brief_alert": f"{greeting},\n\nA review item for {obj} could not be completed in {area}. Please {credential} {urgency} to open the record and review the details.\n\nOpen the review item: {link}\n\nIf the step is not completed, access may be suspended temporarily.\n\nRegards,\n{plan['signature']}",
            "case_followup": f"{greeting},\n\nThis is a final follow-up regarding {obj}. The requested step is still outstanding. Please {credential} {urgency} so you can {action}.\n\nView the case: {link}\n\nAn unanswered request will be returned to the queue and may restrict access.\n\n{plan['signature']}",
            "manager_escalation": f"{greeting},\n\n{obj.title()} has been escalated to {plan['sender']} because no response was recorded. Complete verification by {credential} {urgency}.\n\nReview the escalation: {link}\n\nDelays may suspend your access to this workflow.\n\nThank you,\n{plan['signature']}",
            "service_notice": f"{greeting},\n\nAccess to {obj} was paused during a service update. To restore it, please {credential} {urgency}.\n\nRestore access: {link}\n\nThe record will remain unavailable until the step is completed.\n\n{plan['signature']}",
            "deadline_reminder": f"{greeting},\n\nThe response window for {obj} ends {urgency}. Use the portal to {credential}, then {action}.\n\nComplete the request: {link}\n\nAccess will be withdrawn after the deadline.\n\nRegards,\n{plan['signature']}",
            "workflow_exception": f"{greeting},\n\nThe system flagged an exception against {obj}. To clear it, you must {credential} {urgency} and reopen the item.\n\nResolve the exception: {link}\n\nIgnoring the exception may suspend your permissions.\n\n{plan['signature']}",
            "handover_request": f"{greeting},\n\nA handover item linked to {obj} is waiting before your next shift. Please {credential} to view it and acknowledge receipt.\n\nOpen the handover item: {link}\n\nFailure to acknowledge it may block access during the shift.\n\n{plan['signature']}",
            "record_release": f"{greeting},\n\n{obj.title()} is ready for release, but identity confirmation is required first. Please {credential} {urgency}.\n\nOpen the record: {link}\n\nThe item will be withdrawn if the action is not completed.\n\n{plan['signature']}",
        }
        body = templates[style]
        urgency_ev = next((x for x in ["today", "immediately", "within the next hour", "before your next shift"] if x in body.lower() or x in subject.lower()), urgency)
        indicators = [
            _v30_indicator(1,"domain","Non-official sender domain",f"The sender uses {domain}, not the hospital's official domain.",domain,"from"),
            _v30_indicator(2,"urgency","Strong urgency or threat","The message uses a short deadline or an access consequence to pressure the recipient.",urgency_ev,"subject" if urgency_ev.lower() in subject.lower() else "body"),
            _v30_indicator(3,"credential","Direct credential request","The message directly asks for a password, staff PIN, login credentials, or verification code.",credential,"body"),
            _v30_indicator(4,"link","Visible external link",f"The link points to the non-official domain {domain}.",link,"link"),
            _v30_indicator(5,"greeting","Generic greeting","The message does not address the recipient by name, which may indicate bulk targeting.",greeting,"greeting"),
        ]
        why = "The message combines a non-official domain, a direct credential request, an external link, and strong time pressure—clear beginner-level phishing indicators."
        tip = "Never enter a password, staff PIN, or verification code through an unexpected email link. Open the official system directly or verify the request through a trusted channel."

    return {"from":sender,"to":recipient,"subject":subject,"body":body,"attachment":"",
            "suspicious_link":link,"suspicious_text":credential,"indicators":indicators,
            "why_risky":why,"learning_tip":tip,"is_phishing":True,"email_type":"Phishing",
            "attack_type":"Credential harvesting","risk_level":"easy","scenario_id":plan["fingerprint"],
            "scenario_meta":dict(plan, writer_style=style),"display_time":_v31_pick(["Today, 8:15 AM","Today, 10:42 AM","Monday, 2:31 PM","Yesterday, 4:05 PM","Thursday, 9:18 AM"], index)}


def _v31_compose_phishing(plan, role, index):
    # Easy is fully rewritten with eight distinct rhetorical structures.
    if plan["difficulty"] == "easy":
        used_key = f"v31_{plan['phase']}_phish_styles"
        used = st.session_state.setdefault(used_key, [])
        options = [s for s in V31_PHISHING_STYLES if s not in used] or list(V31_PHISHING_STYLES)
        style = _v31_pick(options, index)
        used.append(style)
        return _v31_easy_phishing(plan, role, index, style)
    # Medium/Hard keep the v30 difficulty mechanics, but their plan structure
    # already rotates and they are not the source of the current easy-template issue.
    result = _v30_compose_phishing(plan, role, index)
    result.setdefault("scenario_meta", {})["writer_style"] = plan.get("structure")
    return result






def _v31_validate(result, plan):
    if not isinstance(result, dict): return False
    body=str(result.get("body", "")); subject=str(result.get("subject", "")); sender=str(result.get("from", ""))
    if not all([body.strip(), subject.strip(), sender.strip()]): return False
    if subject.lower() in body.lower(): return False
    if any(p in body.lower() for p in V31_BANNED_PHRASES): return False
    if plan["role_type"] not in ("clinical","admin","it","other"): return False
    if plan["family_id"] not in {f["id"] for f in V30_KNOWLEDGE[plan["role_type"]]}: return False
    if result.get("is_phishing"):
        count=len(result.get("indicators",[]))
        ranges={"easy":(4,5),"medium":(3,4),"hard":(1,2)}
        lo,hi=ranges[plan["difficulty"]]
        if not (lo <= count <= hi): return False
        fields={"from":sender,"subject":subject,"body":body,"greeting":body,"link":body,"attachment":str(result.get("attachment", ""))}
        for ind in result.get("indicators",[]):
            ev=str(ind.get("evidence", "")); target=ind.get("target")
            if not ev or ev.lower() not in fields.get(target, "").lower(): return False
    else:
        if not sender.lower().endswith("@hospital.org>"): return False
        if re.search(r"password|staff pin|login credentials|verification code",body,re.I): return False
        if result.get("suspicious_link") or result.get("indicators"): return False
    if plan["signature"].lower() not in body.lower(): return False
    return True







# =============================================================
# END RULE-GUIDED ENGINE v31
# =============================================================


# =============================================================
# RULE-GUIDED SCENARIO PLANNER v32
# -------------------------------------------------------------
# Safe upgrade over v31:
#   * changes ONLY the pre-writing scenario plan and the email writer
#   * keeps the stable indicator schema, renderer fields, scoring, UI,
#     provider selection, and results pipeline unchanged
#   * plans coherent role-locked combinations before composing text
#   * stores semantic history so later sessions avoid recent combinations
# =============================================================

_V32_RNG = random.SystemRandom()

V32_PHISH_STRUCTURES = [
    "pending_release", "failed_review", "workflow_handover", "service_interruption",
    "escalation_notice", "case_return", "deadline_reminder", "record_exception",
    "coordination_request", "unresolved_item", "status_change", "final_followup",
]












def _v32_choice(items, used=None):
    items = list(items)
    if used:
        fresh = [x for x in items if x not in used]
        if fresh:
            items = fresh
    return _V32_RNG.choice(items)


def _v32_session_bucket(role_type, language, difficulty):
    cycle = st.session_state.get("v32_cycle_id")
    if cycle is None:
        cycle = _V32_RNG.randrange(10_000_000, 99_999_999)
        st.session_state["v32_cycle_id"] = cycle
    key = f"v32_plan_memory_{cycle}_{role_type}_{language}_{difficulty}"
    return st.session_state.setdefault(key, {
        "semantic": [], "objects": [], "styles": [], "senders": [],
        "openings": [], "deadlines": [], "credentials": [], "subjects": [],
    })








def _v32_legitimate(plan, role, index):
    ar = plan["language"] == "Arabic"
    recipient = _v30_recipient(role, index, plan["language"], plan["phase"])
    person = _v30_display_name(recipient)
    sender_name = plan.get("sender_disp", plan["sender"]) if ar else plan["sender"]
    sender = f'{sender_name} <{_v32_choice(["notifications", "coordination", "quality", "clinical.ops", "records"])}@hospital.org>'
    disp_signature = plan.get("signature_disp", plan["signature"]) if ar else plan["signature"]
    obj, area, action, style = (plan.get("object_disp", plan["object"]), plan.get("area_disp", plan["area"]), plan.get("action_disp", plan["action"]), plan["structure"]) if ar else (plan["object"], plan["area"], plan["action"], plan["structure"])
    if ar:
        subjects = {
            "brief_update": f"تحديث موجز: {obj}", "colleague_note": f"ملاحظة بخصوص {obj}",
            "meeting_preparation": f"{obj} للاجتماع القادم", "record_correction": f"تصحيح مطلوب: {obj}",
            "schedule_change": f"تغيير في الموعد: {obj}", "quality_followup": f"متابعة جودة: {obj}",
            "policy_notice": f"إشعار إرشادي: {obj}", "handover_summary": f"ملخص تسليم: {obj}",
            "request_for_comment": f"طلب ملاحظة: {obj}", "completion_notice": f"اكتمل تحديث {obj}",
            "clarification_request": f"استفسار بسيط: {obj}", "team_briefing": f"إحاطة الفريق: {obj}",
        }
        bodies = {
            "brief_update": f"عزيزي {person}،\n\nتم تحديث {obj} في {area}. يمكنك {action} من خلال مساحة العمل الداخلية عند توفر الوقت.\n\nمع التحية،\n{disp_signature}",
            "colleague_note": f"مرحبًا {person}،\n\nأضفت ملاحظة قصيرة في سجل {obj}. هل يمكنك مراجعتها خلال المناوبة وإخباري إذا كانت هناك نقطة تحتاج تعديلًا؟\n\nشكرًا،\n{disp_signature}",
            "meeting_preparation": f"عزيزي {person}،\n\nنجهز بند {obj} لاجتماع {area} القادم. يرجى إضافة تعليقك في النظام الداخلي قبل ظهر الأربعاء.\n\nشكرًا،\n{disp_signature}",
            "record_correction": f"عزيزي {person}،\n\nلاحظنا اختلافًا بسيطًا في سجل {obj}. هل يمكنك مراجعة الحقل الأخير وتصحيحه في النظام المعتاد؟\n\nمع التقدير،\n{disp_signature}",
            "schedule_change": f"عزيزي {person}،\n\nتم تعديل توقيت {obj}. سيظهر الموعد الجديد في الجدول الداخلي اليوم؛ يرجى الرد فقط إذا تعارض مع مناوبتك.\n\n{disp_signature}",
            "quality_followup": f"عزيزي {person}،\n\nأصبحت خلاصة مراجعة الجودة الخاصة بـ {obj} متاحة في مساحة {area}. توجد ملاحظتان تحتاجان تعليق القسم قبل الاجتماع القادم.\n\nمع التحية،\n{disp_signature}",
            "policy_notice": f"عزيزي {person}،\n\nنُشرت النسخة المحدثة من إرشادات {obj} في مكتبة السياسات الداخلية. سيبدأ تطبيقها الأسبوع القادم وسنغطي التغيير في إحاطة الفريق.\n\n{disp_signature}",
            "handover_summary": f"عزيزي {person}،\n\nأضفت ملخص تسليم يتعلق بـ {obj} للفريق القادم. يرجى مراجعته في السجل وإضافة أي نقطة ناقصة قبل نهاية المناوبة.\n\nمع التحية،\n{disp_signature}",
            "request_for_comment": f"مرحبًا {person}،\n\nهل يمكنك إضافة ملاحظتك على {obj} في مساحة {area} الداخلية قبل نهاية اليوم؟ لا يلزم الرد بالبريد إذا لم توجد تغييرات.\n\nشكرًا،\n{disp_signature}",
            "completion_notice": f"عزيزي {person}،\n\nاكتمل تحديث {obj} وأصبح السجل متاحًا في النظام الداخلي. لا يلزم اتخاذ إجراء ما لم تلاحظ اختلافًا في التفاصيل.\n\nمع التقدير،\n{disp_signature}",
            "clarification_request": f"مرحبًا {person}،\n\nلدينا استفسار بسيط بخصوص {obj}. يرجى مراجعة الملاحظة في السجل الداخلي وإضافة التصحيح عند توفر الوقت.\n\nشكرًا،\n{disp_signature}",
            "team_briefing": f"عزيزي {person}،\n\nسيُناقش {obj} في إحاطة فريق {area} القادمة. أضفنا الملخص إلى مساحة العمل الداخلية للقراءة المسبقة.\n\nمع التحية،\n{disp_signature}",
        }
        subject, body = subjects[style], bodies[style]
        tip = "الرسالة تستخدم نطاق المستشفى الرسمي وسياقًا مهنيًا طبيعيًا وقناة داخلية معروفة دون طلب بيانات حساسة."
    else:
        subjects = {
            "brief_update": f"Update: {obj.title()}", "colleague_note": f"A Note About {obj.title()}",
            "meeting_preparation": f"{obj.title()} for the Next Meeting", "record_correction": f"Correction Needed: {obj.title()}",
            "schedule_change": f"Schedule Change — {obj.title()}", "quality_followup": f"Quality Follow-up: {obj.title()}",
            "policy_notice": f"Guidance Notice: {obj.title()}", "handover_summary": f"Handover Summary: {obj.title()}",
            "request_for_comment": f"Comment Requested: {obj.title()}", "completion_notice": f"{obj.title()} Update Completed",
            "clarification_request": f"Quick Clarification: {obj.title()}", "team_briefing": f"Team Briefing — {obj.title()}",
        }
        bodies = {
            "brief_update": f"Dear {person},\n\nThe {obj} record has been updated in {area}. You can {action} it through the usual internal workspace when convenient.\n\nKind regards,\n{disp_signature}",
            "colleague_note": f"Hello {person},\n\nI added a short note to the {obj} record. Could you review it during your shift and let me know if anything needs changing?\n\nThanks,\n{disp_signature}",
            "meeting_preparation": f"Dear {person},\n\nWe are preparing the {obj} item for the next {area} meeting. Please add your comment in the internal workspace by Wednesday noon.\n\nThank you,\n{disp_signature}",
            "record_correction": f"Dear {person},\n\nWe noticed a minor discrepancy in the {obj} record. Could you check the final field and correct it in the usual system?\n\nBest regards,\n{disp_signature}",
            "schedule_change": f"Dear {person},\n\nThe timing for {obj} has changed. The revised slot will appear in the internal schedule today; please reply only if it conflicts with your shift.\n\nRegards,\n{disp_signature}",
            "quality_followup": f"Dear {person},\n\nThe quality-review summary for {obj} is now available in the {area} workspace. Two observations need the department's comments before the next meeting.\n\nKind regards,\n{disp_signature}",
            "policy_notice": f"Dear {person},\n\nThe revised guidance for {obj} has been published in the internal policy library. It takes effect next week and will be covered at the team briefing.\n\nRegards,\n{disp_signature}",
            "handover_summary": f"Dear {person},\n\nI added a handover summary for {obj} for the incoming team. Please review it in the record and add anything missing before the end of the shift.\n\nKind regards,\n{disp_signature}",
            "request_for_comment": f"Hello {person},\n\nCould you add your comment to {obj} in the internal {area} workspace before the end of the day? There is no need to reply by email if nothing has changed.\n\nThanks,\n{disp_signature}",
            "completion_notice": f"Dear {person},\n\nThe update to {obj} has been completed and the record is available in the internal system. No action is needed unless you notice a discrepancy.\n\nBest regards,\n{disp_signature}",
            "clarification_request": f"Hello {person},\n\nWe have a quick clarification regarding {obj}. Please check the note in the internal record and add the correction when convenient.\n\nThanks,\n{disp_signature}",
            "team_briefing": f"Dear {person},\n\n{obj.title()} will be discussed at the next {area} team briefing. The summary has been added to the internal workspace for advance reading.\n\nKind regards,\n{disp_signature}",
        }
        subject, body = subjects[style], bodies[style]
        tip = "The message uses the official hospital domain, a natural workplace context, and a known internal workflow without asking for sensitive information."
    return {
        "from": sender, "to": recipient, "subject": subject, "body": body,
        "attachment": "", "suspicious_link": "", "suspicious_text": "", "indicators": [],
        "why_risky": "", "learning_tip": tip, "is_phishing": False,
        "email_type": "Legitimate", "attack_type": "None", "risk_level": plan["difficulty"],
        "scenario_id": plan["fingerprint"], "scenario_meta": plan,
        "display_time": _V32_RNG.choice(["Today, 9:05 AM", "Tuesday, 11:20 AM", "Yesterday, 3:40 PM", "Thursday, 1:15 PM"]),
    }


def _v32_validate(result, plan):
    # Preserve all v31 validation and add semantic/session safeguards.
    if not _v31_validate(result, plan):
        return False
    body = str(result.get("body", "")); subject = str(result.get("subject", ""))
    if result.get("is_phishing") and plan["difficulty"] == "easy":
        if len(result.get("indicators", [])) != 5:
            return False
        link = str(result.get("suspicious_link", ""))
        if not link or body.count(link) != 1:
            return False
    # Reject repeated opening/subject inside the active cycle.
    mem = _v32_session_bucket(plan["role_type"], plan["language"], plan["difficulty"])
    opening = next((ln.strip().lower() for ln in body.splitlines() if ln.strip()), "")
    subj_key = re.sub(r"\W+", " ", subject.lower()).strip()
    if opening and opening in mem["openings"][:-1]:
        return False
    if subj_key and subj_key in mem["subjects"]:
        return False
    mem["openings"].append(opening); mem["subjects"].append(subj_key)
    mem["openings"] = mem["openings"][-40:]; mem["subjects"] = mem["subjects"][-40:]
    return True







# =============================================================
# END RULE-GUIDED SCENARIO PLANNER v32
# =============================================================


# =============================================================
# SCENARIO ENGINE v33 — COMBINATORIAL PLANNER (SAFE UPGRADE)
# -------------------------------------------------------------
# This layer changes only scenario planning and message composition.
# It deliberately reuses the stable indicator objects, evidence mapping,
# link field, validator, renderer, scoring, UI, languages, and provider code.
# =============================================================

_V33_HISTORY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scenario_history_v33.json")
_V33_RNG = random.SystemRandom()

V33_PHISH_ARCHETYPES = [
    "release_gate", "exception_resolution", "handover_action", "queue_escalation",
    "deadline_followup", "service_restore", "record_reconciliation", "approval_hold",
    "case_assignment", "status_confirmation", "missing_acknowledgement", "workflow_lock",
    "review_returned", "coverage_change", "safety_followup", "document_release",
]

V33_LEGIT_ARCHETYPES = [
    "brief_update", "colleague_note", "meeting_preparation", "record_correction",
    "schedule_change", "quality_followup", "policy_notice", "handover_summary",
    "request_for_comment", "completion_notice", "clarification_request", "team_briefing",
    "shift_note", "meeting_minutes", "service_update", "reference_notice",
]

V33_TONES = [
    "formal", "concise", "operational", "collegial", "system_notice",
    "departmental", "case_focused", "follow_up",
]

V33_OPENING_MODES = [
    "state_change", "returned_item", "missed_response", "new_exception",
    "release_ready", "queue_notice", "follow_up", "service_event",
]

V33_PRESSURE_STYLES_EN = [
    "today", "within the next hour", "before your next shift", "before 3:00 PM",
    "before the case is returned", "before the queue closes", "before the handover window ends",
    "immediately",
]
V33_PRESSURE_STYLES_AR = [
    "اليوم", "خلال الساعة القادمة", "قبل مناوبتك القادمة", "قبل الساعة الثالثة مساءً",
    "قبل إعادة الحالة", "قبل إغلاق قائمة الانتظار", "قبل انتهاء فترة التسليم", "فورًا",
]

V33_CREDENTIAL_ACTIONS_EN = [
    "confirm your staff PIN", "provide the verification code", "verify your account details",
    "submit your login credentials", "enter your password", "approve the sign-in verification",
]
V33_CREDENTIAL_ACTIONS_AR = [
    "تأكيد الرقم السري للموظف", "إدخال رمز التحقق", "التحقق من بيانات الحساب",
    "إرسال بيانات الدخول", "إدخال كلمة المرور", "اعتماد طلب تسجيل الدخول",
]

V33_CTA_LABELS_EN = [
    "Open the record", "Review the item", "View the case", "Continue to the request",
    "Open the exception", "Review the handover", "Open the review", "View the update",
]
V33_CTA_LABELS_AR = [
    "فتح السجل", "مراجعة العنصر", "عرض الحالة", "متابعة الطلب",
    "فتح الاستثناء", "مراجعة التسليم", "فتح المراجعة", "عرض التحديث",
]

V33_CONSEQUENCES_EN = [
    "The item may be returned to the queue if the step is not completed.",
    "Access to this workflow may be restricted until the request is completed.",
    "The record will remain unavailable until verification is completed.",
    "The request may close automatically if no response is recorded.",
    "The case may be reassigned if the action is not completed.",
    "The workflow may remain locked until the step is finished.",
]
V33_CONSEQUENCES_AR = [
    "قد يُعاد العنصر إلى قائمة الانتظار إذا لم تكتمل الخطوة.",
    "قد يُقيّد الوصول إلى هذا المسار حتى اكتمال الطلب.",
    "سيظل السجل غير متاح حتى اكتمال التحقق.",
    "قد يُغلق الطلب تلقائيًا إذا لم تُسجل استجابة.",
    "قد تُعاد إسناد الحالة إذا لم يكتمل الإجراء.",
    "قد يظل المسار مقفلاً حتى إنهاء الخطوة.",
]

V33_SUBJECT_PATTERNS_EN = {
    "release_gate": ["Release Pending: {obj}", "{obj} Awaiting Release", "Release Hold — {obj}"],
    "exception_resolution": ["Correction Required — {obj}", "Record Exception: {obj}", "Review Exception — {obj}"],
    "handover_action": ["Handover Action Needed: {obj}", "Before the Next Shift — {obj}", "Handover Item — {obj}"],
    "queue_escalation": ["Escalation Notice — {obj}", "{obj} Returned to the Queue", "Escalated Response Needed: {obj}"],
    "deadline_followup": ["Final Follow-up: {obj}", "Outstanding Response — {obj}", "Action Needed Today: {obj}"],
    "service_restore": ["Service Access Notice: {obj}", "Restore Access — {obj}", "Temporary Access Hold: {obj}"],
    "record_reconciliation": ["Reconciliation Needed: {obj}", "Record Review Required — {obj}", "Unresolved Record: {obj}"],
    "approval_hold": ["Approval Hold: {obj}", "Approval Pending — {obj}", "Action Needed Before Release: {obj}"],
    "case_assignment": ["New Case Assignment: {obj}", "Case Response Needed — {obj}", "Assigned Item: {obj}"],
    "status_confirmation": ["Status Confirmation Required: {obj}", "Status Change — {obj}", "Confirm Updated Status: {obj}"],
    "missing_acknowledgement": ["Acknowledgement Missing: {obj}", "Response Not Recorded — {obj}", "Outstanding Acknowledgement: {obj}"],
    "workflow_lock": ["Workflow Locked: {obj}", "Access Blocked — {obj}", "Workflow Access Pending: {obj}"],
    "review_returned": ["Review Returned: {obj}", "{obj} Sent Back for Review", "Returned Review Item — {obj}"],
    "coverage_change": ["Coverage Change Awaiting Confirmation", "Schedule Update — {obj}", "Coverage Review Needed: {obj}"],
    "safety_followup": ["Safety Follow-up: {obj}", "Clinical Safety Review — {obj}", "Safety Response Required: {obj}"],
    "document_release": ["Document Release Pending: {obj}", "Release Review — {obj}", "Document Access Hold: {obj}"],
}

V33_SUBJECT_PATTERNS_AR = {
    k: [
        "إجراء مطلوب: {obj}", "متابعة مطلوبة: {obj}", "تحديث بانتظار الرد: {obj}"
    ] for k in V33_PHISH_ARCHETYPES
}

V33_GENERIC_GREETINGS_EN = [
    "Dear Staff Member", "Dear Healthcare Employee", "Dear Colleague",
    "Hello Team", "Attention Staff", "Dear Team Member",
]
V33_GENERIC_GREETINGS_AR = [
    "عزيزي الموظف", "عزيزي الزميل", "فريق العمل العزيز",
    "الزملاء الأعزاء", "إلى أعضاء الفريق", "عزيزي عضو الفريق",
]

V33_FAKE_DOMAINS = [
    "portal-confirm-now.net", "staff-verify-center.com", "secure-staff-check.net",
    "hospital-access-alert.co", "clinical-update-login.org", "care-review-portal.net",
    "clinical-record-check.com", "staff-access-review.net", "hospital-case-verify.org",
    "care-workflow-alert.co",
]
V33_MAILBOXES = [
    "review", "workflow", "alerts", "coordination", "updates", "case-notice",
    "records", "followup", "service", "response",
]


def _v33_load_history():
    try:
        with open(_V33_HISTORY_PATH, "r", encoding="utf-8") as f:
            rows = json.load(f)
        return rows if isinstance(rows, list) else []
    except Exception:
        return []


def _v33_save_history(rows):
    try:
        with open(_V33_HISTORY_PATH, "w", encoding="utf-8") as f:
            json.dump(rows[-6000:], f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _v33_bucket(role_type, language, difficulty):
    cycle = st.session_state.get("v33_cycle_id")
    if cycle is None:
        cycle = _V33_RNG.randrange(10_000_000, 99_999_999)
        st.session_state["v33_cycle_id"] = cycle
    key = f"v33_memory_{cycle}_{role_type}_{language}_{difficulty}"
    return st.session_state.setdefault(key, {
        "semantic": [], "families": [], "objects": [], "actions": [], "senders": [],
        "archetypes": [], "tones": [], "openings": [], "subjects": [], "credentials": [],
        "deadlines": [], "domains": [], "legit_archetypes": [],
    })


def _v33_pick(items, used=None):
    values = list(items)
    if used:
        fresh = [x for x in values if x not in set(used)]
        if fresh:
            values = fresh
    return _V33_RNG.choice(values)


def _v33_plan(role, index, language, difficulty, phase, is_phishing):
    """Combinatorial scenario plan. It never writes HTML, links, indicators, or tutor text."""
    diff = str(difficulty or "medium").lower()
    if diff not in V30_DIFFICULTY:
        diff = "medium"
    role_type = _v30_role_type(role)
    mem = _v33_bucket(role_type, language, diff)
    history = _v33_load_history()
    recent = {
        row.get("semantic") for row in history[-2500:]
        if row.get("role_type") == role_type and row.get("difficulty") == diff
    }

    families = list(V30_KNOWLEDGE[role_type])
    ar = language == "Arabic"
    for _ in range(240):
        family = _v33_pick(families, [])
        obj = _v33_pick(family["objects"], mem["objects"][-24:])
        action = _v33_pick(family["actions"], mem["actions"][-20:])
        sender = _v33_pick(family["senders"], mem["senders"][-16:])
        sender_idx = family["senders"].index(sender)
        signature = family["signatures"][sender_idx % len(family["signatures"])]
        archetypes = V33_PHISH_ARCHETYPES if is_phishing else V33_LEGIT_ARCHETYPES
        archetype = _v33_pick(archetypes, mem["archetypes"][-12:] if is_phishing else mem["legit_archetypes"][-12:])
        tone = _v33_pick(V33_TONES, mem["tones"][-8:])
        opening_mode = _v33_pick(V33_OPENING_MODES, mem["openings"][-8:])
        semantic = "|".join([
            role_type, family["id"], obj, action, sender, archetype, tone,
            opening_mode, "P" if is_phishing else "L",
        ])
        # Strong session diversity: avoid same family three times in a short window.
        family_recent_count = mem["families"][-8:].count(family["id"])
        if semantic not in recent and semantic not in mem["semantic"] and family_recent_count < 2:
            break

    obj_idx = family["objects"].index(obj)
    action_idx = family["actions"].index(action)
    area_disp = family.get("area_ar", family["area"]) if ar else family["area"]
    event_disp = family.get("event_ar", family["event"]) if ar else family["event"]
    obj_disp = family.get("objects_ar", family["objects"])[obj_idx] if ar else obj
    action_disp = family.get("actions_ar", family["actions"])[action_idx] if ar else action
    sender_disp = family.get("senders_ar", family["senders"])[sender_idx] if ar else sender
    signature_disp = family.get("signatures_ar", family["signatures"])[sender_idx % len(family["signatures"])] if ar else signature

    plan = {
        "role_type": role_type, "language": language, "difficulty": diff,
        "phase": phase, "is_phishing": bool(is_phishing),
        "family_id": family["id"], "area": family["area"], "event": family["event"],
        "object": obj, "action": action, "sender": sender, "signature": signature,
        "area_disp": area_disp, "event_disp": event_disp, "object_disp": obj_disp,
        "action_disp": action_disp, "sender_disp": sender_disp, "signature_disp": signature_disp,
        "structure": archetype, "tone": tone, "opening_mode": opening_mode,
        "channel": _V33_RNG.choice(V30_DIFFICULTY[diff]["allowed_channels"]),
        "semantic": semantic,
    }
    plan["fingerprint"] = "v33:" + _v30_hashlib.sha256(semantic.encode("utf-8")).hexdigest()[:20]

    mem["semantic"].append(semantic); mem["families"].append(family["id"])
    mem["objects"].append(obj); mem["actions"].append(action); mem["senders"].append(sender)
    mem["tones"].append(tone); mem["openings"].append(opening_mode)
    (mem["archetypes"] if is_phishing else mem["legit_archetypes"]).append(archetype)
    for key in mem:
        mem[key] = mem[key][-60:]

    history.append({
        "semantic": semantic, "role_type": role_type, "difficulty": diff,
        "phase": phase, "family_id": family["id"], "object": obj,
        "archetype": archetype, "tone": tone,
    })
    _v33_save_history(history)
    return plan


def _v33_subject(plan):
    ar = plan["language"] == "Arabic"
    bank = V33_SUBJECT_PATTERNS_AR if ar else V33_SUBJECT_PATTERNS_EN
    templates = bank[plan["structure"]]
    return _v33_pick(templates).format(obj=plan["object_disp"] if ar else plan["object_disp"].title())


def _v33_context_sentence(plan, ar=False):
    obj, area, mode = plan["object_disp"], plan["area_disp"], plan["opening_mode"]
    if ar:
        bank = {
            "state_change": f"تغيرت حالة {obj} في {area} بعد تحديث السجل.",
            "returned_item": f"أُعيد {obj} إلى قائمة {area} للمراجعة.",
            "missed_response": f"لم تُسجل استجابة على {obj} ضمن {area}.",
            "new_exception": f"تم اكتشاف استثناء جديد في سجل {obj}.",
            "release_ready": f"أصبح {obj} جاهزًا للإطلاق في {area}.",
            "queue_notice": f"يوجد عنصر متعلق بـ {obj} بانتظارك في قائمة {area}.",
            "follow_up": f"هذه متابعة بخصوص {obj} ضمن {area}.",
            "service_event": f"تأثر الوصول إلى {obj} أثناء تحديث الخدمة.",
        }
    else:
        bank = {
            "state_change": f"The status of {obj} changed after an update in {area}.",
            "returned_item": f"The {obj} item was returned to the {area} queue for review.",
            "missed_response": f"No response has been recorded for {obj} in {area}.",
            "new_exception": f"A new exception was detected in the {obj} record.",
            "release_ready": f"The {obj} item is ready for release in {area}.",
            "queue_notice": f"An item related to {obj} is waiting in the {area} queue.",
            "follow_up": f"This is a follow-up regarding {obj} in {area}.",
            "service_event": f"Access to {obj} was affected during a service update.",
        }
    return bank[mode]


def _v33_easy_phishing(plan, role, index):
    """Stable easy writer: one link field, evidence-grounded indicators, richer scenario wording."""
    ar = plan["language"] == "Arabic"
    recipient = _v30_recipient(role, index, plan["language"], plan["phase"])
    mem = _v33_bucket(plan["role_type"], plan["language"], plan["difficulty"])

    domain = _v33_pick(V33_FAKE_DOMAINS, mem["domains"][-8:])
    mailbox = _v33_pick(V33_MAILBOXES)
    sender_name = plan["sender_disp"] if ar else plan["sender"]
    sender = f'{sender_name} <{mailbox}@{domain}>'
    greeting = _v33_pick(V33_GENERIC_GREETINGS_AR if ar else V33_GENERIC_GREETINGS_EN)
    credential = _v33_pick(V33_CREDENTIAL_ACTIONS_AR if ar else V33_CREDENTIAL_ACTIONS_EN, mem["credentials"][-5:])
    deadline = _v33_pick(V33_PRESSURE_STYLES_AR if ar else V33_PRESSURE_STYLES_EN, mem["deadlines"][-5:])
    cta = _v33_pick(V33_CTA_LABELS_AR if ar else V33_CTA_LABELS_EN)
    consequence = _v33_pick(V33_CONSEQUENCES_AR if ar else V33_CONSEQUENCES_EN)
    subject = _v33_subject(plan)
    slug = re.sub(r"[^a-z0-9]+", "-", plan["family_id"].lower()).strip("-")
    link = f"http://{domain}/{slug}/{_V33_RNG.randrange(1000,9999)}"
    context = _v33_context_sentence(plan, ar)
    obj, action, archetype = plan["object_disp"], plan["action_disp"], plan["structure"]

    if ar:
        action_sentence = f"يرجى {credential} {deadline} حتى تتمكن من {action}."
        connectors = {
            "release_gate": f"{context} يلزم تأكيد الهوية قبل إكمال الإطلاق.",
            "exception_resolution": f"{context} يلزم إكمال خطوة التحقق لتصحيح الحالة.",
            "handover_action": f"{context} يتطلب عنصر التسليم تأكيدًا قبل المتابعة.",
            "queue_escalation": f"{context} تم تصعيد الطلب بعد عدم تسجيل استجابة.",
            "deadline_followup": f"{context} ما زالت الخطوة المطلوبة معلقة.",
            "service_restore": f"{context} يلزم التحقق لاستعادة الوصول.",
            "record_reconciliation": f"{context} يلزم توثيق المراجعة قبل إغلاق العنصر.",
            "approval_hold": f"{context} ما زال الاعتماد متوقفًا حتى اكتمال التحقق.",
            "case_assignment": f"{context} يجب تأكيد الوصول قبل فتح المهمة.",
            "status_confirmation": f"{context} يلزم تأكيد الحالة الجديدة.",
            "missing_acknowledgement": f"{context} لم يتم تسجيل الإقرار المطلوب.",
            "workflow_lock": f"{context} ما زال المسار مقفلاً.",
            "review_returned": f"{context} أُعيدت المراجعة بسبب خطوة غير مكتملة.",
            "coverage_change": f"{context} يحتاج تعديل التغطية إلى تأكيد.",
            "safety_followup": f"{context} يحتاج بند السلامة إلى استجابة موثقة.",
            "document_release": f"{context} يتوقف إصدار المستند على إكمال التحقق.",
        }
        body = f"{greeting}،\n\n{connectors[archetype]} {action_sentence}\n\n{cta}: {link}\n\n{consequence}\n\n{plan['signature_disp']}"
    else:
        action_sentence = f"Please {credential} {deadline} so you can {action}."
        connectors = {
            "release_gate": f"{context} Identity confirmation is required before release can continue.",
            "exception_resolution": f"{context} The verification step must be completed before the status can be corrected.",
            "handover_action": f"{context} The handover item requires confirmation before it can be acknowledged.",
            "queue_escalation": f"{context} The request has been escalated because no response was recorded.",
            "deadline_followup": f"{context} The requested step is still outstanding.",
            "service_restore": f"{context} Verification is required to restore access.",
            "record_reconciliation": f"{context} The review must be documented before the item can close.",
            "approval_hold": f"{context} Approval remains on hold until verification is completed.",
            "case_assignment": f"{context} Access confirmation is required before the task can be opened.",
            "status_confirmation": f"{context} The updated status requires confirmation.",
            "missing_acknowledgement": f"{context} The required acknowledgement has not been recorded.",
            "workflow_lock": f"{context} The workflow remains locked.",
            "review_returned": f"{context} The review was returned because a required step is incomplete.",
            "coverage_change": f"{context} The coverage change requires confirmation.",
            "safety_followup": f"{context} The safety item requires a documented response.",
            "document_release": f"{context} Document release is waiting for verification.",
        }
        body = f"{greeting},\n\n{connectors[archetype]} {action_sentence}\n\n{cta}: {link}\n\n{consequence}\n\n{plan['signature_disp']}"

    first_line = next(x.strip() for x in body.splitlines() if x.strip()).rstrip(",،")
    indicators = [
        _v30_indicator(1, "domain", "Non-official sender domain" if not ar else "نطاق مرسل غير رسمي",
                       (f"The sender uses {domain}, not the hospital's official domain." if not ar else f"يستخدم المرسل النطاق {domain} وليس نطاق المستشفى الرسمي."), domain, "from"),
        _v30_indicator(2, "urgency", "Strong urgency or threat" if not ar else "استعجال أو تهديد واضح",
                       ("The message uses a short deadline or access consequence to pressure the recipient." if not ar else "تستخدم الرسالة مهلة قصيرة أو نتيجة مرتبطة بالوصول للضغط على المستلم."), deadline, "body"),
        _v30_indicator(3, "credential", "Direct credential request" if not ar else "طلب مباشر لبيانات الدخول",
                       ("The message directly asks for a password, staff PIN, login credentials, or verification code." if not ar else "تطلب الرسالة مباشرة كلمة مرور أو رقمًا سريًا أو بيانات دخول أو رمز تحقق."), credential, "body"),
        _v30_indicator(4, "link", "Visible external link" if not ar else "رابط خارجي ظاهر",
                       (f"The link points to the non-official domain {domain}." if not ar else f"يقود الرابط إلى النطاق غير الرسمي {domain}."), link, "link"),
        _v30_indicator(5, "greeting", "Generic greeting" if not ar else "تحية عامة",
                       ("The message does not address the recipient by name, which may indicate bulk targeting." if not ar else "لا تخاطب الرسالة المستلم باسمه، وقد يدل ذلك على إرسال جماعي."), first_line, "greeting"),
    ]
    why = ("The message combines a non-official sender domain, a direct credential request, a visible external link, and strong pressure—clear beginner-level phishing indicators."
           if not ar else "تجمع الرسالة بين نطاق غير رسمي وطلب مباشر لبيانات الدخول ورابط خارجي وضغط واضح، وهي مؤشرات تصيد مناسبة للمستوى المبتدئ.")
    tip = ("Never enter a password, staff PIN, or verification code through an unexpected email link. Open the official system directly or verify the request through a trusted channel."
           if not ar else "لا تُدخل كلمة مرور أو رقمًا سريًا أو رمز تحقق عبر رابط بريد غير متوقع. افتح النظام الرسمي مباشرة أو تحقق من الطلب عبر قناة موثوقة.")

    mem["credentials"].append(credential); mem["deadlines"].append(deadline); mem["domains"].append(domain)
    mem["credentials"] = mem["credentials"][-20:]; mem["deadlines"] = mem["deadlines"][-20:]; mem["domains"] = mem["domains"][-20:]

    return {
        "from": sender, "to": recipient, "subject": subject, "body": body,
        "attachment": "", "suspicious_link": link, "suspicious_text": credential,
        "indicators": indicators, "why_risky": why, "learning_tip": tip,
        "is_phishing": True, "email_type": "Phishing", "attack_type": "Credential harvesting",
        "risk_level": "easy", "scenario_id": plan["fingerprint"], "scenario_meta": plan,
        "display_time": _V33_RNG.choice(["Today, 8:15 AM", "Today, 10:42 AM", "Monday, 2:31 PM", "Yesterday, 4:05 PM", "Thursday, 9:18 AM"]),
    }


def _v33_legitimate(plan, role, index):
    """Legitimate writer delegates to the proven v32 writer after mapping a diverse archetype."""
    # Map the four new legitimate archetypes to safe established structures.
    mapping = {
        "shift_note": "handover_summary",
        "meeting_minutes": "meeting_preparation",
        "service_update": "brief_update",
        "reference_notice": "policy_notice",
    }
    legacy_structure = mapping.get(plan["structure"], plan["structure"])
    legacy_plan = dict(plan, structure=legacy_structure)
    return _v32_legitimate(legacy_plan, role, index)


def _v33_validate(result, plan):
    if not _v32_validate(result, plan):
        return False
    body = str(result.get("body", ""))
    subject = str(result.get("subject", ""))
    if result.get("is_phishing") and plan["difficulty"] == "easy":
        link = str(result.get("suspicious_link", ""))
        if body.count(link) != 1:
            return False
        # Every tutor evidence item must be present in its actual source field.
        source_blob = {
            "from": str(result.get("from", "")), "body": body,
            "link": link, "greeting": body,
        }
        for item in result.get("indicators", []):
            evidence = str(item.get("evidence", "")).strip()
            location = str(item.get("location", "body"))
            if evidence and evidence.lower() not in source_blob.get(location, body).lower():
                return False
    # Avoid lexical near-duplicates inside one cycle.
    mem = _v33_bucket(plan["role_type"], plan["language"], plan["difficulty"])
    subj_key = re.sub(r"\W+", " ", subject.lower()).strip()
    if subj_key in mem["subjects"]:
        return False
    mem["subjects"].append(subj_key)
    mem["subjects"] = mem["subjects"][-60:]
    return True


def _v33_generate(role, index, language, difficulty="medium", is_phishing=True, assessment=False):
    phase = "assess" if assessment else "learn"
    result = None
    for attempt in range(16):
        plan = _v33_plan(role, index + attempt * 104729, language, difficulty, phase, is_phishing)
        if is_phishing and plan["difficulty"] == "easy":
            result = _v33_easy_phishing(plan, role, index)
        elif is_phishing:
            # Preserve the proven medium/hard writer and indicator mapping.
            legacy_plan = dict(plan, structure=_v33_pick(V32_PHISH_STRUCTURES))
            result = _v31_compose_phishing(legacy_plan, role, index)
        else:
            result = _v33_legitimate(plan, role, index)
        if _v33_validate(result, plan):
            try:
                evaluate_and_log_auto_scores(result, plan["difficulty"], language, is_phishing=bool(is_phishing))
            except Exception:
                pass
            return result
    return result





# =============================================================
# END SCENARIO ENGINE v33
# =============================================================


# =============================================================
# SAFE MICRO-IMPROVEMENT PATCH v34
# Scope is intentionally limited to content diversity, grounded
# analysis and Easy-level indicator count. Rendering, link
# placement, badge anchoring, parsing and provider calls are untouched.
# =============================================================
_V34_BUILD_PROMPT = build_prompt
_V34_BUILD_ASSESS_PROMPT = build_assess_prompt
_V34_NORMALIZE = normalize_learning_analysis

_V34_EASY_EN = """

SAFE EASY-DIVERSITY RULES (content-only; do not change the JSON schema):
- Use 3, 4, or 5 clear phishing indicators, selected naturally for this specific email. Do not force the same five indicators into every scenario.
- Across a 10-question cycle, vary the attack construction: credential request, fake document access, MFA/OTP abuse, authority impersonation, external-link workflow, or account/access pressure.
- Vary the event sequence and opening sentence. Do not reuse the same body structure or consequence sentence more than twice in one cycle.
- Every indicator title and description must quote or explicitly refer to evidence that actually appears in THIS email. Never mention an attachment, QR code, patient-data request, typo, or threat unless it is present.
- why_risky must summarize this email's actual combination of clues. learning_tip must address the strongest clue in this email, not use a generic repeated tip.
- Keep Easy obvious for beginners: at least three strong clues, a visible fake URL when a link is used, and no QR code or attachment.
"""

_V34_EASY_AR = """

قواعد تنويع آمنة للمستوى السهل (تعديل محتوى فقط دون تغيير بنية JSON):
- استخدم 3 أو 4 أو 5 مؤشرات تصيد واضحة بحسب محتوى الرسالة نفسها، ولا تفرض المؤشرات الخمسة ذاتها على كل سيناريو.
- نوّع بناء الهجوم داخل دورة الأسئلة: طلب بيانات دخول، وصول مزيف لمستند، إساءة استخدام MFA/OTP، انتحال جهة مسؤولة، سير عمل عبر رابط خارجي، أو ضغط متعلق بالحساب أو الوصول.
- نوّع بداية الرسالة وتسلسل الحدث. لا تكرر قالب الجسم نفسه أو جملة العاقبة نفسها أكثر من مرتين في الدورة.
- يجب أن يشير كل عنوان ووصف في indicators إلى دليل موجود فعلاً في هذه الرسالة. لا تذكر مرفقاً أو QR أو بيانات مرضى أو خطأ لغوياً أو تهديداً إلا إذا ظهر فعلاً.
- يجب أن يلخص why_risky مجموعة الأدلة الفعلية في الرسالة، وأن ترتبط learning_tip بأقوى مؤشر فيها بدلاً من نص عام متكرر.
- أبقِ المستوى السهل واضحاً للمبتدئ: ثلاثة أدلة قوية على الأقل، ورابط مزيف ظاهر عند استخدام الرابط، ومن دون QR أو مرفقات.
"""

def build_prompt(role, index, language):
    base = _V34_BUILD_PROMPT(role, index, language)
    difficulty = st.session_state.get("difficulty", "medium")
    if difficulty == "easy":
        base += _V34_EASY_AR if language == "Arabic" else _V34_EASY_EN
    elif difficulty == "medium":
        mode = get_medium_presentation_mode("learn", index)
        rule = get_medium_channel_instruction(mode, language == "Arabic")
        if language == "Arabic":
            base += f"""

تصحيح إلزامي للمستوى المتوسط:
- {rule}
- استخدم مؤشرين أو 3 مؤشرات فقط.
- أضف للحقل JSON: \"medium_channel\": \"{mode}\".
- لا تترك فراغات كبيرة بين فقرات الرسالة.
"""
        else:
            base += f"""

MANDATORY INTERMEDIATE OVERRIDE:
- {rule}
- Use only 2 or 3 grounded indicators.
- Add this JSON field: \"medium_channel\": \"{mode}\".
- Keep normal compact email paragraph spacing.
"""
    return base

def build_assess_prompt(role, index, is_phishing, language):
    base = _V34_BUILD_ASSESS_PROMPT(role, index, is_phishing, language)
    difficulty = st.session_state.get("difficulty", "medium")
    if is_phishing and difficulty == "easy":
        base += _V34_EASY_AR if language == "Arabic" else _V34_EASY_EN
    elif is_phishing and difficulty == "medium":
        mode = get_medium_presentation_mode("assess", index)
        rule = get_medium_channel_instruction(mode, language == "Arabic")
        if language == "Arabic":
            base += f"""

تصحيح إلزامي للمستوى المتوسط:
- {rule}
- أضف للحقل JSON: \"medium_channel\": \"{mode}\".
- لا تترك فراغات كبيرة بين فقرات الرسالة.
"""
        else:
            base += f"""

MANDATORY INTERMEDIATE OVERRIDE:
- {rule}
- Add this JSON field: \"medium_channel\": \"{mode}\".
- Keep normal compact email paragraph spacing.
"""
    return base

def normalize_learning_analysis(result, role_type, difficulty, is_ar=False):
    result = _V34_NORMALIZE(result, role_type, difficulty, is_ar)
    if not isinstance(result, dict) or "error" in result:
        return result
    # Previous normalizer guarantees three grounded indicators. For Easy,
    # preserve any additional model-generated grounded indicators up to 5,
    # while keeping numbering stable. Never synthesize extra clues.
    if difficulty == "easy":
        raw = result.get("indicators") if isinstance(result.get("indicators"), list) else []
        clean=[]
        seen=set()
        for item in raw[:5]:
            if not isinstance(item, dict):
                continue
            title=str(item.get("title") or "").strip()
            desc=str(item.get("description") or "").strip()
            key=(title.lower(), desc.lower())
            if title and desc and key not in seen:
                seen.add(key)
                clean.append({"number":len(clean)+1,"title":title,"description":desc})
        if len(clean) >= 3:
            result["indicators"] = clean
    return result

# =============================================================
# END SAFE MICRO-IMPROVEMENT PATCH v34
# =============================================================


# =============================================================
# MEDIUM GENERATION ENGINE v35 — SAFE DYNAMIC CHANNEL UPGRADE
# -------------------------------------------------------------
# Scope:
#   * Medium phishing only.
#   * Easy, Hard, legitimate messages, UI, renderer, scoring,
#     badges, results and provider integrations remain unchanged.
#   * Exactly one delivery channel per Medium phishing email:
#       button | visible_link | pdf_attachment | sharepoint_button |
#       internal_reply
#   * Indicators and tutor analysis are generated from the actual
#     evidence in the resulting email.
# =============================================================

_V35_RNG = random.SystemRandom()

V35_MEDIUM_CHANNEL_WEIGHTS = [
    ("button", 24),
    ("visible_link", 22),
    ("pdf_attachment", 18),
    ("sharepoint_button", 18),
    ("internal_reply", 18),
]

V35_MEDIUM_STRUCTURES = [
    "policy_acknowledgement",
    "incident_followup",
    "training_completion",
    "schedule_revision",
    "equipment_notice",
    "audit_request",
    "case_assignment",
    "service_change",
    "quality_review",
    "document_correction",
]

V35_MEDIUM_DEADLINES_EN = [
    "within 24 hours",
    "by tomorrow afternoon",
    "before Friday's departmental review",
    "before the current review window closes",
    "by the end of the next business day",
]
V35_MEDIUM_DEADLINES_AR = [
    "خلال 24 ساعة",
    "قبل ظهر الغد",
    "قبل مراجعة القسم يوم الجمعة",
    "قبل إغلاق فترة المراجعة الحالية",
    "قبل نهاية يوم العمل القادم",
]

V35_MEDIUM_GREETINGS_EN = [
    "Dear {area} Team",
    "Hello {area} Team",
    "Good morning {area} colleagues",
    "Dear Clinical Colleagues",
    "Hello Team",
]
V35_MEDIUM_GREETINGS_AR = [
    "فريق {area} المحترم",
    "مرحبًا فريق {area}",
    "صباح الخير زملاء {area}",
    "الزملاء السريريين الأعزاء",
    "مرحبًا بالفريق",
]

def _v35_weighted_choice(items):
    total = sum(weight for _, weight in items)
    pick = _V35_RNG.uniform(0, total)
    upto = 0
    for value, weight in items:
        upto += weight
        if pick <= upto:
            return value
    return items[-1][0]

def _v35_medium_channel(phase, index):
    """Session-stable weighted channel with anti-streak protection."""
    cycle = st.session_state.get("v18_cycle_id", "default")
    key = f"v35_medium_channels_{cycle}_{phase}"
    planned = st.session_state.setdefault(key, [])
    idx = int(index)
    while len(planned) <= idx:
        candidate = _v35_weighted_choice(V35_MEDIUM_CHANNEL_WEIGHTS)
        if len(planned) >= 2 and planned[-1] == planned[-2] == candidate:
            alternatives = [(v, w) for v, w in V35_MEDIUM_CHANNEL_WEIGHTS if v != candidate]
            candidate = _v35_weighted_choice(alternatives)
        planned.append(candidate)
    return planned[idx]

def _v35_medium_structure(phase, index):
    cycle = st.session_state.get("v18_cycle_id", "default")
    key = f"v35_medium_structures_{cycle}_{phase}"
    used = st.session_state.setdefault(key, [])
    available = [x for x in V35_MEDIUM_STRUCTURES if x not in used[-8:]]
    if not available:
        available = list(V35_MEDIUM_STRUCTURES)
    style = available[(int(index) + len(used)) % len(available)]
    used.append(style)
    return style

def _v35_compact(body):
    body = re.sub(r"[ \t]+\n", "\n", str(body or ""))
    body = re.sub(r"\n{3,}", "\n\n", body)
    return body.strip()

def _v35_medium_phishing(plan, role, index):
    lang = plan["language"]
    ar = lang == "Arabic"
    recipient = _v30_recipient(role, index, lang, plan["phase"])
    domain = _v30_domain("medium")
    link = _v30_link(plan, domain)
    mailbox = _V35_RNG.choice(["workflow", "review", "coordination", "updates", "records"])
    sender = f'{plan["sender"]} <{mailbox}@{domain}>'
    area = plan["area"]
    obj = plan["object"]
    action = plan["action"]
    signature = plan["signature"]
    channel = _v35_medium_channel(plan["phase"], index)
    structure = _v35_medium_structure(plan["phase"], index)
    deadline = _V35_RNG.choice(V35_MEDIUM_DEADLINES_AR if ar else V35_MEDIUM_DEADLINES_EN)
    greeting = _V35_RNG.choice(V35_MEDIUM_GREETINGS_AR if ar else V35_MEDIUM_GREETINGS_EN).format(area=area)
    attachment = ""
    suspicious_link = ""
    medium_channel = "none"

    if ar:
        subjects = {
            "policy_acknowledgement": f"تحديث سياسة {area}: {obj}",
            "incident_followup": f"متابعة حادثة — {obj}",
            "training_completion": f"استكمال تدريب {area}",
            "schedule_revision": f"تعديل جدول: {obj}",
            "equipment_notice": f"إشعار معدات — {obj}",
            "audit_request": f"طلب مراجعة تدقيق: {obj}",
            "case_assignment": f"إسناد حالة — {obj}",
            "service_change": f"تغيير خدمة {area}",
            "quality_review": f"مراجعة جودة: {obj}",
            "document_correction": f"تصحيح مستند: {obj}",
        }
        intros = {
            "policy_acknowledgement": f"نُشرت نسخة محدثة من الإجراء المرتبط بـ {obj} ضمن {area}.",
            "incident_followup": f"أعيد فتح بند متابعة يتعلق بـ {obj} بعد مراجعة أولية من {area}.",
            "training_completion": f"أظهرت قائمة التدريب أن متطلبًا مرتبطًا بـ {obj} ما زال غير مكتمل.",
            "schedule_revision": f"تم تعديل توقيت {obj} ويحتاج القسم إلى تأكيد الاستلام.",
            "equipment_notice": f"سُجل تحديث تشغيلي يتعلق بـ {obj} ضمن {area}.",
            "audit_request": f"اختارت مراجعة الجودة بند {obj} لعينة التدقيق الحالية.",
            "case_assignment": f"أُسند إليك عنصر متعلق بـ {obj} ضمن مسار {area}.",
            "service_change": f"سيُطبق تغيير في خدمة {area} يؤثر على {obj}.",
            "quality_review": f"أضيفت ملاحظة جديدة إلى مراجعة الجودة الخاصة بـ {obj}.",
            "document_correction": f"اكتُشف اختلاف بسيط في مستند {obj} ويحتاج إلى مراجعة.",
        }
        subject = subjects[structure]
        intro = intros[structure]
        closing = _V35_RNG.choice([
            "إذا لم يكن العنصر ضمن مسؤوليتك، أبلغ منسق القسم عبر القناة الرسمية.",
            "لا ترسل أي معلومات سريرية حساسة عبر البريد.",
            "سيبقى البند مفتوحًا حتى تتم مراجعته من القسم.",
        ])
    else:
        subjects = {
            "policy_acknowledgement": f"{area} Policy Update: {obj.title()}",
            "incident_followup": f"Incident Follow-up — {obj.title()}",
            "training_completion": f"{area} Training Completion Notice",
            "schedule_revision": f"Schedule Revision: {obj.title()}",
            "equipment_notice": f"Equipment Notice — {obj.title()}",
            "audit_request": f"Audit Review Request: {obj.title()}",
            "case_assignment": f"Case Assignment — {obj.title()}",
            "service_change": f"{area} Service Change",
            "quality_review": f"Quality Review: {obj.title()}",
            "document_correction": f"Document Correction: {obj.title()}",
        }
        intros = {
            "policy_acknowledgement": f"A revised procedure for {obj} has been published by {area}.",
            "incident_followup": f"A follow-up item for {obj} was reopened after an initial review by {area}.",
            "training_completion": f"The training register shows an outstanding requirement linked to {obj}.",
            "schedule_revision": f"The timing for {obj} has changed and the department needs acknowledgement.",
            "equipment_notice": f"An operational update has been logged for {obj} within {area}.",
            "audit_request": f"The quality review selected {obj} for the current audit sample.",
            "case_assignment": f"An item concerning {obj} has been assigned to you through {area}.",
            "service_change": f"A service change in {area} will affect the workflow for {obj}.",
            "quality_review": f"A new observation has been added to the quality review for {obj}.",
            "document_correction": f"A minor discrepancy was identified in the {obj} document and needs review.",
        }
        subject = subjects[structure]
        intro = intros[structure]
        closing = _V35_RNG.choice([
            "If the item is not assigned to you, notify the approved coordinator through the official directory.",
            "Do not send clinical or patient information by email.",
            "The item will remain open until the department records a response.",
        ])

    # Channel-specific message and grounded indicators.
    indicators = [
        _v30_indicator(
            1, "domain",
            "نطاق مرسل مشابه لكنه غير رسمي" if ar else "Look-alike sender domain",
            (f"النطاق {domain} يشبه نطاق خدمة داخلية لكنه ليس hospital.org."
             if ar else f"The domain {domain} resembles an internal service but is not hospital.org."),
            domain, "from"
        )
    ]

    if channel == "button":
        medium_channel = "button"
        suspicious_link = link
        label = "فتح المهمة" if ar else _V35_RNG.choice(["Open task", "Review item", "View update"])
        request = (f"يرجى {action} {deadline} من خلال صفحة المتابعة."
                   if ar else f"Please {action} {deadline} through the review workspace.")
        body = f"{greeting}،\n\n{intro} {request}\n\n[{label}]({link})\n\n{closing}\n\nمع التحية،\n{signature}" if ar else \
               f"{greeting},\n\n{intro} {request}\n\n[{label}]({link})\n\n{closing}\n\nKind regards,\n{signature}"
        indicators += [
            _v30_indicator(2, "workflow",
                "تسجيل دخول عبر مسار غير معتاد" if ar else "Unusual external workflow",
                "توجّه الرسالة المستخدم إلى مساحة متابعة واردة عبر البريد بدل النظام المحفوظ." if ar else
                "The email directs the recipient to an emailed workspace instead of the bookmarked internal system.",
                request, "body"),
            _v30_indicator(3, "link",
                "زر يقود إلى نطاق غير معتمد" if ar else "Button leads to an unapproved domain",
                f"وجهة الزر هي {domain} وليست نطاق المستشفى الرسمي." if ar else
                f"The button resolves to {domain}, not the hospital's official domain.",
                link, "link"),
        ]
    elif channel == "visible_link":
        medium_channel = "link"
        suspicious_link = link
        request = (f"يمكن مراجعة البند {deadline} عبر عنوان المتابعة أدناه."
                   if ar else f"The item can be reviewed {deadline} using the address below.")
        body = f"{greeting}،\n\n{intro} {request}\n\n{link}\n\n{closing}\n\nمع التحية،\n{signature}" if ar else \
               f"{greeting},\n\n{intro} {request}\n\n{link}\n\n{closing}\n\nRegards,\n{signature}"
        indicators += [
            _v30_indicator(2, "deadline",
                "مهلة مهنية تضغط على المستلم" if ar else "Plausible deadline pressure",
                "المهلة تبدو معقولة لكنها تُستخدم لدفع المستخدم إلى اتخاذ إجراء سريع." if ar else
                "The deadline sounds reasonable but is used to encourage action before independent verification.",
                deadline, "body"),
            _v30_indicator(3, "link",
                "رابط نصي خارجي" if ar else "External text link",
                f"الرابط الظاهر يستخدم النطاق {domain} غير المعتمد." if ar else
                f"The visible URL uses the unapproved domain {domain}.",
                link, "link"),
        ]
    elif channel == "pdf_attachment":
        medium_channel = "none"
        attachment = re.sub(r"[^a-z0-9]+", "_", plan["family_id"].lower()).strip("_") + f"_{_V35_RNG.randrange(10,99)}.pdf"
        request = (f"يرجى مراجعة الملخص المرفق {deadline} وإضافة ملاحظتك في النظام المعتاد."
                   if ar else f"Please review the attached summary {deadline} and record your comment in the usual system.")
        body = f"{greeting}،\n\n{intro} {request}\n\n{closing}\n\nمع التحية،\n{signature}" if ar else \
               f"{greeting},\n\n{intro} {request}\n\n{closing}\n\nBest regards,\n{signature}"
        indicators += [
            _v30_indicator(2, "attachment",
                "مرفق PDF غير متوقع" if ar else "Unexpected PDF attachment",
                "تستخدم الرسالة مرفق PDF لم يطلبه المستلم ويجب التحقق منه قبل فتحه." if ar else
                "The email includes an unsolicited PDF that should be verified before opening.",
                attachment, "attachment"),
            _v30_indicator(3, "deadline",
                "مهلة مراجعة غير متوقعة" if ar else "Unexpected review deadline",
                "تطلب الرسالة إتمام المراجعة ضمن مهلة لم يتم تأكيدها عبر القناة الداخلية." if ar else
                "The message introduces a deadline that was not confirmed through the normal internal workflow.",
                deadline, "body"),
        ]
    elif channel == "sharepoint_button":
        medium_channel = "button"
        suspicious_link = link
        label = "فتح المجلد المشترك" if ar else "Open shared folder"
        request = (f"تمت مشاركة نسخة العمل مع القسم للمراجعة {deadline}."
                   if ar else f"A working copy has been shared with the department for review {deadline}.")
        body = f"{greeting}،\n\n{intro} {request}\n\n[{label}]({link})\n\n{closing}\n\nمع التقدير،\n{signature}" if ar else \
               f"{greeting},\n\n{intro} {request}\n\n[{label}]({link})\n\n{closing}\n\nKind regards,\n{signature}"
        indicators += [
            _v30_indicator(2, "sharepoint",
                "مشاركة ملفات خارج البيئة الرسمية" if ar else "Unverified shared-document workflow",
                "تدّعي الرسالة مشاركة ملف عبر مساحة تشبه SharePoint دون استخدام المسار الرسمي المعروف." if ar else
                "The email claims a shared-document workflow without using the known official SharePoint route.",
                request, "body"),
            _v30_indicator(3, "link",
                "وجهة مشاركة غير رسمية" if ar else "Unapproved shared-folder destination",
                f"زر المجلد المشترك يقود إلى {domain}، وليس نطاق Microsoft أو المستشفى المعتمد." if ar else
                f"The shared-folder button resolves to {domain}, not an approved hospital or Microsoft domain.",
                link, "link"),
        ]
    else:  # internal_reply
        medium_channel = "none"
        requested = "رقم الموظف والقسم" if ar else "your employee ID and department"
        request = (f"لتحديث سجل التوزيع، يرجى الرد على هذه الرسالة بذكر {requested} {deadline}."
                   if ar else f"To update the assignment record, reply with {requested} {deadline}.")
        body = f"{greeting}،\n\n{intro} {request}\n\n{closing}\n\nمع التحية،\n{signature}" if ar else \
               f"{greeting},\n\n{intro} {request}\n\n{closing}\n\nRegards,\n{signature}"
        indicators += [
            _v30_indicator(2, "reply_request",
                "طلب معلومات عبر الرد على البريد" if ar else "Unusual reply-based information request",
                "تطلب الرسالة إرسال معلومات موظف عبر الرد بدل تحديثها في النظام الرسمي." if ar else
                "The message asks for employee information by email reply instead of through the official system.",
                request, "body"),
            _v30_indicator(3, "workflow",
                "مسار عمل غير قياسي" if ar else "Non-standard workflow",
                "طلب تحديث سجل داخلي عبر البريد لا يطابق الإجراء المعتاد." if ar else
                "Updating an internal assignment record by email does not match the normal workflow.",
                requested, "body"),
        ]

    body = _v35_compact(body)
    # Dynamic tutor text based on the actual indicator titles.
    titles = [str(i.get("title", "")).strip() for i in indicators]
    if ar:
        why = "تبدو الرسالة مرتبطة بالعمل، لكن " + "، و".join(titles) + " تجعل الطلب غير موثوق دون تحقق مستقل."
        tip = {
            "button": "افتح النظام الرسمي من الاختصار المحفوظ بدل استخدام زر وصل عبر البريد.",
            "visible_link": "قارن النطاق حرفًا بحرف مع النطاق الرسمي قبل فتح أي رابط.",
            "pdf_attachment": "تحقق من المرفق مع القسم عبر قناة موثوقة قبل فتحه.",
            "sharepoint_button": "افتح SharePoint من بوابة Microsoft الرسمية، لا من زر داخل رسالة غير متوقعة.",
            "internal_reply": "لا ترسل بيانات الموظف عبر الرد؛ استخدم النظام الرسمي أو اتصل بالقسم.",
        }[channel]
    else:
        why = "The message is workplace-relevant, but " + ", ".join(titles) + " make the request unsafe without independent verification."
        tip = {
            "button": "Open the official system from a saved bookmark rather than using an emailed button.",
            "visible_link": "Compare the destination domain character by character with the official hospital domain.",
            "pdf_attachment": "Verify the attachment with the department through a trusted channel before opening it.",
            "sharepoint_button": "Open SharePoint through the official Microsoft or hospital portal, not an unexpected email button.",
            "internal_reply": "Do not send employee details by reply; use the official system or contact the department directly.",
        }[channel]

    return {
        "from": sender,
        "to": recipient,
        "subject": subject,
        "body": body,
        "attachment": attachment,
        "suspicious_link": suspicious_link,
        "suspicious_text": next((i["evidence"] for i in indicators if i.get("target") == "body"), ""),
        "indicators": indicators,
        "why_risky": why,
        "learning_tip": tip,
        "is_phishing": True,
        "email_type": "Phishing",
        "attack_type": {
            "button": "Look-alike portal",
            "visible_link": "External review link",
            "pdf_attachment": "Malicious PDF lure",
            "sharepoint_button": "Fake shared document",
            "internal_reply": "Reply-based information harvesting",
        }[channel],
        "risk_level": "medium",
        "scenario_id": plan["fingerprint"] + ":" + channel + ":" + structure,
        "scenario_meta": dict(plan, medium_channel_type=channel, writer_style=structure),
        "medium_channel": medium_channel,
        "display_time": _V35_RNG.choice([
            "Monday, 9:18 AM", "Tuesday, 1:26 PM", "Wednesday, 11:04 AM",
            "Thursday, 8:47 AM", "Friday, 2:12 PM", "Yesterday, 3:39 PM"
        ]),
    }

def _v35_validate_medium(result):
    if not isinstance(result, dict) or not result.get("is_phishing"):
        return False
    body = str(result.get("body", ""))
    link = str(result.get("suspicious_link", "") or "")
    attachment = str(result.get("attachment", "") or "")
    channel = str(result.get("scenario_meta", {}).get("medium_channel_type", ""))
    inds = result.get("indicators", []) if isinstance(result.get("indicators"), list) else []
    if len(inds) != 3:
        return False
    if "[QR" in body.upper():
        return False
    has_button = bool(re.search(r"\[[^\]]+\]\s*\(\s*https?://", body))
    visible_urls = re.findall(r"https?://[^\s\)\]]+", body)
    if channel in {"button", "sharepoint_button"}:
        if not has_button or not link or len(visible_urls) != 1:
            return False
    elif channel == "visible_link":
        if has_button or not link or body.count(link) != 1:
            return False
    elif channel == "pdf_attachment":
        if has_button or link or not attachment.lower().endswith(".pdf"):
            return False
    elif channel == "internal_reply":
        if has_button or link or attachment:
            return False
    else:
        return False
    # Grounding: every evidence string exists in its declared field.
    sources = {
        "from": str(result.get("from", "")),
        "body": body,
        "link": link,
        "attachment": attachment,
        "greeting": body,
        "subject": str(result.get("subject", "")),
    }
    for item in inds:
        evidence = str(item.get("evidence", "")).strip()
        location = str(item.get("location", item.get("target", "body")))
        if evidence and evidence.lower() not in sources.get(location, body).lower():
            return False
    return True

_V35_BASE_V33_GENERATE = _v33_generate

def _v35_generate(role, index, language, difficulty="medium", is_phishing=True, assessment=False):
    diff = str(difficulty or "medium").lower()
    if diff != "medium" or not is_phishing:
        return _V35_BASE_V33_GENERATE(role, index, language, difficulty, is_phishing, assessment)
    phase = "assess" if assessment else "learn"
    result = None
    for attempt in range(18):
        plan = _v33_plan(role, index + attempt * 104729, language, "medium", phase, True)
        result = _v35_medium_phishing(plan, role, index)
        if _v35_validate_medium(result):
            try:
                evaluate_and_log_auto_scores(result, "medium", language, is_phishing=True)
            except Exception:
                pass
            return result
    return result



# =============================================================
# END MEDIUM GENERATION ENGINE v35
# =============================================================



# =============================================================
# MEDIUM GENERATION ENGINE v36 — PHASE 1
# Scenario Engine + Action Engine + Goal Engine
# -------------------------------------------------------------
# Safety scope:
#   * Medium phishing only.
#   * v35 remains intact and is used as the composer/fallback.
#   * Easy, Hard, legitimate messages, UI and all other systems
#     remain unchanged.
# =============================================================

_V36_RNG = random.SystemRandom()

# Scenario families are semantic building blocks, not email templates.
# Each generated plan combines a family, event, object, sender and signature.
V36_SCENARIO_BANK = {
    "clinical": [
        {"id":"medication_safety","area_en":"Medication Safety","area_ar":"سلامة الدواء",
         "events_en":["a medication reconciliation exception","a high-alert medicine review","a formulary safety update","a dispensing variance follow-up"],
         "events_ar":["استثناء في مطابقة الأدوية","مراجعة دواء عالي الخطورة","تحديث سلامة قائمة الأدوية","متابعة اختلاف في صرف الدواء"],
         "objects_en":["medication reconciliation record","high-alert medication checklist","dose variance note","pharmacy intervention summary"],
         "objects_ar":["سجل مطابقة الأدوية","قائمة تحقق الأدوية عالية الخطورة","ملاحظة اختلاف الجرعة","ملخص تدخل الصيدلية"],
         "senders_en":["Medication Safety Office","Clinical Pharmacy Coordination","Pharmacy Quality Unit"],
         "senders_ar":["مكتب سلامة الدواء","تنسيق الصيدلة السريرية","وحدة جودة الصيدلية"]},
        {"id":"patient_referral","area_en":"Referral Coordination","area_ar":"تنسيق الإحالات",
         "events_en":["a returned referral","a specialist acceptance update","an incomplete referral packet","a transfer-of-care follow-up"],
         "events_ar":["إحالة معادة","تحديث قبول من اختصاصي","ملف إحالة غير مكتمل","متابعة انتقال الرعاية"],
         "objects_en":["patient referral summary","specialist acceptance record","transfer note","referral documentation set"],
         "objects_ar":["ملخص إحالة المريض","سجل قبول الاختصاصي","ملاحظة نقل الرعاية","حزمة مستندات الإحالة"],
         "senders_en":["Referral Coordination","Patient Flow Office","Care Transition Team"],
         "senders_ar":["تنسيق الإحالات","مكتب تدفق المرضى","فريق انتقال الرعاية"]},
        {"id":"infection_control","area_en":"Infection Prevention","area_ar":"مكافحة العدوى",
         "events_en":["an exposure follow-up","an isolation-practice review","a surveillance finding","a hand-hygiene audit exception"],
         "events_ar":["متابعة تعرض مهني","مراجعة ممارسات العزل","نتيجة ترصد","استثناء في تدقيق نظافة اليدين"],
         "objects_en":["exposure review record","isolation checklist","surveillance summary","infection-control audit note"],
         "objects_ar":["سجل مراجعة التعرض","قائمة تحقق العزل","ملخص الترصد","ملاحظة تدقيق مكافحة العدوى"],
         "senders_en":["Infection Prevention Unit","Occupational Health","Clinical Surveillance Team"],
         "senders_ar":["وحدة مكافحة العدوى","الصحة المهنية","فريق الترصد السريري"]},
        {"id":"laboratory","area_en":"Laboratory Services","area_ar":"خدمات المختبر",
         "events_en":["a specimen rejection follow-up","a corrected-result notice","a quality-control exception","a critical-result workflow update"],
         "events_ar":["متابعة رفض عينة","إشعار نتيجة مصححة","استثناء في ضبط الجودة","تحديث مسار نتيجة حرجة"],
         "objects_en":["specimen exception record","corrected laboratory report","quality-control summary","critical-result acknowledgement"],
         "objects_ar":["سجل استثناء العينة","تقرير مختبر مصحح","ملخص ضبط الجودة","إقرار نتيجة حرجة"],
         "senders_en":["Laboratory Quality","Pathology Coordination","Critical Results Desk"],
         "senders_ar":["جودة المختبر","تنسيق علم الأمراض","مكتب النتائج الحرجة"]},
        {"id":"radiology","area_en":"Radiology Services","area_ar":"خدمات الأشعة",
         "events_en":["a protocol clarification","a report amendment","a contrast-safety follow-up","an imaging-priority change"],
         "events_ar":["توضيح بروتوكول","تعديل تقرير","متابعة سلامة مادة التباين","تغيير أولوية تصوير"],
         "objects_en":["imaging protocol note","amended radiology report","contrast screening form","priority imaging request"],
         "objects_ar":["ملاحظة بروتوكول تصوير","تقرير أشعة معدل","نموذج فحص مادة التباين","طلب تصوير ذي أولوية"],
         "senders_en":["Radiology Coordination","Imaging Quality Office","Diagnostic Services"],
         "senders_ar":["تنسيق الأشعة","مكتب جودة التصوير","الخدمات التشخيصية"]},
        {"id":"clinical_training","area_en":"Clinical Education","area_ar":"التعليم السريري",
         "events_en":["a competency renewal","a simulation-session follow-up","a mandatory module update","a skills-validation exception"],
         "events_ar":["تجديد كفاءة","متابعة جلسة محاكاة","تحديث وحدة إلزامية","استثناء في التحقق من المهارة"],
         "objects_en":["competency record","simulation attendance","training module","skills validation form"],
         "objects_ar":["سجل الكفاءة","حضور المحاكاة","وحدة تدريبية","نموذج التحقق من المهارة"],
         "senders_en":["Clinical Education","Nursing Development","Medical Training Office"],
         "senders_ar":["التعليم السريري","تطوير التمريض","مكتب التدريب الطبي"]},
    ],
    "admin": [
        {"id":"hr","area_en":"Human Resources","area_ar":"الموارد البشرية","events_en":["an employee-file correction","a benefits enrolment change","a leave-balance exception","a credentialing renewal"],"events_ar":["تصحيح ملف موظف","تغيير تسجيل المزايا","استثناء في رصيد الإجازات","تجديد اعتماد وظيفي"],"objects_en":["employee profile","benefits selection","leave balance record","credentialing document"],"objects_ar":["ملف الموظف","اختيار المزايا","سجل رصيد الإجازات","مستند الاعتماد الوظيفي"],"senders_en":["HR Operations","Workforce Services","Employee Relations"],"senders_ar":["عمليات الموارد البشرية","خدمات القوى العاملة","علاقات الموظفين"]},
        {"id":"audit","area_en":"Internal Audit","area_ar":"التدقيق الداخلي","events_en":["a sample-selection notice","a control-evidence request","an unresolved audit item","a compliance follow-up"],"events_ar":["إشعار اختيار عينة","طلب دليل رقابي","بند تدقيق غير مغلق","متابعة امتثال"],"objects_en":["audit sample","control evidence pack","open audit observation","compliance response"],"objects_ar":["عينة التدقيق","حزمة الأدلة الرقابية","ملاحظة تدقيق مفتوحة","استجابة الامتثال"],"senders_en":["Internal Audit","Compliance Review","Governance Office"],"senders_ar":["التدقيق الداخلي","مراجعة الامتثال","مكتب الحوكمة"]},
        {"id":"procurement","area_en":"Medical Procurement","area_ar":"المشتريات الطبية","events_en":["a supplier-document correction","a purchase-order hold","a contract-review request","a delivery discrepancy"],"events_ar":["تصحيح مستند مورد","تعليق أمر شراء","طلب مراجعة عقد","اختلاف في التسليم"],"objects_en":["supplier record","purchase order","service contract","delivery note"],"objects_ar":["سجل المورد","أمر الشراء","عقد الخدمة","إشعار التسليم"],"senders_en":["Medical Procurement","Supply Chain","Vendor Management"],"senders_ar":["المشتريات الطبية","سلسلة الإمداد","إدارة الموردين"]},
        {"id":"insurance","area_en":"Insurance Coordination","area_ar":"تنسيق التأمين","events_en":["a rejected-claim follow-up","a coverage verification","a pre-authorisation exception","a payer-document update"],"events_ar":["متابعة مطالبة مرفوضة","تحقق من التغطية","استثناء موافقة مسبقة","تحديث مستند جهة دافعة"],"objects_en":["claim record","coverage profile","pre-authorisation request","payer document"],"objects_ar":["سجل المطالبة","ملف التغطية","طلب الموافقة المسبقة","مستند جهة الدفع"],"senders_en":["Insurance Coordination","Revenue Cycle","Patient Financial Services"],"senders_ar":["تنسيق التأمين","دورة الإيرادات","الخدمات المالية للمرضى"]},
        {"id":"schedule","area_en":"Workforce Scheduling","area_ar":"جدولة القوى العاملة","events_en":["a roster revision","an overtime exception","a coverage-gap notice","a shift-allocation update"],"events_ar":["تعديل جدول","استثناء عمل إضافي","إشعار فجوة تغطية","تحديث توزيع مناوبات"],"objects_en":["department roster","overtime request","coverage plan","shift allocation"],"objects_ar":["جدول القسم","طلب العمل الإضافي","خطة التغطية","توزيع المناوبات"],"senders_en":["Workforce Scheduling","Operations Coordination","Staffing Office"],"senders_ar":["جدولة القوى العاملة","تنسيق العمليات","مكتب التوظيف"]},
    ],
    "it": [
        {"id":"incident","area_en":"Security Operations","area_ar":"عمليات الأمن السيبراني","events_en":["an endpoint alert follow-up","a privileged-access review","a suspicious-session investigation","a service-account exception"],"events_ar":["متابعة تنبيه جهاز","مراجعة وصول مميز","تحقيق جلسة مشبوهة","استثناء حساب خدمة"],"objects_en":["endpoint alert","privileged access record","session investigation","service account"],"objects_ar":["تنبيه جهاز","سجل الوصول المميز","تحقيق الجلسة","حساب الخدمة"],"senders_en":["Security Operations","Cyber Defence","Identity Security"],"senders_ar":["عمليات الأمن السيبراني","الدفاع السيبراني","أمن الهوية"]},
        {"id":"service_change","area_en":"IT Service Management","area_ar":"إدارة خدمات تقنية المعلومات","events_en":["a change-window update","a service restoration notice","a configuration exception","a maintenance follow-up"],"events_ar":["تحديث نافذة تغيير","إشعار استعادة خدمة","استثناء إعداد","متابعة صيانة"],"objects_en":["change record","service restoration task","configuration item","maintenance ticket"],"objects_ar":["سجل التغيير","مهمة استعادة الخدمة","عنصر الإعداد","تذكرة الصيانة"],"senders_en":["IT Service Management","Infrastructure Operations","Technical Support"],"senders_ar":["إدارة خدمات تقنية المعلومات","عمليات البنية التحتية","الدعم التقني"]},
        {"id":"access","area_en":"Identity and Access","area_ar":"الهوية والوصول","events_en":["an access recertification","a role-assignment exception","a dormant-account review","an application-access change"],"events_ar":["إعادة اعتماد وصول","استثناء تعيين صلاحية","مراجعة حساب خامل","تغيير وصول لتطبيق"],"objects_en":["access review","role assignment","dormant account","application entitlement"],"objects_ar":["مراجعة الوصول","تعيين الصلاحية","الحساب الخامل","استحقاق التطبيق"],"senders_en":["Identity and Access","Directory Services","Application Security"],"senders_ar":["الهوية والوصول","خدمات الدليل","أمن التطبيقات"]},
        {"id":"backup","area_en":"Data Protection","area_ar":"حماية البيانات","events_en":["a failed-backup follow-up","a restore-test exception","a retention-policy update","a storage-capacity notice"],"events_ar":["متابعة فشل نسخ احتياطي","استثناء اختبار استعادة","تحديث سياسة الاحتفاظ","إشعار سعة تخزين"],"objects_en":["backup job","restore test","retention setting","storage allocation"],"objects_ar":["مهمة النسخ الاحتياطي","اختبار الاستعادة","إعداد الاحتفاظ","تخصيص التخزين"],"senders_en":["Data Protection","Backup Operations","Infrastructure Reliability"],"senders_ar":["حماية البيانات","عمليات النسخ الاحتياطي","موثوقية البنية التحتية"]},
    ],
}
# Other inherits a balanced mix without changing role selection elsewhere.
V36_SCENARIO_BANK["other"] = (
    V36_SCENARIO_BANK["clinical"][:2] +
    V36_SCENARIO_BANK["admin"][:2] +
    V36_SCENARIO_BANK["it"][:2]
)

V36_ACTION_WEIGHTS = [
    ("button", 18), ("visible_link", 17), ("pdf_attachment", 16),
    ("sharepoint", 16), ("internal_workspace", 17), ("reply_request", 16),
]

V36_GOALS = {
    "button": ["review", "approve", "verify", "confirm", "update", "check"],
    "visible_link": ["review", "verify", "confirm", "read", "check", "update"],
    "pdf_attachment": ["review", "read", "download", "check", "confirm"],
    "sharepoint": ["review", "read", "approve", "confirm", "check"],
    "internal_workspace": ["review", "approve", "update", "check", "confirm"],
    "reply_request": ["confirm", "verify", "acknowledge", "submit", "update"],
}
V36_GOAL_TEXT = {
    "English": {"review":"review", "approve":"approve", "verify":"verify", "confirm":"confirm", "download":"download", "read":"read", "update":"update", "check":"check", "acknowledge":"acknowledge", "submit":"submit"},
    "Arabic": {"review":"مراجعة", "approve":"اعتماد", "verify":"التحقق من", "confirm":"تأكيد", "download":"تنزيل", "read":"قراءة", "update":"تحديث", "check":"فحص", "acknowledge":"الإقرار بـ", "submit":"إرسال"},
}


def _v36_memory(role_type, language, phase):
    cycle = st.session_state.get("v18_cycle_id", "default")
    key = f"v36_phase1_memory_{cycle}_{role_type}_{language}_{phase}"
    return st.session_state.setdefault(key, {
        "families": [], "events": [], "objects": [], "actions": [], "goals": [], "combinations": []
    })


def _v36_fresh_choice(values, recent, window=10):
    vals = list(values)
    blocked = set(recent[-window:])
    fresh = [v for v in vals if v not in blocked]
    return _V36_RNG.choice(fresh or vals)


def _v36_action_engine(mem):
    # Weighted selection with anti-streak and short-window variety.
    candidates = list(V36_ACTION_WEIGHTS)
    if len(mem["actions"]) >= 2 and mem["actions"][-1] == mem["actions"][-2]:
        candidates = [(a, w) for a, w in candidates if a != mem["actions"][-1]]
    recent = set(mem["actions"][-4:])
    fresh = [(a, w) for a, w in candidates if a not in recent]
    return _v35_weighted_choice(fresh or candidates)


def _v36_goal_engine(action_type, mem):
    goals = V36_GOALS[action_type]
    return _v36_fresh_choice(goals, mem["goals"], window=6)


def _v36_scenario_engine(role, index, language, phase):
    role_type = _v30_role_type(role)
    mem = _v36_memory(role_type, language, phase)
    bank = V36_SCENARIO_BANK.get(role_type, V36_SCENARIO_BANK["other"])
    recent_families = set(mem["families"][-5:])
    choices = [f for f in bank if f["id"] not in recent_families] or bank
    family = _V36_RNG.choice(choices)
    ar = language == "Arabic"
    event = _v36_fresh_choice(family["events_ar" if ar else "events_en"], mem["events"], 12)
    obj = _v36_fresh_choice(family["objects_ar" if ar else "objects_en"], mem["objects"], 12)
    sender = _V36_RNG.choice(family["senders_ar" if ar else "senders_en"])
    area = family["area_ar" if ar else "area_en"]
    signature = sender
    action_type = _v36_action_engine(mem)
    goal = _v36_goal_engine(action_type, mem)
    combo = f"{family['id']}|{event}|{obj}|{action_type}|{goal}"
    # Avoid exact semantic combinations in the active cycle.
    for _ in range(30):
        if combo not in mem["combinations"]:
            break
        action_type = _v36_action_engine(mem)
        goal = _v36_goal_engine(action_type, mem)
        event = _V36_RNG.choice(family["events_ar" if ar else "events_en"])
        obj = _V36_RNG.choice(family["objects_ar" if ar else "objects_en"])
        combo = f"{family['id']}|{event}|{obj}|{action_type}|{goal}"
    mem["families"].append(family["id"]); mem["events"].append(event)
    mem["objects"].append(obj); mem["actions"].append(action_type)
    mem["goals"].append(goal); mem["combinations"].append(combo)
    for k in mem:
        mem[k] = mem[k][-80:]
    return {
        "role_type": role_type, "family_id": family["id"], "area": area,
        "event": event, "object": obj, "sender": sender, "signature": signature,
        "action_type": action_type, "goal": goal, "semantic_combo": combo,
    }


def _v36_to_v35_channel(action_type):
    return {
        "button": "button", "visible_link": "visible_link",
        "pdf_attachment": "pdf_attachment", "sharepoint": "sharepoint_button",
        "internal_workspace": "button", "reply_request": "internal_reply",
    }[action_type]


def _v36_apply_goal_text(result, action_type, goal, language):
    """Small, schema-safe goal pass. It changes only CTA wording and metadata."""
    if not isinstance(result, dict):
        return result
    ar = language == "Arabic"
    body = str(result.get("body", ""))
    labels_en = {
        "review":"Review item", "approve":"Approve item", "verify":"Verify details",
        "confirm":"Confirm update", "download":"Download document", "read":"Read notice",
        "update":"Update record", "check":"Check item", "acknowledge":"Acknowledge notice",
        "submit":"Submit response",
    }
    labels_ar = {
        "review":"مراجعة البند", "approve":"اعتماد البند", "verify":"التحقق من التفاصيل",
        "confirm":"تأكيد التحديث", "download":"تنزيل المستند", "read":"قراءة الإشعار",
        "update":"تحديث السجل", "check":"فحص البند", "acknowledge":"الإقرار بالإشعار",
        "submit":"إرسال الرد",
    }
    label = (labels_ar if ar else labels_en)[goal]
    if action_type in {"button", "internal_workspace", "sharepoint"}:
        body = re.sub(r"\[([^\]]+)\](\(https?://[^\)]+\))", lambda m: f"[{label}]{m.group(2)}", body, count=1)
    if action_type == "internal_workspace":
        result["attack_type"] = "Fake internal workspace"
    result["body"] = body
    meta = dict(result.get("scenario_meta") or {})
    meta.update({"v36_action_type": action_type, "v36_goal": goal})
    result["scenario_meta"] = meta
    result["scenario_id"] = str(result.get("scenario_id", "")) + f":{action_type}:{goal}"
    return result


_V36_BASE_V35_GENERATE = _v35_generate


def _v36_generate(role, index, language, difficulty="medium", is_phishing=True, assessment=False):
    diff = str(difficulty or "medium").lower()
    if diff != "medium" or not is_phishing:
        return _V36_BASE_V35_GENERATE(role, index, language, difficulty, is_phishing, assessment)
    phase = "assess" if assessment else "learn"
    last_result = None
    for attempt in range(20):
        scenario = _v36_scenario_engine(role, index + attempt * 7919, language, phase)
        base_plan = _v33_plan(role, index + attempt * 104729, language, "medium", phase, True)
        goal_text = V36_GOAL_TEXT[language][scenario["goal"]]
        plan = dict(base_plan)
        plan.update({
            "role_type": scenario["role_type"], "family_id": scenario["family_id"],
            "area": scenario["area"], "event": scenario["event"],
            "object": scenario["object"], "sender": scenario["sender"],
            "signature": scenario["signature"], "action": goal_text,
            "v36_action_type": scenario["action_type"], "v36_goal": scenario["goal"],
            "v36_semantic_combo": scenario["semantic_combo"],
        })
        # v35 channel selection is session-based. Supply a temporary exact channel
        # for this generation only, then restore state immediately.
        cycle = st.session_state.get("v18_cycle_id", "default")
        ch_key = f"v35_medium_channels_{cycle}_{phase}"
        old_channels = list(st.session_state.get(ch_key, []))
        forced = _v36_to_v35_channel(scenario["action_type"])
        channels = list(old_channels)
        while len(channels) <= int(index):
            channels.append(forced)
        channels[int(index)] = forced
        st.session_state[ch_key] = channels
        try:
            last_result = _v35_medium_phishing(plan, role, index)
        finally:
            st.session_state[ch_key] = old_channels
        last_result = _v36_apply_goal_text(last_result, scenario["action_type"], scenario["goal"], language)
        if _v35_validate_medium(last_result):
            try:
                evaluate_and_log_auto_scores(last_result, "medium", language, is_phishing=True)
            except Exception:
                pass
            return last_result
    # Safe fallback: never expose a failed v36 candidate.
    return _V36_BASE_V35_GENERATE(role, index, language, difficulty, is_phishing, assessment)





# =============================================================
# END MEDIUM GENERATION ENGINE v36 — PHASE 1
# =============================================================


# =============================================================
# MEDIUM GENERATION ENGINE v40 — SINGLE-FILE SAFE REBUILD
# -------------------------------------------------------------
# Scope: Medium phishing only. Easy, Hard, legitimate messages,
# UI, reports, authentication, providers and storage are untouched.
# v36/v35 remain available as automatic fallback.
# =============================================================

V40_ENABLED = True
V40_USE_API_TEXT = True
_V40_RNG = random.SystemRandom()
_V40_FALLBACK = _v36_generate

V40_ACTIONS = {
    "button": 20,
    "visible_link": 17,
    "pdf_attachment": 17,
    "sharepoint": 15,
    "internal_workspace": 16,
    "reply_request": 15,
}

V40_GOALS = {
    "button": ["review", "approve", "verify", "confirm", "acknowledge", "check"],
    "visible_link": ["review", "verify", "read", "update", "check", "confirm"],
    "pdf_attachment": ["review", "read", "download", "check", "approve"],
    "sharepoint": ["review", "read", "approve", "confirm", "acknowledge"],
    "internal_workspace": ["review", "update", "confirm", "check", "approve"],
    "reply_request": ["confirm", "verify", "submit", "acknowledge", "update"],
}

V40_STYLES = [
    "formal_notice", "brief_operational", "follow_up", "system_alert",
    "department_request", "meeting_reference", "handover_note",
    "quality_escalation", "shared_document_notice", "vendor_style",
]

V40_SCENARIOS = {
    "clinical": [
        {"id":"medication_safety","area":"Medication Safety","senders":["Medication Safety Office","Clinical Pharmacy","Pharmacy Quality Unit"],"senders_ar":["مكتب سلامة الدواء","الصيدلية السريرية","وحدة جودة الصيدلة"],
         "events":["dose clarification pending","high-alert medication review","formulary exception follow-up","medication reconciliation discrepancy","antimicrobial approval note"],
         "events_ar":["توضيح جرعة معلق","مراجعة دواء عالي الخطورة","متابعة استثناء بالتشكيلة الدوائية","تباين في مطابقة الأدوية","ملاحظة اعتماد مضاد حيوي"],
         "objects":["insulin order","anticoagulant record","discharge medication list","restricted antibiotic request","infusion concentration"],
         "objects_ar":["طلب الإنسولين","سجل مضاد التخثر","قائمة أدوية الخروج","طلب مضاد حيوي مقيّد","تركيز محلول التسريب"]},
        {"id":"laboratory","area":"Laboratory Services","senders":["Laboratory Quality","Pathology Coordination","Diagnostic Services"],"senders_ar":["جودة المختبر","تنسيق الباثولوجيا","الخدمات التشخيصية"],
         "events":["specimen exception review","critical-result acknowledgement","sample rejection follow-up","reference range update","external test reconciliation"],
         "events_ar":["مراجعة استثناء عينة","إقرار نتيجة حرجة","متابعة رفض عينة","تحديث المعدل المرجعي","مطابقة فحص خارجي"],
         "objects":["blood culture specimen","histopathology slide","coagulation sample","molecular test request","critical potassium result"],
         "objects_ar":["عينة مزرعة الدم","شريحة الأنسجة المرضية","عينة تخثر","طلب فحص جزيئي","نتيجة بوتاسيوم حرجة"]},
        {"id":"radiology","area":"Radiology","senders":["Radiology Operations","Imaging Quality","PACS Coordination"],"senders_ar":["عمليات الأشعة","جودة التصوير","تنسيق نظام الصور"],
         "events":["report addendum pending","contrast checklist correction","image-routing exception","urgent finding acknowledgement","protocol change notice"],
         "events_ar":["ملحق تقرير معلق","تصحيح قائمة تدقيق الصبغة","استثناء توجيه الصورة","إقرار نتيجة عاجلة","إشعار تغيير البروتوكول"],
         "objects":["CT pulmonary angiogram","MRI safety form","portable chest image","ultrasound referral","contrast administration record"],
         "objects_ar":["تصوير مقطعي للشريان الرئوي","نموذج سلامة الرنين المغناطيسي","صورة صدر متنقلة","إحالة الموجات فوق الصوتية","سجل إعطاء الصبغة"]},
        {"id":"infection_control","area":"Infection Prevention","senders":["Infection Prevention","IPC Surveillance","Occupational Health"],"senders_ar":["مكافحة العدوى","ترصد مكافحة العدوى","الصحة المهنية"],
         "events":["exposure follow-up","isolation audit finding","screening list correction","outbreak contact review","hand-hygiene observation"],
         "events_ar":["متابعة تعرض","نتيجة تدقيق العزل","تصحيح قائمة الفحص","مراجعة مخالطي التفشي","ملاحظة نظافة اليدين"],
         "objects":["MRSA screening record","needle-stick exposure","isolation room log","staff vaccination status","contact tracing list"],
         "objects_ar":["سجل فحص المكورات المقاومة","تعرض بوخز إبرة","سجل غرفة العزل","حالة تطعيم الموظف","قائمة تتبع المخالطين"]},
        {"id":"patient_safety","area":"Patient Safety","senders":["Patient Safety Office","Clinical Governance","Risk Management"],"senders_ar":["مكتب سلامة المرضى","الحوكمة السريرية","إدارة المخاطر"],
         "events":["incident follow-up","near-miss clarification","safety action acknowledgement","case-review assignment","root-cause evidence request"],
         "events_ar":["متابعة حادثة","توضيح حالة كادت تقع","إقرار إجراء سلامة","تكليف مراجعة حالة","طلب دليل السبب الجذري"],
         "objects":["fall incident","medication near miss","wrong-label event","handover omission","delayed escalation case"],
         "objects_ar":["حادثة سقوط","خطأ دوائي وشيك","حدث خطأ في الملصق","إغفال بالتسليم","حالة تأخر تصعيد"]},
        {"id":"referral","area":"Referral Coordination","senders":["Referral Management","Care Coordination","Transfer Centre"],"senders_ar":["إدارة الإحالات","تنسيق الرعاية","مركز التحويل"],
         "events":["referral returned for correction","transfer acceptance pending","specialist response requested","appointment pathway update","external facility note"],
         "events_ar":["إحالة معادة للتصحيح","قبول تحويل معلق","رد استشاري مطلوب","تحديث مسار الموعد","ملاحظة منشأة خارجية"],
         "objects":["cardiology referral","oncology transfer","neonatal consultation","rehabilitation request","tertiary-care referral"],
         "objects_ar":["إحالة قلبية","تحويل أورام","استشارة حديثي الولادة","طلب تأهيل","إحالة رعاية متخصصة"]},
        {"id":"equipment","area":"Biomedical Engineering","senders":["Biomedical Engineering","Medical Devices Unit","Clinical Engineering"],"senders_ar":["الهندسة الطبية الحيوية","وحدة الأجهزة الطبية","الهندسة السريرية"],
         "events":["device recall acknowledgement","maintenance slot confirmation","calibration exception","service bulletin review","asset-location verification"],
         "events_ar":["إقرار سحب جهاز","تأكيد موعد صيانة","استثناء معايرة","مراجعة نشرة خدمة","تحقق موقع الأصل"],
         "objects":["infusion pump","ventilator","defibrillator","patient monitor","laboratory analyser"],
         "objects_ar":["مضخة التسريب","جهاز التنفس الصناعي","جهاز الصدمات الكهربائية","جهاز مراقبة المريض","محلل المختبر"]},
        {"id":"training","area":"Clinical Education","senders":["Clinical Education","Simulation Centre","Competency Office"],"senders_ar":["التعليم السريري","مركز المحاكاة","مكتب الكفاءات"],
         "events":["competency evidence pending","mandatory module follow-up","simulation booking change","certificate correction","annual skills validation"],
         "events_ar":["دليل كفاءة معلق","متابعة وحدة إلزامية","تغيير حجز محاكاة","تصحيح شهادة","اعتماد مهارات سنوي"],
         "objects":["basic life support certificate","medication competency","infection-control module","device training record","emergency response drill"],
         "objects_ar":["شهادة الإنعاش الأساسي","كفاءة الأدوية","وحدة مكافحة العدوى","سجل تدريب الأجهزة","تمرين الاستجابة للطوارئ"]},
    ],
    "admin": [
        {"id":"hr","area":"Human Resources","senders":["HR Operations","Workforce Services","Employee Relations"],"senders_ar":["عمليات الموارد البشرية","خدمات القوى العاملة","علاقات الموظفين"],
         "events":["employee record verification","leave balance correction","benefit enrolment follow-up","contract detail confirmation","attendance exception"],
         "events_ar":["تحقق سجل موظف","تصحيح رصيد إجازة","متابعة تسجيل مزايا","تأكيد تفاصيل عقد","استثناء حضور"],
         "objects":["bank information record","annual leave request","housing allowance","employment contract","attendance log"],
         "objects_ar":["سجل بيانات بنكية","طلب إجازة سنوية","بدل سكن","عقد العمل","سجل الحضور"]},
        {"id":"finance","area":"Finance","senders":["Finance Operations","Accounts Payable","Revenue Cycle"],"senders_ar":["عمليات المالية","الحسابات الدائنة","دورة الإيرادات"],
         "events":["invoice exception review","payment batch confirmation","cost-centre correction","refund approval follow-up","reconciliation item"],
         "events_ar":["مراجعة استثناء فاتورة","تأكيد دفعة سداد","تصحيح مركز تكلفة","متابعة اعتماد استرداد","بند مطابقة"],
         "objects":["medical supplier invoice","insurance remittance","patient refund","purchase order","monthly reconciliation file"],
         "objects_ar":["فاتورة مورد طبي","تحويل تأميني","استرداد مريض","أمر شراء","ملف مطابقة شهري"]},
        {"id":"procurement","area":"Procurement","senders":["Procurement Office","Supply Chain","Vendor Management"],"senders_ar":["مكتب المشتريات","سلسلة الإمداد","إدارة الموردين"],
         "events":["supplier document renewal","quotation clarification","delivery discrepancy","contract review","vendor onboarding update"],
         "events_ar":["تجديد مستند مورد","توضيح عرض سعر","تباين تسليم","مراجعة عقد","تحديث تسجيل مورد"],
         "objects":["surgical consumables contract","laboratory reagent order","PPE delivery","maintenance agreement","pharmacy supplier profile"],
         "objects_ar":["عقد مستلزمات جراحية","طلب كواشف مختبر","تسليم معدات وقاية","اتفاقية صيانة","ملف مورد الصيدلية"]},
        {"id":"insurance","area":"Insurance Coordination","senders":["Insurance Office","Claims Management","Revenue Integrity"],"senders_ar":["مكتب التأمين","إدارة المطالبات","نزاهة الإيرادات"],
         "events":["claim documentation correction","coverage verification","pre-authorisation follow-up","rejected claim review","payer rule update"],
         "events_ar":["تصحيح مستندات مطالبة","تحقق تغطية","متابعة تصريح مسبق","مراجعة مطالبة مرفوضة","تحديث قواعد جهة الدفع"],
         "objects":["inpatient claim","day-surgery approval","radiology authorisation","pharmacy claim","emergency admission record"],
         "objects_ar":["مطالبة تنويم","اعتماد جراحة يوم واحد","تصريح أشعة","مطالبة صيدلية","سجل دخول طوارئ"]},
        {"id":"records","area":"Health Information Management","senders":["Medical Records","Document Control","Health Information Management"],"senders_ar":["السجلات الطبية","ضبط المستندات","إدارة المعلومات الصحية"],
         "events":["record completion reminder","scanning discrepancy","release request follow-up","retention notice","coding clarification"],
         "events_ar":["تذكير إكمال سجل","تباين مسح ضوئي","متابعة طلب إفراج","إشعار احتفاظ","توضيح ترميز"],
         "objects":["discharge summary","consent form","patient file index","coding worksheet","record release form"],
         "objects_ar":["ملخص خروج","نموذج موافقة","فهرس ملف مريض","ورقة ترميز","نموذج إفراج سجل"]},
        {"id":"audit","area":"Internal Audit","senders":["Internal Audit","Compliance Office","Quality Assurance"],"senders_ar":["التدقيق الداخلي","مكتب الامتثال","ضمان الجودة"],
         "events":["sample evidence request","audit finding response","control-owner confirmation","policy exception review","closure evidence reminder"],
         "events_ar":["طلب دليل عينة","رد على نتيجة تدقيق","تأكيد مالك الضابط","مراجعة استثناء سياسة","تذكير دليل إغلاق"],
         "objects":["cash handling control","patient identity audit","procurement sample","access review","document retention control"],
         "objects_ar":["ضابط تداول نقدي","تدقيق هوية مريض","عينة مشتريات","مراجعة صلاحيات","ضابط احتفاظ مستندات"]},
        {"id":"schedule","area":"Workforce Scheduling","senders":["Scheduling Office","Operations Planning","Roster Coordination"],"senders_ar":["مكتب الجدولة","تخطيط العمليات","تنسيق الجداول"],
         "events":["roster revision","coverage gap follow-up","overtime confirmation","shift-swap exception","holiday schedule update"],
         "events_ar":["مراجعة جدول","متابعة فجوة تغطية","تأكيد وقت إضافي","استثناء تبديل مناوبة","تحديث جدول إجازة"],
         "objects":["weekend coverage","Ramadan roster","on-call schedule","clinic reception rota","annual leave coverage"],
         "objects_ar":["تغطية عطلة نهاية الأسبوع","جدول رمضان","جدول المناوبة","جدول استقبال العيادة","تغطية إجازة سنوية"]},
    ],
    "it": [
        {"id":"access","area":"Identity and Access","senders":["Identity Services","Access Governance","IT Security"],"senders_ar":["خدمات الهوية","حوكمة الصلاحيات","أمن تقنية المعلومات"],
         "events":["access recertification","privileged account review","inactive account follow-up","role assignment correction","MFA registration notice"],
         "events_ar":["إعادة اعتماد صلاحية","مراجعة حساب مميز","متابعة حساب غير نشط","تصحيح تعيين دور","إشعار تسجيل تحقق ثنائي"],
         "objects":["EMR access","PACS account","VPN profile","shared mailbox permission","administrator role"],
         "objects_ar":["صلاحية النظام الطبي","حساب نظام الصور","ملف VPN","صلاحية بريد مشترك","دور مسؤول"]},
        {"id":"network","area":"Network Operations","senders":["Network Operations","Infrastructure Services","NOC Team"],"senders_ar":["عمليات الشبكة","خدمات البنية التحتية","فريق مركز العمليات"],
         "events":["gateway maintenance notice","certificate update","wireless profile change","firewall exception review","remote access follow-up"],
         "events_ar":["إشعار صيانة بوابة","تحديث شهادة","تغيير ملف لاسلكي","مراجعة استثناء جدار حماية","متابعة وصول عن بعد"],
         "objects":["clinical Wi-Fi profile","VPN gateway","patient portal certificate","firewall rule","remote support session"],
         "objects_ar":["ملف واي فاي سريري","بوابة VPN","شهادة بوابة المريض","قاعدة جدار حماية","جلسة دعم عن بعد"]},
        {"id":"service_desk","area":"IT Service Desk","senders":["IT Service Desk","Technical Support","Endpoint Services"],"senders_ar":["مكتب خدمات تقنية المعلومات","الدعم الفني","خدمات الأجهزة الطرفية"],
         "events":["ticket closure confirmation","device compliance alert","software deployment notice","remote support request","asset assignment correction"],
         "events_ar":["تأكيد إغلاق تذكرة","تنبيه امتثال جهاز","إشعار نشر برنامج","طلب دعم عن بعد","تصحيح تعيين أصل"],
         "objects":["laptop encryption status","clinical workstation","printer driver","antivirus agent","mobile device enrolment"],
         "objects_ar":["حالة تشفير الجهاز المحمول","محطة عمل سريرية","برنامج تشغيل طابعة","برنامج مكافحة فيروسات","تسجيل جهاز محمول"]},
        {"id":"backup","area":"Data Protection","senders":["Backup Operations","Data Protection","Infrastructure Reliability"],"senders_ar":["عمليات النسخ الاحتياطي","حماية البيانات","موثوقية البنية التحتية"],
         "events":["backup failure review","restore test confirmation","retention policy update","storage quota alert","replication exception"],
         "events_ar":["مراجعة فشل نسخ احتياطي","تأكيد اختبار استعادة","تحديث سياسة الاحتفاظ","تنبيه حصة تخزين","استثناء تكرار"],
         "objects":["EMR backup job","radiology archive","shared drive snapshot","database restore test","cloud backup account"],
         "objects_ar":["مهمة نسخ النظام الطبي","أرشيف الأشعة","لقطة قرص مشترك","اختبار استعادة قاعدة بيانات","حساب نسخ سحابي"]},
        {"id":"change","area":"Change Management","senders":["Change Advisory Board","Release Management","Clinical Systems"],"senders_ar":["مجلس استشارة التغيير","إدارة الإصدارات","الأنظمة السريرية"],
         "events":["change approval pending","release window confirmation","rollback plan review","emergency change notice","post-change validation"],
         "events_ar":["اعتماد تغيير معلق","تأكيد نافذة إصدار","مراجعة خطة تراجع","إشعار تغيير طارئ","تحقق ما بعد التغيير"],
         "objects":["EMR upgrade","pharmacy interface","laboratory middleware","patient portal release","network core update"],
         "objects_ar":["ترقية النظام الطبي","واجهة الصيدلية","برمجية وسيطة للمختبر","إصدار بوابة المريض","تحديث الشبكة الأساسية"]},
    ],
    "other": [
        {"id":"policy","area":"Hospital Operations","senders":["Hospital Operations","Policy Office","Corporate Services"],"senders_ar":["عمليات المستشفى","مكتب السياسات","الخدمات المؤسسية"],
         "events":["policy acknowledgement","service update","department contact verification","facility notice","staff information update"],
         "events_ar":["إقرار سياسة","تحديث خدمة","تحقق جهة اتصال قسم","إشعار مرفق","تحديث بيانات موظف"],
         "objects":["visitor policy","parking permit","staff directory","emergency contact list","facility access notice"],
         "objects_ar":["سياسة الزوار","تصريح موقف سيارات","دليل الموظفين","قائمة اتصال الطوارئ","إشعار دخول المرفق"]},
        {"id":"training","area":"Staff Development","senders":["Staff Development","Learning Office","Training Coordination"],"senders_ar":["تطوير الموظفين","مكتب التعلم","تنسيق التدريب"],
         "events":["training completion follow-up","certificate review","course booking change","attendance confirmation","learning record correction"],
         "events_ar":["متابعة إكمال تدريب","مراجعة شهادة","تغيير حجز دورة","تأكيد حضور","تصحيح سجل تعلم"],
         "objects":["mandatory induction","cybersecurity module","fire safety course","patient privacy training","annual competency record"],
         "objects_ar":["تهيئة إلزامية","وحدة الأمن السيبراني","دورة السلامة من الحريق","تدريب خصوصية المريض","سجل كفاءة سنوي"]},
    ],
}

V40_AR_TERMS = {
    "Medication Safety":"سلامة الدواء", "Laboratory Services":"خدمات المختبر", "Radiology":"الأشعة",
    "Infection Prevention":"مكافحة العدوى", "Patient Safety":"سلامة المرضى", "Referral Coordination":"تنسيق الإحالات",
    "Biomedical Engineering":"الهندسة الطبية", "Clinical Education":"التعليم السريري", "Human Resources":"الموارد البشرية",
    "Finance":"المالية", "Procurement":"المشتريات", "Insurance Coordination":"تنسيق التأمين",
    "Health Information Management":"إدارة المعلومات الصحية", "Internal Audit":"التدقيق الداخلي",
    "Workforce Scheduling":"جدولة القوى العاملة", "Identity and Access":"الهوية والصلاحيات",
    "Network Operations":"عمليات الشبكة", "IT Service Desk":"مكتب خدمات تقنية المعلومات",
    "Data Protection":"حماية البيانات", "Change Management":"إدارة التغيير", "Hospital Operations":"عمليات المستشفى",
    "Staff Development":"تطوير الموظفين",
}


def _v40_memory(role_type, language, phase):
    key = f"v40_history_{role_type}_{language}_{phase}"
    return st.session_state.setdefault(key, {k: [] for k in ["family","event","object","action","goal","style","combo","subject"]})


def _v40_weighted_action(mem):
    items = list(V40_ACTIONS.items())
    recent = mem["action"][-3:]
    if len(recent) >= 2 and recent[-1] == recent[-2]:
        items = [(k,w) for k,w in items if k != recent[-1]]
    total = sum(w for _,w in items); pick = _V40_RNG.uniform(0,total); acc=0
    for k,w in items:
        acc += w
        if pick <= acc: return k
    return items[-1][0]


def _v40_choose_fresh(options, history, window=8):
    recent = set(history[-window:])
    fresh = [x for x in options if x not in recent]
    return _V40_RNG.choice(fresh or options)


def _v40_plan(role, index, language, phase):
    role_type = _v30_role_type(role)
    bank = V40_SCENARIOS.get(role_type, V40_SCENARIOS["other"])
    mem = _v40_memory(role_type, language, phase)
    families = [x for x in bank if x["id"] not in set(mem["family"][-5:])] or bank
    family = _V40_RNG.choice(families)
    ar = language == "Arabic"
    ev_idx = family["events"].index(_v40_choose_fresh(family["events"], mem["event"], 14))
    ob_idx = family["objects"].index(_v40_choose_fresh(family["objects"], mem["object"], 14))
    event = family["events"][ev_idx]; obj = family["objects"][ob_idx]
    event_disp = family.get("events_ar", family["events"])[ev_idx] if ar else event
    obj_disp = family.get("objects_ar", family["objects"])[ob_idx] if ar else obj
    action = _v40_weighted_action(mem)
    goal = _v40_choose_fresh(V40_GOALS[action], mem["goal"], 6)
    style = _v40_choose_fresh(V40_STYLES, mem["style"], 7)
    combo = f"{family['id']}|{event}|{obj}|{action}|{goal}|{style}"
    for _ in range(30):
        if combo not in mem["combo"][-100:]: break
        ev_idx = _V40_RNG.randrange(len(family["events"])); ob_idx = _V40_RNG.randrange(len(family["objects"]))
        event = family["events"][ev_idx]; obj = family["objects"][ob_idx]
        event_disp = family.get("events_ar", family["events"])[ev_idx] if ar else event
        obj_disp = family.get("objects_ar", family["objects"])[ob_idx] if ar else obj
        action = _v40_weighted_action(mem); goal = _V40_RNG.choice(V40_GOALS[action]); style = _V40_RNG.choice(V40_STYLES)
        combo = f"{family['id']}|{event}|{obj}|{action}|{goal}|{style}"
    for k,v in [("family",family["id"]),("event",event),("object",obj),("action",action),("goal",goal),("style",style),("combo",combo)]:
        mem[k].append(v); mem[k] = mem[k][-120:]
    return {"role_type":role_type,"family":family,"event":event,"object":obj,"event_disp":event_disp,"object_disp":obj_disp,"action":action,"goal":goal,"style":style,"combo":combo,"phase":phase,"language":language}


def _v40_deadline(language):
    if language == "Arabic":
        return _V40_RNG.choice(["قبل نهاية دوام الغد","خلال يومي عمل","قبل اجتماع القسم القادم","بحلول ظهر الخميس","ضمن نافذة المراجعة الحالية","قبل إغلاق قائمة المتابعة"])
    return _V40_RNG.choice(["by the end of tomorrow's shift","within two business days","before the next department meeting","by Thursday noon","during the current review window","before the follow-up list closes"])


def _v40_domain():
    # Medium-level domains should visually resemble the hospital's real
    # domain family (per spec: "يبدو قريبًا من الرسمي") rather than a
    # fully generic/unrelated word, while still being clearly unofficial.
    return _V40_RNG.choice([
        "hospital-portal.net", "hospital-tasks.org", "hospital-docs.net",
        "hospital-connect.com", "hospital-workspace.org", "moh-hospital.net",
        "hospital-secure.net", "hospitalstaff-services.com", "hospital-review.org",
        "care-hospital.net", "hospital-records.org", "hospital-update.com",
        "clinical-hospital.net", "hospital-hr.org", "hospital-quality.net",
    ])


def _v40_goal_label(goal, language):
    en={"review":"Review","approve":"Approve","verify":"Verify","confirm":"Confirm","download":"Download","read":"Read","update":"Update","check":"Check","acknowledge":"Acknowledge","submit":"Submit"}
    ar={"review":"مراجعة","approve":"اعتماد","verify":"تحقق","confirm":"تأكيد","download":"تنزيل","read":"قراءة","update":"تحديث","check":"فحص","acknowledge":"إقرار","submit":"إرسال"}
    return (ar if language=="Arabic" else en)[goal]


def _v40_local_copy(plan, recipient, domain, link, attachment, evidence_phrase, workflow_note=None):
    ar = plan["language"] == "Arabic"; fam=plan["family"]; area=fam["area"]
    _sidx=_V40_RNG.randrange(len(fam["senders"]))
    sender=(fam.get("senders_ar",fam["senders"])[_sidx] if ar else fam["senders"][_sidx])
    event=plan["event_disp"]; obj=plan["object_disp"]; goal=_v40_goal_label(plan["goal"],plan["language"]); deadline=_v40_deadline(plan["language"])
    if ar:
        area_ar=V40_AR_TERMS.get(area,area)
        # Medium requires a department/role greeting (never a personal name, never fully generic).
        greetings=[f"فريق {area_ar} العزيز",f"مرحبًا فريق {area_ar}",f"زملاء {area_ar} الكرام",f"صباح الخير فريق {area_ar}",f"إلى فريق {area_ar}"]
        g=_V40_RNG.choice(greetings)
        subject=_V40_RNG.choice([f"متابعة مطلوبة: {obj}",f"تحديث {area_ar} — {obj}",f"إجراء معلق بخصوص {obj}",f"مراجعة تشغيلية: {obj}"])
        intro=_V40_RNG.choice([f"بعد مراجعة سجل {obj}، ظهر بند متعلق بـ {event}.",f"أحال فريق {area_ar} بندًا يخص {obj} للمتابعة.",f"تم تسجيل تحديث تشغيلي بخصوص {obj} ضمن {area_ar}.",f"إشارة إلى المتابعة الأخيرة، ما زال بند {obj} بحاجة إلى استكمال."])
        req=f"يرجى {goal} البند {deadline}. {evidence_phrase}"
        sign=_V40_RNG.choice([sender,f"وحدة {area_ar}","فريق التنسيق","مكتب الجودة"])
        closing=_V40_RNG.choice(["إذا لم يكن البند ضمن مسؤوليتك، تحقق من الجهة عبر الدليل الرسمي بدل الرد على هذه الرسالة.","استخدم النظام الداخلي المعتاد لأي معلومات حساسة بدل الرابط أو المرفق هنا.","سيتم تحديث سجل المتابعة داخل النظام الرسمي بعد استلام الإجراء."])
        parts = [f"{g}،", intro, req]
        if workflow_note: parts.append(workflow_note)
        parts += [closing, f"مع التحية،\n{sign}"]
        body = "\n\n".join(parts)
    else:
        greetings=[f"Dear {area} Team",f"Hello {area} Team",f"Dear {area} Colleagues",f"Attention {area} Team",f"Good morning {area} Team"]
        g=_V40_RNG.choice(greetings)
        subject=_V40_RNG.choice([f"Follow-up Required: {obj.title()}",f"{area} Update — {obj.title()}",f"Pending Action: {obj.title()}",f"Operational Review: {obj.title()}"])
        intro=_V40_RNG.choice([f"A review of the {obj} record identified an item related to {event}.",f"The {area} team has referred an item concerning {obj} for follow-up.",f"An operational update was logged for {obj} within {area}.",f"Following the latest handover, the {obj} item still requires completion."])
        req=f"Please {goal.lower()} the item {deadline}. {evidence_phrase}"
        sign=_V40_RNG.choice([sender,f"{area} Coordination","Quality Office","Operations Support"])
        closing=_V40_RNG.choice(["If this is not assigned to you, verify the owner through the official directory instead of replying to this message.","Use the usual internal system for any sensitive information instead of the link or attachment here.","The tracking record will update inside the official system after the requested action is received."])
        parts = [f"{g},", intro, req]
        if workflow_note: parts.append(workflow_note)
        parts += [closing, f"Kind regards,\n{sign}"]
        body = "\n\n".join(parts)
    return subject,body,sender,deadline


def _v40_api_copy(plan, recipient, domain, link, attachment, evidence_phrase, workflow_note=None):
    if not V40_USE_API_TEXT: return None
    ar=plan["language"]=="Arabic"; fam=plan["family"]
    second_fragment_rule = f"\nSECOND MANDATORY sentence to include verbatim, as its own separate paragraph (not next to the first fragment): {workflow_note}" if workflow_note else ""
    instruction = f"""Write ONE realistic {'Arabic' if ar else 'English'} workplace email for a Saudi hospital employee.
Return JSON only with keys subject and body.
Scenario area: {fam['area']}
Specific event: {plan['event']}
Object: {plan['object']}
Goal: {plan['goal']}
Writing style: {plan['style']}
Action type: {plan['action']}
Recipient: {recipient}
MANDATORY exact sentence/fragment to include once in body: {evidence_phrase}{second_fragment_rule}
Rules: 55-135 words; natural healthcare wording; vary opening, paragraph order and closing; no QR; no password/OTP request; do not add any URL other than the exact action fragment; do not mention phishing; no markdown except the supplied action fragment; do not invent a second action.
GREETING RULE (strict, medium difficulty): address the recipient by DEPARTMENT NAME or JOB TITLE only (e.g. "Dear {fam['area']} Team", "Hello Clinical Team") — never by the recipient's personal name and never with a fully generic greeting like "Dear Colleague" with no department reference."""
    try:
        data = call_ai(instruction, max_tokens=900)
        if not isinstance(data, dict) or "error" in data:
            try: _store_debug("v40_api_copy", f"provider error: {str(data)[:300]}")
            except Exception: pass
            return None
        text = ((data.get("choices") or [{}])[0].get("message") or {}).get("content", "")
        if not text:
            return None
        raw = text
        obj=parse_json_response(raw)
        if isinstance(obj,dict):
            subject=str(obj.get("subject","")).strip(); body=str(obj.get("body","")).strip()
            if subject and body and evidence_phrase in body and len(body.split())>=35:
                if workflow_note and workflow_note not in body:
                    return None
                # Enforce the medium-difficulty greeting rule even if the model ignored the
                # instruction: reject a personal-name greeting and fall back to local copy,
                # which is guaranteed to use a department/role greeting.
                opening = body.strip().splitlines()[0] if body.strip() else ""
                person_name = _v30_display_name(recipient)
                name_parts = [p for p in person_name.split() if len(p) > 2]
                if name_parts and any(p.lower() in opening.lower() for p in name_parts):
                    return None
                if fam["area"].lower() not in opening.lower() and not ar:
                    # English opening must reference the department; otherwise it's too generic.
                    generic_only = not any(w in opening for w in ["Team", "Colleagues", fam["area"]])
                    if generic_only:
                        return None
                return subject,body
    except Exception as e:
        try: _store_debug("v40_api_copy",str(e))
        except Exception: pass
    return None


# =============================================================
# v41 patch — scenario-derived indicators & analysis
# -------------------------------------------------------------
# Goal: indicators #2/#3(/#4) and the "why risky" / "learning tip"
# text must be DERIVED from the actual scenario data (family/area/
# event/object/action/deadline), not from a fixed template keyed
# only on action_type. This keeps every generation reproducible
# and schema-safe while making the content vary naturally with the
# scenario, matching the "Indicators -> Analysis" requirement.
# Nothing here touches Easy/Hard, UI, or any file outside this
# engine.
# =============================================================

def _v41_domain_indicator(domain, area, ar):
    if ar:
        variants = [
            (f"نطاق مرسل غير معتمد", f"النطاق {domain} ليس ضمن نطاقات المستشفى الرسمية، رغم أن الرسالة تخص {area}."),
            (f"نطاق مرسل غير معتمد", f"يستخدم المرسل {domain} بدل النطاق الرسمي، وهذا لا يتطابق مع مراسلات {area} المعتادة."),
        ]
    else:
        variants = [
            ("Unapproved sender domain", f"The sender uses {domain}, not the hospital's official domain, even though the message concerns {area}."),
            ("Unapproved sender domain", f"This message claims to be from {area}, but {domain} does not match any approved hospital domain."),
        ]
    title, desc = _V40_RNG.choice(variants)
    return _v30_indicator(1, "domain", title, desc, domain, "from")


def _v41_action_indicators(action, plan, domain, link, attachment, evidence_phrase, deadline, ar, workflow_note=None):
    """Returns the 2-3 action-specific indicator tuples (title, desc, evidence, target, key).
    IMPORTANT: within one email, no indicator's evidence may be a substring of another
    indicator's evidence (or vice versa), and no two indicators may share the same
    rendered UI element (e.g. a markdown button collapses to one highlightable node) —
    otherwise the UI can only show one badge for that whole region. workflow_note is a
    dedicated, separately-rendered sentence used precisely to avoid that collision."""
    _label_match = re.search(r'\[(.*?)\]', evidence_phrase)
    button_label = _label_match.group(1) if _label_match else evidence_phrase
    raw_area = plan["family"]["area"]
    area = V40_AR_TERMS.get(raw_area, raw_area) if ar else raw_area
    event = plan["event_disp"]; obj = plan["object_disp"]
    out = []
    if action == "pdf_attachment":
        if ar:
            out.append(("attachment","مرفق PDF غير متوقع",
                _V40_RNG.choice([f"المرفق {attachment} يخص {obj} ولم يُطلب عبر مسار {area} المعتاد.",
                                 f"يرسل هذا المرفق ملخص {obj} كملف PDF بدل مراجعته داخل النظام."]), attachment, "attachment"))
            out.append(("workflow","طلب مراجعة خارج المسار المعتاد",
                _V40_RNG.choice([f"تُحوّل الرسالة بند {obj} إلى مرفق بريدي بدل نظام {area} الداخلي.",
                                 f"مراجعة {obj} تتم عادة داخل النظام، لا عبر مرفق بريد وارد."]), evidence_phrase, "body"))
        else:
            out.append(("attachment","Unexpected PDF attachment",
                _V40_RNG.choice([f"The attachment {attachment} concerns {obj} but was not requested through the usual {area} workflow.",
                                 f"A summary of {obj} is sent as a PDF instead of being reviewed inside the system."]), attachment, "attachment"))
            out.append(("workflow","Unexpected document workflow",
                _V40_RNG.choice([f"The message moves the {obj} item into an emailed attachment instead of the normal {area} system.",
                                  f"{obj.capitalize()} items are normally reviewed inside the system, not via an emailed attachment."]), evidence_phrase, "body"))
    elif action == "reply_request":
        if ar:
            out.append(("reply_request","طلب معلومات عبر الرد",
                _V40_RNG.choice([f"تطلب الرسالة بيانات الموظف عبر الرد بخصوص {obj} بدل النظام الرسمي.",
                                  f"يُطلب تأكيد الهوية بالرد على رسالة تخص {obj}، بدل القناة المعتمدة."]),
                evidence_phrase, "body"))
            out.append(("workflow","مسار تحقق غير قياسي",
                _V40_RNG.choice([f"لا يُفترض تأكيد رقم الموظف والقسم عبر رسالة غير متوقعة بخصوص {area}.",
                                  f"مسار {area} المعتمد لا يعتمد على الرد ببيانات شخصية داخل البريد."]),
                workflow_note, "body"))
        else:
            out.append(("reply_request","Reply-based information request",
                _V40_RNG.choice([f"The email asks for employee details by reply regarding {obj}, instead of the official system.",
                                  f"Identity confirmation for {obj} is requested by reply rather than the approved channel."]),
                evidence_phrase, "body"))
            out.append(("workflow","Non-standard verification workflow",
                _V40_RNG.choice([f"Employee identity details should not be confirmed through an unexpected email about {area}.",
                                  f"The approved {area} workflow never verifies staff identity by email reply."]),
                workflow_note, "body"))
    elif action == "sharepoint":
        if ar:
            out.append(("sharepoint","إشعار مشاركة غير متوقع",
                _V40_RNG.choice([f"تدّعي الرسالة مشاركة مستند يخص {obj} دون سياق مؤكد من داخل {area}.",
                                  f"لا يوجد إشعار مسبق داخل النظام عن مشاركة مستند هذا البند."]), workflow_note or button_label, "body"))
            out.append(("link","وجهة مشاركة خارجية",
                f"الزر يقود إلى {domain} وليس Microsoft أو نطاق المستشفى.", link, "link"))
        else:
            out.append(("sharepoint","Unexpected shared-document notice",
                _V40_RNG.choice([f"The message claims a shared document about {obj} without confirmed context from {area}.",
                                  f"There is no prior in-system notice about sharing this document."]), workflow_note or button_label, "body"))
            out.append(("link","Unapproved sharing destination",
                f"The shared-document button resolves to {domain}, not an approved Microsoft or hospital domain.", link, "link"))
    elif action == "internal_workspace":
        if ar:
            out.append(("workflow","مساحة عمل غير معروفة",
                _V40_RNG.choice([f"تشير الرسالة إلى مساحة عمل عامة بخصوص {obj} بدل نظام {area} المحفوظ.",
                                  f"لا يوجد اسم نظام معروف يربط هذا البند بهذه المساحة."]), workflow_note or button_label, "body"))
            out.append(("link","زر إلى نطاق خارجي", f"وجهة مساحة العمل هي {domain}.", link, "link"))
        else:
            out.append(("workflow","Unrecognized internal workspace",
                _V40_RNG.choice([f"The email references a generic workspace for {obj} rather than a known {area} system.",
                                  f"No bookmarked hospital system is named for this task."]), workflow_note or button_label, "body"))
            out.append(("link","Workspace button leads externally", f"The workspace button resolves to the external domain {domain}.", link, "link"))
    elif action == "visible_link":
        if ar:
            out.append(("deadline","ضغط زمني معقول ظاهريًا",
                f"المهلة ({deadline}) بخصوص {obj} تبدو مهنية لكنها تشجع على التصرف قبل التحقق المستقل.", deadline, "body"))
            out.append(("link","رابط خارجي ظاهر", f"الرابط الظاهر لمتابعة {obj} يستخدم النطاق غير المعتمد {domain}.", link, "link"))
        else:
            out.append(("deadline","Plausible time pressure",
                f"The deadline ({deadline}) tied to {obj} sounds routine but encourages action before independent verification.", deadline, "body"))
            out.append(("link","Visible external link", f"The visible URL for the {obj} follow-up uses the unapproved domain {domain}.", link, "link"))
    else:  # button
        if ar:
            out.append(("workflow","إجراء عبر زر غير متوقع",
                _V40_RNG.choice([f"يحول الطلب إجراء {area} الداخلي بخصوص {obj} إلى زر داخل رسالة غير متوقعة.",
                                  f"عادة لا يُنجز هذا البند عبر زر داخل بريد وارد."]), workflow_note or button_label, "body"))
            out.append(("link","وجهة زر غير معتمدة", f"الزر يقود إلى {domain} بدل نظام {area} الرسمي.", link, "link"))
        else:
            out.append(("workflow","Unexpected button-based workflow",
                _V40_RNG.choice([f"The message turns an internal {area} process about {obj} into an unexpected emailed button.",
                                  f"This task is not normally completed through an emailed button."]), workflow_note or button_label, "body"))
            out.append(("link","Button leads to an unapproved domain", f"The button resolves to {domain}, not the official {area} system.", link, "link"))
    return out


def _v41_optional_fourth_indicator(action, plan, domain, link, sender_name, deadline, ar):
    """Adds a 4th indicator ~half the time, matching the 3-4 range for Medium.
    Uses a signal not already covered by the action's own indicators."""
    event = plan["event_disp"]
    if action == "visible_link":
        # deadline already covered; add a generic-signer indicator instead
        if ar:
            return ("sender_identity","توقيع عام بدل اسم شخص",
                    f"وقّعت الرسالة باسم جهة عامة ({sender_name}) لا اسم شخص محدد، وهو نمط شائع بالانتحال الداخلي.",
                    sender_name, "from")
        return ("sender_identity","Generic signer identity",
                f"The email is signed by a general office name ({sender_name}), not a specific person — typical of impersonated internal senders.",
                sender_name, "from")
    # otherwise add an urgency/deadline indicator tied to this event
    if ar:
        return ("deadline","مهلة زمنية مرفقة بالطلب",
                f"حددت الرسالة مهلة ({deadline}) مرتبطة بـ {event}، ما يشجع على التصرف السريع قبل التحقق.",
                deadline, "body")
    return ("deadline","Attached time pressure",
            f"The message sets a deadline ({deadline}) tied to {event}, encouraging quick action before verification.",
            deadline, "body")


def _v41_build_analysis(plan, inds, domain, ar):
    area = plan["family"]["area"]; obj = plan["object_disp"]
    goal_label = _v40_goal_label(plan["goal"], plan["language"])
    if ar:
        area_ar = V40_AR_TERMS.get(area, area)
        why_options = [
            f"ترتبط الرسالة بسياق {area_ar} واقعي حول {obj}، لكن الأدلة أعلاه — خصوصًا نطاق المرسل وطريقة طلب {goal_label} — لا تتوافق مع مسار العمل الرسمي، لذلك يجب التحقق منها عبر قناة مستقلة.",
            f"رغم أن الرسالة تبدو منطقية ضمن سيناريو {area_ar} المتعلق بـ {obj}، فإن الأدلة المحددة أعلاه (وخصوصًا {domain}) لا تطابق الطريقة المعتمدة للتعامل مع هذا النوع من الطلبات.",
            f"سياق {obj} يبدو واقعيًا، لكن المؤشرات المحددة أعلاه تُظهر أن الرسالة لم تصدر عن قناة معتمدة داخل المستشفى، لذلك يلزم التحقق المستقل قبل أي إجراء.",
        ]
        tip_options = [
            f"قبل {goal_label} أي بند يخص {obj}، افتح نظام {area_ar} الرسمي من اختصار محفوظ، أو تأكد عبر الدليل الداخلي — وليس عبر الرابط أو المرفق داخل الرسالة.",
            f"عند وصول رسالة غير متوقعة بخصوص {obj}، تحقق من المرسل والطلب عبر قناة {area_ar} الداخلية بدل الضغط أو الرد مباشرة.",
            f"تعامل بحذر مع أي طلب يخص {obj} يصل عبر البريد: راجع نظام المستشفى الداخلي أو اتصل بفريق {area_ar} مباشرة بدل استخدام رابط أو مرفق الرسالة.",
        ]
    else:
        why_options = [
            f"The email fits a realistic {area} scenario around {obj}, but the evidence above — especially the sender domain and how it asks you to {plan['goal']} — does not match the hospital's approved workflow, so it should be verified independently.",
            f"Although the message reads as a plausible {area} update about {obj}, the specific evidence above (particularly {domain}) does not align with how the hospital normally handles this kind of request.",
            f"The context around {obj} feels realistic, but the indicators above show this did not originate from an approved hospital channel, so independent verification is required before acting.",
        ]
        tip_options = [
            f"Before you {plan['goal']} anything related to {obj}, open the official {area} system from a saved bookmark, or confirm through the internal directory — never through the link or attachment in the email itself.",
            f"When an unexpected message about {obj} arrives, verify the sender and the request through the internal {area} channel instead of clicking or replying directly.",
            f"Treat any request about {obj} that arrives by email with caution: check the hospital's internal system or contact the {area} team directly instead of using the email's link or attachment.",
        ]
    return _V40_RNG.choice(why_options), _V40_RNG.choice(tip_options)


def _v40_build(role,index,language,phase):
    plan=_v40_plan(role,index,language,phase); ar=language=="Arabic"
    recipient=_v30_recipient(role,index,language,phase); domain=_v40_domain(); link=f"https://{domain}/{_V40_RNG.choice(['review','workspace','documents','task','shared'])}/{_V40_RNG.randrange(1000,9999)}"
    action=plan["action"]; attachment=""; suspicious_link=""; medium_channel="none"
    label=_v40_goal_label(plan["goal"],language)
    if action=="button": evidence_phrase=f"[{label} item]({link})" if not ar else f"[{label} البند]({link})"; suspicious_link=link; medium_channel="button"
    elif action=="visible_link": evidence_phrase=link; suspicious_link=link; medium_channel="link"
    elif action=="sharepoint": evidence_phrase=f"[Open shared document]({link})" if not ar else f"[فتح المستند المشترك]({link})"; suspicious_link=link; medium_channel="button"
    elif action=="internal_workspace": evidence_phrase=f"[{label} in workspace]({link})" if not ar else f"[{label} في مساحة العمل]({link})"; suspicious_link=link; medium_channel="button"
    elif action=="pdf_attachment":
        attachment=f"{plan['family']['id']}_{_V40_RNG.randrange(100,999)}.pdf"; evidence_phrase=(f"The supporting summary is attached as {attachment}." if not ar else f"الملخص الداعم مرفق باسم {attachment}.")
    else:
        evidence_phrase=("Reply to this email with your employee number and department." if not ar else "يرجى الرد على هذه الرسالة برقم الموظف والقسم.")

    # A second, fully separate sentence (its own paragraph) for the "workflow anomaly"
    # indicator on actions whose first evidence renders as one UI element (a button) —
    # putting two indicators' evidence inside that same element means only one badge
    # can show. Keeping this sentence on its own line guarantees both indicators are
    # independently highlightable, and it is required verbatim from both the local and
    # the API writer so it is present no matter which one produced the final body.
    _raw_area0 = plan["family"]["area"]; _area_disp0 = V40_AR_TERMS.get(_raw_area0,_raw_area0) if ar else _raw_area0
    workflow_note = None
    if action in ("button","sharepoint","internal_workspace"):
        workflow_note = (f"لم يصل هذا الطلب عبر نظام {_area_disp0} المعتمد." if ar
                          else f"This request did not come through the standard {_area_disp0} system.")
    elif action == "reply_request":
        workflow_note = ("لا يتحقق فريقنا من هوية الموظفين عبر الرد على البريد." if ar
                          else "Our team never verifies staff identity by email reply.")

    local_subject,local_body,sender_name,deadline=_v40_local_copy(plan,recipient,domain,link,attachment,evidence_phrase,workflow_note)
    api=_v40_api_copy(plan,recipient,domain,link,attachment,evidence_phrase,workflow_note)
    subject,body=api if api else (local_subject,local_body)
    mailbox=_V40_RNG.choice(["coordination","workflow","quality","operations","records","updates"])
    sender=f"{sender_name} <{mailbox}@{domain}>"

    # --- v41: scenario-derived indicators (3, sometimes 4) ---
    _raw_area = plan["family"]["area"]
    inds=[_v41_domain_indicator(domain, (V40_AR_TERMS.get(_raw_area, _raw_area) if ar else _raw_area), ar)]
    action_tuples = _v41_action_indicators(action, plan, domain, link, attachment, evidence_phrase, deadline, ar, workflow_note)
    n = 2
    for key, title, desc, evidence, target in action_tuples:
        inds.append(_v30_indicator(n, key, title, desc, evidence, target)); n += 1
    if _V40_RNG.random() < 0.5:
        key, title, desc, evidence, target = _v41_optional_fourth_indicator(action, plan, domain, link, sender_name, deadline, ar)
        # avoid duplicating evidence already used by another indicator
        if not any(str(i.get("evidence","")).strip().lower() == str(evidence).strip().lower() for i in inds):
            inds.append(_v30_indicator(n, key, title, desc, evidence, target)); n += 1

    why, tip = _v41_build_analysis(plan, inds, domain, ar)

    return {"from":sender,"to":recipient,"subject":subject,"body":_v35_compact(body),"attachment":attachment,"suspicious_link":suspicious_link,"suspicious_text":next((i.get("evidence","") for i in inds if i.get("target")=="body"),""),"indicators":inds,"why_risky":why,"learning_tip":tip,"is_phishing":True,"email_type":"Phishing","attack_type":{"button":"Look-alike workflow button","visible_link":"External review link","pdf_attachment":"Unexpected PDF lure","sharepoint":"Fake shared document","internal_workspace":"Fake internal workspace","reply_request":"Reply-based information harvesting"}[action],"risk_level":"medium","scenario_id":"v40:"+plan["combo"],"scenario_meta":{"engine":"v41","medium_channel_type":{"button":"button","visible_link":"visible_link","pdf_attachment":"pdf_attachment","sharepoint":"sharepoint_button","internal_workspace":"button","reply_request":"internal_reply"}[action],"action_type":action,"goal":plan["goal"],"style":plan["style"],"family_id":plan["family"]["id"],"event":plan["event"],"object":plan["object"]},"medium_channel":medium_channel,"display_time":_V40_RNG.choice(["Monday, 9:18 AM","Tuesday, 1:26 PM","Wednesday, 11:04 AM","Thursday, 8:47 AM","Friday, 2:12 PM","Yesterday, 3:39 PM"])}


def _v40_validate(result):
    if not isinstance(result,dict) or not result.get("is_phishing"): return False
    body=str(result.get("body",'')); link=str(result.get("suspicious_link",'') or ''); att=str(result.get("attachment",'') or ''); inds=result.get("indicators",[])
    if len(inds) not in (3,4) or "[QR" in body.upper(): return False
    if _DIRECT_PASSWORD_RE.search(body) or _DIRECT_THREAT_RE.search(body): return False
    action=str(result.get("scenario_meta",{}).get("action_type",'')); buttons=re.findall(r"\[[^\]]+\]\(https?://[^\)]+\)",body); urls=re.findall(r"https?://[^\s\)\]]+",body)
    if action in {"button","sharepoint","internal_workspace"} and (len(buttons)!=1 or not link or att): return False
    if action=="visible_link" and (buttons or not link or body.count(link)!=1 or att): return False
    if action=="pdf_attachment" and (buttons or link or not att.lower().endswith('.pdf')): return False
    if action=="reply_request" and (buttons or link or att): return False
    sources={"from":str(result.get("from",'')),"body":body,"link":link,"attachment":att,"subject":str(result.get("subject",''))}
    seen=set()
    for i,item in enumerate(inds,1):
        if int(item.get("number",i)) != i: return False
        ev=str(item.get("evidence",'')).strip(); loc=str(item.get("location",item.get("target","body")))
        if not ev or ev.lower() not in sources.get(loc,body).lower(): return False
        key=(loc,ev.lower())
        if key in seen: return False
        seen.add(key)
    return True


def _v40_generate(role,index,language,difficulty="medium",is_phishing=True,assessment=False):
    diff=str(difficulty or "medium").lower()
    if not V40_ENABLED or diff!="medium" or not is_phishing:
        return _V40_FALLBACK(role,index,language,difficulty,is_phishing,assessment)
    phase="assess" if assessment else "learn"
    for attempt in range(4):
        try:
            result=_v40_build(role,index + attempt*1009,language,phase)
            if _v40_validate(result):
                try: evaluate_and_log_auto_scores(result,"medium",language,True)
                except Exception: pass
                return result
        except Exception as e:
            try: _store_debug("v40_generation",str(e))
            except Exception: pass
    return _V40_FALLBACK(role,index,language,difficulty,is_phishing,assessment)


def generate_email(role,index,language,difficulty="medium"):
    return _v40_generate(role,index,language,difficulty,True,False)


def generate_assess_email(role,index,is_phishing,language,difficulty="medium"):
    return _v40_generate(role,index,language,difficulty,bool(is_phishing),True)

# =============================================================
# END MEDIUM GENERATION ENGINE v40
# =============================================================

# ══════════════════════════════════════════════════════════════
# SIDEBAR — زر القفل السري في الأسفل
# ══════════════════════════════════════════════════════════════
# زر القفل السري — أسفل يسار الصفحة الرئيسية فقط، صغير جداً وشفاف
st.markdown("""
<style>
#MainMenu,header,footer{visibility:hidden;}
[data-testid="stSidebar"]{display:none !important;}
</style>
""", unsafe_allow_html=True)

if st.session_state.get("page","home") == "home":
    if st.button("🔒", key="secret_admin_btn", help=""):
        st.session_state["page"] = "admin"
        st.rerun()

    # نستخدم MutationObserver لأن الزر قد يُعاد رسمه بعد تحميل الصفحة
    import streamlit.components.v1 as components
    components.html("""
    <script>
    function styleSecretLock() {
        const doc = window.parent.document;
        const btns = doc.querySelectorAll('button');
        for (const b of btns) {
            if (b.innerText.trim() === '🔒') {
                b.style.setProperty('background', 'transparent', 'important');
                b.style.setProperty('border', 'none', 'important');
                b.style.setProperty('box-shadow', 'none', 'important');
                b.style.setProperty('font-size', '0.6rem', 'important');
                b.style.setProperty('color', 'rgba(255,255,255,0.10)', 'important');
                b.style.setProperty('padding', '0px', 'important');
                b.style.setProperty('min-height', 'unset', 'important');
                b.style.setProperty('height', '16px', 'important');
                b.style.setProperty('width', '16px', 'important');
                b.style.setProperty('line-height', '1', 'important');
                const wrapper = b.closest('div[data-testid="stButton"]') || b.closest('.element-container');
                if (wrapper) {
                    wrapper.style.setProperty('position', 'fixed', 'important');
                    wrapper.style.setProperty('bottom', '4px', 'important');
                    wrapper.style.setProperty('left', '4px', 'important');
                    wrapper.style.setProperty('z-index', '99999', 'important');
                    wrapper.style.setProperty('width', '18px', 'important');
                    wrapper.style.setProperty('margin', '0', 'important');
                    wrapper.style.setProperty('padding', '0', 'important');
                    // أيضاً نضبط أي عنصر أب يحتوي على padding كبير
                    let parent = wrapper.parentElement;
                    if (parent) {
                        parent.style.setProperty('position', 'static', 'important');
                    }
                }
            }
        }
    }
    styleSecretLock();
    const observer = new MutationObserver(styleSecretLock);
    observer.observe(window.parent.document.body, {childList:true, subtree:true});
    setInterval(styleSecretLock, 500);
    </script>
    """, height=0, width=0)

# ══════════════════════════════════════════════════════════════
# MAIN ROUTING
# ══════════════════════════════════════════════════════════════
pg = st.session_state.get("page", "home")
if pg == "admin":
    page_admin()
else:
    {
        "home":page_home,"login":page_login,"learning":page_learning,
        "complete":page_complete,"assessment":page_assessment,
        "results":page_results,"report":page_report,
        "phishing_caught":page_phishing_caught
    }.get(pg, page_home)()
